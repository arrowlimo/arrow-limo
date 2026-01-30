import os
import psycopg2
import psycopg2.extras


def get_conn():
    return psycopg2.connect(
        host=os.environ.get("DB_HOST", "localhost"),
        dbname=os.environ.get("DB_NAME", "almsdata"),
        user=os.environ.get("DB_USER", "postgres"),
        password=os.environ.get("DB_PASSWORD", "***REDACTED***"),
    )


def main():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    print("Connected.")

    # Query all check/cheque transactions
    print("Fetching all checkbook entries...")
    cur.execute(
        """
        SELECT transaction_date, description, debit_amount, credit_amount, 
               account_number, bank_id, balance, source_file, vendor_extracted
        FROM banking_transactions
        WHERE description ILIKE '%cheque%' 
           OR description ILIKE '%check%'
        ORDER BY transaction_date, transaction_id
        """
    )
    rows = cur.fetchall()
    print(f"Found {len(rows)} check/cheque entries")

    # Export to Excel
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Checkbook Entries"
        
        # Headers
        headers = ["Date", "Payee/Description", "Amount", "Account", "Balance", "Source"]
        ws.append(headers)
        
        # Data rows
        for r in rows:
            amount = r['debit_amount'] or r['credit_amount']
            payee = r['description']
            ws.append([
                r['transaction_date'],
                payee,
                float(amount) if amount else 0,
                r['account_number'],
                float(r['balance']) if r['balance'] else 0,
                r['source_file']
            ])
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        out_path = "l:\\limo\\reports\\exports\\All_Checkbook_Entries.xlsx"
        wb.save(out_path)
        print(f"\nExported to: {out_path}")
        
    except ImportError:
        # Fallback to CSV
        import csv
        out_path = "l:\\limo\\reports\\exports\\All_Checkbook_Entries.csv"
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["Date", "Payee/Description", "Amount", "Account", "Balance", "Source"])
            for r in rows:
                amount = r['debit_amount'] or r['credit_amount']
                writer.writerow([
                    r['transaction_date'],
                    r['description'],
                    amount,
                    r['account_number'],
                    r['balance'],
                    r['source_file']
                ])
        print(f"\nExported to: {out_path}")
    
    conn.close()


if __name__ == "__main__":
    main()
