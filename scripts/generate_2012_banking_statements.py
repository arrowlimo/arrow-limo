"""
Generate banking statement reports for each bank account for 2012.

Creates detailed transaction reports showing:
- Opening balance
- All transactions (debits and credits)
- Running balance
- Closing balance
- Summary statistics

Exports to Excel file with multiple sheets.
"""
import psycopg2
from datetime import datetime
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_banking_statements_2012():
    """Generate banking statements for all accounts in 2012."""
    
    conn = psycopg2.connect(
        host='localhost',
        database='almsdata',
        user='postgres',
        password='***REMOVED***'
    )
    cur = conn.cursor()
    
    print("=" * 100)
    print("2012 BANKING STATEMENTS REPORT")
    print("=" * 100)
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all accounts with 2012 transactions
    cur.execute("""
        SELECT DISTINCT account_number
        FROM banking_transactions
        WHERE EXTRACT(YEAR FROM transaction_date) = 2012
        AND account_number IS NOT NULL
        ORDER BY account_number
    """)
    
    accounts = [row[0] for row in cur.fetchall()]
    
    print(f"Found {len(accounts)} accounts with 2012 transactions")
    print()
    
    # Get account descriptions
    account_names = {
        '0228362': 'CIBC Checking Account',
        '392050': 'CIBC Credit Card',
        # Add more as needed
    }
    
    for account_num in accounts:
        print("=" * 100)
        account_name = account_names.get(account_num, 'Unknown Account')
        print(f"ACCOUNT: {account_num} - {account_name}")
        print("=" * 100)
        print()
        
        # Create sheet for this account
        sheet_name = f"{account_num[:10]}"  # Excel sheet names max 31 chars
        ws = wb.create_sheet(title=sheet_name)
        
        # Get opening balance (last transaction before 2012)
        cur.execute("""
            SELECT balance
            FROM banking_transactions
            WHERE account_number = %s
            AND transaction_date < '2012-01-01'
            ORDER BY transaction_date DESC, transaction_id DESC
            LIMIT 1
        """, (account_num,))
        
        opening_row = cur.fetchone()
        opening_balance = opening_row[0] if opening_row else Decimal('0.00')
        
        print(f"Opening Balance (as of 2011-12-31): ${opening_balance:,.2f}")
        print()
        
        # Get all 2012 transactions
        cur.execute("""
            SELECT 
                transaction_id,
                transaction_date,
                description,
                debit_amount,
                credit_amount,
                balance,
                vendor_extracted,
                category
            FROM banking_transactions
            WHERE account_number = %s
            AND EXTRACT(YEAR FROM transaction_date) = 2012
            ORDER BY transaction_date, transaction_id
        """, (account_num,))
        
        transactions = cur.fetchall()
        
        if not transactions:
            print("No transactions in 2012")
            print()
            continue
        
        # Write account header to Excel
        ws.append([f"Account: {account_num} - {account_name}"])
        ws.append([f"Opening Balance (2011-12-31): ${opening_balance:,.2f}"])
        ws.append([])
        
        # Write transaction headers
        headers = ["Date", "Description", "Debit", "Credit", "Balance", "Vendor", "Category"]
        ws.append(headers)
        
        # Style header row
        header_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Print transactions
        print("TRANSACTIONS:")
        print("-" * 100)
        print(f"{'Date':<12} {'Description':<40} {'Debit':<12} {'Credit':<12} {'Balance':<12}")
        print("-" * 100)
        
        total_debits = Decimal('0.00')
        total_credits = Decimal('0.00')
        
        for txn in transactions:
            txn_id, date, desc, debit, credit, balance, vendor, category = txn
            
            debit = debit or Decimal('0.00')
            credit = credit or Decimal('0.00')
            balance = balance or Decimal('0.00')
            
            total_debits += debit
            total_credits += credit
            
            # Write to Excel
            ws.append([
                date.strftime('%Y-%m-%d'),
                desc or '',
                float(debit) if debit > 0 else None,
                float(credit) if credit > 0 else None,
                float(balance),
                vendor or '',
                category or ''
            ])
            
            # Truncate description if too long
            desc_display = desc[:37] + '...' if desc and len(desc) > 40 else (desc or '')
            
            debit_str = f"${debit:,.2f}" if debit > 0 else ""
            credit_str = f"${credit:,.2f}" if credit > 0 else ""
            
            print(f"{date.strftime('%Y-%m-%d'):<12} {desc_display:<40} {debit_str:<12} {credit_str:<12} ${balance:,.2f}")
        
        print("-" * 100)
        
        # Closing balance (last transaction of 2012)
        closing_balance = transactions[-1][5] or Decimal('0.00')
        
        # Write summary to Excel
        ws.append([])
        ws.append(["SUMMARY"])
        ws.append(["Total Debits:", f"${total_debits:,.2f}"])
        ws.append(["Total Credits:", f"${total_credits:,.2f}"])
        ws.append(["Net Change:", f"${(total_credits - total_debits):,.2f}"])
        ws.append(["Closing Balance:", f"${closing_balance:,.2f}"])
        ws.append(["Transaction Count:", len(transactions)])
        
        # Auto-size columns
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 15
        ws.column_dimensions['B'].width = 50  # Description column wider
        
        # Summary statistics
        print()
        print("SUMMARY:")
        print(f"  Total Debits:        ${total_debits:>15,.2f}")
        print(f"  Total Credits:       ${total_credits:>15,.2f}")
        print(f"  Net Change:          ${(total_credits - total_debits):>15,.2f}")
        print(f"  Closing Balance:     ${closing_balance:>15,.2f}")
        print(f"  Transaction Count:   {len(transactions):>15,}")
        print()
        
        # Monthly breakdown
        print("MONTHLY BREAKDOWN:")
        print("-" * 100)
        print(f"{'Month':<15} {'Debits':<15} {'Credits':<15} {'Net':<15} {'Txn Count':<12}")
        print("-" * 100)
        
        cur.execute("""
            SELECT 
                TO_CHAR(transaction_date, 'YYYY-MM') as month,
                SUM(COALESCE(debit_amount, 0)) as total_debits,
                SUM(COALESCE(credit_amount, 0)) as total_credits,
                COUNT(*) as txn_count
            FROM banking_transactions
            WHERE account_number = %s
            AND EXTRACT(YEAR FROM transaction_date) = 2012
            GROUP BY TO_CHAR(transaction_date, 'YYYY-MM')
            ORDER BY month
        """, (account_num,))
        
        monthly = cur.fetchall()
        for month_data in monthly:
            month, debits, credits, count = month_data
            net = credits - debits
            print(f"{month:<15} ${debits:>12,.2f} ${credits:>12,.2f} ${net:>12,.2f} {count:>11,}")
        
        print("-" * 100)
        print()
        
        # Category breakdown (if available)
        cur.execute("""
            SELECT 
                COALESCE(category, 'Uncategorized') as category,
                COUNT(*) as txn_count,
                SUM(COALESCE(debit_amount, 0)) as total_debits,
                SUM(COALESCE(credit_amount, 0)) as total_credits
            FROM banking_transactions
            WHERE account_number = %s
            AND EXTRACT(YEAR FROM transaction_date) = 2012
            GROUP BY category
            ORDER BY (SUM(COALESCE(debit_amount, 0)) + SUM(COALESCE(credit_amount, 0))) DESC
        """, (account_num,))
        
        categories = cur.fetchall()
        if categories:
            print("CATEGORY BREAKDOWN:")
            print("-" * 100)
            print(f"{'Category':<30} {'Txn Count':<12} {'Debits':<15} {'Credits':<15}")
            print("-" * 100)
            
            for cat_data in categories:
                category, count, debits, credits = cat_data
                category = category or 'Uncategorized'
                category_display = category[:27] + '...' if len(category) > 30 else category
                print(f"{category_display:<30} {count:>11,} ${debits:>12,.2f} ${credits:>12,.2f}")
            
            print("-" * 100)
            print()
        
        # Top vendors (if available)
        cur.execute("""
            SELECT 
                COALESCE(vendor_extracted, 'Unknown') as vendor,
                COUNT(*) as txn_count,
                SUM(COALESCE(debit_amount, 0)) as total_debits,
                SUM(COALESCE(credit_amount, 0)) as total_credits
            FROM banking_transactions
            WHERE account_number = %s
            AND EXTRACT(YEAR FROM transaction_date) = 2012
            AND vendor_extracted IS NOT NULL
            GROUP BY vendor_extracted
            ORDER BY (SUM(COALESCE(debit_amount, 0)) + SUM(COALESCE(credit_amount, 0))) DESC
            LIMIT 20
        """, (account_num,))
        
        vendors = cur.fetchall()
        if vendors:
            print("TOP 20 VENDORS:")
            print("-" * 100)
            print(f"{'Vendor':<40} {'Txn Count':<12} {'Debits':<15} {'Credits':<15}")
            print("-" * 100)
            
            for vendor_data in vendors:
                vendor, count, debits, credits = vendor_data
                vendor_display = vendor[:37] + '...' if len(vendor) > 40 else vendor
                print(f"{vendor_display:<40} {count:>11,} ${debits:>12,.2f} ${credits:>12,.2f}")
            
            print("-" * 100)
            print()
        
        print()
        print()
    
    # Overall summary across all accounts
    print("=" * 100)
    print("2012 OVERALL SUMMARY - ALL ACCOUNTS")
    print("=" * 100)
    print()
    
    cur.execute("""
        SELECT 
            account_number,
            COUNT(*) as txn_count,
            SUM(COALESCE(debit_amount, 0)) as total_debits,
            SUM(COALESCE(credit_amount, 0)) as total_credits
        FROM banking_transactions
        WHERE EXTRACT(YEAR FROM transaction_date) = 2012
        AND account_number IS NOT NULL
        GROUP BY account_number
        ORDER BY account_number
    """)
    
    # Create summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.append(["2012 BANKING STATEMENTS SUMMARY"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    ws_summary.append(["Account", "Account Name", "Txn Count", "Debits", "Credits", "Net"])
    
    # Style summary header
    for col in range(1, 7):
        cell = ws_summary.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    print(f"{'Account':<15} {'Account Name':<30} {'Txn Count':<12} {'Debits':<15} {'Credits':<15} {'Net':<15}")
    print("-" * 100)
    
    grand_total_debits = Decimal('0.00')
    grand_total_credits = Decimal('0.00')
    grand_total_txns = 0
    
    for row in cur.fetchall():
        account, count, debits, credits = row
        net = credits - debits
        account_name = account_names.get(account, 'Unknown')[:27]
        
        grand_total_debits += debits
        grand_total_credits += credits
        grand_total_txns += count
        
        ws_summary.append([
            account,
            account_name,
            count,
            float(debits),
            float(credits),
            float(net)
        ])
        
        print(f"{account:<15} {account_name:<30} {count:>11,} ${debits:>12,.2f} ${credits:>12,.2f} ${net:>12,.2f}")
    
    # Add totals row
    ws_summary.append([
        "TOTAL",
        "",
        grand_total_txns,
        float(grand_total_debits),
        float(grand_total_credits),
        float(grand_total_credits - grand_total_debits)
    ])
    
    # Auto-size summary columns
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 15
    
    print("-" * 100)
    print(f"{'TOTAL':<15} {'':<30} {grand_total_txns:>11,} ${grand_total_debits:>12,.2f} ${grand_total_credits:>12,.2f} ${(grand_total_credits - grand_total_debits):>12,.2f}")
    print()
    
    cur.close()
    conn.close()
    
    # Save Excel file
    output_file = 'reports/2012_Banking_Statements.xlsx'
    wb.save(output_file)
    
    print("=" * 100)
    print("REPORT COMPLETE")
    print("=" * 100)
    print(f"\nExcel file saved: {output_file}")
    print(f"Sheets created: {len(wb.sheetnames)}")

if __name__ == '__main__':
    generate_banking_statements_2012()
