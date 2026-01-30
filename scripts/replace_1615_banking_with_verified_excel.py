import os
import sys
import datetime as dt
from typing import List, Dict, Optional, Tuple

import psycopg2
import psycopg2.extras

VERIFIED_XLSX = r"l:\\limo\\reports\\exports\\CIBC_7461615_2012_2017_VERIFIED.xlsx"
ACCOUNT_NUMBER = "1615"  # use FK-recognized account key
BANK_ID = 4
DATE_START = dt.date(2012, 1, 1)
DATE_END = dt.date(2017, 12, 31)


def get_conn():
    return psycopg2.connect(
        host=os.environ.get("DB_HOST", "localhost"),
        dbname=os.environ.get("DB_NAME", "almsdata"),
        user=os.environ.get("DB_USER", "postgres"),
        password=os.environ.get("DB_PASSWORD", "***REDACTED***"),
    )


def load_verified_rows(path: str) -> List[Dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        print("openpyxl is required to read the verified Excel.")
        raise

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    # Flexible header mapping
    header_map = {}
    for i, h in enumerate(headers):
        key = h.lower()
        if key in ("date", "transaction_date"):
            header_map["date"] = i
        elif key in ("description", "payee/description", "payee", "memo"):
            header_map["description"] = i
        elif key in ("debit", "debit_amount", "withdrawal"):
            header_map["debit"] = i
        elif key in ("credit", "credit_amount", "deposit"):
            header_map["credit"] = i
        elif key in ("amount", "net amount"):
            header_map["amount"] = i
        elif key in ("balance", "running balance", "expected balance", "stored balance"):
            header_map["balance"] = i

    required_any = ["date", "description"]
    for k in required_any:
        if k not in header_map:
            raise ValueError(f"Verified file missing required column: {k} (headers={headers})")

    rows: List[Dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        def get(col: str):
            idx = header_map.get(col)
            return r[idx] if idx is not None else None

        raw_date = get("date")
        # Normalize date
        if isinstance(raw_date, dt.datetime):
            tx_date = raw_date.date()
        elif isinstance(raw_date, dt.date):
            tx_date = raw_date
        elif raw_date is None:
            continue
        else:
            # Try parse from string
            from datetime import datetime
            try:
                tx_date = datetime.strptime(str(raw_date).strip(), "%Y-%m-%d").date()
            except Exception:
                try:
                    tx_date = datetime.strptime(str(raw_date).strip(), "%m/%d/%Y").date()
                except Exception:
                    raise ValueError(f"Unparseable date value: {raw_date}")

        desc = get("description")
        debit = get("debit")
        credit = get("credit")
        amount = get("amount")
        balance = get("balance")

        # Normalize numeric values
        def to_float(x):
            if x is None or x == "":
                return None
            try:
                return float(x)
            except Exception:
                return float(str(x).replace(",", ""))

        debit_f = to_float(debit)
        credit_f = to_float(credit)
        amount_f = to_float(amount)
        balance_f = to_float(balance)

        # If only amount is present, split into debit/credit by sign
        if debit_f is None and credit_f is None and amount_f is not None:
            if amount_f >= 0:
                debit_f = amount_f
                credit_f = None
            else:
                debit_f = None
                credit_f = abs(amount_f)

        rows.append({
            "transaction_date": tx_date,
            "description": desc if desc is not None else "",
            "debit_amount": debit_f,
            "credit_amount": credit_f,
            "balance": balance_f,
        })

    # Sort by date to ensure consistent running order
    rows.sort(key=lambda x: (x["transaction_date"], x["description"]))
    return rows


def backup_existing(cur) -> str:
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_table = f"banking_transactions_1615_backup_{ts}"
    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {backup_table} AS
        SELECT * FROM banking_transactions
        WHERE bank_id = %s
          AND account_number IN ('1615', '74-61615')
          AND transaction_date BETWEEN %s AND %s
        """,
        (BANK_ID, DATE_START, DATE_END),
    )
    return backup_table


def delete_existing(cur) -> int:
    cur.execute(
        """
        DELETE FROM banking_transactions
        WHERE bank_id = %s
          AND account_number IN ('1615', '74-61615')
          AND transaction_date BETWEEN %s AND %s
        """,
        (BANK_ID, DATE_START, DATE_END),
    )
    return cur.rowcount


def insert_verified(cur, rows: List[Dict]) -> int:
    inserted = 0
    for row in rows:
        cur.execute(
            """
            INSERT INTO banking_transactions (
                transaction_date,
                description,
                debit_amount,
                credit_amount,
                balance,
                account_number,
                bank_id,
                source_file
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                row["transaction_date"],
                row["description"],
                row["debit_amount"],
                row["credit_amount"],
                row["balance"],
                ACCOUNT_NUMBER,
                BANK_ID,
                os.path.basename(VERIFIED_XLSX),
            ),
        )
        inserted += 1
    return inserted


def recompute_and_verify(cur) -> Tuple[int, dt.date, dt.date]:
    cur.execute(
        """
        SELECT COUNT(*), MIN(transaction_date), MAX(transaction_date)
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
          AND transaction_date BETWEEN %s AND %s
        """,
        (BANK_ID, ACCOUNT_NUMBER, DATE_START, DATE_END),
    )
    count, min_date, max_date = cur.fetchone()
    return count, min_date, max_date


def main():
    write = True
    if "--dry-run" in sys.argv:
        write = False
    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    print(f"Loading verified rows from: {VERIFIED_XLSX}")
    rows = load_verified_rows(VERIFIED_XLSX)
    print(f"Loaded {len(rows)} rows from verified Excel.")

    # Validate dates
    if not rows:
        print("No rows found in verified Excel; aborting.")
        conn.close()
        return
    if rows[0]["transaction_date"] < DATE_START or rows[-1]["transaction_date"] > DATE_END:
        print("Warning: Verified rows have dates outside 2012–2017. Proceeding anyway.")

    try:
        backup_table = backup_existing(cur)
        print(f"Backup complete: {backup_table}")

        deleted = delete_existing(cur)
        print(f"Deleted {deleted} existing 1615 rows (2012–2017).")

        inserted = insert_verified(cur, rows)
        print(f"Inserted {inserted} verified rows.")

        if write:
            conn.commit()
            print("Committed replacement.")
        else:
            conn.rollback()
            print("Dry-run complete; rolled back.")

        count, min_date, max_date = recompute_and_verify(cur)
        print(f"Post-replacement: {count} rows, date range {min_date} → {max_date}.")

    except Exception as e:
        conn.rollback()
        print(f"Error; rolled back: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
