#!/usr/bin/env python3
"""Sort Scotia file by amount and find $116.00 entry, then sort by date."""

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

# Load the Scotia file
xlsx_path = r"l:\limo\data\2012_scotia_transactions_for_editing.xlsx"

print("=" * 80)
print("SCOTIA FILE ANALYSIS & SORTING")
print("=" * 80)

# Load with pandas to understand structure
df = pd.read_excel(xlsx_path, sheet_name=0)
print(f"\nColumns: {list(df.columns)}")
print(f"Shape: {df.shape}")
print(f"\nFirst 3 rows:")
print(df.head(3))

# Find amount column
amount_col = None
for col in df.columns:
    if any(x in col.lower() for x in ['amount', 'balance', 'debit', 'credit']):
        amount_col = col
        break

if not amount_col:
    # Try numeric columns
    numeric_cols = df.select_dtypes(include=['number']).columns
    if len(numeric_cols) > 0:
        amount_col = numeric_cols[0]

print(f"\n" + "=" * 80)
print(f"SEARCHING FOR $116.00 (using column: {amount_col})")
print("=" * 80)

if amount_col:
    # Search for 116.00
    matches = df[pd.to_numeric(df[amount_col], errors='coerce') == 116.0]
    
    if len(matches) > 0:
        print(f"\n✓ Found {len(matches)} match(es):")
        for idx, row in matches.iterrows():
            print(f"\n  Row {idx + 2}:")
            for col, val in row.items():
                print(f"    {col}: {val}")
    else:
        print(f"\n✗ No exact match for 116.00 found")
        print(f"\nSearching for values close to 116...")
        numeric_col = pd.to_numeric(df[amount_col], errors='coerce')
        close_matches = df[(numeric_col >= 115) & (numeric_col <= 117)]
        if len(close_matches) > 0:
            print(f"Found {len(close_matches)} entries between $115-$117:")
            for idx, row in close_matches.iterrows():
                print(f"\n  Row {idx + 2}:")
                for col, val in row.items():
                    print(f"    {col}: {val}")

# Now sort by amount and export
print(f"\n" + "=" * 80)
print("SORTING BY AMOUNT AND SAVING")
print("=" * 80)

# Load workbook to sort it
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# Get all data with headers
data = []
headers = [cell.value for cell in ws[1]]
print(f"\nHeaders: {headers}")

# Collect data rows
for row in ws.iter_rows(min_row=2, values_only=False):
    row_data = [cell.value for cell in row]
    data.append(row_data)

print(f"Data rows: {len(data)}")

# Find amount column index
amount_col_idx = None
for i, header in enumerate(headers):
    if any(x in str(header).lower() for x in ['amount', 'balance', 'debit', 'credit']):
        amount_col_idx = i
        break

if amount_col_idx is not None:
    print(f"Amount column index: {amount_col_idx} ({headers[amount_col_idx]})")
    
    # Sort by amount (column at amount_col_idx)
    data_sorted = sorted(data, key=lambda x: (x[amount_col_idx] is None, pd.to_numeric(pd.Series(x[amount_col_idx]), errors='coerce')))
    
    # Clear worksheet and rewrite
    ws.delete_rows(2, ws.max_row)
    for row_data in data_sorted:
        ws.append(row_data)
    
    # Save sorted file
    output_path = r"l:\limo\data\2012_scotia_transactions_for_editing_SORTED.xlsx"
    wb.save(output_path)
    print(f"\n✓ Saved sorted file: {output_path}")
    
    # Show entries around $116
    print(f"\nEntries around $116.00 (sorted by amount):")
    for i, row_data in enumerate(data_sorted):
        try:
            amt = float(row_data[amount_col_idx]) if row_data[amount_col_idx] else 0
            if 110 <= amt <= 120:
                print(f"\n  Index {i}: Amount=${amt:.2f}")
                for j, val in enumerate(row_data):
                    print(f"    {headers[j]}: {val}")
        except:
            pass

wb.close()

PYSCRIPT
