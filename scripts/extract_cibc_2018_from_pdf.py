"""
Extract CIBC 8362 2018 transactions from PDF
"""
import pdfplumber
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

pdf_path = Path(r"L:\limo\pdf\2018\CIBC 8362 2018.pdf")
output_path = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

OPENING_BALANCE = Decimal('1058.53')

print("EXTRACTING CIBC 8362 2018 FROM PDF")
print("=" * 80)

if not pdf_path.exists():
    print(f"❌ PDF not found: {pdf_path}")
    exit(1)

trans = []

with pdfplumber.open(pdf_path) as pdf:
    for page_num, page in enumerate(pdf.pages, 1):
        text = page.extract_text()
        if not text:
            continue
        
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            # Look for date pattern at start: Jan 1, Feb 2, etc
            parts = line.split()
            if len(parts) < 2:
                continue
            
            # Try to parse date (format: "Jan 1", "Feb 2", etc.)
            try:
                # Check if first two parts look like "Jan 1"
                month_str = parts[0]
                day_str = parts[1]
                
                # Month mapping
                months = {
                    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                }
                
                if month_str not in months:
                    continue
                
                day = int(day_str)
                month = months[month_str]
                date_obj = datetime(2018, month, day)
                date_str = date_obj.strftime('%Y-%m-%d')
                
                # Rest is description + amounts + balance
                # Format: Jan 1 DESCRIPTION... DEBIT CREDIT BALANCE
                remaining = ' '.join(parts[2:])
                
                # Find last 1-3 numbers (amounts and balance)
                numbers = []
                desc_parts = []
                
                for p in parts[2:]:
                    # Check if looks like number (digits, commas, periods)
                    clean_p = p.replace(',', '').replace('$', '')
                    try:
                        Decimal(clean_p)
                        numbers.append(clean_p)
                    except:
                        if not numbers:  # Still building description
                            desc_parts.append(p)
                
                if len(numbers) < 1:
                    continue
                
                desc = ' '.join(desc_parts)
                
                # Parse based on number count
                # 1 number = balance only (opening)
                # 2 numbers = amount + balance
                # 3 numbers = debit + credit + balance (or debit + amount + balance)
                
                balance = Decimal(numbers[-1])
                
                if len(numbers) == 1:
                    # Opening balance line
                    debit = Decimal('0')
                    credit = Decimal('0')
                elif len(numbers) == 2:
                    # One amount + balance
                    amt = Decimal(numbers[0])
                    # Guess: if balance increased, it's credit; if decreased, debit
                    # (We'd need previous balance to know for sure)
                    debit = amt if 'WITHDRAWAL' in desc.upper() or 'PURCHASE' in desc.upper() else Decimal('0')
                    credit = amt if debit == 0 else Decimal('0')
                elif len(numbers) >= 3:
                    # Debit, Credit, Balance
                    debit = Decimal(numbers[0]) if numbers[0] else Decimal('0')
                    credit = Decimal(numbers[1]) if numbers[1] else Decimal('0')
                
                trans.append({
                    'date': date_str,
                    'date_obj': date_obj,
                    'desc': desc.strip(),
                    'debit': debit,
                    'credit': credit,
                    'balance': balance
                })
                
            except (ValueError, IndexError):
                continue

print(f"Extracted {len(trans)} transactions from PDF\n")

if not trans:
    print("❌ No transactions found - trying alternative parsing...")
    exit(1)

# Sort chronologically
trans.sort(key=lambda x: x['date_obj'])

# Recalculate running balance from opening
print("RECALCULATING RUNNING BALANCE...")
print("=" * 80)
print(f"Opening balance: ${OPENING_BALANCE:,.2f}\n")

running = OPENING_BALANCE
for t in trans:
    # Recalculate balance
    running = running - t['debit'] + t['credit']
    t['balance'] = running

total_cr = sum(t['credit'] for t in trans)
total_db = sum(t['debit'] for t in trans)

print(f"Total Credits: ${total_cr:,.2f}")
print(f"Total Debits: ${total_db:,.2f}")
print(f"Net Change: ${total_cr - total_db:,.2f}")
print(f"Closing: ${running:,.2f}")

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Transactions"

headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
ws.append(headers)

fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
font = Font(bold=True, color="FFFFFF")
for c in range(1, 6):
    ws.cell(1, c).fill = fill
    ws.cell(1, c).font = font

for t in trans:
    ws.append([
        t['date'], t['desc'],
        float(t['debit']), float(t['credit']),
        float(t['balance'])
    ])

ws.column_dimensions['B'].width = 50

for c in [3, 4, 5]:
    for r in range(2, len(trans) + 2):
        ws.cell(r, c).number_format = '$#,##0.00'

# Summary
s = wb.create_sheet("Summary")
s['A1'] = "CIBC 8362 2018"
s['A1'].font = Font(bold=True, size=14)
s['A3'] = "Opening:"
s['B3'] = float(OPENING_BALANCE)
s['B3'].number_format = '$#,##0.00'
s['A4'] = "Transactions:"
s['B4'] = len(trans)
s['A5'] = "Credits:"
s['B5'] = float(total_cr)
s['B5'].number_format = '$#,##0.00'
s['A6'] = "Debits:"
s['B6'] = float(total_db)
s['B6'].number_format = '$#,##0.00'
s['A7'] = "Closing:"
s['B7'] = float(running)
s['B7'].number_format = '$#,##0.00'

wb.save(output_path)
print(f"\n✅ {output_path}")
print(f"   {len(trans)} transactions")
