import os
from datetime import date, timedelta
from typing import List, Dict, Tuple

import psycopg2
import psycopg2.extras

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", os.environ.get("DB_PASSWORD"))

DATE_START = date(2012, 1, 1)
DATE_END = date(2017, 12, 31)

ACC_1615 = {"bank_id": 4, "account_number": "1615"}
ACC_8362 = {"bank_id": 1, "account_number": "0228362"}

KEYWORDS = ["transfer", "xfer", "x-fer", "xfr", "moving", "to 0228362", "to 1615", "from 0228362", "from 1615"]


def get_conn():
    return psycopg2.connect(host=DB_HOST, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)


def fetch_tx(cur, bank_id: int, account_number: str) -> List[Dict]:
    cur.execute(
        """
        SELECT transaction_id, transaction_date, description,
               debit_amount, credit_amount, balance, source_file
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
          AND transaction_date BETWEEN %s AND %s
        ORDER BY transaction_date, transaction_id
        """,
        (bank_id, account_number, DATE_START, DATE_END),
    )
    return [dict(r) for r in cur.fetchall()]


def amt_out(row: Dict) -> float:
    d = row.get("debit_amount")
    return float(d) if d is not None else 0.0


def amt_in(row: Dict) -> float:
    c = row.get("credit_amount")
    return float(c) if c is not None else 0.0


def has_kw(desc: str) -> bool:
    if not desc:
        return False
    low = desc.lower()
    return any(k in low for k in KEYWORDS)


def match_transfers(out_list: List[Dict], in_list: List[Dict], day_window: int = 3) -> Tuple[List[Tuple[Dict, Dict, int, int]], List[Dict], List[Dict]]:
    # Build amount-bucket for in_list by amount (2 decimals)
    from collections import defaultdict
    bucket = defaultdict(list)
    for r in in_list:
        a = round(amt_in(r), 2)
        if a > 0:
            bucket[a].append(r)

    matched = []
    used_in_ids = set()

    for o in out_list:
        a = round(amt_out(o), 2)
        if a <= 0:
            continue
        candidates = [r for r in bucket.get(a, []) if r["transaction_id"] not in used_in_ids]
        if not candidates:
            continue
        # choose by min date diff, then by keyword presence, then by closest transaction_id
        best = None
        best_score = None
        for cand in candidates:
            dd = abs((cand["transaction_date"] - o["transaction_date"]).days)
            if dd > day_window:
                continue
            score = (dd, -(1 if has_kw(o.get("description")) else 0) - (1 if has_kw(cand.get("description")) else 0), abs(cand["transaction_id"] - o["transaction_id"]))
            if best_score is None or score < best_score:
                best = cand
                best_score = score
        if best is not None:
            used_in_ids.add(best["transaction_id"])
            # score breakdown
            dd = abs((best["transaction_date"] - o["transaction_date"]).days)
            kw_score = (1 if has_kw(o.get("description")) else 0) + (1 if has_kw(best.get("description")) else 0)
            matched.append((o, best, dd, kw_score))

    unmatched_out = [o for o in out_list if round(amt_out(o), 2) > 0 and all(o["transaction_id"] != m[0]["transaction_id"] for m in matched)]
    unmatched_in = [i for i in in_list if round(amt_in(i), 2) > 0 and i["transaction_id"] not in used_in_ids]
    return matched, unmatched_out, unmatched_in


def export_excel(pairs: List[Tuple[Dict, Dict, int, int]], out_path: str):
    if Workbook is None:
        print("openpyxl not available; skipping Excel export")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Transfers"
    headers = [
        "From_Account", "From_TxID", "From_Date", "From_Desc", "Out_Amount",
        "To_Account", "To_TxID", "To_Date", "To_Desc", "In_Amount",
        "Date_Diff_Days", "Keyword_Score"
    ]
    ws.append(headers)
    for o, i, dd, kws in pairs:
        ws.append([
            "1615" if o.get("debit_amount") else "0228362",
            o["transaction_id"], o["transaction_date"], o.get("description"), round(amt_out(o), 2),
            "0228362" if i.get("credit_amount") else "1615",
            i["transaction_id"], i["transaction_date"], i.get("description"), round(amt_in(i), 2),
            dd, kws
        ])
    # Autosize columns
    from openpyxl.utils import get_column_letter
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


def main():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Fetch transactions
    t1615 = fetch_tx(cur, ACC_1615["bank_id"], ACC_1615["account_number"])
    t8362 = fetch_tx(cur, ACC_8362["bank_id"], ACC_8362["account_number"])

    # 1615 -> 8362 (out from 1615 debit, into 8362 credit)
    out_1615 = [r for r in t1615 if r["debit_amount"] is not None]
    in_8362 = [r for r in t8362 if r["credit_amount"] is not None]
    pairs_1615_to_8362, uo_1615, ui_8362 = match_transfers(out_1615, in_8362)

    # 8362 -> 1615 (out from 8362 debit, into 1615 credit)
    out_8362 = [r for r in t8362 if r["debit_amount"] is not None]
    in_1615 = [r for r in t1615 if r["credit_amount"] is not None]
    pairs_8362_to_1615, uo_8362, ui_1615 = match_transfers(out_8362, in_1615)

    total_pairs = len(pairs_1615_to_8362) + len(pairs_8362_to_1615)
    print(f"Matched transfers: 1615→8362={len(pairs_1615_to_8362)}, 8362→1615={len(pairs_8362_to_1615)}, total={total_pairs}")
    print(f"Unmatched candidates: out_1615={len(uo_1615)}, in_8362={len(ui_8362)}, out_8362={len(uo_8362)}, in_1615={len(ui_1615)}")

    # Export
    out_dir = r"l:\\limo\\reports\\exports"
    os.makedirs(out_dir, exist_ok=True)
    export_excel(pairs_1615_to_8362, os.path.join(out_dir, "Transfers_1615_to_0228362_2012_2017.xlsx"))
    export_excel(pairs_8362_to_1615, os.path.join(out_dir, "Transfers_0228362_to_1615_2012_2017.xlsx"))

    # CSV fallback
    import csv
    with open(os.path.join(out_dir, "Transfers_1615_to_0228362_2012_2017.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["From_Account","From_TxID","From_Date","From_Desc","Out_Amount","To_Account","To_TxID","To_Date","To_Desc","In_Amount","Date_Diff_Days","Keyword_Score"])
        for o,i,dd,kws in pairs_1615_to_8362:
            w.writerow(["1615", o["transaction_id"], o["transaction_date"], o.get("description"), round(amt_out(o),2), "0228362", i["transaction_id"], i["transaction_date"], i.get("description"), round(amt_in(i),2), dd, kws])
    with open(os.path.join(out_dir, "Transfers_0228362_to_1615_2012_2017.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["From_Account","From_TxID","From_Date","From_Desc","Out_Amount","To_Account","To_TxID","To_Date","To_Desc","In_Amount","Date_Diff_Days","Keyword_Score"])
        for o,i,dd,kws in pairs_8362_to_1615:
            w.writerow(["0228362", o["transaction_id"], o["transaction_date"], o.get("description"), round(amt_out(o),2), "1615", i["transaction_id"], i["transaction_date"], i.get("description"), round(amt_in(i),2), dd, kws])

    conn.close()


if __name__ == "__main__":
    main()
