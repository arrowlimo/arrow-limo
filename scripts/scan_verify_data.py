#!/usr/bin/env python3
"""
Scan the folder 'L:\\limo\\CIBC UPLOADS\\verify this data' recursively and build an index of files:
- path, name, extension, size, modified time, sha256, access_ok
- basic content extraction for txt/csv/eml
- write two CSVs under reports/: verify_index.csv and verify_errors.csv

Usage:
  python scripts/scan_verify_data.py --index-only   # fast index without content read
  python scripts/scan_verify_data.py --with-content # includes basic extraction for text-like files

Notes:
- This script will NOT move or modify files.
- For .eml, it extracts headers (From, To, Date, Subject) and plain text body if available.
- For .csv/.txt, it stores a short head sample only.
"""
import os
import csv
import sys
import hashlib
from datetime import datetime
from email import policy
from email.parser import BytesParser

BASE_DIR = r"L:\\limo\\CIBC UPLOADS\\verify this data"
REPORTS_DIR = r"L:\\limo\\reports"

os.makedirs(REPORTS_DIR, exist_ok=True)

INDEX_CSV = os.path.join(REPORTS_DIR, 'verify_index.csv')
ERRORS_CSV = os.path.join(REPORTS_DIR, 'verify_errors.csv')
CONTENT_CSV = os.path.join(REPORTS_DIR, 'verify_content_samples.csv')

SAFE_TEXT_EXT = {'.txt', '.csv', '.tsv', '.log'}
EMAIL_EXT = {'.eml'}
PDF_EXT = {'.pdf'}
DOCX_EXT = {'.docx'}
XLSX_EXT = {'.xlsx'}
MSG_EXT = {'.msg'}

HEAD_BYTES = 20000  # read at most 20KB for sampling


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def walk_files(base: str):
    for root, dirs, files in os.walk(base):
        for name in files:
            yield os.path.join(root, name)


def safe_stat(path: str):
    try:
        st = os.stat(path)
        return dict(size=st.st_size, mtime=int(st.st_mtime), ok=True, err=None)
    except Exception as e:
        return dict(size=None, mtime=None, ok=False, err=str(e))


def read_text_head(path: str) -> str:
    try:
        with open(path, 'rb') as f:
            data = f.read(HEAD_BYTES)
        try:
            return data.decode('utf-8', errors='replace')
        except Exception:
            return data.decode('latin-1', errors='replace')
    except Exception as e:
        return f"<error reading text head: {e}>"


def read_eml_head(path: str) -> str:
    try:
        with open(path, 'rb') as f:
            msg = BytesParser(policy=policy.default).parse(f)
        headers = {
            'from': str(msg.get('From', '')),
            'to': str(msg.get('To', '')),
            'date': str(msg.get('Date', '')),
            'subject': str(msg.get('Subject', '')),
        }
        # Extract plain text body sample
        body = ''
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == 'text/plain':
                    body = part.get_content()
                    break
        else:
            if msg.get_content_type() == 'text/plain':
                body = msg.get_content()
        body_sample = (body or '')
        if len(body_sample) > 2000:
            body_sample = body_sample[:2000] + '...'
        return f"headers={headers}; body_sample={body_sample}"
    except Exception as e:
        return f"<error reading eml: {e}>"

def read_pdf_text(path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        texts = []
        for i, page in enumerate(reader.pages[:10]):  # limit 10 pages
            try:
                texts.append(page.extract_text() or '')
            except Exception:
                pass
        sample = "\n".join(texts)
        return sample[:20000]
    except Exception as e:
        return f"<error reading pdf: {e}>"

def read_docx_text(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        texts = [p.text for p in doc.paragraphs]
        sample = "\n".join(texts)
        return sample[:20000]
    except Exception as e:
        return f"<error reading docx: {e}>"

def read_xlsx_text(path: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        texts = []
        for ws in wb.worksheets[:5]:  # first 5 sheets
            rows = []
            for r, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
                rows.append(",".join('' if v is None else str(v) for v in row))
                if r >= 50:
                    break
            texts.append("\n".join(rows))
        sample = "\n\n".join(texts)
        return sample[:20000]
    except Exception as e:
        return f"<error reading xlsx: {e}>"

def read_msg_head(path: str) -> str:
    try:
        import extract_msg
        msg = extract_msg.Message(path)
        headers = {
            'from': msg.sender or '',
            'to': msg.to or '',
            'date': str(msg.date) if msg.date else '',
            'subject': msg.subject or '',
        }
        body = (msg.body or '')
        if len(body) > 2000:
            body = body[:2000] + '...'
        return f"headers={headers}; body_sample={body}"
    except Exception as e:
        return f"<error reading msg: {e}>"

def heuristics_summary(text: str) -> str:
    import re
    if not text:
        return ''
    # basic patterns
    vendor = None
    # Look for common vendor keywords (Heffner, CIBC, Honda, Toyota, etc.)
    for v in ["HEFFNER", "CIBC", "HONDA", "TOYOTA", "FORD", "CHEV", "CHEVROLET", "KIA", "HYUNDAI", "NISSAN"]:
        if v in text.upper():
            vendor = v
            break
    dates = re.findall(r"\b(20[0-9]{2}[-/][01]?[0-9][-/.][0-3]?[0-9]|[01]?[0-9]/[0-3]?[0-9]/20[0-9]{2})\b", text)[:5]
    amounts = re.findall(r"\$\s*[0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?", text)[:5]
    invoice = re.findall(r"\bINV[-\s]?[0-9A-Z-]{3,}\b", text)[:3]
    vin = re.findall(r"\b[0-9A-HJ-NPR-Z]{17}\b", text)[:2]
    parts = []
    if vendor: parts.append(f"vendor={vendor}")
    if dates: parts.append(f"dates={dates}")
    if amounts: parts.append(f"amounts={amounts}")
    if invoice: parts.append(f"invoice={invoice}")
    if vin: parts.append(f"vin={vin}")
    return "; ".join(parts)


def main():
    import argparse
    p = argparse.ArgumentParser(description='Index and sample files in verify-this-data folder')
    g = p.add_mutually_exclusive_group()
    g.add_argument('--index-only', dest='index_only', action='store_true', help='Do not attempt content sampling')
    g.add_argument('--with-content', dest='with_content', action='store_true', help='Attempt sampling for txt/csv/eml')
    args = p.parse_args()

    index_rows = []
    error_rows = []
    content_rows = []

    for path in walk_files(BASE_DIR):
        rel = os.path.relpath(path, BASE_DIR)
        name = os.path.basename(path)
        ext = os.path.splitext(name)[1].lower()
        st = safe_stat(path)
        size = st['size']
        mtime = st['mtime']
        access_ok = st['ok']
        sha = ''
        if access_ok and size is not None:
            try:
                sha = sha256_file(path)
            except Exception as e:
                access_ok = False
                error_rows.append({'path': path, 'error': f'hash: {e}'})
        else:
            error_rows.append({'path': path, 'error': st['err'] or 'stat failed'})

        index_rows.append({
            'path': path,
            'relative_path': rel,
            'name': name,
            'ext': ext,
            'size': size,
            'modified': datetime.fromtimestamp(mtime).isoformat() if mtime else '',
            'sha256': sha,
            'access_ok': access_ok,
        })

        # Content sampling
        if args.with_content and access_ok:
            sample = ''
            if ext in SAFE_TEXT_EXT:
                sample = read_text_head(path)
            elif ext in EMAIL_EXT:
                sample = read_eml_head(path)
            elif ext in PDF_EXT:
                sample = read_pdf_text(path)
            elif ext in DOCX_EXT:
                sample = read_docx_text(path)
            elif ext in XLSX_EXT:
                sample = read_xlsx_text(path)
            elif ext in MSG_EXT:
                sample = read_msg_head(path)
            else:
                # Not attempting PDFs/MSG/images here
                sample = ''
            if sample:
                summary = heuristics_summary(sample)
                content_rows.append({'path': path, 'sample': sample, 'summary': summary})

    # Write CSVs
    with open(INDEX_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['path','relative_path','name','ext','size','modified','sha256','access_ok'])
        w.writeheader()
        w.writerows(index_rows)
    with open(ERRORS_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['path','error'])
        w.writeheader()
        w.writerows(error_rows)
    if content_rows:
        with open(CONTENT_CSV, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=['path','summary','sample'])
            w.writeheader()
            w.writerows(content_rows)

    print(f"Indexed {len(index_rows)} files. Errors: {len(error_rows)}. Samples: {len(content_rows)}.\n"
          f"Index: {INDEX_CSV}\nErrors: {ERRORS_CSV}\nSamples: {CONTENT_CSV if content_rows else 'n/a'}")

if __name__ == '__main__':
    main()
