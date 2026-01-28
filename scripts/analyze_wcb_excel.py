#!/usr/bin/env python3
"""
Analyze the WCB_2011_to_Dec_2012.xlsx Excel file
"""

import openpyxl
from decimal import Decimal

excel_path = r"L:\limo\reports\WCB_2011_to_Dec_2012.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

print("="*70)
print("ANALYZING: WCB_2011_to_Dec_2012.xlsx")
print("="*70)

# Read headers first
headers = [cell.value for cell in ws[1]]
print(f"\nHeaders: {headers}")

# Read all rows
rows = list(ws.iter_rows(min_row=2, values_only=True))

print(f"\nTotal rows: {len(rows)}")
print(f"Sample row: {rows[0] if rows else 'No data'}")
# Show first 10 rows to understand structure
print(f"\nFirst 10 rows:")
for i, row in enumerate(rows[:10], 1):
    print(f"{i:2}. {row}")
# Separate by type
invoices = []
payments = []
refunds = []

for row in rows:
    # Columns: Date | Type | Invoice/Check # | Description | Amount | Running Balance | Status
    if len(row) < 5:
        continue
    
    date = row[0]
    row_type = row[1] if len(row) > 1 else None
    ref = row[2] if len(row) > 2 else None
    desc = row[3] if len(row) > 3 else None
    amount = row[4] if len(row) > 4 else None  # Amount is column 5
    running_bal = row[5] if len(row) > 5 else None
    status = row[6] if len(row) > 6 else None
    
    # Determine fiscal year from date
    if date:
        try:
            fiscal_year = date.year if hasattr(date, 'year') else None
        except:
            fiscal_year = None
    else:
        fiscal_year = None
    
    # Convert amount to Decimal
    if amount:
        try:
            amount_val = Decimal(str(amount).replace(',', '').replace('$', ''))
        except:
            amount_val = None
    else:
        amount_val = None
    
    if not amount_val:
        continue
    
    # Categorize by type
    if row_type and 'PAYMENT' in str(row_type).upper():
        payments.append((date, ref, abs(amount_val), desc, fiscal_year))
    elif row_type and 'refund' in str(row_type).lower():
        refunds.append((date, ref, abs(amount_val), desc, fiscal_year))
    elif row_type and 'INVOICE' in str(row_type).upper():
        invoices.append((date, ref, amount_val, desc, fiscal_year))
    elif amount_val < 0:
        payments.append((date, ref, abs(amount_val), desc, fiscal_year))
    else:
        invoices.append((date, ref, amount_val, desc, fiscal_year))

print(f"\nBreakdown:")
print(f"  Invoices: {len(invoices)}")
print(f"  Payments: {len(payments)}")
print(f"  Refunds: {len(refunds)}")

# Group by fiscal year
invoices_2011 = [inv for inv in invoices if inv[4] == 2011]
invoices_2012 = [inv for inv in invoices if inv[4] == 2012]
payments_2011 = [pmt for pmt in payments if pmt[4] == 2011]
payments_2012 = [pmt for pmt in payments if pmt[4] == 2012]

total_inv_2011 = sum(Decimal(str(inv[2])) for inv in invoices_2011)
total_inv_2012 = sum(Decimal(str(inv[2])) for inv in invoices_2012)
total_pmt_2011 = sum(Decimal(str(pmt[2])) for pmt in payments_2011)
total_pmt_2012 = sum(Decimal(str(pmt[2])) for pmt in payments_2012)

print(f"\n2011:")
print(f"  Invoices: {len(invoices_2011):2} = ${total_inv_2011:>10,.2f}")
print(f"  Payments: {len(payments_2011):2} = ${total_pmt_2011:>10,.2f}")

print(f"\n2012:")
print(f"  Invoices: {len(invoices_2012):2} = ${total_inv_2012:>10,.2f}")
print(f"  Payments: {len(payments_2012):2} = ${total_pmt_2012:>10,.2f}")

# Show all 2012 invoices
print(f"\n{'='*70}")
print("2012 INVOICES IN EXCEL")
print("="*70)

for i, inv in enumerate(invoices_2012, 1):
    date, ref, amount, desc, fiscal_year = inv
    desc_display = (desc[:40] + "...") if desc and len(desc) > 40 else (desc or "")
    print(f"{i:2}. {date} | {ref or 'N/A':12} | ${amount:>10,.2f} | {desc_display}")

print(f"\n   Total: {len(invoices_2012)} invoices = ${total_inv_2012:,.2f}")

# Show all 2012 payments
print(f"\n{'='*70}")
print("2012 PAYMENTS IN EXCEL")
print("="*70)

for i, pmt in enumerate(payments_2012, 1):
    date, ref, amount, desc, fiscal_year = pmt
    desc_display = (desc[:40] + "...") if desc and len(desc) > 40 else (desc or "")
    print(f"{i:2}. {date} | {ref or 'N/A':12} | ${amount:>10,.2f} | {desc_display}")

print(f"\n   Total: {len(payments_2012)} payments = ${total_pmt_2012:,.2f}")

# Final balance
print(f"\n{'='*70}")
print("EXCEL FINAL BALANCE")
print("="*70)

# Get last running balance
last_row = rows[-1]
final_balance = last_row[4] if len(last_row) > 4 else None

print(f"\n2012 Calculation:")
print(f"  Invoices: ${total_inv_2012:>10,.2f}")
print(f"  Payments: ${total_pmt_2012:>10,.2f}")
print(f"  Balance:  ${total_inv_2012 - total_pmt_2012:>10,.2f}")
print(f"\nRunning Balance (last row): ${final_balance:,.2f}" if final_balance else "\nNo final balance found")

wb.close()
