#!/usr/bin/env python3
"""
Create comprehensive Fibrenew report combining Excel data (2012-2017) 
with 2025 statement (2019-2025).
"""

import pandas as pd
from decimal import Decimal
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_FILE = r'L:\limo\pdf\2012\fibrenew.xlsx'
OUTPUT_FILE = r'L:\limo\reports\fibrenew_complete_2012_2025_reconciliation.xlsx'

def parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except:
                continue
    return None

def parse_amount(val):
    if pd.isna(val):
        return None
    try:
        return Decimal(str(val))
    except:
        return None

# Read 2012-2017 Excel data
df = pd.read_excel(EXCEL_FILE, header=None)

print("Reading 2012-2017 Excel data...")

# Parse unique invoices from Excel
excel_invoices = []
seen_invoices = {}

for idx, row in df.iterrows():
    col0 = str(row[0]).strip()
    if col0 and col0 not in ['inv', 'pmt', 'statement', 'nan'] and not 'balance' in str(row[1]).lower():
        inv_date = parse_date(row[1])
        inv_amt = parse_amount(row[2])
        notes = str(row[3]) if not pd.isna(row[3]) else ''
        
        if inv_date and col0 not in seen_invoices:
            amount = inv_amt if inv_amt else Decimal('0')
            
            # Categorize
            if amount >= Decimal('1000'):
                category = 'RENT'
            elif Decimal('100') <= amount < Decimal('1000'):
                category = 'UTILITIES'
            elif amount < Decimal('0'):
                category = 'CREDIT'
            else:
                category = 'OTHER'
            
            excel_invoices.append({
                'Invoice #': col0,
                'Date': inv_date,
                'Category': category,
                'Amount': float(amount),
                'Open Amount': 0.00,  # Assumed paid from Excel file
                'Source': 'Excel 2012-2017',
                'Notes': notes
            })
            seen_invoices[col0] = True

# Parse 2019-2025 statement data
statement_invoices = [
    {'Invoice #': '8696', 'Date': datetime(2019, 2, 1).date(), 'Category': 'UTILITIES', 'Amount': 301.48, 'Open Amount': 193.84},
    {'Invoice #': '8693', 'Date': datetime(2019, 3, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8697', 'Date': datetime(2019, 3, 1).date(), 'Category': 'UTILITIES', 'Amount': 345.88, 'Open Amount': 345.88},
    {'Invoice #': '8695', 'Date': datetime(2019, 4, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8690', 'Date': datetime(2019, 5, 7).date(), 'Category': 'UTILITIES', 'Amount': 295.69, 'Open Amount': 295.69},
    {'Invoice #': '8691', 'Date': datetime(2019, 5, 7).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8743', 'Date': datetime(2019, 5, 31).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8744', 'Date': datetime(2019, 5, 31).date(), 'Category': 'UTILITIES', 'Amount': 254.32, 'Open Amount': 254.32},
    {'Invoice #': '8832', 'Date': datetime(2019, 7, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8833', 'Date': datetime(2019, 7, 11).date(), 'Category': 'UTILITIES', 'Amount': 153.13, 'Open Amount': 153.13},
    {'Invoice #': '8894', 'Date': datetime(2019, 8, 6).date(), 'Category': 'UTILITIES', 'Amount': 144.89, 'Open Amount': 144.89},
    {'Invoice #': '8895', 'Date': datetime(2019, 8, 6).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8942', 'Date': datetime(2019, 9, 4).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8943', 'Date': datetime(2019, 9, 4).date(), 'Category': 'UTILITIES', 'Amount': 183.91, 'Open Amount': 183.91},
    {'Invoice #': '8979', 'Date': datetime(2019, 10, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '8980', 'Date': datetime(2019, 10, 1).date(), 'Category': 'UTILITIES', 'Amount': 152.62, 'Open Amount': 152.62},
    {'Invoice #': '9025', 'Date': datetime(2019, 11, 1).date(), 'Category': 'UTILITIES', 'Amount': 163.46, 'Open Amount': 163.46},
    {'Invoice #': '9066', 'Date': datetime(2019, 11, 6).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '9067', 'Date': datetime(2019, 11, 6).date(), 'Category': 'UTILITIES', 'Amount': 157.88, 'Open Amount': 157.88},
    {'Invoice #': '9103', 'Date': datetime(2019, 12, 4).date(), 'Category': 'UTILITIES', 'Amount': 126.60, 'Open Amount': 126.60},
    {'Invoice #': '9135', 'Date': datetime(2020, 1, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '9139', 'Date': datetime(2020, 1, 8).date(), 'Category': 'UTILITIES', 'Amount': 190.20, 'Open Amount': 190.20},
    {'Invoice #': '9172', 'Date': datetime(2020, 2, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '9201', 'Date': datetime(2020, 2, 14).date(), 'Category': 'UTILITIES', 'Amount': 228.12, 'Open Amount': 228.12},
    {'Invoice #': '9239', 'Date': datetime(2020, 3, 2).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '9288', 'Date': datetime(2020, 3, 30).date(), 'Category': 'UTILITIES', 'Amount': 304.47, 'Open Amount': 304.47},
    {'Invoice #': '9287', 'Date': datetime(2020, 4, 1).date(), 'Category': 'RENT', 'Amount': 682.50, 'Open Amount': 682.50},
    {'Invoice #': '9325', 'Date': datetime(2020, 5, 14).date(), 'Category': 'UTILITIES', 'Amount': 199.26, 'Open Amount': 199.26},
    {'Invoice #': '9392', 'Date': datetime(2020, 6, 23).date(), 'Category': 'UTILITIES', 'Amount': 156.64, 'Open Amount': 156.64},
    {'Invoice #': '9407', 'Date': datetime(2020, 7, 2).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9436', 'Date': datetime(2020, 7, 22).date(), 'Category': 'UTILITIES', 'Amount': 134.81, 'Open Amount': 134.81},
    {'Invoice #': '9490', 'Date': datetime(2020, 8, 5).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9542', 'Date': datetime(2020, 9, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9561', 'Date': datetime(2020, 9, 10).date(), 'Category': 'UTILITIES', 'Amount': 142.63, 'Open Amount': 142.63},
    {'Invoice #': '9609', 'Date': datetime(2020, 10, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9623', 'Date': datetime(2020, 10, 8).date(), 'Category': 'UTILITIES', 'Amount': 145.20, 'Open Amount': 145.20},
    {'Invoice #': '9670', 'Date': datetime(2020, 11, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9694', 'Date': datetime(2020, 11, 18).date(), 'Category': 'UTILITIES', 'Amount': 162.21, 'Open Amount': 162.21},
    {'Invoice #': '9727', 'Date': datetime(2020, 12, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9742', 'Date': datetime(2020, 12, 7).date(), 'Category': 'UTILITIES', 'Amount': 191.25, 'Open Amount': 191.25},
    {'Invoice #': '9767', 'Date': datetime(2021, 1, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9772', 'Date': datetime(2021, 1, 18).date(), 'Category': 'UTILITIES', 'Amount': 201.35, 'Open Amount': 201.35},
    {'Invoice #': '9800', 'Date': datetime(2021, 2, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9815', 'Date': datetime(2021, 2, 5).date(), 'Category': 'UTILITIES', 'Amount': 169.44, 'Open Amount': 169.44},
    {'Invoice #': '9866', 'Date': datetime(2021, 3, 1).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '9885', 'Date': datetime(2021, 3, 8).date(), 'Category': 'UTILITIES', 'Amount': 220.34, 'Open Amount': 220.34},
    {'Invoice #': '9956', 'Date': datetime(2021, 4, 6).date(), 'Category': 'RENT', 'Amount': 840.00, 'Open Amount': 840.00},
    {'Invoice #': '12131', 'Date': datetime(2024, 1, 2).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12132', 'Date': datetime(2024, 2, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12133', 'Date': datetime(2024, 3, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12177', 'Date': datetime(2024, 4, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12226', 'Date': datetime(2024, 5, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12419', 'Date': datetime(2024, 8, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 167.06},
    {'Invoice #': '12494', 'Date': datetime(2024, 9, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12540', 'Date': datetime(2024, 10, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12601', 'Date': datetime(2024, 11, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12664', 'Date': datetime(2024, 12, 2).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12714', 'Date': datetime(2025, 1, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12775', 'Date': datetime(2025, 2, 3).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12835', 'Date': datetime(2025, 3, 3).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12909', 'Date': datetime(2025, 4, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '12973', 'Date': datetime(2025, 5, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '13041', 'Date': datetime(2025, 6, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '13103', 'Date': datetime(2025, 7, 1).date(), 'Category': 'RENT', 'Amount': 1102.50, 'Open Amount': 1102.50},
    {'Invoice #': '13180', 'Date': datetime(2025, 8, 1).date(), 'Category': 'RENT', 'Amount': 1260.00, 'Open Amount': 1260.00},
    {'Invoice #': '13248', 'Date': datetime(2025, 9, 1).date(), 'Category': 'RENT', 'Amount': 1260.00, 'Open Amount': 1260.00},
    {'Invoice #': '13310', 'Date': datetime(2025, 10, 1).date(), 'Category': 'RENT', 'Amount': 1260.00, 'Open Amount': 1260.00},
    {'Invoice #': '13379', 'Date': datetime(2025, 11, 1).date(), 'Category': 'RENT', 'Amount': 1260.00, 'Open Amount': 1260.00},
]

# Add source and notes to statement invoices
for inv in statement_invoices:
    inv['Source'] = '2025 Statement'
    inv['Notes'] = ''

# Parse 2023 adjustments
adjustments = [
    {'Date': datetime(2023, 7, 31).date(), 'Description': 'Journal Entry #21: shareholders Earning', 'Amount': -3508.25, 'Applied': -1458.58},
    {'Date': datetime(2023, 7, 31).date(), 'Description': 'Journal Entry #22: Shareholder - wedding', 'Amount': -2767.50, 'Applied': -746.42},
]

# Parse payments from 2025 statement
statement_payments = [
    {'Date': datetime(2023, 9, 22).date(), 'Amount': 500.00, 'Applied': 380.64, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2023, 10, 12).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2023, 10, 26).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2023, 11, 1).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2023, 12, 1).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2023, 12, 31).date(), 'Amount': 400.00, 'Applied': 400.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 2, 20).date(), 'Amount': 300.00, 'Applied': 300.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 3, 13).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 4, 18).date(), 'Amount': 400.00, 'Applied': 400.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 4, 18).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 5, 13).date(), 'Amount': 1200.00, 'Applied': 1200.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 7, 2).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 9, 4).date(), 'Amount': 2100.00, 'Applied': 2100.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 9, 11).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 9, 26).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 10, 15).date(), 'Amount': 1000.00, 'Applied': 1000.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 11, 4).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2024, 12, 5).date(), 'Amount': 1500.00, 'Applied': 1500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 1, 7).date(), 'Amount': 1200.00, 'Applied': 1200.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 2, 4).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 3, 10).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 4, 8).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 5, 14).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 6, 10).date(), 'Amount': 1102.50, 'Applied': 1102.50, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 7, 4).date(), 'Amount': 800.00, 'Applied': 800.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 7, 4).date(), 'Amount': 400.00, 'Applied': 400.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 7, 31).date(), 'Amount': 2500.00, 'Applied': 2500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 8, 15).date(), 'Amount': 300.00, 'Applied': 300.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 9, 16).date(), 'Amount': 500.00, 'Applied': 500.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 10, 2).date(), 'Amount': 2000.00, 'Applied': 2000.00, 'Method': 'Payment', 'Notes': ''},
    {'Date': datetime(2025, 11, 10).date(), 'Amount': 900.00, 'Applied': 900.00, 'Method': 'Payment', 'Notes': '#9/100'},
    {'Date': datetime(2025, 11, 17).date(), 'Amount': 1100.00, 'Applied': 200.00, 'Method': 'Payment', 'Notes': ''},
]

# Combine all invoices
all_invoices = excel_invoices + statement_invoices

print(f"Found {len(excel_invoices)} invoices from Excel (2012-2017)")
print(f"Found {len(statement_invoices)} invoices from 2025 Statement (2019-2025)")
print(f"Total invoices: {len(all_invoices)}")

# Calculate summary by year
summary_by_year = defaultdict(lambda: {'rent': Decimal('0'), 'utilities': Decimal('0'), 
                                        'credits': Decimal('0'), 'total_invoiced': Decimal('0'),
                                        'open_balance': Decimal('0')})

for inv in all_invoices:
    year = inv['Date'].year
    amount = Decimal(str(inv['Amount']))
    open_amt = Decimal(str(inv['Open Amount']))
    
    if inv['Category'] == 'RENT':
        summary_by_year[year]['rent'] += amount
    elif inv['Category'] == 'UTILITIES':
        summary_by_year[year]['utilities'] += amount
    elif inv['Category'] == 'CREDIT':
        summary_by_year[year]['credits'] += amount
    
    summary_by_year[year]['total_invoiced'] += amount
    summary_by_year[year]['open_balance'] += open_amt

summary_list = []
for year in sorted(summary_by_year.keys()):
    data = summary_by_year[year]
    summary_list.append({
        'Year': year,
        'Rent': float(data['rent']),
        'Utilities': float(data['utilities']),
        'Credits': float(data['credits']),
        'Total Invoiced': float(data['total_invoiced']),
        'Open Balance': float(data['open_balance'])
    })

# Calculate final totals
total_invoiced = sum(Decimal(str(inv['Amount'])) for inv in all_invoices)
total_open = sum(Decimal(str(inv['Open Amount'])) for inv in all_invoices)
total_payments_2025 = sum(Decimal(str(p['Amount'])) for p in statement_payments)
total_adjustments = sum(Decimal(str(adj['Amount'])) for adj in adjustments)

final_summary = [{
    'Description': 'Total Invoiced (2012-2025)',
    'Amount': float(total_invoiced)
}, {
    'Description': 'Total Payments (2023-2025 Statement)',
    'Amount': float(total_payments_2025)
}, {
    'Description': 'Total Adjustments (2023 wedding/shareholderearnings)',
    'Amount': float(total_adjustments)
}, {
    'Description': 'Current Balance (per Nov 26, 2025 statement)',
    'Amount': float(total_open)
}, {
    'Description': 'Statement Balance Breakdown:',
    'Amount': None
}, {
    'Description': '  Current Due',
    'Amount': 1260.00
}, {
    'Description': '  1-30 Days',
    'Amount': 160.00
}, {
    'Description': '  31-60 Days',
    'Amount': -740.00
}, {
    'Description': '  61-90 Days',
    'Amount': 760.00
}, {
    'Description': '  90+ Days',
    'Amount': 13294.56
}, {
    'Description': '  TOTAL AMOUNT DUE',
    'Amount': 14734.56
}]

# Create Excel report
print(f"\nCreating comprehensive Excel report: {OUTPUT_FILE}")

with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    # Sheet 1: All Invoices
    df_invoices = pd.DataFrame(all_invoices)
    df_invoices = df_invoices.sort_values('Date')
    df_invoices.to_excel(writer, sheet_name='All Invoices', index=False)
    
    # Sheet 2: 2025 Statement Payments
    df_payments = pd.DataFrame(statement_payments)
    df_payments.to_excel(writer, sheet_name='Payments 2023-2025', index=False)
    
    # Sheet 3: 2023 Adjustments
    df_adj = pd.DataFrame(adjustments)
    df_adj.to_excel(writer, sheet_name='2023 Adjustments', index=False)
    
    # Sheet 4: Summary by Year
    df_summary = pd.DataFrame(summary_list)
    df_summary.to_excel(writer, sheet_name='Summary by Year', index=False)
    
    # Sheet 5: Final Summary
    df_final = pd.DataFrame(final_summary)
    df_final.to_excel(writer, sheet_name='Final Summary', index=False)
    
    # Sheet 6: Outstanding Invoices Only
    outstanding = [inv for inv in all_invoices if inv['Open Amount'] > 0]
    df_outstanding = pd.DataFrame(outstanding)
    df_outstanding.to_excel(writer, sheet_name='Outstanding Nov 2025', index=False)

# Format the Excel file
wb = openpyxl.load_workbook(OUTPUT_FILE)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format currency columns
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, (int, float)):
                if 'Amount' in str(ws.cell(1, cell.column).value):
                    cell.number_format = '$#,##0.00'

wb.save(OUTPUT_FILE)

print(f"\n{'='*80}")
print("COMPREHENSIVE REPORT CREATED SUCCESSFULLY")
print(f"{'='*80}")
print(f"\nFile: {OUTPUT_FILE}")
print(f"\nSheets created:")
print(f"  1. All Invoices: {len(all_invoices)} invoices (2012-2025)")
print(f"  2. Payments 2023-2025: {len(statement_payments)} payments")
print(f"  3. 2023 Adjustments: {len(adjustments)} adjustments")
print(f"  4. Summary by Year: {len(summary_list)} years")
print(f"  5. Final Summary: Complete reconciliation")
print(f"  6. Outstanding Nov 2025: {len(outstanding)} unpaid invoices")
print(f"\nKey Totals:")
print(f"  Total Invoiced (2012-2025): ${total_invoiced:,.2f}")
print(f"  Total Payments (2023-2025): ${total_payments_2025:,.2f}")
print(f"  Total Adjustments (2023): ${total_adjustments:,.2f}")
print(f"  Current Balance (Nov 26, 2025): ${total_open:,.2f}")
print(f"  Statement Total Due: $14,734.56")
print(f"\nRent Evolution:")
print(f"  2019-2020 Q1: $682.50/month")
print(f"  2020 Jul-2021 Apr: $840.00/month")
print(f"  2024-2025 Jul: $1,102.50/month")
print(f"  2025 Aug-Nov: $1,260.00/month")
