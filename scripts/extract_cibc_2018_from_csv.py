"""
CIBC 8362 2018 - Parse CSV and calculate running balance
CSV format: Date, Description, Debit, Credit (NO balance column - we calculate it)
"""
import csv
from pathlib import Path
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

csv_path = Path(r"L:\limo\CIBC UPLOADS\0228362 (CIBC checking account)\cibc 8362 2018.csv")
output_path = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

# Opening balance from statement (user can adjust if needed)
OPENING_BALANCE = Decimal('1058.53')  # From PDF "Balance forward $418.93" was wrong, trying from sheet

print("PARSING CIBC 8362 2018 CSV + BALANCE AUDIT")
print("=" * 80)

if not csv_path.exists():
    print(f"❌ CSV not found: {csv_path}")
    print("   User typed data shows format: Date | Description | Debit | Credit | Balance")
    print("   If CSV exists, place it at the path above")
    exit(1)

trans = []

with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    
    for row in reader:
        if not row or len(row) < 2:
            continue
        
        # Skip header rows
        if 'date' in row[0].lower() or 'transaction' in ' '.join(row).lower():
            continue
        
        # Parse transaction row
        # CSV format: Date, Description, Debit, Credit (4 columns only)
        try:
            date_str = row[0].strip()
            desc = row[1].strip() if len(row) > 1 else ""
            debit_str = row[2].strip() if len(row) > 2 else ""
            credit_str = row[3].strip() if len(row) > 3 else ""
            
            # Parse amounts (empty string = 0)
            debit = Decimal(debit_str.replace(',', '').replace('$', '')) if debit_str else Decimal('0')
            credit = Decimal(credit_str.replace(',', '').replace('$', '')) if credit_str else Decimal('0')
            
            # Net amount (credit positive, debit negative)
            amount = credit - debit
            
            trans.append({
                'date': date_str,
                'desc': desc,
                'debit': debit,
                'credit': credit,
                'amount': amount
            })
            
        except Exception as e:
            print(f"⚠️  Skipping row: {row[:3]} - {e}")
            continue

print(f"Extracted {len(trans)} transactions\n")

if not trans:
    print("❌ No transactions extracted")
    exit(1)

# Sort by date to ensure chronological order
trans.sort(key=lambda x: x['date'])

# CALCULATE RUNNING BALANCE
print("CALCULATING RUNNING BALANCE...")
print("=" * 80)
print(f"Opening balance: ${OPENING_BALANCE:,.2f}\n")

running_bal = OPENING_BALANCE

for i, t in enumerate(trans):
    running_bal += t['amount']
    t['calc'] = running_bal

total_credits = sum(t['credit'] for t in trans)
total_debits = sum(t['debit'] for t in trans)

print(f"Total Credits: ${total_credits:,.2f}")
print(f"Total Debits: ${total_debits:,.2f}")
print(f"Net Change: ${total_credits - total_debits:,.2f}")
print(f"Closing balance (calculated): ${running_bal:,.2f}")

# CREATE EXCEL WITH AUDIT
wb = Workbook()
ws = wb.active
ws.title = "Transactions + Audit"

headers = ['Date', 'Description', 'Debit', 'Credit', 'Net Amount', 'Stated Balance', 'Calculated Balance', 'Diff', 'Error?']
ws.append(headers)

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col in range(1, len(headers) + 1):
    ws.cell(1, col).fill = header_fill
    ws.cell(1, col).font = header_font

for t in trans:
    row_data = [
        t['date'],
        t['desc'],
        float(t['debit']),
        float(t['credit']),
        float(t['amount']),
        float(t['calc'])
    ]
    ws.append(row_data)

ws.column_dimensions['B'].width = 50

for col in [3, 4, 5, 6]:
    pass
summary['B3'].number_format = '$#,##0.00'

summary['A4'] = "Total Transactions:"
summary['B4'] = len(trans)
summary['A1'].font = Font(bold=True, size=14)

summary['A3'] = "Opening Balance:"
summary['B3'] = float(OPENING_BALANCE)
summary['B3'].number_format = '$#,##0.00'

summary['A4'] = "Total Transactions:"
summary['B4'] = len(trans)

summary['A5'] = "Total Deposits:"
summary['B5'] = float(total_credits)
summary['B5'].number_format = '$#,##0.00'

summary['A6'] = "Total Withdrawals:"
summary['B6'] = float(total_debits)
summary['B6'].number_format = '$#,##0.00'

summary['A7'] = "Net Change:"
summary['B7'] = float(total_credits - total_debits)
summary['B7'].number_format = '$#,##0.00'

summary['A8'] = "Closing Balance:"
summary['B8'] = float(running_bal)
summary['B8'].number_format = '$#,##0.00'

wb.save(output_path)

print(f"\n✅ Saved: {output_path}")
print(f"   {len(trans)} transactions with calculated running balances")