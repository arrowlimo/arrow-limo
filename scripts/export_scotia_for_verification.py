import os
from datetime import date

import psycopg2
import psycopg2.extras

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", os.environ.get("DB_PASSWORD"))

SCOTIA_BANK_ID = 2
SCOTIA_ACCOUNT = "903990106011"
DATE_START = date(2012, 1, 1)
DATE_END = date(2017, 12, 31)


def get_conn():
    return psycopg2.connect(host=DB_HOST, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)


def main():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    print("="*80)
    print("SCOTIA BANK DATA EXPORT FOR VERIFICATION")
    print("="*80)

    # Check what Scotia data exists
    cur.execute("""
        SELECT COUNT(*), MIN(transaction_date), MAX(transaction_date)
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
    """, (SCOTIA_BANK_ID, SCOTIA_ACCOUNT))
    
    count, min_date, max_date = cur.fetchone()
    print(f"\nScotia Bank transactions in database:")
    print(f"  Total: {count:,}")
    print(f"  Date range: {min_date} to {max_date}")

    if count == 0:
        print("\n⚠️  No Scotia Bank transactions found!")
        print("   Checking for alternative account numbers...")
        
        cur.execute("""
            SELECT DISTINCT account_number, bank_id, COUNT(*) 
            FROM banking_transactions
            WHERE account_number ILIKE '%scotia%' 
               OR account_number ILIKE '%903990%'
            GROUP BY account_number, bank_id
        """)
        
        for acc, bid, cnt in cur.fetchall():
            print(f"   Found: {acc} (bank_id={bid}): {cnt} transactions")
        
        conn.close()
        return

    # Yearly breakdown
    cur.execute("""
        SELECT EXTRACT(YEAR FROM transaction_date) as year, COUNT(*)
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
        GROUP BY year
        ORDER BY year
    """, (SCOTIA_BANK_ID, SCOTIA_ACCOUNT))
    
    print(f"\nBy Year:")
    for year, cnt in cur.fetchall():
        print(f"  {int(year)}: {cnt:,} transactions")

    # Export to Excel for verification
    cur.execute("""
        SELECT transaction_id, transaction_date, description,
               debit_amount, credit_amount, balance, source_file,
               vendor_extracted, is_transfer
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
          AND transaction_date BETWEEN %s AND %s
        ORDER BY transaction_date, transaction_id
    """, (SCOTIA_BANK_ID, SCOTIA_ACCOUNT, DATE_START, DATE_END))
    
    rows = cur.fetchall()
    
    if Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Scotia 2012-2017"
        
        headers = ["TxID", "Date", "Description", "Debit", "Credit", "Balance", 
                   "Source", "Vendor", "IsTransfer"]
        ws.append(headers)
        
        for r in rows:
            ws.append([
                r['transaction_id'], r['transaction_date'], r['description'],
                float(r['debit_amount']) if r['debit_amount'] else None,
                float(r['credit_amount']) if r['credit_amount'] else None,
                float(r['balance']) if r['balance'] else None,
                r['source_file'], r['vendor_extracted'],
                r['is_transfer']
            ])
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        out_path = r"l:\limo\reports\exports\Scotia_903990106011_2012_2017_Current.xlsx"
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)
        print(f"\n✅ Exported to: {out_path}")
        print(f"   Please verify this file and save as: Scotia_903990106011_2012_2017_VERIFIED.xlsx")
    
    conn.close()


if __name__ == "__main__":
    main()
