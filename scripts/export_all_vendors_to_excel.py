#!/usr/bin/env python3
"""
Export all vendors to Excel with counts and totals.
"""

import psycopg2
import pandas as pd
from datetime import datetime

conn = psycopg2.connect(
    host="localhost",
    database="almsdata",
    user="postgres",
    password="***REDACTED***"
)

cur = conn.cursor()

print("=" * 80)
print("EXPORTING ALL VENDORS TO EXCEL")
print("=" * 80)

# Get all vendors from receipts
print("\nQuerying receipts vendors...")
cur.execute("""
    SELECT 
        vendor_name,
        COUNT(*) as receipt_count,
        SUM(gross_amount) as total_amount,
        MIN(receipt_date) as first_date,
        MAX(receipt_date) as last_date,
        COUNT(DISTINCT EXTRACT(YEAR FROM receipt_date)) as years_active
    FROM receipts
    WHERE vendor_name IS NOT NULL
    GROUP BY vendor_name
    ORDER BY total_amount DESC NULLS LAST
""")

receipts_data = cur.fetchall()
print(f"Found {len(receipts_data)} unique vendors in receipts")

# Create DataFrame
df_receipts = pd.DataFrame(receipts_data, columns=[
    'Vendor Name', 'Receipt Count', 'Total Amount', 
    'First Date', 'Last Date', 'Years Active'
])

# Get all vendors from banking
print("Querying banking vendors...")
cur.execute("""
    SELECT 
        vendor_extracted,
        COUNT(*) as transaction_count,
        SUM(debit_amount) as total_debit,
        SUM(credit_amount) as total_credit,
        MIN(transaction_date) as first_date,
        MAX(transaction_date) as last_date
    FROM banking_transactions
    WHERE vendor_extracted IS NOT NULL
    GROUP BY vendor_extracted
    ORDER BY total_debit DESC NULLS LAST
""")

banking_data = cur.fetchall()
print(f"Found {len(banking_data)} unique vendors in banking")

df_banking = pd.DataFrame(banking_data, columns=[
    'Vendor Name', 'Transaction Count', 'Total Debit', 
    'Total Credit', 'First Date', 'Last Date'
])

# Get top vendors by category
print("Analyzing vendor categories...")
cur.execute("""
    SELECT 
        category,
        vendor_name,
        COUNT(*) as count,
        SUM(gross_amount) as total
    FROM receipts
    WHERE vendor_name IS NOT NULL
      AND category IS NOT NULL
    GROUP BY category, vendor_name
    ORDER BY category, total DESC
""")

category_data = cur.fetchall()
df_categories = pd.DataFrame(category_data, columns=[
    'Category', 'Vendor Name', 'Receipt Count', 'Total Amount'
])

# Create summary stats
summary = {
    'Metric': [
        'Total Unique Vendors (Receipts)',
        'Total Unique Vendors (Banking)',
        'Total Receipts',
        'Total Receipt Amount',
        'Total Banking Transactions',
        'Total Banking Debits',
        'Total Banking Credits'
    ],
    'Value': [
        len(receipts_data),
        len(banking_data),
        df_receipts['Receipt Count'].sum(),
        df_receipts['Total Amount'].sum(),
        df_banking['Transaction Count'].sum(),
        df_banking['Total Debit'].sum(),
        df_banking['Total Credit'].sum()
    ]
}
df_summary = pd.DataFrame(summary)

# Export to Excel
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_file = f"l:\\limo\\reports\\all_vendors_{timestamp}.xlsx"

print(f"\nWriting to Excel: {output_file}")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='Summary', index=False)
    df_receipts.to_excel(writer, sheet_name='Receipt Vendors', index=False)
    df_banking.to_excel(writer, sheet_name='Banking Vendors', index=False)
    df_categories.to_excel(writer, sheet_name='Vendors by Category', index=False)

print("✅ Excel file created")

# Format the Excel file
print("Formatting Excel file...")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = load_workbook(output_file)

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Number formatting for amount columns
    if sheet_name == 'Receipt Vendors':
        for row in range(2, ws.max_row + 1):
            ws[f'C{row}'].number_format = '$#,##0.00'
    elif sheet_name == 'Banking Vendors':
        for row in range(2, ws.max_row + 1):
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'].number_format = '$#,##0.00'
    elif sheet_name == 'Vendors by Category':
        for row in range(2, ws.max_row + 1):
            ws[f'D{row}'].number_format = '$#,##0.00'
    elif sheet_name == 'Summary':
        for row in range(2, ws.max_row + 1):
            if 'Amount' in str(ws[f'A{row}'].value) or 'Debit' in str(ws[f'A{row}'].value) or 'Credit' in str(ws[f'A{row}'].value):
                ws[f'B{row}'].number_format = '$#,##0.00'

wb.save(output_file)
print("✅ Formatting complete")

# Print summary
print("\n" + "=" * 80)
print("EXPORT SUMMARY")
print("=" * 80)
print(f"\nFile: {output_file}")
print(f"\nSheets created:")
print(f"  1. Summary - Overall statistics")
print(f"  2. Receipt Vendors - {len(receipts_data)} unique vendors")
print(f"  3. Banking Vendors - {len(banking_data)} unique vendors")
print(f"  4. Vendors by Category - {len(df_categories)} category-vendor combinations")

print(f"\nTop 10 vendors by amount:")
for idx, row in df_receipts.head(10).iterrows():
    print(f"  {row['Vendor Name']:40} ${row['Total Amount']:>12,.2f} ({row['Receipt Count']:>4} receipts)")

cur.close()
conn.close()

print("\n✅ COMPLETE")
