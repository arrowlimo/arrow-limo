#!/usr/bin/env python3
"""
Match Outlook calendar appointments to charters and update database.
- Matches by reserve_number
- Updates dispatcher_notes with calendar information
- Reconciles driver assignments
- Generates Excel report for mismatches
"""

import json
import sys
import os
import argparse
from datetime import datetime
import re
import psycopg2

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def normalize_name(name):
    """Normalize a name for fuzzy matching."""
    if not name:
        return ""
    # Remove extra whitespace, lowercase
    return ' '.join(name.lower().split())


def fuzzy_match_driver(calendar_driver, employees):
    """
    Fuzzy match a driver name from calendar to employees table.
    Returns (employee_id, full_name, confidence) or None
    """
    if not calendar_driver:
        return None
    
    cal_norm = normalize_name(calendar_driver)
    cal_parts = set(cal_norm.split())
    
    best_match = None
    best_score = 0
    
    for emp in employees:
        emp_id = emp[0]
        emp_full = emp[1] or ""
        emp_first = emp[2] or ""
        emp_last = emp[3] or ""
        
        # Try full name
        emp_norm = normalize_name(emp_full)
        if emp_norm == cal_norm:
            return (emp_id, emp_full, 1.0)  # Perfect match
        
        # Try first + last
        emp_fl = normalize_name(f"{emp_first} {emp_last}")
        if emp_fl == cal_norm:
            return (emp_id, emp_full, 1.0)
        
        # Partial match: check if calendar parts are in employee name
        emp_parts = set(emp_norm.split())
        if emp_parts and cal_parts:
            overlap = len(cal_parts & emp_parts)
            score = overlap / max(len(cal_parts), len(emp_parts))
            
            if score > best_score and score >= 0.5:  # At least 50% match
                best_match = (emp_id, emp_full, score)
                best_score = score
    
    return best_match


def merge_dispatcher_notes(existing_notes, calendar_info):
    """
    Merge calendar information into booking_notes (dispatcher notes) without overwriting.
    Appends new calendar data with clear separator.
    """
    if not existing_notes:
        existing_notes = ""
    
    separator = "\n\n--- Calendar Import (Outlook 'arrow new') ---\n"
    
    # Check if calendar info already present
    if "Calendar Import (Outlook 'arrow new')" in existing_notes:
        # Already imported, append with date
        separator = f"\n\n--- Calendar Update ({datetime.now().strftime('%Y-%m-%d %H:%M')}) ---\n"
    
    return existing_notes + separator + calendar_info


def match_and_update_charters(json_file, dry_run=True, write=False):
    """
    Main processing function:
    1. Load calendar JSON
    2. Match to charters by reserve_number
    3. Update dispatcher_notes
    4. Reconcile driver assignments
    5. Generate Excel report for issues
    """
    
    # Load calendar data
    print(f"Loading calendar data from {json_file}...")
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    appointments = data.get('appointments', [])
    print(f"Loaded {len(appointments)} appointments")
    
    # Filter to those with reserve numbers
    with_reserve = [a for a in appointments if a.get('reserve_number')]
    print(f"  {len(with_reserve)} appointments have reserve numbers")
    
    # Connect to database
    conn = psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        database=os.getenv('DB_NAME', 'almsdata'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', '***REDACTED***')
    )
    cur = conn.cursor()
    
    # Load all employees (chauffeurs) for driver matching
    print("Loading employees...")
    cur.execute("""
        SELECT employee_id, full_name, first_name, last_name
        FROM employees
        WHERE status = 'active' OR is_chauffeur = true
        ORDER BY full_name
    """)
    employees = cur.fetchall()
    print(f"  Loaded {len(employees)} employees")
    
    # Process each appointment
    matched_count = 0
    updated_notes_count = 0
    updated_driver_count = 0
    updated_cancelled_count = 0
    
    mismatches = {
        'no_charter_found': [],
        'driver_mismatch': [],
        'dual_assignment': [],
        'missing_calendar_driver': [],
        'missing_db_driver': [],
        'missing_vehicle': [],
        'name_phone_candidates': [],
        'client_email_matched': [],
        'needs_client_creation': []
    }

    def find_candidates_by_phone(cur, phones, base_start_time):
        """Return candidate charters based on client phone digits near the start date."""
        if not phones:
            return []
        # Normalize to digits
        digits_list = []
        for p in phones:
            d = re.sub(r"\D", "", p or "")
            if len(d) >= 7:
                digits_list.append(d)
        if not digits_list:
            return []

        from datetime import timedelta
        date_from = None
        date_to = None
        try:
            if base_start_time:
                base = datetime.strptime(base_start_time[:10], '%Y-%m-%d')
                date_from = (base - timedelta(days=3)).date()
                date_to = (base + timedelta(days=7)).date()
        except Exception:
            pass

        candidates = []
        for d in digits_list:
            if date_from and date_to:
                cur.execute(
                    """
                    SELECT ch.reserve_number, ch.charter_date, cl.client_name
                    FROM charters ch
                    LEFT JOIN clients cl ON cl.client_id = ch.client_id
                    WHERE ch.charter_date BETWEEN %s AND %s
                      AND (
                        regexp_replace(coalesce(cl.primary_phone,''),'\\D','','g') = %s
                        OR cl.contact_info ILIKE %s
                      )
                    ORDER BY ch.charter_date
                    """,
                    (date_from, date_to, d, f"%{d}%"),
                )
            else:
                cur.execute(
                    """
                    SELECT ch.reserve_number, ch.charter_date, cl.client_name
                    FROM charters ch
                    LEFT JOIN clients cl ON cl.client_id = ch.client_id
                    WHERE (
                        regexp_replace(coalesce(cl.primary_phone,''),'\\D','','g') = %s
                        OR cl.contact_info ILIKE %s
                    )
                    ORDER BY ch.charter_date DESC
                    LIMIT 50
                    """,
                    (d, f"%{d}%"),
                )
            for r in cur.fetchall():
                candidates.append({
                    'reserve_number': r[0],
                    'charter_date': str(r[1]) if r[1] else None,
                    'client_name': r[2],
                    'match_type': 'phone',
                    'match_value': d,
                })
        return candidates

    def find_candidates_by_name(cur, subject_text, base_start_time):
        """Return candidate charters based on tokens in subject near the start date."""
        if not subject_text:
            return []
        tokens = [t.lower() for t in re.findall(r"[A-Za-z]{3,}", subject_text)]
        stop = set(['quote','pickup','dropoff','trip','bus','limo','party','return','am','pm','for','and','the','or'])
        tokens = [t for t in tokens if t not in stop][:4]
        if not tokens:
            return []

        from datetime import timedelta
        date_from = None
        date_to = None
        try:
            if base_start_time:
                base = datetime.strptime(base_start_time[:10], '%Y-%m-%d')
                date_from = (base - timedelta(days=3)).date()
                date_to = (base + timedelta(days=7)).date()
        except Exception:
            pass

        ilike = " OR ".join(["cl.client_name ILIKE %s" for _ in tokens])
        params = [f"%{t}%" for t in tokens]
        if date_from and date_to:
            sql = (
                "SELECT ch.reserve_number, ch.charter_date, cl.client_name "
                "FROM charters ch LEFT JOIN clients cl ON cl.client_id=ch.client_id "
                "WHERE ch.charter_date BETWEEN %s AND %s AND (" + ilike + ") "
                "ORDER BY ch.charter_date"
            )
            cur.execute(sql, [date_from, date_to] + params)
        else:
            sql = (
                "SELECT ch.reserve_number, ch.charter_date, cl.client_name "
                "FROM charters ch LEFT JOIN clients cl ON cl.client_id=ch.client_id "
                "WHERE " + ilike + " ORDER BY ch.charter_date DESC LIMIT 50"
            )
            cur.execute(sql, params)

        candidates = []
        for r in cur.fetchall():
            candidates.append({
                'reserve_number': r[0],
                'charter_date': str(r[1]) if r[1] else None,
                'client_name': r[2],
                'match_type': 'name',
                'match_value': ','.join(tokens)
            })
        return candidates

    def find_candidates_by_datetime(cur, cal_start_time):
        """Return charters with matching or nearby date/time."""
        if not cal_start_time:
            return []
        
        from datetime import timedelta
        try:
            cal_dt = datetime.strptime(cal_start_time[:19], '%Y-%m-%d %H:%M:%S')
            cal_date = cal_dt.date()
            cal_time = cal_dt.time()
            
            # Find charters on same day, or within 6 hours on adjacent days
            date_from = (cal_dt - timedelta(hours=6)).date()
            date_to = (cal_dt + timedelta(hours=6)).date()
            
            cur.execute(
                """
                SELECT ch.reserve_number, ch.charter_date, ch.pickup_time, 
                       cl.client_name, ch.pickup_address
                FROM charters ch
                LEFT JOIN clients cl ON cl.client_id = ch.client_id
                WHERE ch.charter_date BETWEEN %s AND %s
                ORDER BY ch.charter_date, ch.pickup_time
                """,
                (date_from, date_to)
            )
            
            candidates = []
            for r in cur.fetchall():
                charter_date = r[1]
                pickup_time = r[2]
                
                # Calculate time difference
                time_diff_hours = None
                if charter_date and pickup_time:
                    charter_dt = datetime.combine(charter_date, pickup_time)
                    time_diff = abs((charter_dt - cal_dt).total_seconds() / 3600)
                    time_diff_hours = round(time_diff, 1)
                    
                    # Only include if within 6 hours
                    if time_diff <= 6:
                        candidates.append({
                            'reserve_number': r[0],
                            'charter_date': str(charter_date) if charter_date else None,
                            'pickup_time': str(pickup_time) if pickup_time else None,
                            'client_name': r[3],
                            'pickup_address': r[4],
                            'match_type': 'datetime',
                            'match_value': f'{time_diff_hours}h diff',
                            'time_diff_hours': time_diff_hours
                        })
                elif charter_date == cal_date:
                    # Same day but no pickup time in DB
                    candidates.append({
                        'reserve_number': r[0],
                        'charter_date': str(charter_date),
                        'pickup_time': str(pickup_time) if pickup_time else None,
                        'client_name': r[3],
                        'pickup_address': r[4],
                        'match_type': 'datetime',
                        'match_value': 'same day',
                        'time_diff_hours': None
                    })
            
            # Sort by time difference (closest first)
            candidates.sort(key=lambda x: x['time_diff_hours'] if x['time_diff_hours'] is not None else 999)
            return candidates[:20]  # Limit to top 20 closest matches
            
        except Exception as e:
            return []
    
    for appt in with_reserve:
        reserve_num = appt['reserve_number']
        
        # Find charter by reserve_number
        cur.execute("""
            SELECT charter_id, reserve_number, assigned_driver_id, driver_name,
                   notes, booking_notes, client_notes, driver_notes,
                   charter_date, pickup_time, status, vehicle, vehicle_id,
                   pickup_address, dropoff_address
            FROM charters
            WHERE reserve_number = %s
        """, (reserve_num,))
        
        charter = cur.fetchone()
        
        if not charter:
            mismatches['no_charter_found'].append({
                'reserve_number': reserve_num,
                'calendar_subject': appt['subject'],
                'calendar_date': appt['start_time'],
                'calendar_driver': appt['driver_name']
            })
            # Attempt client match by email from calendar notes
            try:
                emails = appt.get('emails') or []
                matched_clients = []
                for em in emails[:3]:  # limit queries
                    cur.execute(
                        """
                        SELECT client_id, COALESCE(company_name, client_name) AS display_name, email
                        FROM clients
                        WHERE LOWER(email) = LOWER(%s)
                        """,
                        (em,)
                    )
                    rows = cur.fetchall()
                    for r in rows:
                        matched_clients.append({'client_id': r[0], 'client_name': r[1], 'email': r[2]})

                if matched_clients:
                    mismatches['client_email_matched'].append({
                        'calendar_subject': appt.get('subject',''),
                        'calendar_start': appt.get('start_time'),
                        'emails': emails,
                        'matches': matched_clients
                    })
                else:
                    # Suggest creating a client with email/name from calendar subject/body
                    client_guess = None
                    subj = appt.get('subject','')
                    loc = appt.get('location','')
                    # crude guess: text before '-' as client name
                    m = re.match(r"^([^\-\|]+)", subj)
                    if m:
                        client_guess = m.group(1).strip()
                    mismatches['needs_client_creation'].append({
                        'calendar_subject': subj,
                        'calendar_start': appt.get('start_time'),
                        'emails': emails,
                        'suggested_client_name': client_guess or '',
                        'location': loc,
                    })
            except Exception:
                pass
            # Suggest candidates: 1) phone, 2) datetime proximity, 3) name tokens
            phone_cands = find_candidates_by_phone(cur, appt.get('phone_numbers', []), appt.get('start_time'))
            if phone_cands:
                for c in phone_cands:
                    mismatches['name_phone_candidates'].append({
                        'calendar_subject': appt.get('subject',''),
                        'calendar_start': appt.get('start_time'),
                        'phones': ', '.join(appt.get('phone_numbers', [])),
                        **c
                    })
            else:
                # Try datetime proximity match
                datetime_cands = find_candidates_by_datetime(cur, appt.get('start_time'))
                if datetime_cands:
                    for c in datetime_cands:
                        mismatches['name_phone_candidates'].append({
                            'calendar_subject': appt.get('subject',''),
                            'calendar_start': appt.get('start_time'),
                            'phones': ', '.join(appt.get('phone_numbers', [])),
                            **c
                        })
                else:
                    # Fallback to name token matching
                    name_cands = find_candidates_by_name(cur, appt.get('subject',''), appt.get('start_time'))
                    for c in name_cands:
                        mismatches['name_phone_candidates'].append({
                            'calendar_subject': appt.get('subject',''),
                            'calendar_start': appt.get('start_time'),
                            'phones': ', '.join(appt.get('phone_numbers', [])),
                            **c
                        })
            continue
        
        matched_count += 1
        (charter_id, charter_reserve, charter_driver_id, charter_driver_name, 
         notes, booking_notes, client_notes, driver_notes,
         charter_date, pickup_time, status, vehicle, vehicle_id,
         pickup_address, dropoff_address) = charter
        
        # Extract vehicle code(s) from calendar subject/title
        # Handle patterns like: "L-24 or 11 or 10", "L24/L23", "L 7, 8"
        # Normalize to L<num> within bounds 1..25; keep all options, choose first as primary
        subject = appt.get('subject', '')

        def parse_vehicle_options(text: str):
            if not text:
                return []

            options = []

            # First, look for an explicit L-coded vehicle to anchor the list
            first_match = re.search(r"\bL\s*-?\s*(\d{1,2})\b", text, re.IGNORECASE)
            search_start = 0
            if first_match:
                try:
                    n = int(first_match.group(1))
                    if 1 <= n <= 25:
                        options.append(f"L{n}")
                except Exception:
                    pass
                search_start = first_match.end()

                # Collect trailing "or <num>" or ", <num>" items (with or without L)
                for m in re.finditer(r"\b(?:or|/|,|and)\b\s*(?:L\s*-?\s*)?(\d{1,2})\b", text[search_start:], re.IGNORECASE):
                    n2 = int(m.group(1))
                    if 1 <= n2 <= 25:
                        code = f"L{n2}"
                        if code not in options:
                            options.append(code)
            else:
                # Fallback: look for a compact list like "L-24/L-23/L-22" anywhere
                for m in re.finditer(r"\bL\s*-?\s*(\d{1,2})\b", text, re.IGNORECASE):
                    n3 = int(m.group(1))
                    if 1 <= n3 <= 25:
                        code = f"L{n3}"
                        if code not in options:
                            options.append(code)

                # As a conservative fallback, if we saw at least two small bare numbers separated by 'or',
                # treat them as vehicle candidates (avoids picking times/years)
                bare_numbers = []
                for m in re.finditer(r"\bor\b\s*(\d{1,2})\b", text, re.IGNORECASE):
                    n4 = int(m.group(1))
                    if 1 <= n4 <= 25:
                        code = f"L{n4}"
                        if code not in bare_numbers:
                            bare_numbers.append(code)
                # Prepend the first bare number if any and not already included
                for code in bare_numbers:
                    if code not in options:
                        options.append(code)

            return options

        vehicle_options = parse_vehicle_options(subject)
        vehicle_from_calendar = vehicle_options[0] if vehicle_options else None
        
        # Extract client name from subject (usually formatted like "Client Name - AC 177")
        client_from_calendar = None
        if subject:
            # Remove vehicle info and get client name
            client_match = re.match(r'^(.+?)(?:\s*-\s*(?:AC|Unit)?\s*\d{2,3})?$', subject)
            if client_match:
                client_from_calendar = client_match.group(1).strip()
        
        # Detect cancellation from calendar
        is_cancelled_in_calendar = False
        cancellation_note = None
        for field in ['location', 'subject', 'body', 'categories']:
            text = str(appt.get(field, '')).lower()
            if 'cancel' in text:
                is_cancelled_in_calendar = True
                # Extract the specific text mentioning cancellation
                if field == 'location' and 'cancel' in appt.get('location', '').lower():
                    cancellation_note = appt.get('location', '')
                break
        
        # Build calendar info text for booking_notes (dispatcher notes)
        calendar_info_parts = []
        
        # Add cancellation notice if detected
        if is_cancelled_in_calendar:
            if cancellation_note:
                calendar_info_parts.append(f"🚫 CANCELLED (Calendar: {cancellation_note})")
            else:
                calendar_info_parts.append(f"🚫 CANCELLED (per calendar)")
        
        # Add merge notice if applicable
        if appt.get('is_merged') and appt.get('merged_count', 0) > 1:
            calendar_info_parts.append(f"[WARN] MERGED FROM {appt['merged_count']} SPLIT CALENDAR ENTRIES (past midnight)")
        
        # Add client name from calendar
        if client_from_calendar:
            calendar_info_parts.append(f"Client (Calendar): {client_from_calendar}")
        
        # Add vehicle info from calendar (show all options if present)
        if vehicle_options:
            if len(vehicle_options) == 1:
                calendar_info_parts.append(f"Vehicle (Calendar): {vehicle_from_calendar}")
            else:
                calendar_info_parts.append(
                    f"Vehicle options (Calendar): {', '.join(vehicle_options)}"
                )
        
        # Always include subject for full context
        if appt.get('subject'):
            calendar_info_parts.append(f"Subject: {appt['subject']}")

        if appt.get('start_time'):
            calendar_info_parts.append(f"Start: {appt['start_time']}")
        if appt.get('end_time'):
            calendar_info_parts.append(f"End: {appt['end_time']}")
        if appt.get('location'):
            calendar_info_parts.append(f"Location: {appt['location']}")
        if appt.get('driver_name'):
            calendar_info_parts.append(f"Driver (Calendar): {appt['driver_name']}")
        
        # Add phone numbers
        if appt.get('phone_numbers'):
            phones_str = ', '.join(appt['phone_numbers'])
            calendar_info_parts.append(f"Phone: {phones_str}")
        
        # Add passenger info
        if appt.get('passenger_info'):
            pax_info = appt['passenger_info']
            if pax_info.get('count'):
                calendar_info_parts.append(f"Passengers: {pax_info['count']}")

        # Additional Outlook metadata for completeness
        if appt.get('all_day_event') is not None:
            calendar_info_parts.append(f"All-day: {bool(appt['all_day_event'])}")
        if appt.get('categories'):
            calendar_info_parts.append(f"Categories: {appt['categories']}")
        if appt.get('organizer'):
            calendar_info_parts.append(f"Organizer: {appt['organizer']}")
        if appt.get('importance') is not None:
            imp_map = {0: 'Low', 1: 'Normal', 2: 'High'}
            calendar_info_parts.append(f"Importance: {imp_map.get(appt['importance'], appt['importance'])}")
        if appt.get('reminder_set') is not None:
            calendar_info_parts.append(f"Reminder set: {bool(appt['reminder_set'])}")
        
        # Add pickup/dropoff if available
        if pickup_address:
            calendar_info_parts.append(f"Pickup: {pickup_address}")
        if dropoff_address:
            calendar_info_parts.append(f"Dropoff: {dropoff_address}")
        
        if appt.get('body'):
            # Include full body text (booking_notes is TEXT type, unlimited)
            body_text = appt['body']
            calendar_info_parts.append(f"Details: {body_text}")
        
        calendar_info = '\n'.join(calendar_info_parts)
        
        # Update booking_notes (dispatcher notes), vehicle, and cancellation status
        updates_needed = []
        update_values = []
        
        if calendar_info:
            new_booking_notes = merge_dispatcher_notes(booking_notes, calendar_info)
            
            if new_booking_notes != booking_notes:
                updates_needed.append("booking_notes = %s")
                update_values.append(new_booking_notes)
                updated_notes_count += 1
        
        # Update cancelled status if detected in calendar and not already cancelled in DB
        if is_cancelled_in_calendar:
            current_status = str(charter[10] or '').lower()
            if 'cancel' not in current_status:
                updates_needed.append("cancelled = TRUE")
                updates_needed.append("status = %s")
                update_values.append('cancelled')
                updated_cancelled_count += 1
        
        # Update vehicle if calendar has vehicle and DB missing or incorrectly set to flight (e.g., 'AC 177')
        if vehicle_from_calendar:
            vehicle_str = (vehicle or '').strip() if vehicle is not None else ''
            is_flight_value = bool(re.match(r'^AC\s*\d{2,4}$', vehicle_str, re.IGNORECASE))
            # Treat non-fleet strings as missing (e.g., empty, AC ###, misc text)
            is_fleet_value = bool(re.match(r'^L\s*-?\s*\d{1,2}$', vehicle_str, re.IGNORECASE))
            if (not vehicle_str) or is_flight_value or (not is_fleet_value):
                updates_needed.append("vehicle = %s")
                update_values.append(vehicle_from_calendar)
        
        # Record missing vehicle cases for review if no vehicle could be determined
        if not vehicle_from_calendar:
            vehicle_str = (vehicle or '').strip() if vehicle is not None else ''
            if not vehicle_str:
                mismatches['missing_vehicle'].append({
                    'reserve_number': reserve_num,
                    'charter_date': str(charter_date) if charter_date else None,
                    'calendar_subject': subject,
                    'calendar_location': appt.get('location', '')
                })

        # Apply updates if any
        if write and updates_needed:
            update_values.append(charter_id)
            update_sql = f"""
                UPDATE charters
                SET {', '.join(updates_needed)},
                    updated_at = CURRENT_TIMESTAMP
                WHERE charter_id = %s
            """
            cur.execute(update_sql, update_values)
        
        # Driver reconciliation
        calendar_driver = appt.get('driver_name')
        
        if calendar_driver:
            # Try to match calendar driver to employee
            driver_match = fuzzy_match_driver(calendar_driver, employees)
            
            if driver_match:
                matched_emp_id, matched_emp_name, confidence = driver_match
                
                # Check if charter already has a different driver assigned
                if charter_driver_id and charter_driver_id != matched_emp_id:
                    # Driver mismatch
                    mismatches['driver_mismatch'].append({
                        'reserve_number': reserve_num,
                        'charter_date': str(charter_date) if charter_date else None,
                        'db_driver_id': charter_driver_id,
                        'db_driver_name': charter_driver_name,
                        'calendar_driver': calendar_driver,
                        'matched_employee_id': matched_emp_id,
                        'matched_employee_name': matched_emp_name,
                        'confidence': confidence
                    })
                elif not charter_driver_id:
                    # No driver assigned in DB, we can assign
                    if write:
                        cur.execute("""
                            UPDATE charters
                            SET assigned_driver_id = %s,
                                driver_name = %s,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE charter_id = %s
                        """, (matched_emp_id, matched_emp_name, charter_id))
                        updated_driver_count += 1
                    else:
                        mismatches['missing_db_driver'].append({
                            'reserve_number': reserve_num,
                            'charter_date': str(charter_date) if charter_date else None,
                            'calendar_driver': calendar_driver,
                            'matched_employee_id': matched_emp_id,
                            'matched_employee_name': matched_emp_name,
                            'confidence': confidence
                        })
            else:
                # Could not match calendar driver to any employee
                mismatches['missing_calendar_driver'].append({
                    'reserve_number': reserve_num,
                    'charter_date': str(charter_date) if charter_date else None,
                    'calendar_driver': calendar_driver,
                    'db_driver_name': charter_driver_name
                })
        elif charter_driver_id:
            # Charter has driver but calendar doesn't mention one
            pass  # Not necessarily a problem
    
    # For appointments with no reserve at all, try suggesting candidates as well
    without_reserve = [a for a in appointments if not a.get('reserve_number')]
    for appt in without_reserve:
        # Priority: 1) phone, 2) datetime proximity, 3) name tokens
        phone_cands = find_candidates_by_phone(cur, appt.get('phone_numbers', []), appt.get('start_time'))
        if phone_cands:
            for c in phone_cands:
                mismatches['name_phone_candidates'].append({
                    'calendar_subject': appt.get('subject',''),
                    'calendar_start': appt.get('start_time'),
                    'phones': ', '.join(appt.get('phone_numbers', [])),
                    **c
                })
        else:
            # Try datetime proximity match
            datetime_cands = find_candidates_by_datetime(cur, appt.get('start_time'))
            if datetime_cands:
                for c in datetime_cands:
                    mismatches['name_phone_candidates'].append({
                        'calendar_subject': appt.get('subject',''),
                        'calendar_start': appt.get('start_time'),
                        'phones': ', '.join(appt.get('phone_numbers', [])),
                        **c
                    })
            else:
                # Fallback to name token matching
                name_cands = find_candidates_by_name(cur, appt.get('subject',''), appt.get('start_time'))
                for c in name_cands:
                    mismatches['name_phone_candidates'].append({
                        'calendar_subject': appt.get('subject',''),
                        'calendar_start': appt.get('start_time'),
                        'phones': ', '.join(appt.get('phone_numbers', [])),
                        **c
                    })

    # Commit changes if write mode
    if write:
        conn.commit()
        print(f"\n✓ COMMITTED CHANGES TO DATABASE")
    else:
        conn.rollback()
        print(f"\n✓ DRY RUN - No changes committed")
    
    cur.close()
    conn.close()
    
    # Print summary
    print(f"\n=== MATCHING SUMMARY ===")
    print(f"Appointments with reserve numbers: {len(with_reserve)}")
    print(f"Matched to charters: {matched_count}")
    print(f"Dispatcher notes updated: {updated_notes_count}")
    print(f"Drivers assigned/updated: {updated_driver_count}")
    print(f"Charters marked cancelled: {updated_cancelled_count}")
    
    print(f"\n=== ISSUES FOUND ===")
    print(f"No charter found: {len(mismatches['no_charter_found'])}")
    print(f"Driver mismatches: {len(mismatches['driver_mismatch'])}")
    print(f"Missing DB driver (would assign): {len(mismatches['missing_db_driver'])}")
    print(f"Missing calendar driver: {len(mismatches['missing_calendar_driver'])}")
    print(f"Name/Phone candidate suggestions: {len(mismatches['name_phone_candidates'])}")
    
    return mismatches


def generate_excel_report(mismatches, output_file):
    """Generate Excel workbook with mismatch sheets."""
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # Sheet 1: No Charter Found
    if mismatches['no_charter_found']:
        ws = wb.create_sheet("No Charter Found")
        ws.append(['Reserve Number', 'Calendar Subject', 'Calendar Date', 'Calendar Driver'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        for item in mismatches['no_charter_found']:
            ws.append([
                item['reserve_number'],
                item['calendar_subject'],
                item['calendar_date'],
                item['calendar_driver']
            ])
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 2: Driver Mismatches
    if mismatches['driver_mismatch']:
        ws = wb.create_sheet("Driver Mismatches")
        ws.append(['Reserve Number', 'Charter Date', 'DB Driver ID', 'DB Driver Name', 
                   'Calendar Driver', 'Matched Employee ID', 'Matched Employee Name', 'Confidence'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        for item in mismatches['driver_mismatch']:
            ws.append([
                item['reserve_number'],
                item['charter_date'],
                item['db_driver_id'],
                item['db_driver_name'],
                item['calendar_driver'],
                item['matched_employee_id'],
                item['matched_employee_name'],
                f"{item['confidence']:.1%}"
            ])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 3: Missing DB Driver (can assign)
    if mismatches['missing_db_driver']:
        ws = wb.create_sheet("Can Assign Driver")
        ws.append(['Reserve Number', 'Charter Date', 'Calendar Driver', 
                   'Matched Employee ID', 'Matched Employee Name', 'Confidence'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        for item in mismatches['missing_db_driver']:
            ws.append([
                item['reserve_number'],
                item['charter_date'],
                item['calendar_driver'],
                item['matched_employee_id'],
                item['matched_employee_name'],
                f"{item['confidence']:.1%}"
            ])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Sheet 4: Missing Calendar Driver
    if mismatches['missing_calendar_driver']:
        ws = wb.create_sheet("Unmatched Calendar Drivers")
        ws.append(['Reserve Number', 'Charter Date', 'Calendar Driver', 'DB Driver Name'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        for item in mismatches['missing_calendar_driver']:
            ws.append([
                item['reserve_number'],
                item['charter_date'],
                item['calendar_driver'],
                item['db_driver_name']
            ])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Sheet 5: Missing Vehicle (no calendar determination and DB empty)
    if mismatches.get('missing_vehicle'):
        ws = wb.create_sheet("Missing Vehicle")
        ws.append(['Reserve Number', 'Charter Date', 'Subject', 'Location'])

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for item in mismatches['missing_vehicle']:
            ws.append([
                item['reserve_number'],
                item['charter_date'],
                item['calendar_subject'],
                item['calendar_location']
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Sheet 6: Name/Phone/DateTime Candidate Matches
    if mismatches.get('name_phone_candidates'):
        ws = wb.create_sheet("Name-Phone-DateTime Candidates")
        ws.append(['Calendar Subject', 'Calendar Start', 'Phones', 'Candidate Reserve', 'Charter Date', 'Pickup Time', 'Client Name', 'Pickup Address', 'Match Type', 'Match Value'])

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for item in mismatches['name_phone_candidates']:
            ws.append([
                item.get('calendar_subject',''),
                item.get('calendar_start',''),
                item.get('phones',''),
                item.get('reserve_number',''),
                item.get('charter_date',''),
                item.get('pickup_time',''),
                item.get('client_name',''),
                item.get('pickup_address',''),
                item.get('match_type',''),
                item.get('match_value',''),
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Excel report saved: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Match Outlook calendar to charters')
    parser.add_argument('--input', required=True,
                        help='Input JSON file from extract_outlook_calendar.py')
    parser.add_argument('--write', action='store_true',
                        help='Apply changes to database (default: dry-run)')
    parser.add_argument('--excel', default='reports/outlook_calendar_mismatches.xlsx',
                        help='Output Excel file for mismatches')
    parser.add_argument('--json', default='reports/outlook_calendar_mismatches.json',
                        help='Also write mismatches JSON for UI workflows')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)
    
    # Process matches and updates
    mismatches = match_and_update_charters(args.input, dry_run=not args.write, write=args.write)
    
    # Generate Excel report if any mismatches
    total_issues = sum(len(v) for v in mismatches.values())
    if total_issues > 0:
        generate_excel_report(mismatches, args.excel)
        # Also write JSON for desktop UI to act on (e.g., open Add Client form)
        try:
            os.makedirs(os.path.dirname(args.json), exist_ok=True)
            with open(args.json, 'w', encoding='utf-8') as jf:
                json.dump(mismatches, jf, indent=2, ensure_ascii=False)
            print(f"\n✓ JSON mismatches saved: {args.json}")
        except Exception as e:
            print(f"WARNING: Failed to write JSON mismatches: {e}")
    else:
        print("\n✓ No mismatches found - no Excel report needed")


if __name__ == '__main__':
    main()
