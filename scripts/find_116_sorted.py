#!/usr/bin/env python3
"""Sort Scotia file by amount and find $116.00 entry."""

import openpyxl
import pandas as pd

xlsx_path = r"l:\limo\data\2012_scotia_transactions_for_editing.xlsx"

print("=" * 80)
print("FOUND $116.00 ENTRY WITH MISSING DATE")
print("=" * 80)

# Load with pandas
df = pd.read_excel(xlsx_path, sheet_name=0)

# Find the $116.00 entry
matches = df[df['debit/withdrawal'] == 116.0]
print(f"\nFound {len(matches)} entry with $116.00 debit:")
for idx, row in matches.iterrows():
    print(f"\n  Row {idx + 2}:")
    print(f"    Date: {row['date']} (TYPE: {type(row['date']).__name__})")
    print(f"    Description: {row['Description']}")
    print(f"    Debit/Withdrawal: ${row['debit/withdrawal']:.2f}")
    print(f"    Deposit/Credit: {row['deposit/credit']}")
    print(f"    Balance: ${row['balance']:.2f}")

print(f"\n" + "=" * 80)
print("ANALYSIS")
print("=" * 80)
print(f"\n⚠️  This entry has:")
print(f"  - MISSING DATE (NaT)")
print(f"  - MISSING DESCRIPTION")
print(f"  - Amount: $116.00 (debit/withdrawal)")
print(f"  - Balance after: $1,460.14")

print(f"\n" + "=" * 80)
print("SORTING BY AMOUNT TO SHOW CONTEXT")
print("=" * 80)

# Sort by debit/withdrawal amount
df_sorted = df.sort_values('debit/withdrawal', ascending=False, na_position='last')

print(f"\nTop 20 debit entries by amount:")
for i, (idx, row) in enumerate(df_sorted.head(20).iterrows()):
    debit = row['debit/withdrawal'] if pd.notna(row['debit/withdrawal']) else 0
    date_str = str(row['date']) if pd.notna(row['date']) else "MISSING"
    desc = row['Description'] if pd.notna(row['Description']) else "MISSING"
    print(f"  {i+1:2d}. ${debit:>8.2f} | {date_str:12s} | {desc}")

# Save sorted file
print(f"\n" + "=" * 80)
print("SAVING SORTED FILE")
print("=" * 80)

# Load workbook and sort
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

# Read all data
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(row)

# Sort by debit/withdrawal (column index 2)
data_sorted = sorted(data, key=lambda x: (pd.isna(x[2]), -float(x[2]) if pd.notna(x[2]) else 0))

# Clear and rewrite
ws.delete_rows(2, ws.max_row)
for row_data in data_sorted:
    ws.append(row_data)

output_path = r"l:\limo\data\2012_scotia_transactions_for_editing_SORTED_BY_AMOUNT.xlsx"
wb.save(output_path)
print(f"✓ Saved: {output_path}")
print(f"\nYou can now open this file and manually find the $116.00 entry")
print(f"to fill in the missing DATE and DESCRIPTION.")

wb.close()

PYSCRIPT
