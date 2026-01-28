"""
CIBC 8362 2018 - COLUMN-AWARE PARSER
Properly handles Withdrawals ($) vs Deposits ($) columns
"""
import re
from pathlib import Path
from decimal import Decimal
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

pdf_path = Path(r"L:\limo\pdf\2018\CIBC 8362 2018.pdf")
output = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

trans = []
opening = None
month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
             'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}

print("EXTRACTING CIBC 8362 2018 - COLUMN-AWARE PARSER + BALANCE AUDIT")
print("=" * 80)

with pdfplumber.open(pdf_path) as pdf:
    for pg_num, pg in enumerate(pdf.pages, 1):
        txt = pg.extract_text() or ""
        lines = txt.split('\n')
        
        # Find opening balance
        if opening is None:
            for line in lines:
                if 'forward' in line.lower():
                    m = re.search(r'\$?([\d,]+\.\d{2})', line)
                    if m:
                        opening = Decimal(m.group(1).replace(',', ''))
                        print(f"Page {pg_num}: Opening balance ${opening:,.2f}")
                        break
        
        # Process line by line, accumulating multi-line descriptions
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Start of transaction: Month Day Description...
            m = re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\s+(.+)', line, re.I)
            if not m:
                i += 1
                continue
            
            month_str, day, rest = m.groups()
            month = month_map[month_str.lower()]
            
            # Accumulate continuation lines
            full_text = rest
            i += 1
            while i < len(lines):
                next_line = lines[i].strip()
                if not next_line or re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d', next_line, re.I):
                    break
                if any(skip in next_line for skip in ['CIBC', 'Account', 'Page ', 'Transaction details']):
                    i += 1
                    break
                full_text += " " + next_line
                i += 1
            
            # Clean whitespace
            full_text = re.sub(r'\s+', ' ', full_text).strip()
            
            # Extract numbers
            nums = re.findall(r'([\d,]+\.\d{2})', full_text)
            if not nums:
                continue
            
            # Pattern: Description [Withdrawal] [Deposit] ~ Balance
            # OR: Description Amount ~ Balance
            
            # Find balance (after ~)
            balance = None
            if '~' in full_text:
                parts = full_text.split('~')
                balance_part = parts[-1].strip()
                bal_match = re.search(r'([\d,]+\.\d{2})', balance_part)
                if bal_match:
                    balance = Decimal(bal_match.group(1).replace(',', ''))
                
                # Everything before ~ contains description + amounts
                pre_tilde = parts[0]
                amounts = re.findall(r'([\d,]+\.\d{2})', pre_tilde)
            else:
                # No tilde: balance is last number
                if len(nums) >= 2:
                    balance = Decimal(nums[-1].replace(',', ''))
                    amounts = nums[:-1]
                else:
                    continue
            
            if not balance or not amounts:
                continue
            
            # Determine withdrawal vs deposit
            # If there's one amount: check keywords
            # If there are two amounts: first is withdrawal, second is deposit
            withdrawal = None
            deposit = None
            
            if len(amounts) == 2:
                # Likely: withdrawal deposit columns
                withdrawal = Decimal(amounts[0].replace(',', ''))
                deposit = Decimal(amounts[1].replace(',', ''))
            elif len(amounts) == 1:
                # Single amount: determine type by keywords
                amt = Decimal(amounts[0].replace(',', ''))
                if any(kw in full_text.upper() for kw in [
                    'DEPOSIT', 'CREDIT', 'E-TRANSFER', 'SQUARE', 'PAYROLL'
                ]):
                    deposit = amt
                else:
                    withdrawal = amt
            else:
                # Multiple amounts: take last two before balance
                withdrawal = Decimal(amounts[-2].replace(',', ''))
                deposit = Decimal(amounts[-1].replace(',', ''))
            
            # Calculate net amount (deposit positive, withdrawal negative)
            if withdrawal and not deposit:
                amount = -withdrawal
            elif deposit and not withdrawal:
                amount = deposit
            elif withdrawal and deposit:
                # Both present: net is deposit - withdrawal (unusual but handle it)
                amount = deposit - withdrawal
            else:
                continue
            
            # Description is everything before first number
            desc_end = full_text.find(amounts[0])
            desc = full_text[:desc_end].strip() if desc_end > 0 else full_text[:60]
            
            trans.append({
                'date': f'2018-{month:02d}-{int(day):02d}',
                'desc': desc[:100],
                'withdrawal': withdrawal if withdrawal else Decimal('0'),
                'deposit': deposit if deposit else Decimal('0'),
                'amount': amount,
                'balance': balance,
                'page': pg_num
            })

print(f"\nExtracted {len(trans)} transactions")

if not trans:
    print("No transactions found")
    exit(1)

# Audit balance
print("\nAUDITING BALANCES...")
print("=" * 80)

# Infer opening if needed
if opening is None and trans:
    opening = trans[0]['balance'] - trans[0]['amount']
    print(f"Inferred opening balance: ${opening:,.2f}")

calc = opening
errors = []

for i, t in enumerate(trans):
    calc += t['amount']
    diff = abs(calc - t['balance'])
    t['calc'] = calc
    t['diff'] = diff
    t['err'] = diff > Decimal('0.01')
    if t['err']:
        errors.append(i)

print(f"Opening: ${opening:,.2f}")
print(f"Closing (stated): ${trans[-1]['balance']:,.2f}")
print(f"Closing (calculated): ${calc:,.2f}")
print(f"Balance errors: {len(errors)}")

if errors:
    print(f"\nFirst 5 errors:")
    for idx in errors[:5]:
        t = trans[idx]
        print(f"  {t['date']} | {t['desc'][:40]:<40} | Amt: ${t['amount']:>10,.2f} | Stated: ${t['balance']:>10,.2f} | Calc: ${t['calc']:>10,.2f} | Diff: ${t['diff']:>8,.2f}")

# Write Excel
wb = Workbook()
ws = wb.active
ws.title = "Transactions + Audit"
ws.append(['Date', 'Description', 'Withdrawal', 'Deposit', 'Net Amount', 'Stated Bal', 'Calc Bal', 'Diff', 'Error', 'Page'])

for c in range(1, 11):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for t in trans:
    row_data = [
        t['date'], t['desc'],
        float(t['withdrawal']), float(t['deposit']), float(t['amount']),
        float(t['balance']), float(t['calc']), float(t['diff']),
        'YES' if t['err'] else '', t['page']
    ]
    ws.append(row_data)
    if t['err']:
        for c in range(1, 11):
            ws.cell(ws.max_row, c).fill = err_fill

for c in [3, 4, 5, 6, 7, 8]:
    for r in range(2, len(trans) + 2):
        ws.cell(r, c).number_format = '$#,##0.00'

ws.column_dimensions['B'].width = 50

# Summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "CIBC 8362 2018 Extraction Summary"
summary['A1'].font = Font(bold=True, size=14)
summary['A3'] = "Opening Balance:"
summary['B3'] = float(opening)
summary['B3'].number_format = '$#,##0.00'
summary['A4'] = "Total Transactions:"
summary['B4'] = len(trans)
summary['A5'] = "Total Deposits:"
summary['B5'] = float(sum(t['amount'] for t in trans if t['amount'] > 0))
summary['B5'].number_format = '$#,##0.00'
summary['A6'] = "Total Withdrawals:"
summary['B6'] = float(abs(sum(t['amount'] for t in trans if t['amount'] < 0)))
summary['B6'].number_format = '$#,##0.00'
summary['A7'] = "Closing (stated):"
summary['B7'] = float(trans[-1]['balance']) if trans else 0
summary['B7'].number_format = '$#,##0.00'
summary['A8'] = "Closing (calculated):"
summary['B8'] = float(calc)
summary['B8'].number_format = '$#,##0.00'
summary['A9'] = "Balance Errors:"
summary['B9'] = len(errors)
if errors:
    summary['B9'].font = Font(color="FF0000", bold=True)

wb.save(output)
print(f"\n✅ Saved: {output}")
print(f"   Review highlighted rows for balance discrepancies")
