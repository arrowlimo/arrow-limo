"""
Comprehensive Receipt-to-Banking Reconciliation with Excel Export
==================================================================

This script:
1. Matches receipts to banking statements (amount, date, vendor fuzzy matching)
2. Detects duplicate receipts
3. Generates color-coded Excel workbook with all receipt columns
4. Supports manual editing and re-import workflow

Color Coding:
- GREEN: Matched to banking statement
- YELLOW: Potential duplicate receipt
- RED: Unmatched to banking (needs review)
- BLUE: Marked as personal expense
- PURPLE: Split receipt (parent or child)
- ORANGE: Needs vehicle/card/fuel data

Author: AI Agent
Date: December 19, 2025
"""

import psycopg2
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from difflib import SequenceMatcher
import re
import argparse

# Database configuration
DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "***REMOVED***")

# Color scheme
COLOR_MATCHED = "C6EFCE"  # Green
COLOR_DUPLICATE = "FFEB9C"  # Yellow
COLOR_UNMATCHED = "FFC7CE"  # Red
COLOR_PERSONAL = "D9E1F2"  # Blue
COLOR_SPLIT = "E4DFEC"  # Purple
COLOR_NEEDS_DATA = "FCE4D6"  # Orange

def fuzzy_match_vendor(vendor1, vendor2, threshold=0.8):
    """Fuzzy match two vendor names."""
    if not vendor1 or not vendor2:
        return 0.0
    
    # Normalize
    v1 = re.sub(r'[^a-z0-9]', '', vendor1.lower())
    v2 = re.sub(r'[^a-z0-9]', '', vendor2.lower())
    
    # Check if one contains the other
    if v1 in v2 or v2 in v1:
        return 1.0
    
    # Sequence matcher
    return SequenceMatcher(None, v1, v2).ratio()


def get_receipts_data(conn, year_start=2007, year_end=2025):
    """Fetch all receipts with their current banking linkage."""
    query = """
        SELECT 
            r.receipt_id,
            r.receipt_date,
            r.vendor_name,
            r.canonical_vendor,
            r.description,
            r.gross_amount,
            r.gst_amount,
            r.net_amount,
            r.expense_account,
            r.payment_method,
            r.canonical_pay_method,
            r.card_type,
            r.card_number,
            r.vehicle_id,
            r.vehicle_number,
            r.fuel_amount,
            r.fuel,
            r.category,
            r.classification,
            r.sub_classification,
            r.banking_transaction_id,
            r.created_from_banking,
            r.is_personal_purchase,
            r.business_personal,
            r.is_split_receipt,
            r.parent_receipt_id,
            r.is_driver_reimbursement,
            r.reimbursed_via,
            r.reimbursement_date,
            r.cash_box_transaction_id,
            r.source_system,
            r.source_reference,
            r.source_file,
            r.validation_status,
            r.comment,
            r.expense,
            r.revenue,
            r.deductible_status,
            r.owner_personal_amount,
            r.currency,
            r.amount_usd,
            r.fx_rate,
            r.is_transfer,
            r.gl_account_code,
            r.gl_account_name,
            r.gl_subcategory,
            -- Banking transaction details if linked
            bt.transaction_date as banking_date,
            bt.description as banking_description,
            bt.debit_amount as banking_debit,
            bt.credit_amount as banking_credit,
            bt.account_number as banking_account,
            bt.vendor_extracted as banking_vendor,
            bt.category as banking_category,
            bt.source_file as banking_source_file,
            bt.reconciliation_status
        FROM receipts r
        LEFT JOIN banking_transactions bt ON r.banking_transaction_id = bt.transaction_id
        WHERE EXTRACT(YEAR FROM r.receipt_date) >= %s
        AND EXTRACT(YEAR FROM r.receipt_date) <= %s
        ORDER BY r.receipt_date, r.receipt_id
    """
    
    df = pd.read_sql_query(query, conn, params=(year_start, year_end))
    return df


def get_unmatched_banking_debits(conn, year_start=2007, year_end=2025):
    """Fetch unmatched banking debit transactions."""
    query = """
        SELECT 
            bt.transaction_id,
            bt.transaction_date,
            bt.description,
            bt.debit_amount,
            bt.account_number,
            bt.vendor_extracted,
            bt.category,
            bt.source_file,
            bt.balance
        FROM banking_transactions bt
        WHERE bt.debit_amount > 0
        AND EXTRACT(YEAR FROM bt.transaction_date) >= %s
        AND EXTRACT(YEAR FROM bt.transaction_date) <= %s
        AND NOT EXISTS (
            SELECT 1 FROM receipts r 
            WHERE r.banking_transaction_id = bt.transaction_id
        )
        AND NOT EXISTS (
            SELECT 1 FROM banking_receipt_matching_ledger brl
            WHERE brl.banking_transaction_id = bt.transaction_id
        )
        ORDER BY bt.transaction_date, bt.transaction_id
    """
    
    df = pd.read_sql_query(query, conn, params=(year_start, year_end))
    return df


def detect_duplicates(df):
    """Detect potential duplicate receipts."""
    # Group by date, amount, vendor
    duplicates = df.groupby(['receipt_date', 'gross_amount', 'vendor_name']).filter(
        lambda x: len(x) > 1 and x['vendor_name'].notna().all()
    )
    
    return set(duplicates['receipt_id'].tolist())


def attempt_matching(receipts_df, banking_df, date_tolerance=3):
    """Attempt to match unlinked receipts to banking transactions."""
    matches = []
    
    for _, receipt in receipts_df[receipts_df['banking_transaction_id'].isna()].iterrows():
        receipt_date = receipt['receipt_date']
        receipt_amount = float(receipt['gross_amount']) if pd.notna(receipt['gross_amount']) else 0
        receipt_vendor = receipt['vendor_name'] or receipt['canonical_vendor'] or ''
        
        # Find banking transactions within date range and amount match
        date_min = receipt_date - timedelta(days=date_tolerance)
        date_max = receipt_date + timedelta(days=date_tolerance)
        
        candidates = banking_df[
            (banking_df['transaction_date'] >= date_min) &
            (banking_df['transaction_date'] <= date_max) &
            (abs(banking_df['debit_amount'] - receipt_amount) < 0.01)
        ]
        
        # Try vendor matching
        best_match = None
        best_score = 0.0
        
        for _, banking in candidates.iterrows():
            banking_vendor = banking['vendor_extracted'] or banking['description'] or ''
            score = fuzzy_match_vendor(receipt_vendor, banking_vendor)
            
            if score > best_score:
                best_score = score
                best_match = banking
        
        if best_match is not None and best_score >= 0.6:
            matches.append({
                'receipt_id': receipt['receipt_id'],
                'banking_transaction_id': best_match['transaction_id'],
                'match_score': best_score,
                'date_diff': abs((receipt_date - best_match['transaction_date']).days)
            })
    
    return pd.DataFrame(matches)


def create_excel_workbook(receipts_df, duplicate_ids, output_file):
    """Create color-coded Excel workbook."""
    print(f"\nCreating Excel workbook: {output_file}")
    
    # Add withdrawal/deposit columns (receipts are always withdrawals, negative values)
    receipts_df['withdrawal'] = receipts_df['gross_amount'].abs()
    receipts_df['deposit'] = 0.0
    
    # Reorder columns to put withdrawal/deposit after gross_amount
    cols = list(receipts_df.columns)
    gross_idx = cols.index('gross_amount')
    # Insert withdrawal and deposit right after gross_amount
    new_cols = cols[:gross_idx+1] + ['withdrawal', 'deposit'] + [c for c in cols[gross_idx+1:] if c not in ['withdrawal', 'deposit']]
    receipts_df = receipts_df[new_cols]
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write receipts data
        receipts_df.to_excel(writer, sheet_name='Receipts', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Receipts']
        
        # Define fills
        fill_matched = PatternFill(start_color=COLOR_MATCHED, end_color=COLOR_MATCHED, fill_type="solid")
        fill_duplicate = PatternFill(start_color=COLOR_DUPLICATE, end_color=COLOR_DUPLICATE, fill_type="solid")
        fill_unmatched = PatternFill(start_color=COLOR_UNMATCHED, end_color=COLOR_UNMATCHED, fill_type="solid")
        fill_personal = PatternFill(start_color=COLOR_PERSONAL, end_color=COLOR_PERSONAL, fill_type="solid")
        fill_split = PatternFill(start_color=COLOR_SPLIT, end_color=COLOR_SPLIT, fill_type="solid")
        fill_needs_data = PatternFill(start_color=COLOR_NEEDS_DATA, end_color=COLOR_NEEDS_DATA, fill_type="solid")
        
        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply color coding to data rows
        for idx, row in receipts_df.iterrows():
            excel_row = idx + 2  # Excel is 1-indexed, +1 for header
            
            # Determine color
            fill = None
            
            if row['is_split_receipt'] or pd.notna(row['parent_receipt_id']):
                fill = fill_split
            elif row['is_personal_purchase'] or row['business_personal'] == 'personal':
                fill = fill_personal
            elif row['receipt_id'] in duplicate_ids:
                fill = fill_duplicate
            elif pd.notna(row['banking_transaction_id']):
                fill = fill_matched
            else:
                fill = fill_unmatched
            
            # Check if needs additional data (fuel receipts without vehicle)
            if pd.notna(row['fuel_amount']) and pd.isna(row['vehicle_id']):
                fill = fill_needs_data
            
            # Apply fill to entire row
            for cell in worksheet[excel_row]:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Auto-size columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Additional filtered sheets: Unmatched, Duplicates, Personal, Split, Summary
        def write_filtered_sheet(name, df, fill):
            if df.empty:
                return
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            # Header style
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # Row fill
            for r in range(2, len(df) + 2):
                for cell in ws[r]:
                    cell.fill = fill
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            # Auto-size
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        unmatched_df = receipts_df[receipts_df['banking_transaction_id'].isna()]
        duplicate_df = receipts_df[receipts_df['receipt_id'].isin(duplicate_ids)]
        personal_df = receipts_df[(receipts_df['is_personal_purchase'] == True) | (receipts_df['business_personal'] == 'personal')]
        split_df = receipts_df[(receipts_df['is_split_receipt'] == True) | (receipts_df['parent_receipt_id'].notna())]

        write_filtered_sheet('Unmatched', unmatched_df, fill_unmatched)
        write_filtered_sheet('Duplicates', duplicate_df, fill_duplicate)
        write_filtered_sheet('Personal', personal_df, fill_personal)
        write_filtered_sheet('Split', split_df, fill_split)

        # Summary sheet
        summary_rows = []
        summary_rows.append({'Metric': 'Total Receipts', 'Value': len(receipts_df)})
        linked_count = len(receipts_df[receipts_df['banking_transaction_id'].notna()])
        summary_rows.append({'Metric': 'Linked Receipts', 'Value': linked_count})
        summary_rows.append({'Metric': 'Unlinked Receipts', 'Value': len(unmatched_df)})
        summary_rows.append({'Metric': 'Duplicate Receipts', 'Value': len(duplicate_df)})
        summary_rows.append({'Metric': 'Personal Receipts', 'Value': len(personal_df)})
        summary_rows.append({'Metric': 'Split Receipts', 'Value': len(split_df)})
        # Year breakdown for unmatched
        if not unmatched_df.empty and 'receipt_date' in unmatched_df.columns:
            unmatched_df['receipt_date'] = pd.to_datetime(unmatched_df['receipt_date'])
            by_year = unmatched_df.groupby(unmatched_df['receipt_date'].dt.year).agg({'receipt_id': 'count', 'gross_amount': 'sum'})
            for year, row in by_year.iterrows():
                summary_rows.append({'Metric': f'Unlinked {int(year)}', 'Value': f"{int(row['receipt_id'])} receipts, ${row['gross_amount']:,.2f}"})
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        ws_sum = writer.sheets['Summary']
        for cell in ws_sum[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column in ws_sum.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_sum.column_dimensions[column_letter].width = adjusted_width

        # Add legend worksheet
        legend_data = {
            'Color': ['Green', 'Yellow', 'Red', 'Blue', 'Purple', 'Orange'],
            'Meaning': [
                'Matched to banking statement',
                'Potential duplicate',
                'Unmatched - needs review',
                'Personal expense',
                'Split receipt',
                'Needs vehicle/card/fuel data'
            ],
            'Action Required': [
                'Verify match is correct',
                'Review if legitimate recurring payment or true duplicate',
                'Find matching banking transaction or add receipt manually',
                'Confirm personal vs business classification',
                'Review split allocation',
                'Add missing vehicle, card number, or fuel volume'
            ]
        }
        
        legend_df = pd.DataFrame(legend_data)
        legend_df.to_excel(writer, sheet_name='Legend', index=False)
        
        legend_ws = writer.sheets['Legend']
        
        # Color the legend color column
        legend_ws['A2'].fill = PatternFill(start_color=COLOR_MATCHED, end_color=COLOR_MATCHED, fill_type="solid")
        legend_ws['A3'].fill = PatternFill(start_color=COLOR_DUPLICATE, end_color=COLOR_DUPLICATE, fill_type="solid")
        legend_ws['A4'].fill = PatternFill(start_color=COLOR_UNMATCHED, end_color=COLOR_UNMATCHED, fill_type="solid")
        legend_ws['A5'].fill = PatternFill(start_color=COLOR_PERSONAL, end_color=COLOR_PERSONAL, fill_type="solid")
        legend_ws['A6'].fill = PatternFill(start_color=COLOR_SPLIT, end_color=COLOR_SPLIT, fill_type="solid")
        legend_ws['A7'].fill = PatternFill(start_color=COLOR_NEEDS_DATA, end_color=COLOR_NEEDS_DATA, fill_type="solid")
        
        # Format legend
        for cell in legend_ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for col in legend_ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            legend_ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)
    
    print(f"✅ Excel file created: {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Generate receipt reconciliation Excel report')
    parser.add_argument('--year', type=int, help='Filter by specific year (e.g., 2012)')
    parser.add_argument('--year-start', type=int, default=2007, help='Start year (default: 2007)')
    parser.add_argument('--year-end', type=int, default=2025, help='End year (default: 2025)')
    args = parser.parse_args()
    
    # Determine year range
    if args.year:
        year_start = args.year
        year_end = args.year
        year_label = str(args.year)
    else:
        year_start = args.year_start
        year_end = args.year_end
        year_label = f"{year_start}-{year_end}" if year_start != year_end else str(year_start)
    
    print("="*70)
    print(f"RECEIPT-TO-BANKING RECONCILIATION & EXCEL EXPORT ({year_label})")
    print("="*70)
    
    # Connect to database
    print("\n1. Connecting to database...")
    conn = psycopg2.connect(
        host=DB_HOST,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )
    
    # Fetch receipts data
    print("2. Fetching receipts data...")
    receipts_df = get_receipts_data(conn, year_start=year_start, year_end=year_end)
    print(f"   Found {len(receipts_df):,} receipts")
    
    # Fetch unmatched banking
    print("3. Fetching unmatched banking transactions...")
    banking_df = get_unmatched_banking_debits(conn, year_start=year_start, year_end=year_end)
    print(f"   Found {len(banking_df):,} unmatched banking debits")
    
    # Detect duplicates
    print("4. Detecting duplicate receipts...")
    duplicate_ids = detect_duplicates(receipts_df)
    print(f"   Found {len(duplicate_ids):,} potential duplicate receipts")
    
    # Attempt matching (dry-run)
    print("5. Attempting to match unlinked receipts to banking...")
    matches_df = attempt_matching(receipts_df, banking_df)
    print(f"   Found {len(matches_df):,} potential new matches")
    
    if len(matches_df) > 0:
        print(f"   Match score distribution:")
        print(f"     Excellent (>0.9): {len(matches_df[matches_df['match_score'] >= 0.9]):,}")
        print(f"     Good (0.8-0.9): {len(matches_df[(matches_df['match_score'] >= 0.8) & (matches_df['match_score'] < 0.9)]):,}")
        print(f"     Fair (0.6-0.8): {len(matches_df[(matches_df['match_score'] >= 0.6) & (matches_df['match_score'] < 0.8)]):,}")
    
    # Generate Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"l:\\limo\\reports\\receipts_reconciliation_{year_label}_{timestamp}.xlsx"
    
    print(f"\n6. Generating Excel workbook...")
    create_excel_workbook(receipts_df, duplicate_ids, output_file)
    
    # Summary statistics
    print("\n" + "="*70)
    print("SUMMARY STATISTICS")
    print("="*70)
    
    total_receipts = len(receipts_df)
    linked_receipts = len(receipts_df[receipts_df['banking_transaction_id'].notna()])
    unlinked_receipts = total_receipts - linked_receipts
    personal_receipts = len(receipts_df[
        (receipts_df['is_personal_purchase'] == True) | 
        (receipts_df['business_personal'] == 'personal')
    ])
    split_receipts = len(receipts_df[
        (receipts_df['is_split_receipt'] == True) | 
        (receipts_df['parent_receipt_id'].notna())
    ])
    
    print(f"\nTotal Receipts: {total_receipts:,}")
    print(f"  ✅ Linked to banking: {linked_receipts:,} ({linked_receipts/total_receipts*100:.1f}%)")
    print(f"  ❌ Unlinked: {unlinked_receipts:,} ({unlinked_receipts/total_receipts*100:.1f}%)")
    print(f"  🔄 Duplicates: {len(duplicate_ids):,}")
    print(f"  👤 Personal: {personal_receipts:,}")
    print(f"  ➗ Split receipts: {split_receipts:,}")
    
    print(f"\nUnmatched Banking Debits: {len(banking_df):,}")
    print(f"Potential New Matches: {len(matches_df):,}")
    
    # Year breakdown
    print("\n" + "="*70)
    print("UNLINKED RECEIPTS BY YEAR")
    print("="*70)
    
    receipts_df['receipt_date'] = pd.to_datetime(receipts_df['receipt_date'])
    unlinked_by_year = receipts_df[receipts_df['banking_transaction_id'].isna()].groupby(
        receipts_df['receipt_date'].dt.year
    ).agg({
        'receipt_id': 'count',
        'gross_amount': 'sum'
    }).round(2)
    
    for year, row in unlinked_by_year.iterrows():
        print(f"{int(year)}: {int(row['receipt_id']):,} receipts, ${row['gross_amount']:,.2f}")
    
    print("\n" + "="*70)
    print("NEXT STEPS")
    print("="*70)
    print(f"\n1. Open Excel file: {output_file}")
    print("2. Review color-coded receipts:")
    print("   - RED rows need banking statement matches")
    print("   - YELLOW rows may be duplicates (verify if recurring payments)")
    print("   - ORANGE rows need vehicle/card/fuel data")
    print("3. Add or edit receipts in Excel")
    print("4. Save and run re-import script (to be created)")
    print("5. For completed years (2012-2023), banking data should be locked")
    
    conn.close()
    print("\n✅ Reconciliation complete!")


if __name__ == '__main__':
    main()
