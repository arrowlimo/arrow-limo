"""
Extract 2012 Scotia Bank statements from PDF to Excel with validation.
Handles OCR irregularities, multi-row descriptions, and balance verification.
Uses OCR for scanned PDFs.
"""

import pdfplumber
import pytesseract
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
import os
import io

# Vendor normalization patterns
VENDOR_PATTERNS = {
    r'MOHAWK': 'Mohawk',
    r'CENTEX': 'Centex',
    r'GEORGE[\'S]*\s*PIZZA': 'George\'s Pizza',
    r'FAS\s*GAS': 'Fas Gas',
    r'PETRO[\s-]*CAN(?:ADA)?': 'Petro-Canada',
    r'CO[\s-]*OP': 'Co-op',
    r'SHELL': 'Shell',
    r'ESSO': 'Esso',
    r'7[\s-]*ELEVEN': '7-Eleven',
    r'TELUS': 'Telus',
    r'ENMAX': 'Enmax',
    r'DIRECT\s*ENERGY': 'Direct Energy',
    r'BEST\s*BUY': 'Best Buy',
    r'VISA': 'Visa',
    r'INTERAC': 'Interac',
    r'PAYROLL': 'Payroll',
    r'TRANSFER': 'Transfer',
    r'SERVICE\s*CHARGE': 'Service Charge',
    r'NSF': 'NSF Fee',
    r'CHEQUE\s*ORDER': 'Cheque Order (GST included)',
    r'SBAP\s*FEE': 'SBAP Fee',
    r'DEBIT\s*MEMO': 'Debit Memo',
    r'MONEY\s*ORDER': 'Money Order',
    r'CHASE\s*PAYMENTECH': 'Chase Paymentech',
    r'AMEX': 'American Express',
}

def clean_vendor_name(description):
    """Normalize vendor names using patterns."""
    if not description:
        return description
    
    desc_upper = description.upper()
    for pattern, replacement in VENDOR_PATTERNS.items():
        if re.search(pattern, desc_upper):
            return replacement
    
    return description.strip()

def extract_tables_from_pdf(pdf_path):
    """Extract transaction lines from PDF using OCR."""
    print(f"Opening PDF: {pdf_path}")
    all_lines = []
    
    # Set tesseract path if on Windows
    if os.name == 'nt':
        # Common Windows installation paths
        for possible_path in [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            r'C:\Users\pdric\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
        ]:
            if os.path.exists(possible_path):
                pytesseract.pytesseract.tesseract_cmd = possible_path
                print(f"Found Tesseract at: {possible_path}")
                break
    
    print("Processing PDF pages with OCR...")
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"Total pages: {total_pages}")
        
        for page_num, page in enumerate(pdf.pages, 1):
            if page_num % 5 == 0 or page_num == 1:
                print(f"  OCR processing page {page_num}/{total_pages}...")
            
            # Convert PDF page to image
            try:
                im = page.to_image(resolution=300)
                pil_image = im.original
                
                # Perform OCR
                text = pytesseract.image_to_string(pil_image, config='--psm 6')
                
                if not text:
                    continue
                    
                # Split into lines
                lines = text.split('\n')
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Skip header lines and footers
                    skip_patterns = [
                        'SCOTIABANK', 'ACCOUNT NUMBER', 'STATEMENT OF', 
                        'DESCRIPTION', 'WITHDRAWAL', 'DEPOSIT', 'BALANCE FORWARD',
                        'TOTAL AMOUNT', 'NO. OF', 'DEBITS', 'CREDITS',
                        'ENCLOSURES', 'UNCOLLECTED FEES', 'ODI OWING'
                    ]
                    if any(x in line.upper() for x in skip_patterns):
                        continue
                    
                    # Skip summary/balance lines (have "FORWARD" or just numbers)
                    if 'FORWARD' in line.upper():
                        continue
                    
                    # Look for transaction lines (have description keywords)
                    transaction_keywords = [
                        'PURCHASE', 'DEPOSIT', 'WITHDRAWAL', 'SERVICE CHARGE', 
                        'FEE', 'TRANSFER', 'MEMO', 'PAYMENT', 'ABM', 'DEBIT', 
                        'VISA', 'MCARD', 'SBAP', 'AMEX', 'CHASE', 'MOHAWK',
                        'CENTEX', 'BEST BUY', 'PIZZA'
                    ]
                    
                    if any(x in line.upper() for x in transaction_keywords):
                        all_lines.append((page_num, line))
            except Exception as e:
                print(f"    Error on page {page_num}: {e}")
                continue
                
    print(f"  Total transaction lines extracted: {len(all_lines)}")
    return all_lines

def parse_transaction_line(line, prev_balance=None):
    """Parse a single transaction line."""
    # Pattern: DATE DESCRIPTION [CHEQUE#] [DEBIT] [CREDIT] BALANCE
    # Handle various date formats: MM/DD, MM-DD, MM/DD/YY
    date_pattern = r'^(\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?)\s+'
    
    # Try to extract date at start
    match = re.match(date_pattern, line.strip())
    if not match:
        return None
    
    date_str = match.group(1)
    remaining = line[match.end():].strip()
    
    # Extract all numbers (including potential cheque numbers)
    numbers = re.findall(r'[\d,]+\.?\d{0,2}', remaining)
    
    if len(numbers) < 1:
        return None
    
    # Last number is always balance
    balance_str = numbers[-1].replace(',', '')
    try:
        balance = float(balance_str)
    except:
        return None
    
    # Determine debit/credit from balance change
    debit = None
    credit = None
    amount = None
    
    if len(numbers) >= 2:
        # Second-to-last might be transaction amount
        amount_str = numbers[-2].replace(',', '')
        try:
            amount = float(amount_str)
            # Determine if debit or credit based on balance change
            if prev_balance is not None:
                if balance < prev_balance:
                    debit = amount
                else:
                    credit = amount
        except:
            pass
    
    # Extract cheque number (3-6 digits, not an amount)
    cheque_num = None
    for num in numbers[:-2]:  # Exclude amount and balance
        if len(num) >= 3 and len(num) <= 6 and '.' not in num:
            cheque_num = num
            break
    
    # Description is everything between date and first number
    desc_part = remaining
    for num in numbers:
        desc_part = desc_part.rsplit(num, 1)[0]
    
    description = desc_part.strip()
    
    return {
        'date': date_str,
        'description': description,
        'cheque_num': cheque_num,
        'debit': debit,
        'credit': credit,
        'balance': balance
    }

def parse_table_rows(transaction_lines):
    """
    Parse extracted transaction lines.
    Format: DESCRIPTION | WITHDRAWAL | DEPOSIT | DATE (MM/DD) | BALANCE
    Note: Monetary columns separated by | which OCR reads as space
          Values appear as space-separated (e.g., "8 4123" = 84.23)
          Dates are 4-digit MMDD only on last row of transaction groups
    """
    transactions = []
    prev_balance = None
    current_date = None
    
    for page_num, line in transaction_lines:
        # Clean line
        line = line.replace('|', ' ').replace('  ', ' ').strip()
        
        # Extract description - find where numbers start
        # Numbers will be space-separated digits: "8 4123" or standalone: "11250"
        num_start = re.search(r'\s+\d+\s+\d+|\s+\d{3,}', line)
        if not num_start:
            continue
        
        description = line[:num_start.start()].strip()
        remaining = line[num_start.start():].strip()
        
        # Find all numbers in remaining text
        numbers = re.findall(r'\d+', remaining)
        if not numbers:
            continue
        
        # Look for 4-digit date (MMDD format)
        date_str = current_date
        date_found = False
        for i, num in enumerate(numbers):
            if len(num) == 4:
                mm = int(num[:2])
                dd = int(num[2:])
                if 1 <= mm <= 12 and 1 <= dd <= 31:
                    current_date = f"{num[:2]}/{num[2:]}"
                    date_str = current_date
                    date_found = True
                    # Remove date from numbers list for amount parsing
                    numbers = numbers[:i] + numbers[i+1:]
                    break
        
        # Parse amounts - OCR shows as space-separated digits
        # Pattern: withdrawal_dollars withdrawal_cents deposit_dollars deposit_cents balance_dollars balance_cents
        # Or: amount_dollars amount_cents date(4) balance_dollars balance_cents
        
        withdrawal = None
        deposit = None
        balance = None
        cheque_num = None
        
        if len(numbers) >= 2:
            # Last 1-2 numbers are balance
            # Check if last number is 2-4 digits (likely cents)
            if len(numbers) >= 2 and len(numbers[-1]) <= 4:
                # Space-separated format: last two are balance
                bal_dollars = numbers[-2]
                bal_cents = numbers[-1].ljust(2, '0')[:2]  # Pad/trim to 2 digits
                try:
                    balance = float(f"{bal_dollars}.{bal_cents}")
                    numbers = numbers[:-2]
                except:
                    pass
            
            # Remaining numbers are withdrawal/deposit
            if len(numbers) >= 2:
                # First amount (withdrawal or deposit)
                amt1_dollars = numbers[0]
                amt1_cents = numbers[1].ljust(2, '0')[:2] if len(numbers) > 1 else '00'
                amt1 = float(f"{amt1_dollars}.{amt1_cents}")
                
                # Check if there's a second amount
                if len(numbers) >= 4:
                    amt2_dollars = numbers[2]
                    amt2_cents = numbers[3].ljust(2, '0')[:2]
                    amt2 = float(f"{amt2_dollars}.{amt2_cents}")
                    
                    # Both withdrawal and deposit present
                    withdrawal = amt1
                    deposit = amt2
                else:
                    # Only one amount - determine if debit or credit from description or balance change
                    if 'DEPOSIT' in description.upper() or 'CREDIT' in description.upper():
                        deposit = amt1
                    elif 'PURCHASE' in description.upper() or 'WITHDRAWAL' in description.upper() or 'FEE' in description.upper() or 'CHARGE' in description.upper():
                        withdrawal = amt1
                    elif balance and prev_balance:
                        if balance > prev_balance:
                            deposit = amt1
                        else:
                            withdrawal = amt1
                    else:
                        withdrawal = amt1
        
        # Extract cheque number from description
        cheque_match = re.search(r'\b(\d{3,6})\b', description)
        if cheque_match:
            cheque_num = cheque_match.group(1)
        
        # Clean description - extract vendor name from purchase descriptions
        # Pattern: "POINT OF SALE PURCHASE VENDOR NAME" -> "Purchase - Vendor Name"
        if 'POINT OF SALE PURCHASE' in description.upper():
            vendor = description.upper().replace('POINT OF SALE PURCHASE', '').strip()
            description = f"Purchase - {vendor.title()}"
        elif 'ABM DEPOSIT' in description.upper():
            location = description.upper().replace('ABM DEPOSIT', '').strip()
            description = f"ATM Deposit - {location.title()}" if location else "ATM Deposit"
        elif 'ABM WITHDRAWAL' in description.upper():
            location = description.upper().replace('ABM WITHDRAWAL', '').strip()
            description = f"ATM Withdrawal - {location.title()}" if location else "ATM Withdrawal"
        
        # Apply vendor name cleanup
        description = clean_vendor_name(description)
        
        transactions.append({
            'date': date_str or '',
            'description': description,
            'cheque_num': cheque_num,
            'debit': withdrawal,
            'credit': deposit,
            'balance': balance
        })
        
        if balance:
            prev_balance = balance
    
    return transactions

def create_excel_with_validation(transactions, output_path):
    """Create Excel file with auto-balance calculation and validation."""
    
    # Create DataFrame
    df_data = []
    running_balance = None
    
    for i, txn in enumerate(transactions):
        # Clean vendor name
        clean_desc = clean_vendor_name(txn['description'])
        
        # Note if cheque order (GST included)
        gst_note = ''
        if 'CHEQUE ORDER' in txn['description'].upper():
            gst_note = 'GST charged - match in receipts'
        
        # Parse amounts
        debit = float(txn['debit']) if txn['debit'] else None
        credit = float(txn['credit']) if txn['credit'] else None
        hard_balance = float(txn['balance']) if txn['balance'] else None
        
        # Calculate running balance
        if i == 0:
            running_balance = hard_balance if hard_balance else 0
        else:
            if running_balance is not None:
                if debit:
                    running_balance -= debit
                if credit:
                    running_balance += credit
            elif hard_balance:
                running_balance = hard_balance
        
        # Verify balance
        balance_match = ''
        if hard_balance and running_balance:
            diff = abs(hard_balance - running_balance)
            if diff > 0.01:  # Allow 1 cent rounding
                balance_match = f'ERROR: Δ{diff:.2f}'
            else:
                balance_match = 'OK'
        
        df_data.append({
            'Date': txn['date'],
            'Description': clean_desc,
            'Cheque #': txn['cheque_num'] or '',
            'Cheque Payee Lookup': 'LOOKUP NEEDED' if txn['cheque_num'] else '',
            'Debit': debit,
            'Credit': credit,
            'Hard Balance': hard_balance,
            'Auto Balance': running_balance,
            'Verification': balance_match,
            'GST Note': gst_note,
            'Notes': ''
        })
    
    df = pd.DataFrame(df_data)
    
    # Create Excel with formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scotia 2012 Transactions"
    
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Format columns
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 40  # Description
    ws.column_dimensions['C'].width = 10  # Cheque #
    ws.column_dimensions['D'].width = 20  # Cheque Payee
    ws.column_dimensions['E'].width = 12  # Debit
    ws.column_dimensions['F'].width = 12  # Credit
    ws.column_dimensions['G'].width = 14  # Hard Balance
    ws.column_dimensions['H'].width = 14  # Auto Balance
    ws.column_dimensions['I'].width = 18  # Verification
    ws.column_dimensions['J'].width = 25  # GST Note
    ws.column_dimensions['K'].width = 30  # Notes
    
    # Format currency columns
    currency_format = '#,##0.00'
    for row in range(2, len(df) + 2):
        for col in ['E', 'F', 'G', 'H']:
            cell = ws[f'{col}{row}']
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal='right')
    
    # Highlight errors
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_font = Font(color="9C0006", bold=True)
    
    for row in range(2, len(df) + 2):
        verification_cell = ws[f'I{row}']
        if verification_cell.value and 'ERROR' in str(verification_cell.value):
            verification_cell.fill = error_fill
            verification_cell.font = error_font
    
    # Highlight GST notes
    gst_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for row in range(2, len(df) + 2):
        gst_cell = ws[f'J{row}']
        if gst_cell.value:
            gst_cell.fill = gst_fill
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Excel file created: {output_path}")
    print(f"   Rows: {len(df)}")
    print(f"   Ready for manual editing")

def main():
    pdf_path = Path(r"L:\limo\pdf\2012\2012 scotiabank statements all.pdf")
    output_path = Path(r"L:\limo\data\2012_scotia_transactions_for_editing.xlsx")
    
    print("=" * 80)
    print("2012 Scotia Bank Statement Extraction")
    print("=" * 80)
    
    # Extract tables
    print("\n📄 Extracting tables from PDF...")
    table_rows = extract_tables_from_pdf(pdf_path)
    
    # Parse transactions
    print("\n🔍 Parsing transactions from tables...")
    transactions = parse_table_rows(table_rows)
    print(f"   Found {len(transactions)} transactions")
    
    # Create Excel
    print("\n📊 Creating Excel file with validation...")
    create_excel_with_validation(transactions, output_path)
    
    print("\n" + "=" * 80)
    print("INSTRUCTIONS FOR MANUAL EDITING:")
    print("=" * 80)
    print("1. Review 'Verification' column for balance errors")
    print("2. Fill in 'Cheque Payee Lookup' for all cheques")
    print("3. Check 'GST Note' - match cheque orders in receipts table")
    print("4. Use 'Notes' column for any observations")
    print("5. Auto Balance formula validates against Hard Balance")
    print("=" * 80)

if __name__ == '__main__':
    main()
