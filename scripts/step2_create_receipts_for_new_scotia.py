#!/usr/bin/env python3
"""
STEP 2: Recreate receipts for all new 2012 Scotia banking rows.
Auto-categorizes and links to banking transactions.
"""

import os
from datetime import datetime
import hashlib
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DB_SETTINGS = {
    "host": os.getenv("DB_HOST", "localhost"),
    "database": os.getenv("DB_NAME", "almsdata"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", "***REMOVED***"),
}

SCOTIA_ACCOUNT_ID = 2

def generate_hash(date, description, amount):
    """Generate SHA256 hash for deduplication."""
    text = f"{date}|{description}|{amount:.2f}"
    return hashlib.sha256(text.encode()).hexdigest()

def categorize_banking(description, amount):
    """Auto-categorize Scotia transactions."""
    desc_upper = str(description).upper()
    
    if any(x in desc_upper for x in ['CHQ', 'CHEQUE', 'DRAFT']):
        return 'payment_out', 'Cheque payment'
    if any(x in desc_upper for x in ['ATM', 'CASH', 'WITHDRAWAL']):
        return 'cash_withdrawal', 'Cash withdrawal'
    if any(x in desc_upper for x in ['CHARGE', 'FEE', 'INTEREST']):
        return 'bank_fees', 'Bank service charge'
    if 'DEPOSIT' in desc_upper:
        return 'deposit', 'Deposit'
    if any(x in desc_upper for x in ['TRANSFER', 'FROM CIBC', 'CIBC']):
        return 'inter_account_transfer', 'Inter-account transfer'
    if any(x in desc_upper for x in ['DEBIT', 'PAYMENT', 'ELECTRONIC']):
        return 'debit_card_purchase', 'Debit card transaction'
    
    return 'uncategorized', 'Bank transaction'

def create_receipts_for_scotia(dry_run=False):
    """Create receipts for 2012 Scotia banking rows."""
    print("\n" + "="*80)
    print("STEP 2: RECREATE RECEIPTS FOR NEW SCOTIA BANKING ROWS")
    print("="*80)
    print(f"Mode: {'DRY-RUN' if dry_run else 'WRITE'}\n")
    
    conn = psycopg2.connect(**DB_SETTINGS)
    cur = conn.cursor()
    
    try:
        # Fetch 2012 Scotia transactions
        cur.execute("""
            SELECT transaction_id, transaction_date, description, debit_amount, credit_amount
            FROM banking_transactions
            WHERE account_number = '903990106011'
              AND EXTRACT(YEAR FROM transaction_date) = 2012
            ORDER BY transaction_date, transaction_id
        """)
        
        scotia_rows = cur.fetchall()
        print(f"Found {len(scotia_rows)} 2012 Scotia transactions\n")
        
        receipts_created = []
        
        for tx_id, tx_date, description, debit, credit in scotia_rows:
            # Only create receipts for debits (expenses)
            if debit == 0:
                continue
            
            # Convert Decimal to float
            debit = float(debit)
            
            # Auto-categorize
            category, cat_desc = categorize_banking(description, debit)
            
            # Calculate GST (5% included)
            gst_amount = debit * 0.05 / 1.05
            net_amount = debit - gst_amount
            
            # Create hash
            source_hash = generate_hash(tx_date, description, debit)
            
            receipt_data = {
                'tx_id': tx_id,
                'receipt_date': tx_date,
                'vendor_name': description[:100] if description else 'Bank Transaction',
                'description': description,
                'gross_amount': round(debit, 2),
                'gst_amount': round(gst_amount, 2),
                'net_amount': round(net_amount, 2),
                'category': category,
                'source_hash': source_hash,
            }
            
            if not dry_run:
                # Insert receipt (skip if duplicate hash)
                cur.execute("""
                    INSERT INTO receipts
                    (receipt_date, vendor_name, description, gross_amount, gst_amount, net_amount, 
                     category, mapped_bank_account_id, source_hash, created_from_banking)
                    VALUES (%(receipt_date)s, %(vendor_name)s, %(description)s, %(gross_amount)s, 
                            %(gst_amount)s, %(net_amount)s, %(category)s, %(bank_account_id)s, 
                            %(source_hash)s, true)
                    ON CONFLICT (source_hash) DO NOTHING
                    RETURNING receipt_id
                """, {
                    'receipt_date': tx_date,
                    'vendor_name': receipt_data['vendor_name'],
                    'description': description,
                    'gross_amount': receipt_data['gross_amount'],
                    'gst_amount': receipt_data['gst_amount'],
                    'net_amount': receipt_data['net_amount'],
                    'category': category,
                    'bank_account_id': SCOTIA_ACCOUNT_ID,
                    'source_hash': source_hash,
                })
                
                receipt_row = cur.fetchone()
                if not receipt_row:
                    continue
                
                receipt_id = receipt_row[0]
                
                # Link to banking
                cur.execute("""
                    INSERT INTO banking_receipt_matching_ledger
                    (banking_transaction_id, receipt_id, match_type)
                    VALUES (%s, %s, %s)
                    ON CONFLICT DO NOTHING
                """, (tx_id, receipt_id, 'auto_created_from_scotia'))
                
                receipt_data['receipt_id'] = receipt_id
            
            receipts_created.append(receipt_data)
        
        if not dry_run:
            conn.commit()
            print(f"[OK] Created {len(receipts_created)} receipts for Scotia transactions")
        else:
            print(f"[DRY-RUN] Would create {len(receipts_created)} receipts")
        
        # Export to Excel
        if receipts_created:
            export_receipts_to_excel(receipts_created, Path('reports/scotia_auto_created_receipts.xlsx'))
        
        return len(receipts_created)
    
    except Exception as e:
        if not dry_run:
            conn.rollback()
        print(f"[ERROR] {e}")
        raise
    finally:
        cur.close()
        conn.close()

def export_receipts_to_excel(receipts, output_path):
    """Export to Excel with bright yellow highlighting."""
    print(f"Exporting {len(receipts)} receipts to {output_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Scotia Receipts"
    
    headers = ['Receipt ID', 'Date', 'Vendor', 'Amount', 'GST', 'Net', 'Category', 'TX ID']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for receipt in receipts:
        ws.append([
            receipt.get('receipt_id', ''),
            receipt['receipt_date'],
            receipt['vendor_name'][:50],
            receipt['gross_amount'],
            receipt['gst_amount'],
            receipt['net_amount'],
            receipt['category'],
            receipt['tx_id'],
        ])
    
    for row in range(2, len(receipts) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = yellow_fill
    
    for row in range(2, len(receipts) + 2):
        for col in [4, 5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 10
    
    wb.save(output_path)
    print(f"[OK] Exported to {output_path}\n")

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Create receipts for new Scotia 2012')
    parser.add_argument('--dry-run', action='store_true', help='Preview without writing')
    parser.add_argument('--write', action='store_true', help='Write to database')
    args = parser.parse_args()
    
    if not args.write:
        args.dry_run = True
    
    count = create_receipts_for_scotia(dry_run=args.dry_run)
    
    print("="*80)
    print("NEXT STEPS:")
    print("1. Link CIBC->Scotia split deposits (parent/child)")
    print("2. Dedup receipts (QuickBooks artifacts)")
    print("3. Re-export receipt_lookup_and_entry_2012.xlsx")
    print("="*80)

if __name__ == '__main__':
    main()
