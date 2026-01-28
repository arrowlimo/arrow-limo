"""
Extract CIBC 8362 2018 using table extraction
"""
import pdfplumber
from pathlib import Path
from decimal import Decimal
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

pdf_path = Path(r"L:\limo\pdf\2018\CIBC 8362 2018.pdf")
output_path = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

OPENING_BALANCE = Decimal('1058.53')

print("EXTRACTING CIBC 8362 2018 USING TABLE EXTRACTION")
print("=" * 80)

trans = []

with pdfplumber.open(pdf_path) as pdf:
    for page_num, page in enumerate(pdf.pages, 1):
        print(f"Page {page_num}...")
        
        # Extract tables
        tables = page.extract_tables()
        
        for table in tables:
            for row in table:
                if not row or len(row) < 4:
                    continue
                
                # Skip headers
                if 'Date' in str(row[0]) or 'Transaction' in str(row):
                    continue
                
                try:
                    date_str = str(row[0]).strip()
                    desc = str(row[1]).strip() if len(row) > 1 else ""
                    
                    # Skip empty or opening balance
                    if not date_str or 'Opening' in desc:
                        continue
                    
                    # Parse date (format: "Jan 2" or similar)
                    months = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                             'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
                    
                    parts = date_str.split()
                    if len(parts) < 2:
                        continue
                    
                    month_str = parts[0]
                    day = int(parts[1])
                    
                    if month_str not in months:
                        continue
                    
                    date_obj = datetime(2018, months[month_str], day)
                    date_fmt = date_obj.strftime('%Y-%m-%d')
                    
                    # Columns: Date, Description, Withdrawals, Deposits, Balance
                    withdrawal_str = str(row[2]).strip() if len(row) > 2 else ""
                    deposit_str = str(row[3]).strip() if len(row) > 3 else ""
                    balance_str = str(row[4]).strip() if len(row) > 4 else ""
                    
                    # Parse amounts
                    debit = Decimal(withdrawal_str.replace(',', '').replace('$', '')) if withdrawal_str and withdrawal_str != 'None' else Decimal('0')
                    credit = Decimal(deposit_str.replace(',', '').replace('$', '')) if deposit_str and deposit_str != 'None' else Decimal('0')
                    balance = Decimal(balance_str.replace(',', '').replace('$', '')) if balance_str and balance_str != 'None' else Decimal('0')
                    
                    trans.append({
                        'date': date_fmt,
                        'date_obj': date_obj,
                        'desc': desc,
                        'debit': debit,
                        'credit': credit,
                        'balance': balance
                    })
                    
                except (ValueError, IndexError, AttributeError) as e:
                    continue

print(f"\nExtracted {len(trans)} transactions\n")

if not trans:
    print("❌ No transactions - table extraction failed")
    exit(1)

# Sort chronologically
trans.sort(key=lambda x: x['date_obj'])

total_cr = sum(t['credit'] for t in trans)
total_db = sum(t['debit'] for t in trans)

print(f"Total Credits: ${total_cr:,.2f}")
print(f"Total Debits: ${total_db:,.2f}")
print(f"First: {trans[0]['date']} - {trans[0]['desc']}")
print(f"Last: {trans[-1]['date']} - {trans[-1]['desc']}")

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Transactions"

headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
ws.append(headers)

fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
font = Font(bold=True, color="FFFFFF")
for c in range(1, 6):
    ws.cell(1, c).fill = fill
    ws.cell(1, c).font = font

for t in trans:
    ws.append([
        t['date'], t['desc'],
        float(t['debit']), float(t['credit']),
        float(t['balance'])
    ])

ws.column_dimensions['B'].width = 50

for c in [3, 4, 5]:
    for r in range(2, len(trans) + 2):
        ws.cell(r, c).number_format = '$#,##0.00'

# Summary
s = wb.create_sheet("Summary")
s['A1'] = "CIBC 8362 2018"
s['A1'].font = Font(bold=True, size=14)
s['A3'] = "Opening:"
s['B3'] = float(OPENING_BALANCE)
s['B3'].number_format = '$#,##0.00'
s['A4'] = "Transactions:"
s['B4'] = len(trans)
s['A5'] = "Credits:"
s['B5'] = float(total_cr)
s['B5'].number_format = '$#,##0.00'
s['A6'] = "Debits:"
s['B6'] = float(total_db)
s['B6'].number_format = '$#,##0.00'

wb.save(output_path)
print(f"\n✅ {output_path}")
print(f"   {len(trans)} transactions")
