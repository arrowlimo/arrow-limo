#!/usr/bin/env python3
"""
Extract detailed invoice information from fibrenew_0001.xlsx sheets 2 and 3.
Looking for invoice amounts, payments, and balances.
"""

import openpyxl
import re

file_path = r'L:\limo\receipts\fibrenew_0001.xlsx'

print("DETAILED INVOICE EXTRACTION FROM FIBRENEW_0001.XLSX")
print("=" * 80)

wb = openpyxl.load_workbook(file_path)

for sheet_name in ['Sheet2', 'Sheet3']:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print("=" * 80)
    
    ws = wb[sheet_name]
    
    invoice_number = None
    invoice_date = None
    amounts = []
    
    # Search all cells for invoice data
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_str = str(cell.value)
                
                # Look for invoice number
                if 'Invoice #' in cell_str or (cell.column in [4, 5] and re.match(r'^\d{4}', cell_str)):
                    # Check next cell or cells around for the number
                    if cell.column < ws.max_column:
                        next_cell = ws.cell(row=cell.row, column=cell.column + 1).value
                        if next_cell and re.match(r'^\d{4}', str(next_cell)):
                            invoice_number = str(next_cell).split('.')[0]
                
                # Look for amounts (currency values)
                amount_match = re.search(r'\$?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2}))', cell_str)
                if amount_match:
                    amount = float(amount_match.group(1).replace(',', ''))
                    amounts.append({
                        'row': cell.row,
                        'col': cell.column,
                        'amount': amount,
                        'context': cell_str
                    })
                
                # Look for dates
                if '2017' in cell_str and '-' in cell_str:
                    invoice_date = cell_str
    
    print(f"Invoice Number: {invoice_number if invoice_number else 'Not found'}")
    print(f"Invoice Date: {invoice_date if invoice_date else 'Not found'}")
    print(f"\nAmounts found ({len(amounts)}):")
    for amt in amounts:
        print(f"  Row {amt['row']}, Col {amt['col']}: ${amt['amount']:,.2f} - {amt['context'][:50]}")
    
    # Try to identify: invoice total, payment, balance
    if amounts:
        print(f"\nLikely interpretation:")
        print(f"  Amounts: {', '.join([f'${a['amount']:,.2f}' for a in amounts])}")

wb.close()
