#!/usr/bin/env python
"""Inspect by-month Scotia XLSX files (2012-2014) and show sheet names + last 5 rows per sheet."""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed. Install with pip install openpyxl\n")
    sys.exit(1)

files = [
    Path(r'L:\limo\reports\scotia_903990106011_2012_by_month.xlsx'),
    Path(r'L:\limo\reports\scotia_903990106011_2013_by_month.xlsx'),
    Path(r'L:\limo\reports\scotia_903990106011_2014_by_month.xlsx'),
]

for f in files:
    if not f.exists():
        print(f"Missing {f}")
        continue
    wb = openpyxl.load_workbook(f, data_only=True)
    print(f"\n{f.name} sheets: {wb.sheetnames}")
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        maxr = ws.max_row
        maxc = ws.max_column
        print(f" Sheet {ws_name}: last 5 rows (up to {maxc} cols):")
        start = max(1, maxr - 4)
        for row in range(start, maxr + 1):
            vals = [ws.cell(row=row, column=col).value for col in range(1, maxc + 1)]
            print(vals)
