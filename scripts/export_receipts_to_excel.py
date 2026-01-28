import argparse
import os
import pathlib
import sys
from typing import Iterable, List, Any
import datetime as _dt

# Ensure repository root is on sys.path to import api
ROOT = pathlib.Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from api import get_db_connection  # type: ignore


def batched(cursor, size: int = 5000) -> Iterable[list]:
    while True:
        rows = cursor.fetchmany(size)
        if not rows:
            break
        yield rows


def main() -> int:
    ap = argparse.ArgumentParser(description='Export receipts table to Excel (.xlsx)')
    ap.add_argument('--out', default=str(pathlib.Path('reports') / 'receipts_export.xlsx'), help='Output .xlsx path')
    ap.add_argument('--where', default='', help="Optional WHERE clause without the word WHERE (e.g. 'receipt_date>=''2012-01-01'' AND receipt_date<=''2012-12-31''')")
    ap.add_argument('--order-by', default='receipt_date, id', help='ORDER BY clause')
    ap.add_argument('--batch', type=int, default=5000, help='Fetch batch size')
    args = ap.parse_args()

    out_path = pathlib.Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Lazy import to avoid requiring openpyxl unless we actually run export
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        print('[ERROR] openpyxl not installed. Install with: pip install openpyxl')
        raise

    where_clause = f" WHERE {args.where}" if args.where else ''
    sql = f"SELECT * FROM receipts{where_clause} ORDER BY {args.order_by}"

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.itersize = args.batch
        cur.execute(sql)
        colnames = [d[0] for d in cur.description]

        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title='receipts')
        # Remove default sheet if present
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            std = wb['Sheet']
            wb.remove(std)

        # Header row
        ws.append(colnames)

        total = 0
        def _norm(v: Any) -> Any:
            if isinstance(v, _dt.datetime):
                # Convert tz-aware to naive UTC; Excel doesn't support tzinfo
                if v.tzinfo is not None:
                    return v.astimezone(_dt.timezone.utc).replace(tzinfo=None)
                return v
            if isinstance(v, _dt.time):
                if v.tzinfo is not None:
                    return v.replace(tzinfo=None)
                return v
            # Convert memoryview/bytes to string for safety
            if isinstance(v, (bytes, memoryview)):
                try:
                    return bytes(v).decode('utf-8', errors='replace')
                except Exception:
                    return str(v)
            return v

        for chunk in batched(cur, size=args.batch):
            for row in chunk:
                ws.append([_norm(x) for x in row])
            total += len(chunk)
            if total % (args.batch * 10) == 0:
                print(f"[INFO] wrote {total} rows...")

        # Freeze header row
        try:
            ws.freeze_panes = 'A2'
        except Exception:
            pass

        wb.save(str(out_path))
        print(f"[OK] Exported {total} rows to {out_path}")
        return 0
    finally:
        conn.close()


if __name__ == '__main__':
    raise SystemExit(main())
