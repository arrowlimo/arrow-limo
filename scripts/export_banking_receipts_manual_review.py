"""
Export all receipts and banking transactions into a single Excel workbook for manual review.
- Sheet 1: All receipts (all columns), vendor names capitalized, fuzzy vendor match suggestion with banking descriptions, color-coded by match score.
- Sheet 2: All banking transactions (all columns), combined across bank accounts, sortable by account_number, vendor-like fields normalized.
Color coding:
  - Match score >= 85: light green
  - 70-84: light yellow
  - Below 70 or no match: no color
"""
import re
import sys
from datetime import datetime
import psycopg2
import pandas as pd
from difflib import SequenceMatcher, get_close_matches
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

DB_CONFIG = {
    "host": "localhost",
    "database": "almsdata",
    "user": "postgres",
    "password": "***REMOVED***",
}

# Colors
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

def normalize_vendor(text: str) -> str:
    if not text:
        return ""
    text = text.upper()
    text = re.sub(r"QBO|QUICKBOOKS|QUICK BOOKS", "", text)
    text = re.sub(r"CHEQUE|CHECK|CHQ", "", text)
    text = re.sub(r"E[- ]?TRANSFER|EFT|ETRANSFER", "", text)
    text = re.sub(r"DD", "", text)
    text = re.sub(r"X+", "", text)
    text = re.sub(r"\bNO\.\s*\d+\b", "", text)
    text = re.sub(r"\d{4,}", "", text)  # remove long numeric strings
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def best_match(target: str, candidates: list[str]) -> tuple[str, float]:
    if not target or not candidates:
        return "", 0.0
    # quick exact
    if target in candidates:
        return target, 1.0
    # use difflib close matches for speed
    close = get_close_matches(target, candidates, n=5, cutoff=0.6)
    best_score = 0.0
    best_val = ""
    for c in close or candidates[:50]:  # fallback small sample
        score = SequenceMatcher(None, target, c).ratio()
        if score > best_score:
            best_score = score
            best_val = c
    return best_val, best_score

def fetch_df(conn, query):
    return pd.read_sql_query(query, conn)

def apply_color_coding(file_path: str, sheet_name: str, score_col: str):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    headers = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    score_idx = headers.get(score_col)
    if not score_idx:
        wb.save(file_path)
        return
    for row in ws.iter_rows(min_row=2):
        cell = row[score_idx - 1]
        try:
            score = float(cell.value)
        except (TypeError, ValueError):
            score = None
        if score is None:
            continue
        if score >= 85:
            for c in row:
                c.fill = FILL_GREEN
        elif score >= 70:
            for c in row:
                c.fill = FILL_YELLOW
    wb.save(file_path)

def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"l:/limo/reports/banking_receipts_manual_review_{timestamp}.xlsx"
    print("=" * 80)
    print("EXPORTING BANKING & RECEIPTS FOR MANUAL REVIEW")
    print("=" * 80)

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        # Fetch receipts (all columns)
        print("\nFetching receipts...")
        receipts_df = fetch_df(conn, "SELECT * FROM receipts ORDER BY receipt_date, receipt_id")
        print(f"  Loaded {len(receipts_df):,} receipts")
        # Capitalize vendor names
        if 'vendor_name' in receipts_df.columns:
            receipts_df['vendor_name'] = receipts_df['vendor_name'].fillna('').str.upper()
            receipts_df['vendor_normalized'] = receipts_df['vendor_name'].apply(normalize_vendor)
        else:
            receipts_df['vendor_normalized'] = ''

        # Fetch banking (all columns)
        print("Fetching banking transactions...")
        banking_df = fetch_df(conn, "SELECT * FROM banking_transactions ORDER BY account_number, transaction_date, transaction_id")
        print(f"  Loaded {len(banking_df):,} banking transactions")
        # Add normalized vendor/description
        vendor_like_cols = [c for c in ['vendor_extracted', 'vendor_truncated', 'description', 'memo', 'payee_name', 'vendor_name'] if c in banking_df.columns]
        if vendor_like_cols:
            banking_df['vendor_candidate'] = banking_df[vendor_like_cols].astype(str).fillna('').agg(' '.join, axis=1).str.upper()
            banking_df['vendor_normalized'] = banking_df['vendor_candidate'].apply(normalize_vendor)
        else:
            banking_df['vendor_candidate'] = ''
            banking_df['vendor_normalized'] = ''

        # Prepare candidate list for matching
        banking_candidates = sorted(set([v for v in banking_df['vendor_normalized'] if isinstance(v, str) and v]))

        # Fuzzy match suggestions for receipts
        print("Computing fuzzy vendor suggestions (receipts -> banking)...")
        suggestions = []
        scores = []
        for v in receipts_df['vendor_normalized']:
            match, score = best_match(v, banking_candidates)
            suggestions.append(match)
            scores.append(round(score * 100, 1))
        receipts_df['match_vendor_suggestion'] = suggestions
        receipts_df['match_score'] = scores

        # Write to Excel
        print(f"\nWriting Excel: {output_file}")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            receipts_df.to_excel(writer, sheet_name='Receipts', index=False)
            banking_df.to_excel(writer, sheet_name='Banking', index=False)

        # Apply color coding based on match_score in receipts
        apply_color_coding(output_file, 'Receipts', 'match_score')

        print("\nSUMMARY:")
        print(f"  Receipts: {len(receipts_df):,}")
        print(f"  Banking: {len(banking_df):,}")
        print(f"  Output: {output_file}")
        print("\nColor coding applied: >=85 green, 70-84 yellow")
        print("✅ EXPORT COMPLETE")
    finally:
        conn.close()

if __name__ == '__main__':
    main()
