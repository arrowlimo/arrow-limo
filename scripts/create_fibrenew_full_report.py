#!/usr/bin/env python3
"""
Create comprehensive Excel report of Fibrenew invoices with payments, balances, and 2015 summary.
"""

import pandas as pd
from decimal import Decimal
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = r'L:\limo\pdf\2012\fibrenew.xlsx'
OUTPUT_FILE = r'L:\limo\reports\fibrenew_complete_reconciliation.xlsx'

def parse_date(val):
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except:
                continue
    return None

def parse_amount(val):
    if pd.isna(val):
        return None
    try:
        return Decimal(str(val))
    except:
        return None

def extract_invoice_references(notes):
    """Extract invoice numbers and amounts from payment notes."""
    if pd.isna(notes):
        return []
    
    notes_str = str(notes)
    refs = []
    
    # Pattern: "5797,104.89,5802,895.11" - comma-separated invoice#,amount pairs
    parts = notes_str.split(',')
    i = 0
    while i < len(parts) - 1:
        try:
            inv_num = parts[i].strip()
            amount = Decimal(parts[i+1].strip())
            refs.append({'invoice': inv_num, 'amount': amount})
            i += 2
        except:
            i += 1
    
    return refs

def main():
    # Read Excel file
    df = pd.read_excel(EXCEL_FILE, header=None)
    
    print("Reading Fibrenew data...")
    
    # Parse unique invoices (first occurrence only)
    invoice_entries = []
    seen_invoices = {}
    
    for idx, row in df.iterrows():
        col0 = str(row[0]).strip()
        if col0 and col0 not in ['inv', 'pmt', 'statement', 'nan'] and not 'balance' in str(row[1]).lower():
            inv_date = parse_date(row[1])
            inv_amt = parse_amount(row[2])
            notes = str(row[3]) if not pd.isna(row[3]) else ''
            
            # Include ALL years (no date filtering)
            if inv_date and col0 not in seen_invoices:
                seen_invoices[col0] = {
                    'invoice_number': col0,
                    'date': inv_date,
                    'amount': inv_amt if inv_amt else Decimal('0'),
                    'notes': notes
                }
    
    # Parse payment records
    payments = []
    for idx, row in df.iterrows():
        if str(row[0]).strip().lower() == 'pmt':
            pmt_date = parse_date(row[1])
            pmt_amt = parse_amount(row[2])
            pmt_notes = str(row[3]) if not pd.isna(row[3]) else ''
            
            # Handle special case: "auditors note paid 2400 cash" in amount column
            if pmt_amt is None and pmt_notes and 'paid' in pmt_notes.lower():
                # Try to extract amount from notes
                import re
                match = re.search(r'(\d+(?:\.\d+)?)', pmt_notes)
                if match:
                    pmt_amt = Decimal(match.group(1))
            
            if pmt_date and pmt_amt:
                # Extract invoice references
                refs = extract_invoice_references(pmt_notes)
                
                payments.append({
                    'date': pmt_date,
                    'amount': abs(pmt_amt),  # Make positive
                    'notes': pmt_notes,
                    'invoice_refs': refs
                })
    
    # Parse Aug 2015 balance summary (row 90)
    balance_summary = {}
    for idx, row in df.iterrows():
        if 'balance' in str(row[1]).lower():
            summary_date = parse_date(row[0])
            if summary_date and summary_date.year == 2015 and summary_date.month == 8:
                # Parse invoice list from row[3]
                detail = str(row[3])
                # Format: "5977,383.84,5978,144.13,6030,141.3,..."
                parts = detail.split(',')
                i = 0
                while i < len(parts) - 1:
                    try:
                        inv_num = parts[i].strip()
                        balance_amt = Decimal(parts[i+1].strip())
                        balance_summary[inv_num] = balance_amt
                        i += 2
                    except:
                        i += 1
                break
    
    print(f"Found {len(seen_invoices)} unique invoices")
    print(f"Found {len(payments)} payments")
    print(f"Found {len(balance_summary)} invoices in Aug 2015 balance")
    
    # Build invoice list with categorization
    invoice_list = []
    for inv_num, inv_data in sorted(seen_invoices.items(), key=lambda x: x[1]['date']):
        amount = inv_data['amount']
        
        # Categorize
        if amount >= Decimal('1000'):
            category = 'RENT'
        elif Decimal('100') <= amount < Decimal('1000'):
            category = 'UTILITIES'
        elif amount < Decimal('0'):
            category = 'CREDIT'
        elif amount == Decimal('0'):
            category = 'NO AMOUNT'
        else:
            category = 'OTHER'
        
        # Check if in Aug 2015 balance (unpaid)
        balance_status = 'PAID'
        balance_owing = Decimal('0')
        if inv_num in balance_summary:
            balance_status = 'UNPAID'
            balance_owing = balance_summary[inv_num]
        
        invoice_list.append({
            'Invoice #': inv_num,
            'Date': inv_data['date'],
            'Category': category,
            'Amount': float(amount),
            'Status': balance_status,
            'Balance Owing (Aug 2015)': float(balance_owing),
            'Notes': inv_data['notes']
        })
    
    # Build payment list
    payment_list = []
    for pmt in sorted(payments, key=lambda x: x['date']):
        # Extract payment method from notes
        notes = pmt['notes'].lower()
        if 'cash' in notes:
            method = 'CASH'
        elif 'cheque' in notes or 'chq' in notes:
            method = 'CHEQUE'
        else:
            method = 'OTHER'
        
        # Get invoice references
        inv_refs = ', '.join([f"#{ref['invoice']} (${ref['amount']})" for ref in pmt['invoice_refs']])
        if not inv_refs:
            inv_refs = pmt['notes']
        
        payment_list.append({
            'Date': pmt['date'],
            'Amount': float(pmt['amount']),
            'Method': method,
            'Applied To': inv_refs,
            'Notes': pmt['notes']
        })
    
    # Calculate summary by year
    summary_by_year = defaultdict(lambda: {'rent': Decimal('0'), 'utilities': Decimal('0'), 
                                            'credits': Decimal('0'), 'total': Decimal('0')})
    
    for inv in invoice_list:
        year = inv['Date'].year
        amount = Decimal(str(inv['Amount']))
        
        if inv['Category'] == 'RENT':
            summary_by_year[year]['rent'] += amount
        elif inv['Category'] == 'UTILITIES':
            summary_by_year[year]['utilities'] += amount
        elif inv['Category'] == 'CREDIT':
            summary_by_year[year]['credits'] += amount
        
        summary_by_year[year]['total'] += amount
    
    summary_list = []
    for year in sorted(summary_by_year.keys()):
        data = summary_by_year[year]
        summary_list.append({
            'Year': year,
            'Rent': float(data['rent']),
            'Utilities': float(data['utilities']),
            'Credits': float(data['credits']),
            'Total Invoiced': float(data['total'])
        })
    
    # Calculate payments by year
    payments_by_year = defaultdict(Decimal)
    for pmt in payments:
        payments_by_year[pmt['date'].year] += pmt['amount']
    
    # Add payments to summary
    for item in summary_list:
        year = item['Year']
        item['Payments'] = float(payments_by_year.get(year, Decimal('0')))
        item['Net (Invoiced - Paid)'] = item['Total Invoiced'] - item['Payments']
    
    # Add 2015 final balance
    balance_2015 = sum(balance_summary.values())
    
    # Create final summary
    total_invoiced = sum(Decimal(str(inv['Amount'])) for inv in invoice_list)
    total_paid = sum(pmt['amount'] for pmt in payments)
    
    final_summary = [{
        'Description': 'Total Invoiced (2012-2017)',
        'Amount': float(total_invoiced)
    }, {
        'Description': 'Total Payments (2013-2016)',
        'Amount': float(total_paid)
    }, {
        'Description': 'Balance per Aug 2015 Summary',
        'Amount': float(balance_2015)
    }, {
        'Description': 'Calculated Balance (Invoiced - Paid)',
        'Amount': float(total_invoiced - total_paid)
    }, {
        'Description': 'Discrepancy (payments not tracked in file)',
        'Amount': float(balance_2015 - (total_invoiced - total_paid))
    }, {
        'Description': '',
        'Amount': None
    }, {
        'Description': '2025 CURRENT BALANCE (per recent report)',
        'Amount': 14000.00
    }, {
        'Description': 'Note: 2017-2025 invoices not in this file',
        'Amount': None
    }]
    
    # Create Excel workbook
    print(f"\nCreating Excel report: {OUTPUT_FILE}")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Sheet 1: All Invoices
        df_invoices = pd.DataFrame(invoice_list)
        df_invoices.to_excel(writer, sheet_name='All Invoices', index=False)
        
        # Sheet 2: All Payments
        df_payments = pd.DataFrame(payment_list)
        df_payments.to_excel(writer, sheet_name='Payments', index=False)
        
        # Sheet 3: Summary by Year
        df_summary = pd.DataFrame(summary_list)
        df_summary.to_excel(writer, sheet_name='Summary by Year', index=False)
        
        # Sheet 4: Final Summary
        df_final = pd.DataFrame(final_summary)
        df_final.to_excel(writer, sheet_name='Final Summary', index=False)
        
        # Sheet 5: Outstanding Invoices (Aug 2015)
        outstanding = [inv for inv in invoice_list if inv['Status'] == 'UNPAID']
        df_outstanding = pd.DataFrame(outstanding)
        df_outstanding.to_excel(writer, sheet_name='Outstanding Aug 2015', index=False)
    
    # Format the Excel file
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Header formatting
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format currency columns
        if sheet_name in ['All Invoices', 'Payments', 'Summary by Year', 'Final Summary', 'Outstanding Aug 2015']:
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter in ['D', 'F'] or (sheet_name == 'Payments' and cell.column_letter == 'B'):
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
                    elif sheet_name == 'Summary by Year' and cell.column_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
                    elif sheet_name == 'Final Summary' and cell.column_letter == 'B':
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
    
    wb.save(OUTPUT_FILE)
    
    print(f"\n{'='*80}")
    print("REPORT CREATED SUCCESSFULLY")
    print(f"{'='*80}")
    print(f"\nFile: {OUTPUT_FILE}")
    print(f"\nSheets created:")
    print(f"  1. All Invoices: {len(invoice_list)} invoices")
    print(f"  2. Payments: {len(payment_list)} payments")
    print(f"  3. Summary by Year: {len(summary_list)} years")
    print(f"  4. Final Summary: Reconciliation totals")
    print(f"  5. Outstanding Aug 2015: {len(outstanding)} unpaid invoices")
    print(f"\nKey Totals:")
    print(f"  Total Invoiced: ${total_invoiced:,.2f}")
    print(f"  Total Paid: ${total_paid:,.2f}")
    print(f"  Balance per Aug 2015: ${balance_2015:,.2f}")
    print(f"  Calculated Balance: ${total_invoiced - total_paid:,.2f}")
    
    # Show payment breakdown
    cash_payments = [p for p in payment_list if p['Method'] == 'CASH']
    cheque_payments = [p for p in payment_list if p['Method'] == 'CHEQUE']
    other_payments = [p for p in payment_list if p['Method'] == 'OTHER']
    
    print(f"\nPayment Methods:")
    print(f"  Cash: {len(cash_payments)} payments (${sum(p['Amount'] for p in cash_payments):,.2f})")
    print(f"  Cheque: {len(cheque_payments)} payments (${sum(p['Amount'] for p in cheque_payments):,.2f})")
    print(f"  Other: {len(other_payments)} payments (${sum(p['Amount'] for p in other_payments):,.2f})")

if __name__ == '__main__':
    main()
