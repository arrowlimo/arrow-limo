"""
Generate Complete Receipts Workbook - ALL Years, ALL Banking Data
Includes parent-child split support and color coding
"""

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def get_db_connection():
    """Connect to PostgreSQL database."""
    return psycopg2.connect(
        host=os.environ.get('DB_HOST', 'localhost'),
        database=os.environ.get('DB_NAME', 'almsdata'),
        user=os.environ.get('DB_USER', 'postgres'),
        password=os.environ.get('DB_PASSWORD', '***REMOVED***')
    )

def get_color_for_source(receipt_source, display_color):
    """Return Excel fill color based on receipt source/color."""
    # Priority: display_color > receipt_source
    if display_color:
        color_map = {
            'GREEN': 'C6EFCE',   # Light green - matched to banking
            'YELLOW': 'FFEB9C',  # Light yellow - cash payments
            'ORANGE': 'FFC7CE',  # Light orange - reimbursements
            'BLUE': 'BDD7EE',    # Light blue - manual entry
            'RED': 'FFC7CE'      # Light red - unmatched
        }
        return color_map.get(display_color, 'FFFFFF')
    
    # Fallback to receipt_source if no display_color
    if receipt_source:
        source_map = {
            'BANKING': 'C6EFCE',      # Green
            'CASH': 'FFEB9C',         # Yellow
            'REIMBURSEMENT': 'FFC7CE', # Orange
            'MANUAL': 'BDD7EE',       # Blue
            'UNMATCHED': 'FFC7CE'     # Red
        }
        return source_map.get(receipt_source, 'FFFFFF')
    
    return 'FFFFFF'  # White default

def get_bank_account_color(mapped_bank_account_id):
    """Return color for bank account."""
    if mapped_bank_account_id == 1:
        return 'C6EFCE'  # Green for CIBC
    elif mapped_bank_account_id == 2:
        return 'FFEB9C'  # Yellow/orange for Scotia
    return 'FFFFFF'

def apply_header_style(ws, row=1):
    """Apply consistent header styling."""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def create_all_receipts_sheet(wb, conn):
    """Create sheet with ALL receipts from all years."""
    ws = wb.create_sheet("All Receipts", 0)
    
    # Headers
    headers = [
        'Receipt ID', 'Parent Receipt #', 'Receipt Date', 'Vendor', 'Amount', 'GST', 'Net',
        'Category', 'Payment Method', 'Bank Account', 
        'Vehicle ID', 'Business/Personal', 'Description',
        'Receipt Source', 'Display Color', 'Created From Banking'
    ]
    
    ws.append(headers)
    apply_header_style(ws)
    
    # Query ALL receipts
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            r.receipt_id,
            r.parent_receipt_id,
            r.receipt_date,
            r.vendor_name,
            r.gross_amount,
            r.gst_amount,
            r.net_amount,
            r.category,
            r.payment_method,
            CASE 
                WHEN r.mapped_bank_account_id = 1 THEN 'CIBC 0228362'
                WHEN r.mapped_bank_account_id = 2 THEN 'Scotia 903990106011'
                ELSE NULL
            END as bank_account,
            r.vehicle_id,
            CASE WHEN r.is_personal_purchase THEN 'Personal' ELSE 'Business' END as business_personal,
            r.description,
            r.receipt_source,
            r.display_color,
            r.created_from_banking,
            r.mapped_bank_account_id
        FROM receipts r
        ORDER BY r.receipt_date DESC, r.receipt_id DESC
    """)
    
    print(f"Loading ALL receipts...")
    row_count = 0
    
    for row_data in cur.fetchall():
        receipt_id, parent_id, receipt_date, vendor, gross, gst, net, category, payment_method, \
        bank_account, vehicle_id, bus_pers, description, \
        receipt_source, display_color, created_from_banking, mapped_bank_id = row_data
        
        row = [
            receipt_id,
            parent_id if parent_id else '',
            receipt_date.strftime('%Y-%m-%d') if receipt_date else '',
            vendor or '',
            float(gross) if gross else 0.0,
            float(gst) if gst else 0.0,
            float(net) if net else 0.0,
            category or '',
            payment_method or '',
            bank_account or '',
            vehicle_id or '',
            bus_pers or '',
            description or '',
            receipt_source or '',
            display_color or '',
            'Yes' if created_from_banking else 'No'
        ]
        
        ws.append(row)
        row_count += 1
        
        # Apply color coding to the row
        current_row = ws.max_row
        fill_color = get_color_for_source(receipt_source, display_color)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for cell in ws[current_row]:
            cell.fill = fill
            if cell.column in [5, 6, 7]:  # Amount, GST, Net columns
                cell.number_format = '$#,##0.00'
        
        # Color bank account column
        bank_col = 11  # Bank Account column
        if mapped_bank_id:
            bank_color = get_bank_account_color(mapped_bank_id)
            ws.cell(row=current_row, column=bank_col).fill = PatternFill(
                start_color=bank_color, end_color=bank_color, fill_type='solid'
            )
        
        if row_count % 1000 == 0:
            print(f"  Loaded {row_count} receipts...")
    
    print(f"  Total receipts loaded: {row_count}")
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    cur.close()
    return row_count

def create_banking_transactions_sheet(wb, conn):
    """Create sheet with ALL banking transactions from all years."""
    ws = wb.create_sheet("Banking Transactions")
    
    # Headers
    headers = [
        'Transaction ID', 'Account Number', 'Transaction Date', 'Description', 
        'Debit', 'Credit', 'Balance', 'Vendor Extracted', 'Category',
        'Linked Receipt ID', 'Bank Account Name'
    ]
    
    ws.append(headers)
    apply_header_style(ws)
    
    # Query ALL banking transactions
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            bt.transaction_id,
            bt.account_number,
            bt.transaction_date,
            bt.description,
            bt.debit_amount,
            bt.credit_amount,
            bt.balance,
            bt.vendor_extracted,
            bt.category,
            bt.receipt_id,
            CASE 
                WHEN bt.account_number = '0228362' THEN 'CIBC Checking'
                WHEN bt.account_number = '903990106011' THEN 'Scotia Bank'
                ELSE bt.account_number
            END as account_name
        FROM banking_transactions bt
        ORDER BY bt.transaction_date DESC, bt.transaction_id DESC
    """)
    
    print(f"Loading ALL banking transactions...")
    row_count = 0
    
    for row_data in cur.fetchall():
        txn_id, acct_num, txn_date, description, debit, credit, balance, vendor, \
        category, receipt_id, acct_name = row_data
        
        row = [
            txn_id,
            acct_num or '',
            txn_date.strftime('%Y-%m-%d') if txn_date else '',
            description or '',
            float(debit) if debit else 0.0,
            float(credit) if credit else 0.0,
            float(balance) if balance else 0.0,
            vendor or '',
            category or '',
            receipt_id if receipt_id else '',
            acct_name or ''
        ]
        
        ws.append(row)
        row_count += 1
        
        # Apply color coding
        current_row = ws.max_row
        
        # Color by account
        if acct_num == '0228362':
            fill_color = 'C6EFCE'  # Green for CIBC
        elif acct_num == '903990106011':
            fill_color = 'FFEB9C'  # Yellow for Scotia
        else:
            fill_color = 'FFFFFF'
        
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for cell in ws[current_row]:
            cell.fill = fill
            if cell.column in [5, 6, 7]:  # Debit, Credit, Balance
                cell.number_format = '$#,##0.00'
        
        if row_count % 1000 == 0:
            print(f"  Loaded {row_count} banking transactions...")
    
    print(f"  Total banking transactions loaded: {row_count}")
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    cur.close()
    return row_count

def create_reference_sheet(wb, conn):
    """Create reference data sheet."""
    ws = wb.create_sheet("Reference Data")
    
    # Section 1: Vehicles
    ws.append(['VEHICLES'])
    apply_header_style(ws, ws.max_row)
    ws.append(['Vehicle ID', 'Unit Number', 'Type', 'Make', 'Model', 'Year', 'License Plate'])
    
    cur = conn.cursor()
    cur.execute("""
        SELECT vehicle_id, unit_number, vehicle_type, make, model, year, license_plate
        FROM vehicles
        ORDER BY unit_number
    """)
    
    for row in cur.fetchall():
        ws.append(list(row))
    
    ws.append([])  # Blank row
    
    # Section 2: Bank Accounts
    ws.append(['BANK ACCOUNTS'])
    apply_header_style(ws, ws.max_row)
    ws.append(['Account ID', 'Account Number', 'Account Name', 'Color Code'])
    
    accounts = [
        [1, '0228362', 'CIBC Checking', 'Green'],
        [2, '903990106011', 'Scotia Bank', 'Yellow/Orange']
    ]
    
    for acct in accounts:
        row_num = ws.max_row + 1
        ws.append(acct)
        if acct[0] == 1:
            fill_color = 'C6EFCE'  # Green
        else:
            fill_color = 'FFEB9C'  # Yellow
        
        for cell in ws[row_num]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    
    ws.append([])  # Blank row
    
    # Section 3: Payment Methods
    ws.append(['PAYMENT METHODS'])
    apply_header_style(ws, ws.max_row)
    ws.append(['Method', 'Description'])
    
    methods = [
        ['cash', 'Cash payment'],
        ['check', 'Cheque payment'],
        ['credit_card', 'Credit card'],
        ['debit_card', 'Debit card'],
        ['bank_transfer', 'E-transfer or bank transfer'],
        ['trade_of_services', 'Trade arrangement'],
        ['unknown', 'Unknown method']
    ]
    
    for method in methods:
        ws.append(method)
    
    # Auto-size columns
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    cur.close()

def create_instructions_sheet(wb):
    """Create instructions and color legend sheet."""
    ws = wb.create_sheet("Instructions")
    
    # Title
    ws.append(['RECEIPT WORKBOOK INSTRUCTIONS'])
    ws['A1'].font = Font(size=16, bold=True)
    ws.append([])
    
    # Color Legend
    ws.append(['COLOR LEGEND - Receipt Sources'])
    ws[ws.max_row][0].font = Font(size=14, bold=True)
    ws.append(['Color', 'Source', 'Meaning'])
    apply_header_style(ws, ws.max_row)
    
    legend = [
        ['GREEN', 'BANKING', 'Matched to banking transaction - fully reconciled'],
        ['YELLOW', 'CASH', 'Cash payment through cash box - no banking match expected'],
        ['ORANGE', 'REIMBURSEMENT', 'Employee reimbursement - handled through payroll'],
        ['BLUE', 'MANUAL', 'Manually entered - may need matching'],
        ['RED', 'UNMATCHED', 'Created from banking but link lost - needs investigation']
    ]
    
    for color_name, source, meaning in legend:
        row_num = ws.max_row + 1
        ws.append([color_name, source, meaning])
        
        if color_name == 'GREEN':
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif color_name == 'YELLOW':
            fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        elif color_name == 'ORANGE':
            fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif color_name == 'BLUE':
            fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        else:  # RED
            fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        for cell in ws[row_num]:
            cell.fill = fill
    
    ws.append([])
    
    # Split Receipt Instructions
    ws.append(['SPLIT RECEIPT WORKFLOW - Parent-Child Linking'])
    ws[ws.max_row][0].font = Font(size=14, bold=True)
    ws.append([])
    
    ws.append(['What is a Split Receipt?'])
    ws.append(['A combined receipt where some items are business expenses and others are personal.'])
    ws.append([])
    
    ws.append(['How to Enter Split Receipts:'])
    ws.append(['1. Enter the FULL receipt first (entire amount from receipt)'])
    ws.append(['2. Note the Receipt ID that gets assigned (e.g., 12345)'])
    ws.append(['3. Enter each SPLIT PORTION as separate receipts'])
    ws.append(['4. In the "Parent Receipt #" column, enter the Receipt ID from step 2'])
    ws.append(['5. Mark each split as "Business" or "Personal" in Business/Personal column'])
    ws.append([])
    
    ws.append(['Example: Costco Receipt for $200 (mixed business/personal)'])
    ws.append([])
    ws.append(['Receipt ID', 'Parent Receipt #', 'Vendor', 'Amount', 'Category', 'Business/Personal', 'Notes'])
    apply_header_style(ws, ws.max_row)
    
    ws.append([12345, '', 'Costco', 200.00, 'Mixed', 'Business', 'Full receipt - combined'])
    ws.append([12346, 12345, 'Costco', 60.00, 'Fuel', 'Business', 'Gas for vehicle'])
    ws.append([12347, 12345, 'Costco', 80.00, 'Office Supplies', 'Business', 'Printer paper'])
    ws.append([12348, 12345, 'Costco', 60.00, 'Groceries', 'Personal', 'Personal food'])
    
    ws.append([])
    ws.append(['Note: Split portions ($60 + $80 + $60 = $200) must equal parent total'])
    ws.append([])
    
    # GST Calculation
    ws.append(['GST CALCULATION - Tax INCLUDED in Amounts'])
    ws[ws.max_row][0].font = Font(size=14, bold=True)
    ws.append([])
    
    ws.append(['Alberta GST Rate: 5%'])
    ws.append(['IMPORTANT: GST is already INCLUDED in the total amount, not added on top.'])
    ws.append([])
    
    ws.append(['Example: You paid $20 total for fuel'])
    ws.append(['  • Gross Amount = $20.00 (what you actually paid)'])
    ws.append(['  • GST = $20.00 × 0.05 ÷ 1.05 = $0.95'])
    ws.append(['  • Net Amount = $20.00 - $0.95 = $19.05'])
    ws.append([])
    
    ws.append(['Formula: GST = Gross × 0.05 ÷ 1.05'])
    ws.append(['Formula: Net = Gross - GST'])
    ws.append([])
    
    # Bank Account Colors
    ws.append(['BANK ACCOUNT COLOR CODING'])
    ws[ws.max_row][0].font = Font(size=14, bold=True)
    ws.append(['Account', 'Number', 'Color'])
    apply_header_style(ws, ws.max_row)
    
    row_num = ws.max_row + 1
    ws.append(['CIBC Checking', '0228362', 'Green'])
    for cell in ws[row_num]:
        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    row_num = ws.max_row + 1
    ws.append(['Scotia Bank', '903990106011', 'Yellow/Orange'])
    for cell in ws[row_num]:
        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Auto-size columns
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 25

def main():
    """Generate complete receipts workbook with all years and all banking data."""
    print("\n" + "="*60)
    print("GENERATING COMPLETE RECEIPTS WORKBOOK")
    print("ALL RECEIPTS + ALL BANKING TRANSACTIONS")
    print("="*60 + "\n")
    
    conn = get_db_connection()
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all sheets
    print("\nCreating ALL RECEIPTS sheet...")
    receipt_count = create_all_receipts_sheet(wb, conn)
    
    print("\nCreating ALL BANKING TRANSACTIONS sheet...")
    banking_count = create_banking_transactions_sheet(wb, conn)
    
    print("\nCreating REFERENCE DATA sheet...")
    create_reference_sheet(wb, conn)
    
    print("\nCreating INSTRUCTIONS sheet...")
    create_instructions_sheet(wb)
    
    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f'l:/limo/reports/complete_receipts_workbook_{timestamp}.xlsx'
    
    wb.save(output_path)
    conn.close()
    
    print("\n" + "="*60)
    print("✅ WORKBOOK CREATED SUCCESSFULLY")
    print("="*60)
    print(f"\nLocation: {output_path}")
    print(f"\nSHEETS INCLUDED:")
    print(f"  1. All Receipts - {receipt_count:,} receipts (ALL years)")
    print(f"  2. Banking Transactions - {banking_count:,} transactions (ALL years)")
    print(f"  3. Reference Data - vehicles, bank accounts, payment methods")
    print(f"  4. Instructions - color legend and split receipt workflow")
    print(f"\nFEATURES:")
    print(f"  ✅ Receipt ID column (for parent-child linking)")
    print(f"  ✅ Parent Receipt # column (link split receipts)")
    print(f"  ✅ Color coding by receipt source (GREEN/YELLOW/ORANGE/BLUE/RED)")
    print(f"  ✅ Bank account color coding (CIBC green, Scotia yellow)")
    print(f"  ✅ GST and Net amount calculations")
    print(f"  ✅ Business/Personal indicator")
    print(f"  ✅ All banking transactions with receipt linkage")
    print("\n" + "="*60 + "\n")

if __name__ == '__main__':
    main()
