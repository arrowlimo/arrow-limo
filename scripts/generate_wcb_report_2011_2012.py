#!/usr/bin/env python3
"""
Generate WCB report for 2011-Dec 2012 in Excel format.
Includes invoices, payments, and running balance.
"""
import psycopg2
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Installing...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", os.environ.get("DB_PASSWORD"))

try:
    conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    cur = conn.cursor()
    
    print("\n" + "="*80)
    print("GENERATING WCB REPORT (2011 - Dec 2012)")
    print("="*80)
    
    # Get all WCB invoices
    print("\nQuerying WCB invoices...")
    cur.execute("""
        SELECT 
            'INVOICE' as type,
            receipt_date as transaction_date,
            receipt_id,
            source_reference as invoice_ref,
            gross_amount as amount,
            description,
            NULL as payment_method,
            banking_transaction_id
        FROM receipts
        WHERE vendor_name ILIKE '%wcb%'
          AND receipt_date >= '2011-01-01'
          AND receipt_date <= '2012-12-31'
          AND gross_amount > 0
        
        UNION ALL
        
        SELECT 
            'PAYMENT' as type,
            transaction_date,
            transaction_id,
            check_number,
            debit_amount as amount,
            description,
            NULL,
            transaction_id
        FROM banking_transactions
        WHERE description ILIKE '%wcb%'
          AND transaction_date >= '2011-01-01'
          AND transaction_date <= '2012-12-31'
          AND debit_amount > 0
        
        ORDER BY transaction_date ASC, type DESC
    """)
    
    rows = cur.fetchall()
    print(f"Found {len(rows)} WCB transactions")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WCB 2011-2012"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    invoice_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    payment_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Date", "Type", "Invoice/Check #", "Description", "Amount", "Running Balance", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Add data rows
    running_balance = 0.0
    row_num = 2
    
    for row_data in rows:
        trans_type, trans_date, trans_id, ref, amount, desc, payment_method, banking_id = row_data
        amount_f = float(amount)
        
        # Calculate running balance (invoices add, payments subtract)
        if trans_type == "INVOICE":
            running_balance += amount_f
            fill = invoice_fill
        else:
            running_balance -= amount_f
            fill = payment_fill
        
        # Determine status
        if trans_type == "INVOICE":
            # Check if this invoice is linked to a payment
            if banking_id:
                status = "PAID"
            else:
                status = "UNPAID"
        else:
            status = "RECEIVED"
        
        # Format date
        if isinstance(trans_date, str):
            date_str = trans_date
        else:
            date_str = trans_date.strftime('%Y-%m-%d')
        
        # Format description
        desc_str = (desc[:50] if desc else "")
        ref_str = str(ref) if ref else ""
        
        # Write row
        ws.cell(row=row_num, column=1).value = date_str
        ws.cell(row=row_num, column=2).value = trans_type
        ws.cell(row=row_num, column=3).value = ref_str
        ws.cell(row=row_num, column=4).value = desc_str
        ws.cell(row=row_num, column=5).value = amount_f
        ws.cell(row=row_num, column=6).value = running_balance
        ws.cell(row=row_num, column=7).value = status
        
        # Apply styles
        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Format numbers
            if col in [5, 6]:  # Amount columns
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        row_num += 1
    
    # Add summary row
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = total_font
    ws.cell(row=summary_row, column=1).fill = total_fill
    
    ws.cell(row=summary_row, column=5).value = "Final Balance:"
    ws.cell(row=summary_row, column=5).font = total_font
    ws.cell(row=summary_row, column=5).fill = total_fill
    ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")
    
    ws.cell(row=summary_row, column=6).value = running_balance
    ws.cell(row=summary_row, column=6).font = total_font
    ws.cell(row=summary_row, column=6).fill = total_fill
    ws.cell(row=summary_row, column=6).number_format = '$#,##0.00'
    
    # Count transactions
    invoice_count = sum(1 for r in rows if r[0] == "INVOICE")
    payment_count = sum(1 for r in rows if r[0] == "PAYMENT")
    total_invoiced = sum(float(r[4]) for r in rows if r[0] == "INVOICE")
    total_paid = sum(float(r[4]) for r in rows if r[0] == "PAYMENT")
    
    # Add statistics
    stats_row = summary_row + 2
    ws.cell(row=stats_row, column=1).value = "Invoices:"
    ws.cell(row=stats_row, column=2).value = invoice_count
    
    ws.cell(row=stats_row+1, column=1).value = "Total Invoiced:"
    ws.cell(row=stats_row+1, column=2).value = total_invoiced
    ws.cell(row=stats_row+1, column=2).number_format = '$#,##0.00'
    
    ws.cell(row=stats_row+2, column=1).value = "Payments:"
    ws.cell(row=stats_row+2, column=2).value = payment_count
    
    ws.cell(row=stats_row+3, column=1).value = "Total Paid:"
    ws.cell(row=stats_row+3, column=2).value = total_paid
    ws.cell(row=stats_row+3, column=2).number_format = '$#,##0.00'
    
    ws.cell(row=stats_row+4, column=1).value = "Remaining Balance:"
    ws.cell(row=stats_row+4, column=2).value = running_balance
    ws.cell(row=stats_row+4, column=2).number_format = '$#,##0.00'
    
    # Save file
    output_path = r"L:\limo\reports\WCB_2011_to_Dec_2012.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"\n" + "="*80)
    print("REPORT GENERATED")
    print("="*80)
    print(f"\nFile saved: {output_path}")
    print(f"\nReport Summary:")
    print(f"  Invoices: {invoice_count}")
    print(f"  Total Invoiced: ${total_invoiced:,.2f}")
    print(f"  Payments: {payment_count}")
    print(f"  Total Paid: ${total_paid:,.2f}")
    print(f"  Final Balance: ${running_balance:,.2f}")
    
    cur.close()
    conn.close()
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()
