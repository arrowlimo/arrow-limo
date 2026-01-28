#!/usr/bin/env python3
"""
Generate improved Excel report with color coding and vendor name extraction.
"""
import psycopg2
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

DB_HOST = 'localhost'
DB_NAME = 'almsdata'
DB_USER = 'postgres'
DB_PASSWORD = os.environ.get('DB_PASSWORD')

def extract_vendor_from_description(description):
    """Extract vendor name from banking description."""
    if not description:
        return "Unknown"
    
    # Common patterns to clean up
    desc = description.strip()
    
    # Remove common prefixes
    prefixes = [
        'DEBIT PURCHASE - ',
        'PURCHASE - ',
        'POS PURCHASE - ',
        'INTERNET BANKING E-TRANSFER ',
        'E-TRANSFER ',
        'INTERAC E-TRF- ',
        'VISA DEBIT PUR-',
        'INTERAC PUR ',
        'POS ',
    ]
    
    for prefix in prefixes:
        if desc.upper().startswith(prefix.upper()):
            desc = desc[len(prefix):].strip()
    
    # Take first part before common delimiters
    for delimiter in [' - ', ' #', ' REF:', ' CARD ', ' Date:', ' Time:']:
        if delimiter in desc:
            desc = desc.split(delimiter)[0]
    
    # Limit length
    if len(desc) > 50:
        desc = desc[:50].strip()
    
    return desc if desc else "Unknown"

def main():
    conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    
    print("="*70)
    print("GENERATING IMPROVED RECEIPTS EXCEL REPORT")
    print("="*70)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"L:\\limo\\reports\\receipts_color_coded_{timestamp}.xlsx"
    
    # Query with improved vendor names
    query = """
        SELECT 
            r.receipt_id,
            r.receipt_date as "Date",
            COALESCE(
                NULLIF(r.vendor_name, 'Unknown'),
                NULLIF(r.vendor_name, ''),
                bt.vendor_extracted,
                'Unknown'
            ) as "Vendor Name",
            r.description as "Description",
            
            -- Withdrawal (expense) and Deposit (revenue) columns
            CASE 
                WHEN r.gross_amount > 0 THEN r.gross_amount 
                ELSE NULL 
            END as "Withdrawal",
            CASE 
                WHEN r.revenue > 0 THEN r.revenue
                WHEN r.gross_amount < 0 THEN ABS(r.gross_amount)
                ELSE NULL 
            END as "Deposit",
            
            -- GST columns
            r.gst_amount as "GST Amount",
            r.net_amount as "Amount Less GST",
            CASE 
                WHEN r.gst_amount > 0 THEN 'GST'
                WHEN r.gst_amount = 0 THEN 'No GST'
                ELSE 'Exempt'
            END as "GST Category",
            
            -- Business classification
            COALESCE(r.business_personal, 'Business') as "Business/Personal",
            
            -- Categories
            r.category as "Category",
            r.sub_classification as "Subcategory",
            
            -- Payment details
            r.card_number as "Card Number",
            r.canonical_pay_method as "Payment Method",
            
            -- Vehicle info
            r.vehicle_number as "Vehicle Number",
            r.vehicle_id as "Vehicle ID",
            r.fuel_amount as "Fuel Amount (LTS)",
            
            -- Split receipt tracking
            r.is_split_receipt as "Is Split Receipt",
            r.parent_receipt_id as "Parent Receipt ID",
            r.split_key as "Split Group Key",
            
            -- Banking info
            CASE 
                WHEN bt.bank_id = 1 THEN 'CIBC 0228362'
                WHEN bt.bank_id = 2 THEN 'Scotia 903990106011'
                WHEN bt.bank_id = 4 THEN 'CIBC 1615'
                WHEN bt.source_file = 'cibc 8362 all.csv' THEN 'CIBC 8362 (QB)'
                ELSE COALESCE(bt.account_number, 'No Banking')
            END as "Bank Account",
            bt.bank_id,
            bt.source_file,
            
            -- Verification status
            CASE 
                WHEN r.is_verified_banking IS TRUE THEN 'VERIFIED'
                WHEN bt.source_file = 'cibc 8362 all.csv' THEN 'QB 8362 - REVIEW'
                WHEN r.potential_duplicate IS TRUE THEN 'DUPLICATE'
                ELSE 'OK'
            END as "Status",
            r.verified_source as "Verified Source",
            
            -- Additional fields
            r.comment as "Comments",
            r.source_system as "Source System",
            r.deductible_status as "Deductible Status"
            
        FROM receipts r
        LEFT JOIN banking_transactions bt ON r.banking_transaction_id = bt.transaction_id
        ORDER BY 
            bt.bank_id NULLS LAST,
            r.receipt_date,
            r.receipt_id
    """
    
    print("\nQuerying receipts data...")
    df = pd.read_sql_query(query, conn)
    
    # Extract better vendor names where needed
    print("Improving vendor names...")
    mask = (df['Vendor Name'] == 'Unknown') | (df['Vendor Name'].isna())
    df.loc[mask, 'Vendor Name'] = df.loc[mask, 'Description'].apply(extract_vendor_from_description)
    
    print(f"✅ Retrieved {len(df):,} receipts\n")
    
    # Create Excel writer
    print(f"Creating Excel file: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: All Receipts
        print("  Creating 'All Receipts' sheet...")
        all_df = df.drop(columns=['receipt_id', 'bank_id', 'source_file'])
        all_df.to_excel(writer, sheet_name='All Receipts', index=False)
        
        # Sheet 2: Verified Only
        print("  Creating 'Verified Banking Only' sheet...")
        verified = df[df['Status'] == 'VERIFIED'].drop(columns=['receipt_id', 'bank_id', 'source_file'])
        if len(verified) > 0:
            verified.to_excel(writer, sheet_name='Verified Banking Only', index=False)
            print(f"    → {len(verified):,} receipts")
        
        # Sheet 3: QB 8362 (needs review)
        print("  Creating 'QB 8362 - Review' sheet...")
        qb_8362 = df[df['Status'] == 'QB 8362 - REVIEW'].drop(columns=['receipt_id', 'bank_id', 'source_file'])
        if len(qb_8362) > 0:
            qb_8362.to_excel(writer, sheet_name='QB 8362 - Review', index=False)
            print(f"    → {len(qb_8362):,} receipts")
        
        # Sheet 4: By bank account
        for bank_name, bank_filter in [
            ('CIBC 0228362', df['bank_id'] == 1),
            ('Scotia 903990106011', df['bank_id'] == 2),
            ('CIBC 1615', df['bank_id'] == 4),
            ('No Banking Link', df['bank_id'].isna())
        ]:
            print(f"  Creating '{bank_name}' sheet...")
            bank_df = df[bank_filter].drop(columns=['receipt_id', 'bank_id', 'source_file'])
            if len(bank_df) > 0:
                bank_df.to_excel(writer, sheet_name=bank_name, index=False)
                print(f"    → {len(bank_df):,} receipts")
    
    # Apply color coding
    print("\nApplying color coding...")
    wb = load_workbook(output_file)
    
    # Color definitions
    colors = {
        'VERIFIED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Light green
        'QB 8362 - REVIEW': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Light yellow
        'OK': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),  # White
    }
    
    # Apply to "All Receipts" sheet
    if 'All Receipts' in wb.sheetnames:
        ws = wb['All Receipts']
        
        # Find Status column
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Status':
                status_col = idx
                break
        
        if status_col:
            for row in range(2, ws.max_row + 1):
                status = ws.cell(row, status_col).value
                if status in colors:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = colors[status]
    
    wb.save(output_file)
    
    file_size = os.path.getsize(output_file) / (1024*1024)
    
    print(f"\n{'='*70}")
    print("✅ IMPROVED EXCEL REPORT GENERATED!")
    print(f"{'='*70}")
    print(f"File: {output_file}")
    print(f"Size: {file_size:.2f} MB")
    print(f"Total receipts: {len(df):,}")
    print(f"\nColor Coding:")
    print(f"  🟢 Green = VERIFIED (clean from PDF banking)")
    print(f"  🟡 Yellow = QB 8362 - REVIEW (needs manual review)")
    print(f"  ⚪ White = OK (unique, no issues)")
    print(f"\nVendor names improved from banking descriptions where missing")
    
    conn.close()

if __name__ == '__main__':
    main()
