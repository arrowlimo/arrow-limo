#!/usr/bin/env python3
"""
Validate charter date/client against legacy LMS data.

Load Reserve.xlsx and verify each charter in database matches
the exact date and client from LMS for the same reserve number.
"""
import os
import psycopg2
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

DB_HOST = "localhost"
DB_NAME = "almsdata"
DB_USER = "postgres"
DB_PASSWORD = os.environ.get("DB_PASSWORD")

LMS_FILE = r"L:\limo\data\Reserve.xlsx"

def load_lms_data():
    """Load reserve numbers with date and client from LMS Excel file."""
    print(f"Loading LMS data from {LMS_FILE}...")
    
    if not os.path.exists(LMS_FILE):
        print(f"❌ File not found: {LMS_FILE}")
        return {}
    
    lms_data = {}
    
    try:
        wb = load_workbook(LMS_FILE)
        ws = wb.active
        
        # Print first few rows to understand structure
        print(f"\nSheet: {ws.title}")
        print(f"First 5 rows to determine structure:")
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            print(f"  Row {i}: {row}")
        
        # Try to identify columns
        headers = {}
        header_row = ws[1]
        for col_idx, cell in enumerate(header_row, 1):
            if cell.value:
                headers[cell.value.lower().strip()] = col_idx
                print(f"  Column {col_idx}: {cell.value}")
        
        print(f"\nDetected columns: {headers}")
        
        # Look for reserve, date, client columns
        reserve_col = None
        date_col = None
        client_col = None
        
        for key in headers:
            if 'reserve' in key:
                reserve_col = headers[key]
            if 'pu_date' in key:  # Pickup date
                date_col = headers[key]
            if 'name' in key and 'client' not in key:  # First match for Name is typically client
                client_col = headers[key]
        
        if not reserve_col:
            print("⚠️  Could not identify reserve column")
            return {}
        
        print(f"\nParsing: Reserve col={reserve_col}, Date col={date_col}, Client col={client_col}")
        
        # Load data
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                reserve = row[reserve_col - 1]
                date_val = row[date_col - 1] if date_col else None
                client_val = row[client_col - 1] if client_col else None
                
                if reserve:
                    # Normalize reserve to 6-digit string
                    if isinstance(reserve, int):
                        reserve = f"{reserve:06d}"
                    else:
                        reserve = str(reserve).strip().zfill(6)
                    
                    # Normalize date
                    if date_val:
                        if isinstance(date_val, datetime):
                            date_val = date_val.date()
                        else:
                            try:
                                date_val = datetime.strptime(str(date_val), "%Y-%m-%d").date()
                            except:
                                date_val = None
                    
                    client_val = str(client_val).strip() if client_val else ""
                    
                    lms_data[reserve] = {
                        'date': date_val,
                        'client': client_val,
                        'row': row_idx
                    }
            except Exception as e:
                if row_idx <= 10:
                    print(f"  Row {row_idx} parse error: {e}")
        
        print(f"[OK] Loaded {len(lms_data)} reserve records from LMS")
        return lms_data
        
    except Exception as e:
        print(f"[ERROR] Error loading LMS file: {e}")
        import traceback
        traceback.print_exc()
        return {}

def main():
    # Load LMS data
    lms_data = load_lms_data()
    if not lms_data:
        print("Cannot proceed without LMS data")
        return
    
    # Connect to database
    try:
        conn = psycopg2.connect(host=DB_HOST, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)
        cur = conn.cursor()
        
        print("\n" + "=" * 100)
        print("VALIDATING CHARTERS AGAINST LMS DATA")
        print("=" * 100)
        
        # Get all charters with client info
        cur.execute("""
            SELECT c.charter_id, c.reserve_number, c.charter_date, 
                   COALESCE(cl.company_name, cl.first_name || ' ' || cl.last_name, 'UNKNOWN') as client_name
            FROM charters c
            LEFT JOIN clients cl ON c.client_id = cl.client_id
            ORDER BY c.reserve_number, c.charter_date
        """)
        
        charters = cur.fetchall()
        print(f"\nLoaded {len(charters):,} charters from database\n")
        
        # Validate each charter
        matches = 0
        mismatches = defaultdict(list)
        not_in_lms = []
        
        for charter_id, reserve_number, charter_date, client_name in charters:
            if not reserve_number:
                continue
                
            reserve_str = str(reserve_number).strip().zfill(6)
            
            if reserve_str not in lms_data:
                not_in_lms.append({
                    'charter_id': charter_id,
                    'reserve': reserve_str,
                    'date': charter_date,
                    'client': client_name
                })
                continue
            
            lms_record = lms_data[reserve_str]
            lms_date = lms_record['date']
            lms_client = lms_record['client']
            
            # Check for mismatches
            date_match = (charter_date == lms_date) if lms_date else True
            client_match = (client_name or "").strip().upper() == lms_client.upper() if lms_client else True
            
            if date_match and client_match:
                matches += 1
            else:
                mismatches[reserve_str].append({
                    'charter_id': charter_id,
                    'db_date': charter_date,
                    'lms_date': lms_date,
                    'date_match': date_match,
                    'db_client': client_name,
                    'lms_client': lms_client,
                    'client_match': client_match
                })
        
        # Report results
        print(f"RESULTS")
        print("-" * 100)
        print(f"[OK] Matching (date & client): {matches:>6}")
        print(f"[ERROR] Date/Client mismatches:  {len(mismatches):>6}")
        print(f"[WARN] Not in LMS at all:       {len(not_in_lms):>6}")
        print(f"   TOTAL CHARTERS:          {len(charters):>6}")
        
        # Show mismatches
        if mismatches:
            print(f"\n{'MISMATCHED CHARTERS':<100}")
            print("-" * 100)
            for reserve, records in sorted(mismatches.items())[:30]:
                for rec in records:
                    print(f"\nReserve: {reserve}")
                    print(f"  Charter ID: {rec['charter_id']}")
                    print(f"  Date:  DB={rec['db_date']}  LMS={rec['lms_date']}  {'[OK]' if rec['date_match'] else '[MISMATCH]'}")
                    print(f"  Client: DB='{rec['db_client']}'  LMS='{rec['lms_client']}'  {'[OK]' if rec['client_match'] else '[MISMATCH]'}")
        
        # Show not in LMS
        if not_in_lms:
            print(f"\n{'CHARTERS NOT IN LMS (sample of 30)':<100}")
            print("-" * 100)
            for rec in not_in_lms[:30]:
                print(f"{rec['reserve']}  {rec['date']}  {rec['client']:<30}  (charter_id={rec['charter_id']})")
            if len(not_in_lms) > 30:
                print(f"  ... and {len(not_in_lms) - 30} more")
        
        # Summary statistics
        print(f"\n{'SUMMARY STATISTICS':<100}")
        print("-" * 100)
        if matches + len(mismatches) > 0:
            match_pct = 100 * matches / (matches + len(mismatches))
            print(f"Match rate (valid reserves only): {match_pct:.1f}%")
        print(f"Charters with invalid reserves: {len(not_in_lms):,}")
        print(f"Total invalid + mismatched: {len(not_in_lms) + len(mismatches):,}")
        
        cur.close()
        conn.close()
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
