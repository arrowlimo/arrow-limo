#!/usr/bin/env python3
"""
Generate comprehensive receipt Excel workbook with all relevant business columns.
Includes fuel receipts with vehicle/charter linking, payment method tracking, and account matching.
"""

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

conn = psycopg2.connect(host='localhost', database='almsdata', user='postgres', password='***REMOVED***')
cur = conn.cursor()

print('='*80)
print('GENERATING COMPREHENSIVE RECEIPT WORKBOOK')
print('='*80)
print()

wb = Workbook()

# ============================================================================
# SHEET 1: FUEL RECEIPTS (for fuel entry and tracking)
# ============================================================================
print('Creating FUEL RECEIPTS sheet...')
ws_fuel = wb.active
ws_fuel.title = 'Fuel Receipts'

fuel_headers = [
    'Receipt Date',
    'Vendor',
    'Amount',
    'Fuel Type',
    'Litres',
    'Price/Litre',
    'Vehicle ID/Unit',
    'Vehicle Type',
    'Odometer Reading',
    'Charter #',
    'Driver',
    'Card Used (Last 4)',
    'Bank Account',
    'Category',
    'Notes'
]

# Add header row
for col_num, header in enumerate(fuel_headers, 1):
    cell = ws_fuel.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Get some sample fuel receipts
cur.execute("""
    SELECT 
        r.receipt_id,
        r.receipt_date,
        r.vendor_name,
        r.gross_amount,
        r.category,
        r.description,
        r.display_color,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC'
            WHEN bt.account_number = '903990106011' THEN 'Scotia'
            ELSE 'Unknown'
        END as bank_account
    FROM receipts r
    LEFT JOIN banking_transactions bt ON bt.transaction_id = r.banking_transaction_id
    WHERE (
        r.vendor_name ILIKE '%CENTEX%' 
        OR r.vendor_name ILIKE '%FAS GAS%'
        OR r.vendor_name ILIKE '%SHELL%'
        OR r.category = 'fuel'
    )
    AND r.receipt_date >= '2025-01-01'
    ORDER BY r.receipt_date DESC
    LIMIT 50
""")

fuel_data = cur.fetchall()

for row_num, (receipt_id, receipt_date, vendor, amount, category, description, color, bank_acct) in enumerate(fuel_data, 2):
    ws_fuel.cell(row=row_num, column=1).value = receipt_date
    ws_fuel.cell(row=row_num, column=2).value = vendor or ''
    ws_fuel.cell(row=row_num, column=3).value = float(amount) if amount else 0
    ws_fuel.cell(row=row_num, column=4).value = 'Gasoline'  # Default, user can change
    ws_fuel.cell(row=row_num, column=5).value = None  # Litres - user fills in
    ws_fuel.cell(row=row_num, column=6).value = None  # Price/Litre - auto calculate
    ws_fuel.cell(row=row_num, column=7).value = None  # Vehicle ID - user fills in
    ws_fuel.cell(row=row_num, column=8).value = None  # Vehicle Type - user fills in
    ws_fuel.cell(row=row_num, column=9).value = None  # Odometer - user fills in
    ws_fuel.cell(row=row_num, column=10).value = None  # Charter # - user fills in
    ws_fuel.cell(row=row_num, column=11).value = None  # Driver - user fills in
    ws_fuel.cell(row=row_num, column=12).value = None  # Card Last 4 - user fills in
    ws_fuel.cell(row=row_num, column=13).value = bank_acct
    ws_fuel.cell(row=row_num, column=14).value = category or ''
    ws_fuel.cell(row=row_num, column=15).value = f'Receipt ID: {receipt_id}'
    
    # Color code by bank account
    if bank_acct == 'CIBC':
        ws_fuel.cell(row=row_num, column=13).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    elif bank_acct == 'Scotia':
        ws_fuel.cell(row=row_num, column=13).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

# Set column widths
ws_fuel.column_dimensions['A'].width = 12
ws_fuel.column_dimensions['B'].width = 20
ws_fuel.column_dimensions['C'].width = 10
ws_fuel.column_dimensions['D'].width = 12
ws_fuel.column_dimensions['E'].width = 10
ws_fuel.column_dimensions['F'].width = 12
ws_fuel.column_dimensions['G'].width = 15
ws_fuel.column_dimensions['H'].width = 15
ws_fuel.column_dimensions['I'].width = 15
ws_fuel.column_dimensions['J'].width = 12
ws_fuel.column_dimensions['K'].width = 15
ws_fuel.column_dimensions['L'].width = 15
ws_fuel.column_dimensions['M'].width = 12
ws_fuel.column_dimensions['N'].width = 12
ws_fuel.column_dimensions['O'].width = 20

print(f'  ✅ Added {len(fuel_data)} fuel receipts')

# ============================================================================
# SHEET 2: EXPENSE RECEIPTS (for general expenses)
# ============================================================================
print('Creating EXPENSE RECEIPTS sheet...')
ws_expense = wb.create_sheet('Expense Receipts')

expense_headers = [
    'Receipt Date',
    'Vendor',
    'Amount',
    'GST Amount',
    'Net Amount',
    'Category',
    'Payment Method',
    'Card Last 4 / Cash / NSF',
    'Bank Account',
    'Employee',
    'Charter #',
    'Vehicle ID',
    'Description',
    'Notes'
]

for col_num, header in enumerate(expense_headers, 1):
    cell = ws_expense.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Get expense receipts
cur.execute("""
    SELECT 
        r.receipt_id,
        r.receipt_date,
        r.vendor_name,
        r.gross_amount,
        r.gst_amount,
        r.net_amount,
        r.category,
        r.receipt_source,
        r.description,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC'
            WHEN bt.account_number = '903990106011' THEN 'Scotia'
            ELSE 'Unknown'
        END as bank_account,
        r.display_color
    FROM receipts r
    LEFT JOIN banking_transactions bt ON bt.transaction_id = r.banking_transaction_id
    WHERE r.receipt_date >= '2025-01-01'
    AND (r.category != 'fuel' OR r.category IS NULL)
    ORDER BY r.receipt_date DESC
    LIMIT 100
""")

expense_data = cur.fetchall()

for row_num, (receipt_id, receipt_date, vendor, gross, gst, net, category, source, description, bank_acct, color) in enumerate(expense_data, 2):
    ws_expense.cell(row=row_num, column=1).value = receipt_date
    ws_expense.cell(row=row_num, column=2).value = vendor or ''
    ws_expense.cell(row=row_num, column=3).value = float(gross) if gross else 0
    ws_expense.cell(row=row_num, column=4).value = float(gst) if gst else 0
    ws_expense.cell(row=row_num, column=5).value = float(net) if net else 0
    ws_expense.cell(row=row_num, column=6).value = category or ''
    ws_expense.cell(row=row_num, column=7).value = source or ''
    ws_expense.cell(row=row_num, column=8).value = None  # Card/Cash/NSF - user fills in
    ws_expense.cell(row=row_num, column=9).value = bank_acct
    ws_expense.cell(row=row_num, column=10).value = None  # Employee - user fills in
    ws_expense.cell(row=row_num, column=11).value = None  # Charter # - user fills in
    ws_expense.cell(row=row_num, column=12).value = None  # Vehicle ID - user fills in
    ws_expense.cell(row=row_num, column=13).value = description or ''
    ws_expense.cell(row=row_num, column=14).value = f'Receipt ID: {receipt_id}'
    
    # Color code by display_color
    color_map = {
        'GREEN': "E2EFDA",
        'YELLOW': "FEF2CC",
        'ORANGE': "FCE4D6",
        'BLUE': "DEEBF7",
        'RED': "FCE4E4"
    }
    if color in color_map:
        for col in range(1, 15):
            ws_expense.cell(row=row_num, column=col).fill = PatternFill(start_color=color_map[color], end_color=color_map[color], fill_type="solid")

# Set column widths
ws_expense.column_dimensions['A'].width = 12
ws_expense.column_dimensions['B'].width = 20
ws_expense.column_dimensions['C'].width = 10
ws_expense.column_dimensions['D'].width = 10
ws_expense.column_dimensions['E'].width = 10
ws_expense.column_dimensions['F'].width = 15
ws_expense.column_dimensions['G'].width = 15
ws_expense.column_dimensions['H'].width = 18
ws_expense.column_dimensions['I'].width = 12
ws_expense.column_dimensions['J'].width = 15
ws_expense.column_dimensions['K'].width = 12
ws_expense.column_dimensions['L'].width = 12
ws_expense.column_dimensions['M'].width = 20
ws_expense.column_dimensions['N'].width = 20

print(f'  ✅ Added {len(expense_data)} expense receipts')

# ============================================================================
# SHEET 3: REFERENCE DATA (Vehicles, Cards, Payment Methods)
# ============================================================================
print('Creating REFERENCE DATA sheet...')
ws_ref = wb.create_sheet('Reference Data')

ws_ref['A1'] = 'VEHICLES'
ws_ref['A1'].font = Font(bold=True, size=12, color="FFFFFF")
ws_ref['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

# Get vehicles
cur.execute("""
    SELECT vehicle_id, unit_number, vehicle_type, make, model, license_plate
    FROM vehicles
    ORDER BY unit_number
    LIMIT 30
""")

row = 2
for vehicle_id, unit_num, vtype, make, model, plate in cur.fetchall():
    ws_ref.cell(row=row, column=1).value = f'{unit_num}: {make} {model} ({plate})'
    ws_ref.cell(row=row, column=2).value = vehicle_id
    row += 1

ws_ref['D1'] = 'PAYMENT METHODS'
ws_ref['D1'].font = Font(bold=True, size=12, color="FFFFFF")
ws_ref['D1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

payment_methods = [
    'CIBC Card (Last 4)',
    'Scotia Card (Last 4)',
    'Cash',
    'NSF Charge',
    'E-Transfer',
    'Check'
]

for row, method in enumerate(payment_methods, 2):
    ws_ref.cell(row=row, column=4).value = method

ws_ref['G1'] = 'BANK ACCOUNTS'
ws_ref['G1'].font = Font(bold=True, size=12, color="FFFFFF")
ws_ref['G1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

accounts = ['CIBC 0228362', 'Scotia 903990106011']
for row, acct in enumerate(accounts, 2):
    ws_ref.cell(row=row, column=7).value = acct

ws_ref.column_dimensions['A'].width = 35
ws_ref.column_dimensions['B'].width = 12
ws_ref.column_dimensions['D'].width = 20
ws_ref.column_dimensions['G'].width = 20

print('  ✅ Added reference data')

# ============================================================================
# SHEET 4: COLOR LEGEND
# ============================================================================
print('Creating COLOR LEGEND sheet...')
ws_legend = wb.create_sheet('Color Legend')

ws_legend['A1'] = 'RECEIPT COLOR CODING'
ws_legend['A1'].font = Font(bold=True, size=14)

colors = [
    ('GREEN', 'E2EFDA', 'Matched to banking transaction - fully reconciled'),
    ('YELLOW', 'FEF2CC', 'Cash payment - no banking match expected'),
    ('ORANGE', 'FCE4D6', 'Employee reimbursement'),
    ('BLUE', 'DEEBF7', 'Manually entered - may need matching'),
    ('RED', 'FCE4E4', 'Unmatched - needs investigation')
]

for row, (color_name, hex_color, description) in enumerate(colors, 3):
    cell = ws_legend.cell(row=row, column=1)
    cell.value = color_name
    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    cell.font = Font(bold=True)
    
    ws_legend.cell(row=row, column=2).value = description

ws_legend['A10'] = 'BANK ACCOUNT MAPPING'
ws_legend['A10'].font = Font(bold=True, size=12)

ws_legend['A12'].value = 'CIBC'
ws_legend['A12'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ws_legend['B12'].value = '0228362 (Primary operating account)'

ws_legend['A13'].value = 'Scotia'
ws_legend['A13'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ws_legend['B13'].value = '903990106011 (Secondary account)'

ws_legend.column_dimensions['A'].width = 15
ws_legend.column_dimensions['B'].width = 50

print('  ✅ Added color legend')

# Save workbook
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_path = f'l:/limo/reports/receipts_workbook_{timestamp}.xlsx'

wb.save(output_path)

print()
print('='*80)
print('✅ WORKBOOK CREATED SUCCESSFULLY')
print('='*80)
print(f'Location: {output_path}')
print()
print('SHEETS INCLUDED:')
print('  1. Fuel Receipts - Track fuel purchases with vehicle/charter/litres')
print('  2. Expense Receipts - General expenses with payment method tracking')
print('  3. Reference Data - Vehicles, payment methods, bank accounts')
print('  4. Color Legend - Color coding meanings')
print()
print('COLUMNS INCLUDE:')
print('  ✅ Receipt date, vendor, amount')
print('  ✅ Fuel litres, price per litre (for fuel sheet)')
print('  ✅ Vehicle ID/Unit and vehicle type')
print('  ✅ Charter number for linking')
print('  ✅ Driver name')
print('  ✅ Card Last 4 / Cash / NSF / other payment method')
print('  ✅ Bank account (color coded)')
print('  ✅ GST/net amounts (for expense sheet)')
print('  ✅ Employee/charter/vehicle linking')
print()

cur.close()
conn.close()
