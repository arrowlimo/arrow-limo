"""
Extract 2012 Scotia Bank statements from PDF to Excel with validation.
Handles OCR irregularities, multi-row descriptions, and balance verification.
Uses OCR for scanned PDFs.
"""

from pdf2image import convert_from_path
import pdfplumber
import pytesseract
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PIL import Image
import os
import io

# Vendor normalization patterns
VENDOR_PATTERNS = {
    r'MOHAWK': 'Mohawk',
    r'CENTEX': 'Centex',
    r'GEORGE[\'S]*\s*PIZZA': 'George\'s Pizza',
    r'FAS\s*GAS': 'Fas Gas',
    r'PETRO[\s-]*CAN(?:ADA)?': 'Petro-Canada',
    r'CO[\s-]*OP': 'Co-op',
    r'SHELL': 'Shell',
    r'ESSO': 'Esso',
    r'7[\s-]*ELEVEN': '7-Eleven',
    r'TELUS': 'Telus',
    r'ENMAX': 'Enmax',
    r'DIRECT\s*ENERGY': 'Direct Energy',
    r'BEST\s*BUY': 'Best Buy',
    r'VISA': 'Visa',
    r'INTERAC': 'Interac',
    r'PAYROLL': 'Payroll',
    r'TRANSFER': 'Transfer',
    r'SERVICE\s*CHARGE': 'Service Charge',
    r'NSF': 'NSF Fee',
    r'CHEQUE\s*ORDER': 'Cheque Order (GST included)',
    r'SBAP\s*FEE': 'SBAP Fee',
    r'DEBIT\s*MEMO': 'Debit Memo',
    r'MONEY\s*ORDER': 'Money Order',
    r'CHASE\s*PAYMENTECH': 'Chase Paymentech',
    r'AMEX': 'American Express',
}

def clean_vendor_name(description):
    """Normalize vendor names using patterns."""
    if not description:
        return description
    
    desc_upper = description.upper()
    for pattern, replacement in VENDOR_PATTERNS.items():
        if re.search(pattern, desc_upper):
            return replacement
    
    return description.strip()

def extract_tables_from_pdf(pdf_path):
    """Extract transaction lines from PDF using OCR."""
    print(f"Opening PDF: {pdf_path}")
    all_lines = []
    
    # Set tesseract path if on Windows
    if os.name == 'nt':
        # Common Windows installation paths
        for possible_path in [
            r'C:\Program Files\Tesseract-OCR\tesseract.exe',
            r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
            r'C:\Users\pdric\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'
        ]:
            if os.path.exists(possible_path):
                pytesseract.pytesseract.tesseract_cmd = possible_path
                print(f"Found Tesseract at: {possible_path}")
                break
    
    print("Processing PDF pages with OCR...")
    
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"Total pages: {total_pages}")
        
        for page_num, page in enumerate(pdf.pages, 1):
            if page_num % 5 == 0 or page_num == 1:
                print(f"  OCR processing page {page_num}/{total_pages}...")
            
            # Convert PDF page to image
            try:
                im = page.to_image(resolution=300)
                pil_image = im.original
                
                # Perform OCR
                text = pytesseract.image_to_string(pil_image, config='--psm 6')
                
                if not text:
                    continue
                    
                # Split into lines
                lines = text.split('\n')
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Keep all non-header lines for grouping
                    # We'll filter later
                    if not any(x in line.upper() for x in ['SCOTIABANK', 'ACCOUNT', 'STATEMENT', 'TOTAL AMOUNT', 'NO. OF']):
                        all_lines.append((page_num, line))
                        
            except Exception as e:
                print(f"    Error on page {page_num}: {e}")
                continue
                
    print(f"  Total lines extracted: {len(all_lines)}")
    return all_lines

def parse_table_rows(all_lines):
    """
    Parse extracted lines into transactions.
    Handles multi-line descriptions where vendor info is on separate OCR lines.
    """
    transactions = []
    prev_balance = None
    current_date = None
    transaction_buffer = []  # Buffer for multi-line transactions
    
    for page_num, line in all_lines:
        # Check if this is a header or footer
        if any(x in line.upper() for x in ['BALANCE FORWARD', 'DESCRIPTION', 'WITHDRAWAL', 'DEPOSIT', 'ENCLOSURES', 'UNCOLLECTED', 'ODI OWING', 'FORWARD']):
            # Finalize previous transaction if exists
            if transaction_buffer:
                txn = finalize_transaction(transaction_buffer, current_date, prev_balance)
                if txn and txn['description']:
                    transactions.append(txn)
                    prev_balance = txn['balance']
                    if txn['date']:
                        current_date = txn['date']
            transaction_buffer = []
            continue
        
        # Check if line starts a new transaction (has numbers that look like amounts)
        has_amounts = bool(re.search(r'\d{2,}', line))
        
        if has_amounts and not transaction_buffer:
            # Start new transaction
            transaction_buffer = [line]
        elif has_amounts and transaction_buffer:
            # Line with amounts - might be continuation or new transaction
            # Check if it looks like a continuation (no transaction keyword at start, or location info)
            is_location = any(x in line.upper() for x in ['RED DEER', 'ABCA', 'ABCD', 'BRANCH', 'DEPOT', 'STREET', 'GRCRY', 'DEERPARK'])
            
            if is_location or (not re.match(r'^[A-Z][A-Z\s]*[A-Z]', line)):
                # Continuation - add to buffer
                transaction_buffer.append(line)
            else:
                # New transaction - finalize previous
                txn = finalize_transaction(transaction_buffer, current_date, prev_balance)
                if txn and txn['description']:
                    transactions.append(txn)
                    prev_balance = txn['balance']
                    if txn['date']:
                        current_date = txn['date']
                transaction_buffer = [line]
        else:
            # No amounts - continuation line (like vendor address)
            if transaction_buffer:
                transaction_buffer.append(line)
    
    # Finalize last transaction
    if transaction_buffer:
        txn = finalize_transaction(transaction_buffer, current_date, prev_balance)
        if txn and txn['description']:
            transactions.append(txn)
    
    return transactions

def finalize_transaction(buffer, current_date, prev_balance):
    """
    Finalize a transaction from buffered lines.
    Combines multi-line descriptions and extracts amounts.
    """
    if not buffer:
        return None
    
    # Join all lines
    full_text = ' '.join(buffer)
    full_text = full_text.replace('|', ' ').replace('  ', ' ').strip()
    
    # Extract description (text before numbers)
    desc_match = re.match(r'^([^0-9]+?)(\d|$)', full_text)
    if not desc_match:
        return None
    
    description = desc_match.group(1).strip()
    remaining = full_text[len(description):].strip()
    
    # Skip if description is too short or is a header
    if len(description) < 3 or description.upper() in ['DESCRIPTION', 'BALANCE FORWARD']:
        return None
    
    # Extract all numbers
    numbers = re.findall(r'\d+', remaining)
    if not numbers:
        return None
    
    # Look for 4-digit date (MMDD)
    date_str = current_date
    date_found = False
    for i, num in enumerate(numbers):
        if len(num) == 4:
            mm = int(num[:2])
            dd = int(num[2:])
            if 1 <= mm <= 12 and 1 <= dd <= 31:
                date_str = f"{num[:2]}/{num[2:]}"
                current_date = date_str
                date_found = True
                # Remove date from list for amount parsing
                numbers = numbers[:i] + numbers[i+1:]
                break
    
    # Parse amounts
    withdrawal = None
    deposit = None
    balance = None
    
    if len(numbers) >= 2:
        # Last 1-2 numbers are balance
        if len(numbers[-1]) <= 4:
            bal_dollars = numbers[-2]
            bal_cents = numbers[-1].ljust(2, '0')[:2]
            try:
                balance = float(f"{bal_dollars}.{bal_cents}")
                numbers = numbers[:-2]
            except:
                pass
        
        # Remaining are withdrawal/deposit
        if len(numbers) >= 2:
            amt1_dollars = numbers[0]
            amt1_cents = numbers[1].ljust(2, '0')[:2] if len(numbers) > 1 else '00'
            try:
                amt1 = float(f"{amt1_dollars}.{amt1_cents}")
                
                if len(numbers) >= 4:
                    amt2_dollars = numbers[2]
                    amt2_cents = numbers[3].ljust(2, '0')[:2]
                    amt2 = float(f"{amt2_dollars}.{amt2_cents}")
                    withdrawal = amt1
                    deposit = amt2
                else:
                    # Determine debit vs credit from description
                    if 'DEPOSIT' in description.upper() or 'CREDIT' in description.upper():
                        deposit = amt1
                    else:
                        withdrawal = amt1
            except:
                pass
    
    # Extract cheque number from description
    cheque_num = None
    cheque_match = re.search(r'\b(\d{3,6})\b', description)
    if cheque_match:
        cheque_num = cheque_match.group(1)
    
    # Clean description - remove vendor location info that got concatenated
    description = clean_vendor_name(description)
    
    return {
        'date': date_str or '',
        'description': description,
        'cheque_num': cheque_num,
        'debit': withdrawal,
        'credit': deposit,
        'balance': balance
    }

def create_excel_with_validation(transactions, output_path):
    """Create Excel file with auto-balance calculation and validation."""
    
    # Create DataFrame
    df_data = []
    running_balance = None
    
    for i, txn in enumerate(transactions):
        # Clean vendor name
        clean_desc = clean_vendor_name(txn['description'])
        
        # Note if cheque order (GST included)
        gst_note = ''
        if 'CHEQUE ORDER' in txn['description'].upper():
            gst_note = 'GST charged - match in receipts'
        
        # Parse amounts
        debit = float(txn['debit']) if txn['debit'] else None
        credit = float(txn['credit']) if txn['credit'] else None
        hard_balance = float(txn['balance']) if txn['balance'] else None
        
        # Calculate running balance
        if i == 0:
            running_balance = hard_balance if hard_balance else 0
        else:
            if running_balance is not None:
                if debit:
                    running_balance -= debit
                if credit:
                    running_balance += credit
            elif hard_balance:
                running_balance = hard_balance
        
        # Verify balance
        balance_match = ''
        if hard_balance and running_balance:
            diff = abs(hard_balance - running_balance)
            if diff > 0.01:  # Allow 1 cent rounding
                balance_match = f'ERROR: Δ{diff:.2f}'
            else:
                balance_match = 'OK'
        
        df_data.append({
            'Date': txn['date'],
            'Description': clean_desc,
            'Cheque #': txn['cheque_num'] or '',
            'Cheque Payee Lookup': 'LOOKUP NEEDED' if txn['cheque_num'] else '',
            'Debit': debit,
            'Credit': credit,
            'Hard Balance': hard_balance,
            'Auto Balance': running_balance,
            'Verification': balance_match,
            'GST Note': gst_note,
            'Notes': ''
        })
    
    df = pd.DataFrame(df_data)
    
    # Create Excel with formatting
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scotia 2012 Transactions"
    
    # Write headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row in openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    # Format columns
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 40  # Description
    ws.column_dimensions['C'].width = 10  # Cheque #
    ws.column_dimensions['D'].width = 20  # Cheque Payee
    ws.column_dimensions['E'].width = 12  # Debit
    ws.column_dimensions['F'].width = 12  # Credit
    ws.column_dimensions['G'].width = 14  # Hard Balance
    ws.column_dimensions['H'].width = 14  # Auto Balance
    ws.column_dimensions['I'].width = 18  # Verification
    ws.column_dimensions['J'].width = 25  # GST Note
    ws.column_dimensions['K'].width = 30  # Notes
    
    # Format currency columns
    currency_format = '#,##0.00'
    for row in range(2, len(df) + 2):
        for col in ['E', 'F', 'G', 'H']:
            cell = ws[f'{col}{row}']
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal='right')
    
    # Highlight errors
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    error_font = Font(color="9C0006", bold=True)
    
    for row in range(2, len(df) + 2):
        verification_cell = ws[f'I{row}']
        if verification_cell.value and 'ERROR' in str(verification_cell.value):
            verification_cell.fill = error_fill
            verification_cell.font = error_font
    
    # Highlight GST notes
    gst_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for row in range(2, len(df) + 2):
        gst_cell = ws[f'J{row}']
        if gst_cell.value:
            gst_cell.fill = gst_fill
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Excel file created: {output_path}")
    print(f"   Rows: {len(df)}")
    print(f"   Ready for manual editing")

def main():
    pdf_path = Path(r"L:\limo\pdf\2012\2012 scotiabank statements all.pdf")
    output_path = Path(r"L:\limo\data\2012_scotia_transactions_for_editing.xlsx")
    
    print("=" * 80)
    print("2012 Scotia Bank Statement Extraction")
    print("=" * 80)
    
    # Extract lines
    print("\n📄 Extracting lines from PDF...")
    all_lines = extract_tables_from_pdf(pdf_path)
    
    # Parse transactions
    print("\n🔍 Parsing transactions from lines...")
    transactions = parse_table_rows(all_lines)
    print(f"   Found {len(transactions)} transactions")
    
    # Create Excel
    print("\n📊 Creating Excel file with validation...")
    create_excel_with_validation(transactions, output_path)
    
    print("\n" + "=" * 80)
    print("INSTRUCTIONS FOR MANUAL EDITING:")
    print("=" * 80)
    print("1. Review 'Verification' column for balance errors")
    print("2. Fill in 'Cheque Payee Lookup' for all cheques")
    print("3. Check 'GST Note' - match cheque orders in receipts table")
    print("4. Use 'Notes' column for any observations")
    print("5. Auto Balance formula validates against Hard Balance")
    print("=" * 80)

if __name__ == '__main__':
    main()
