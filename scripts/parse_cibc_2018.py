"""
CIBC 8362 2018 - Parse CSV and calculate running balance
CSV format: Date, Description, Debit, Credit (4 columns, NO balance)
"""
import csv
from pathlib import Path
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

csv_path = Path(r"L:\limo\CIBC UPLOADS\0228362 (CIBC checking account)\cibc 8362 2018.csv")
output_path = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

# Opening balance from PDF: Jan 1, 2018 opening balance
OPENING_BALANCE = Decimal('1058.53')  # From CIBC statement summary

print("PARSING CIBC 8362 2018 CSV + CALCULATING RUNNING BALANCE")
print("=" * 80)

if not csv_path.exists():
    print(f"❌ CSV not found: {csv_path}")
    exit(1)

trans = []

with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    
    for row in reader:
        if not row or len(row) < 2:
            continue
        
        # Skip header rows
        if 'date' in row[0].lower() or 'transaction' in ' '.join(row).lower():
            continue
        
        # CSV format: Date, Description, Debit, Credit
        try:
            date_str = row[0].strip()
            desc = row[1].strip() if len(row) > 1 else ""
            debit_str = row[2].strip() if len(row) > 2 else ""
            credit_str = row[3].strip() if len(row) > 3 else ""
            
            # Parse amounts (empty = 0)
            debit = Decimal(debit_str.replace(',', '').replace('$', '')) if debit_str else Decimal('0')
            credit = Decimal(credit_str.replace(',', '').replace('$', '')) if credit_str else Decimal('0')
            
            # Net: credit positive, debit negative
            amount = credit - debit
            
            trans.append({
                'date': date_str,
                'desc': desc,
                'debit': debit,
                'credit': credit,
                'amount': amount
            })
            
        except Exception as e:
            print(f"⚠️  Skip row: {row[:2]} - {e}")
            continue

print(f"Extracted {len(trans)} transactions\n")

if not trans:
    print("❌ No transactions")
    exit(1)

# Sort chronologically by date object
from datetime import datetime
for t in trans:
    t['date_obj'] = datetime.strptime(t['date'], '%Y-%m-%d')

trans.sort(key=lambda x: x['date_obj'])

# Calculate running balance
print("CALCULATING RUNNING BALANCE...")
print("=" * 80)
print(f"Opening balance: ${OPENING_BALANCE:,.2f}\n")

running = OPENING_BALANCE

for t in trans:
    running += t['amount']
    t['balance'] = running

total_cr = sum(t['credit'] for t in trans)
total_db = sum(t['debit'] for t in trans)

print(f"Total Credits: ${total_cr:,.2f}")
print(f"Total Debits: ${total_db:,.2f}")
print(f"Net Change: ${total_cr - total_db:,.2f}")
print(f"Closing: ${running:,.2f}")

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Transactions"

headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
ws.append(headers)

fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
font = Font(bold=True, color="FFFFFF")
for c in range(1, 6):
    ws.cell(1, c).fill = fill
    ws.cell(1, c).font = font

for t in trans:
    ws.append([
        t['date'], t['desc'],
        float(t['debit']), float(t['credit']),
        float(t['balance'])
    ])

ws.column_dimensions['B'].width = 50

for c in [3, 4, 5]:
    for r in range(2, len(trans) + 2):
        ws.cell(r, c).number_format = '$#,##0.00'

# Summary
s = wb.create_sheet("Summary")
s['A1'] = "CIBC 8362 2018"
s['A1'].font = Font(bold=True, size=14)
s['A3'] = "Opening:"
s['B3'] = float(OPENING_BALANCE)
s['B3'].number_format = '$#,##0.00'
s['A4'] = "Transactions:"
s['B4'] = len(trans)
s['A5'] = "Credits:"
s['B5'] = float(total_cr)
s['B5'].number_format = '$#,##0.00'
s['A6'] = "Debits:"
s['B6'] = float(total_db)
s['B6'].number_format = '$#,##0.00'
s['A7'] = "Closing:"
s['B7'] = float(running)
s['B7'].number_format = '$#,##0.00'

wb.save(output_path)
print(f"\n✅ {output_path}")
print(f"   {len(trans)} transactions")
