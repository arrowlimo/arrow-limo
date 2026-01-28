import os
import re
from datetime import date, timedelta
from typing import List, Dict, Tuple, Optional
from collections import defaultdict

import psycopg2
import psycopg2.extras

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", os.environ.get("DB_PASSWORD"))

DATE_START = date(2012, 1, 1)
DATE_END = date(2017, 12, 31)

ACC_1615 = {"bank_id": 4, "account_number": "1615"}
ACC_8362 = {"bank_id": 1, "account_number": "0228362"}

TRANSFER_KEYWORDS = ["transfer", "xfer", "x-fer", "xfr", "moving", "to 0228362", "to 1615", "from 0228362", "from 1615"]
CASH_KEYWORDS = ["cash", "withdrawal", "atm", "cash box", "petty cash"]


def get_conn():
    return psycopg2.connect(host=DB_HOST, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)


def fetch_tx(cur, bank_id: int, account_number: str) -> List[Dict]:
    cur.execute(
        """
        SELECT transaction_id, transaction_date, description,
               debit_amount, credit_amount, balance, source_file
        FROM banking_transactions
        WHERE bank_id = %s AND account_number = %s
          AND transaction_date BETWEEN %s AND %s
        ORDER BY transaction_date, transaction_id
        """,
        (bank_id, account_number, DATE_START, DATE_END),
    )
    return [dict(r) for r in cur.fetchall()]


def amt_out(row: Dict) -> float:
    d = row.get("debit_amount")
    return float(d) if d is not None else 0.0


def amt_in(row: Dict) -> float:
    c = row.get("credit_amount")
    return float(c) if c is not None else 0.0


def has_kw(desc: str, keywords: List[str]) -> bool:
    if not desc:
        return False
    low = desc.lower()
    return any(k in low for k in keywords)


def is_arrow_cheque(desc: str) -> bool:
    """Detect CHQ xxx PAID TO ARROW LIMOUSINE patterns"""
    if not desc:
        return False
    low = desc.lower()
    return ('chq' in low or 'cheque' in low or 'check' in low) and 'arrow' in low and 'limousine' in low


def is_cash_withdrawal(desc: str) -> bool:
    """Detect cash withdrawal patterns"""
    if not desc:
        return False
    return has_kw(desc, CASH_KEYWORDS)


def has_pennies(amount: float) -> bool:
    """Check if amount has pennies (not round dollar amount)"""
    if amount is None:
        return False
    return round(amount % 1.0, 2) not in [0.0, 0.00]


def classify_transaction(row: Dict) -> str:
    """Classify transaction type"""
    desc = row.get("description", "")
    amt_d = amt_out(row)
    amt_c = amt_in(row)
    amount = amt_d if amt_d > 0 else amt_c
    
    if is_arrow_cheque(desc):
        return "arrow_cheque_transfer"
    if is_cash_withdrawal(desc):
        return "cash_withdrawal"
    if has_kw(desc, TRANSFER_KEYWORDS):
        return "explicit_transfer"
    if has_pennies(amount):
        return "non_round_amount"  # Possible cheque or mixed deposit
    if amount > 0 and round(amount % 100, 2) == 0:
        return "round_100s"  # Could be cash
    return "other"


def match_transfers(out_list: List[Dict], in_list: List[Dict], day_window: int = 5) -> Tuple[List[Tuple[Dict, Dict, int, str]], List[Dict], List[Dict]]:
    """Match outgoing to incoming transactions"""
    bucket = defaultdict(list)
    for r in in_list:
        a = round(amt_in(r), 2)
        if a > 0:
            bucket[a].append(r)

    matched = []
    used_in_ids = set()

    for o in out_list:
        a = round(amt_out(o), 2)
        if a <= 0:
            continue
        
        candidates = [r for r in bucket.get(a, []) if r["transaction_id"] not in used_in_ids]
        if not candidates:
            continue
        
        # Score candidates
        best = None
        best_score = None
        match_reason = ""
        
        for cand in candidates:
            dd = abs((cand["transaction_date"] - o["transaction_date"]).days)
            if dd > day_window:
                continue
            
            # Build score (lower is better)
            score = dd * 100  # Base score on date difference
            reason_parts = []
            
            # Arrow cheque bonus
            if is_arrow_cheque(o.get("description")) or is_arrow_cheque(cand.get("description")):
                score -= 500
                reason_parts.append("arrow_cheque")
            
            # Transfer keyword bonus
            if has_kw(o.get("description"), TRANSFER_KEYWORDS):
                score -= 300
                reason_parts.append("xfer_kw_out")
            if has_kw(cand.get("description"), TRANSFER_KEYWORDS):
                score -= 300
                reason_parts.append("xfer_kw_in")
            
            # Cash withdrawal
            if is_cash_withdrawal(o.get("description")):
                score -= 200
                reason_parts.append("cash_withdrawal")
            
            # Pennies (likely cheque/mixed)
            if has_pennies(a):
                score -= 100
                reason_parts.append("pennies")
            
            # Transaction ID proximity (tie-breaker)
            score += abs(cand["transaction_id"] - o["transaction_id"]) / 10000
            
            reason = "+".join(reason_parts) if reason_parts else "amount_date_match"
            
            if best_score is None or score < best_score:
                best = cand
                best_score = score
                match_reason = reason
        
        if best is not None:
            used_in_ids.add(best["transaction_id"])
            dd = abs((best["transaction_date"] - o["transaction_date"]).days)
            matched.append((o, best, dd, match_reason))

    unmatched_out = [o for o in out_list if round(amt_out(o), 2) > 0 and all(o["transaction_id"] != m[0]["transaction_id"] for m in matched)]
    unmatched_in = [i for i in in_list if round(amt_in(i), 2) > 0 and i["transaction_id"] not in used_in_ids]
    
    return matched, unmatched_out, unmatched_in


def export_excel(pairs: List[Tuple[Dict, Dict, int, str]], out_path: str, from_acc: str, to_acc: str):
    """Export transfer pairs to Excel with color coding"""
    if Workbook is None:
        print("openpyxl not available; skipping Excel export")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transfers"
    
    headers = [
        "From_Account", "From_TxID", "From_Date", "From_Desc", "Out_Amount", "From_Type",
        "To_Account", "To_TxID", "To_Date", "To_Desc", "In_Amount", "To_Type",
        "Date_Diff_Days", "Match_Reason"
    ]
    ws.append(headers)
    
    # Color fills
    arrow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    cash_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    pennies_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
    
    for o, i, dd, reason in pairs:
        o_type = classify_transaction(o)
        i_type = classify_transaction(i)
        
        row = [
            from_acc, o["transaction_id"], o["transaction_date"], o.get("description"), round(amt_out(o), 2), o_type,
            to_acc, i["transaction_id"], i["transaction_date"], i.get("description"), round(amt_in(i), 2), i_type,
            dd, reason
        ]
        ws.append(row)
        
        row_idx = ws.max_row
        
        # Apply color coding
        if "arrow_cheque" in reason:
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).fill = arrow_fill
        elif "cash" in reason:
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).fill = cash_fill
        elif "pennies" in reason:
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).fill = pennies_fill
    
    # Autosize columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"✅ Exported: {out_path}")


def export_unmatched(unmatched: List[Dict], out_path: str, label: str):
    """Export unmatched transactions for review"""
    if Workbook is None:
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Unmatched {label}"
    
    headers = ["TxID", "Date", "Description", "Debit", "Credit", "Amount", "Type"]
    ws.append(headers)
    
    for row in unmatched:
        tx_type = classify_transaction(row)
        amt = amt_out(row) if amt_out(row) > 0 else amt_in(row)
        ws.append([
            row["transaction_id"], row["transaction_date"], row.get("description"),
            row.get("debit_amount"), row.get("credit_amount"), round(amt, 2), tx_type
        ])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    wb.save(out_path)
    print(f"✅ Exported unmatched: {out_path}")


def main():
    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Fetch transactions
    print("Loading transactions...")
    t1615 = fetch_tx(cur, ACC_1615["bank_id"], ACC_1615["account_number"])
    t8362 = fetch_tx(cur, ACC_8362["bank_id"], ACC_8362["account_number"])
    print(f"  1615: {len(t1615)} transactions")
    print(f"  0228362: {len(t8362)} transactions")

    # Classify all transactions
    print("\nClassifying transactions...")
    stats_1615 = defaultdict(int)
    stats_8362 = defaultdict(int)
    
    for tx in t1615:
        stats_1615[classify_transaction(tx)] += 1
    for tx in t8362:
        stats_8362[classify_transaction(tx)] += 1
    
    print(f"\n1615 Classification:")
    for k, v in sorted(stats_1615.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    
    print(f"\n0228362 Classification:")
    for k, v in sorted(stats_8362.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    # Match transfers: 1615 -> 8362
    print("\n\nMatching 1615 → 0228362...")
    out_1615 = [r for r in t1615 if r["debit_amount"] is not None]
    in_8362 = [r for r in t8362 if r["credit_amount"] is not None]
    pairs_1615_to_8362, uo_1615, ui_8362 = match_transfers(out_1615, in_8362)
    
    # Match transfers: 8362 -> 1615
    print("Matching 0228362 → 1615...")
    out_8362 = [r for r in t8362 if r["debit_amount"] is not None]
    in_1615 = [r for r in t1615 if r["credit_amount"] is not None]
    pairs_8362_to_1615, uo_8362, ui_1615 = match_transfers(out_8362, in_1615)

    total_pairs = len(pairs_1615_to_8362) + len(pairs_8362_to_1615)
    print(f"\n✅ Matched transfers:")
    print(f"  1615 → 0228362: {len(pairs_1615_to_8362)}")
    print(f"  0228362 → 1615: {len(pairs_8362_to_1615)}")
    print(f"  Total: {total_pairs}")
    
    print(f"\nUnmatched candidates:")
    print(f"  1615 outgoing: {len(uo_1615)}")
    print(f"  0228362 incoming: {len(ui_8362)}")
    print(f"  0228362 outgoing: {len(uo_8362)}")
    print(f"  1615 incoming: {len(ui_1615)}")

    # Breakdown by match reason
    print(f"\n1615 → 0228362 Match Reasons:")
    reason_counts = defaultdict(int)
    for _, _, _, reason in pairs_1615_to_8362:
        reason_counts[reason] += 1
    for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}")
    
    print(f"\n0228362 → 1615 Match Reasons:")
    reason_counts = defaultdict(int)
    for _, _, _, reason in pairs_8362_to_1615:
        reason_counts[reason] += 1
    for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
        print(f"  {reason}: {count}")

    # Export
    print("\nExporting results...")
    out_dir = r"l:\\limo\\reports\\exports"
    os.makedirs(out_dir, exist_ok=True)
    
    export_excel(pairs_1615_to_8362, os.path.join(out_dir, "Transfers_1615_to_0228362_ENHANCED.xlsx"), "1615", "0228362")
    export_excel(pairs_8362_to_1615, os.path.join(out_dir, "Transfers_0228362_to_1615_ENHANCED.xlsx"), "0228362", "1615")
    
    # Export unmatched for review
    export_unmatched(uo_1615, os.path.join(out_dir, "Unmatched_1615_Outgoing.xlsx"), "1615_out")
    export_unmatched(ui_8362, os.path.join(out_dir, "Unmatched_0228362_Incoming.xlsx"), "8362_in")
    export_unmatched(uo_8362, os.path.join(out_dir, "Unmatched_0228362_Outgoing.xlsx"), "8362_out")
    export_unmatched(ui_1615, os.path.join(out_dir, "Unmatched_1615_Incoming.xlsx"), "1615_in")

    print("\n✅ All exports complete!")

    conn.close()


if __name__ == "__main__":
    main()
