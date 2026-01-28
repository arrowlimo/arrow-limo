#!/usr/bin/env python3
"""
Read the corrected WCB Excel report and update database records.
"""
import openpyxl
import psycopg2
import os
from datetime import datetime

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "***REMOVED***")

excel_path = r"L:\limo\reports\WCB_2011_to_Dec_2012.xlsx"

try:
    # Load Excel file
    print("\n" + "="*80)
    print("READING CORRECTED WCB EXCEL REPORT")
    print("="*80)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    print(f"\nReading from: {excel_path}")
    
    # Extract data from Excel
    updates = []
    row_num = 2
    
    while True:
        date_cell = ws.cell(row=row_num, column=1).value
        if not date_cell or date_cell == "SUMMARY":
            break
        
        trans_type = ws.cell(row=row_num, column=2).value
        invoice_ref = ws.cell(row=row_num, column=3).value
        description = ws.cell(row=row_num, column=4).value
        amount = ws.cell(row=row_num, column=5).value
        status = ws.cell(row=row_num, column=7).value
        
        # Parse date
        if isinstance(date_cell, datetime):
            date_str = date_cell.strftime('%Y-%m-%d')
        else:
            date_str = str(date_cell).strip()
        
        updates.append({
            'row': row_num,
            'date': date_str,
            'type': trans_type,
            'invoice_ref': invoice_ref,
            'description': description,
            'amount': float(amount) if amount else 0.0,
            'status': status
        })
        
        row_num += 1
    
    print(f"\nFound {len(updates)} records to process:")
    
    # Connect to database
    conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    cur = conn.cursor()
    
    # Process each record
    updated_count = 0
    not_found_count = 0
    error_count = 0
    
    for update in updates:
        invoice_ref = str(update['invoice_ref']).strip() if update['invoice_ref'] else None
        date_str = update['date']
        desc = update['description']
        amount_abs = abs(update['amount'])
        trans_type = update['type']
        
        print(f"\nRow {update['row']}: {date_str} | {trans_type} | Ref: {invoice_ref} | ${amount_abs:.2f}")
        
        try:
            if trans_type == "INVOICE":
                # Update receipts table
                # Find by reference
                if invoice_ref:
                    cur.execute("""
                        SELECT receipt_id FROM receipts
                        WHERE source_reference = %s
                        LIMIT 1
                    """, (invoice_ref,))
                    
                    result = cur.fetchone()
                    if result:
                        receipt_id = result[0]
                        cur.execute("""
                            UPDATE receipts
                            SET receipt_date = %s,
                                description = %s
                            WHERE receipt_id = %s
                        """, (date_str, desc, receipt_id))
                        
                        print(f"  ✓ Updated receipt {receipt_id}")
                        updated_count += 1
                    else:
                        print(f"  ✗ Receipt not found (ref: {invoice_ref})")
                        not_found_count += 1
                else:
                    print(f"  ✗ No invoice reference provided")
                    not_found_count += 1
                    
            elif trans_type in ["PAYMENT", "payment"]:
                # Update banking_transactions table
                if invoice_ref:
                    cur.execute("""
                        SELECT transaction_id FROM banking_transactions
                        WHERE check_number = %s OR transaction_id = %s
                        LIMIT 1
                    """, (invoice_ref, invoice_ref))
                    
                    result = cur.fetchone()
                    if result:
                        tx_id = result[0]
                        cur.execute("""
                            UPDATE banking_transactions
                            SET transaction_date = %s,
                                description = %s
                            WHERE transaction_id = %s
                        """, (date_str, desc, tx_id))
                        
                        print(f"  ✓ Updated banking transaction {tx_id}")
                        updated_count += 1
                    else:
                        print(f"  ✗ Banking transaction not found (ref: {invoice_ref})")
                        not_found_count += 1
                else:
                    print(f"  ✗ No transaction reference provided")
                    not_found_count += 1
                    
            elif trans_type == "refunded":
                # This is a reversal/credit - find the matching receipt
                if invoice_ref:
                    cur.execute("""
                        SELECT receipt_id FROM receipts
                        WHERE source_reference = %s
                        ORDER BY receipt_date DESC
                        LIMIT 1
                    """, (invoice_ref,))
                    
                    result = cur.fetchone()
                    if result:
                        receipt_id = result[0]
                        cur.execute("""
                            UPDATE receipts
                            SET receipt_date = %s,
                                description = %s
                            WHERE receipt_id = %s
                        """, (date_str, desc, receipt_id))
                        
                        print(f"  ✓ Updated refund receipt {receipt_id}")
                        updated_count += 1
                    else:
                        print(f"  ✗ Receipt not found (ref: {invoice_ref})")
                        not_found_count += 1
        
        except Exception as e:
            print(f"  ✗ Error: {e}")
            error_count += 1
    
    conn.commit()
    cur.close()
    conn.close()
    
    print(f"\n" + "="*80)
    print("UPDATE SUMMARY")
    print("="*80)
    print(f"\nRecords updated: {updated_count}")
    print(f"Records not found: {not_found_count}")
    print(f"Errors: {error_count}")
    print(f"Total processed: {len(updates)}")
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()
