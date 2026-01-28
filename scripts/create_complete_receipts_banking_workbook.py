#!/usr/bin/env python3
"""
Create comprehensive workbook with ALL receipts and banking data.
Color-coded by bank account for easy matching and editing.

Color scheme:
- CIBC (account_id=1): Light Blue (#ADD8E6)
- Scotia (account_id=2): Light Green (#90EE90)
- Unmatched receipts: Light Yellow (#FFFFE0)
- Matched receipts: White
"""

import os
import sys
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Database connection
DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "***REMOVED***")

# Color definitions
CIBC_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
SCOTIA_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
UNMATCHED_FILL = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light Yellow
MATCHED_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Dark Blue

BOLD_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_all_receipts(conn):
    """Get all receipts with banking match status."""
    query = """
    SELECT 
        r.receipt_id,
        r.receipt_date,
        r.vendor_name,
        r.category,
        r.gross_amount,
        r.gst_amount,
        r.net_amount,
        r.description,
        r.payment_method,
        r.mapped_bank_account_id,
        r.banking_transaction_id,
        r.created_from_banking,
        CASE 
            WHEN r.mapped_bank_account_id = 1 THEN 'CIBC 0228362'
            WHEN r.mapped_bank_account_id = 2 THEN 'Scotia 903990106011'
            ELSE 'Unknown'
        END as bank_account,
        CASE 
            WHEN r.banking_transaction_id IS NOT NULL THEN 'Matched'
            ELSE 'Unmatched'
        END as match_status,
        bt.transaction_date as banking_date,
        bt.debit_amount as banking_amount,
        bt.description as banking_description
    FROM receipts r
    LEFT JOIN banking_transactions bt ON r.banking_transaction_id = bt.transaction_id
    ORDER BY r.receipt_date, r.receipt_id
    """
    
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(query)
        return cur.fetchall()


def get_all_banking(conn):
    """Get all banking transactions with match status."""
    query = """
    SELECT 
        bt.transaction_id,
        bt.transaction_date,
        bt.description,
        bt.debit_amount,
        bt.credit_amount,
        bt.balance,
        bt.account_number,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC 0228362'
            WHEN bt.account_number = '903990106011' THEN 'Scotia 903990106011'
            ELSE bt.account_number
        END as bank_account,
        CASE 
            WHEN EXISTS (
                SELECT 1 FROM receipts r 
                WHERE r.banking_transaction_id = bt.transaction_id
            ) THEN 'Matched'
            ELSE 'Unmatched'
        END as match_status,
        r.receipt_id as matched_receipt_id,
        r.vendor_name as matched_vendor_name,
        r.gross_amount as matched_amount
    FROM banking_transactions bt
    LEFT JOIN receipts r ON r.banking_transaction_id = bt.transaction_id
    ORDER BY bt.transaction_date, bt.transaction_id
    """
    
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(query)
        return cur.fetchall()


def get_unmatched_banking_by_account(conn):
    """Get unmatched banking transactions grouped by account."""
    query = """
    SELECT 
        bt.transaction_id,
        bt.transaction_date,
        bt.description,
        bt.debit_amount,
        bt.credit_amount,
        bt.balance,
        bt.account_number,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC 0228362'
            WHEN bt.account_number = '903990106011' THEN 'Scotia 903990106011'
            ELSE bt.account_number
        END as bank_account
    FROM banking_transactions bt
    WHERE bt.debit_amount > 0  -- Only debits (expenses)
    AND NOT EXISTS (
        SELECT 1 FROM receipts r 
        WHERE r.banking_transaction_id = bt.transaction_id
    )
    ORDER BY bt.account_number, bt.transaction_date, bt.transaction_id
    """
    
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(query)
        return cur.fetchall()


def apply_row_color(ws, row_num, fill, include_header=False):
    """Apply color to entire row."""
    for cell in ws[row_num]:
        cell.fill = fill
        cell.border = BORDER
        if include_header:
            cell.font = BOLD_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')


def create_receipts_sheet(wb, receipts):
    """Create receipts sheet with color coding."""
    ws = wb.create_sheet("All Receipts", 0)
    
    # Headers
    headers = [
        "Receipt ID", "Date", "Vendor", "Category", "Amount", 
        "GST", "Net Amount", "Description", "Payment Method",
        "Bank Account", "Match Status", "Banking Date", 
        "Banking Amount", "Banking Description"
    ]
    ws.append(headers)
    apply_row_color(ws, 1, HEADER_FILL, include_header=True)
    
    # Data rows
    row_colors = []  # Store colors to apply after all data is added
    for idx, receipt in enumerate(receipts, start=2):
        ws.append([
            receipt['receipt_id'],
            receipt['receipt_date'],
            receipt['vendor_name'],
            receipt['category'],
            float(receipt['gross_amount']) if receipt['gross_amount'] else 0,
            float(receipt['gst_amount']) if receipt['gst_amount'] else 0,
            float(receipt['net_amount']) if receipt['net_amount'] else 0,
            receipt['description'],
            receipt['payment_method'],
            receipt['bank_account'],
            receipt['match_status'],
            receipt['banking_date'],
            float(receipt['banking_amount']) if receipt['banking_amount'] else None,
            receipt['banking_description']
        ])
        
        # Determine color based on bank account and match status
        if receipt['match_status'] == 'Unmatched':
            row_colors.append((idx, UNMATCHED_FILL))
        elif receipt['mapped_bank_account_id'] == 1:
            row_colors.append((idx, CIBC_FILL))
        elif receipt['mapped_bank_account_id'] == 2:
            row_colors.append((idx, SCOTIA_FILL))
        else:
            row_colors.append((idx, MATCHED_FILL))
    
    # Apply colors to all rows at once
    print(f"      Applying color formatting to {len(row_colors)} rows...")
    for row_num, fill in row_colors:
        apply_row_color(ws, row_num, fill)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 35
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    return len(receipts)


def create_banking_sheet(wb, banking_txns):
    """Create banking sheet with color coding."""
    ws = wb.create_sheet("All Banking")
    
    # Headers
    headers = [
        "Transaction ID", "Date", "Description", "Debit", "Credit", 
        "Balance", "Bank Account", "Match Status", 
        "Matched Receipt ID", "Matched Vendor", "Matched Amount"
    ]
    ws.append(headers)
    apply_row_color(ws, 1, HEADER_FILL, include_header=True)
    
    # Data rows
    row_colors = []  # Store colors to apply after all data is added
    for idx, txn in enumerate(banking_txns, start=2):
        ws.append([
            txn['transaction_id'],
            txn['transaction_date'],
            txn['description'],
            float(txn['debit_amount']) if txn['debit_amount'] else None,
            float(txn['credit_amount']) if txn['credit_amount'] else None,
            float(txn['balance']) if txn['balance'] else None,
            txn['bank_account'],
            txn['match_status'],
            txn['matched_receipt_id'],
            txn['matched_vendor_name'],
            float(txn['matched_amount']) if txn['matched_amount'] else None
        ])
        
        # Determine color based on bank account and match status
        if txn['match_status'] == 'Unmatched' and txn['debit_amount'] and float(txn['debit_amount']) > 0:
            row_colors.append((idx, UNMATCHED_FILL))
        elif 'CIBC' in txn['bank_account']:
            row_colors.append((idx, CIBC_FILL))
        elif 'Scotia' in txn['bank_account']:
            row_colors.append((idx, SCOTIA_FILL))
        else:
            row_colors.append((idx, MATCHED_FILL))
    
    # Apply colors to all rows at once
    print(f"      Applying color formatting to {len(row_colors)} rows...")
    for row_num, fill in row_colors:
        apply_row_color(ws, row_num, fill)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    return len(banking_txns)


def create_unmatched_banking_sheet(wb, unmatched_txns):
    """Create sheet for unmatched banking transactions."""
    ws = wb.create_sheet("Unmatched Banking")
    
    # Headers
    headers = [
        "Transaction ID", "Date", "Description", "Debit Amount", 
        "Bank Account", "Action Needed"
    ]
    ws.append(headers)
    apply_row_color(ws, 1, HEADER_FILL, include_header=True)
    
    # Data rows grouped by account
    current_account = None
    for idx, txn in enumerate(unmatched_txns, start=2):
        # Add separator row when account changes
        if current_account and current_account != txn['bank_account']:
            idx += 1
        
        ws.append([
            txn['transaction_id'],
            txn['transaction_date'],
            txn['description'],
            float(txn['debit_amount']) if txn['debit_amount'] else 0,
            txn['bank_account'],
            "Create receipt or match existing"
        ])
        
        # Apply color based on bank account
        if 'CIBC' in txn['bank_account']:
            apply_row_color(ws, idx, CIBC_FILL)
        elif 'Scotia' in txn['bank_account']:
            apply_row_color(ws, idx, SCOTIA_FILL)
        else:
            apply_row_color(ws, idx, UNMATCHED_FILL)
        
        current_account = txn['bank_account']
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    return len(unmatched_txns)


def create_summary_sheet(wb, receipt_count, banking_count, unmatched_count):
    """Create summary sheet with statistics."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "Complete Receipts & Banking Workbook"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:D2')
    
    # Statistics
    ws['A4'] = "Statistics:"
    ws['A4'].font = Font(bold=True, size=12)
    
    ws['A5'] = "Total Receipts:"
    ws['B5'] = receipt_count
    
    ws['A6'] = "Total Banking Transactions:"
    ws['B6'] = banking_count
    
    ws['A7'] = "Unmatched Banking (Debits):"
    ws['B7'] = unmatched_count
    
    # Color legend
    ws['A9'] = "Color Legend:"
    ws['A9'].font = Font(bold=True, size=12)
    
    ws['A10'] = "CIBC Account"
    apply_row_color(ws, 10, CIBC_FILL)
    
    ws['A11'] = "Scotia Account"
    apply_row_color(ws, 11, SCOTIA_FILL)
    
    ws['A12'] = "Unmatched Items"
    apply_row_color(ws, 12, UNMATCHED_FILL)
    
    ws['A13'] = "Matched Items"
    apply_row_color(ws, 13, MATCHED_FILL)
    
    # Instructions
    ws['A15'] = "Instructions:"
    ws['A15'].font = Font(bold=True, size=12)
    
    instructions = [
        "1. All Receipts - Complete receipt records color-coded by bank account",
        "2. All Banking - Complete banking transactions with match status",
        "3. Unmatched Banking - Banking debits that need receipts created or matched",
        "4. Use color coding to quickly identify which account each transaction belongs to",
        "5. Yellow highlighted items need attention (unmatched)",
        "6. Edit or add receipt data as needed, then import back to database"
    ]
    
    for idx, instruction in enumerate(instructions, start=16):
        ws[f'A{idx}'] = instruction
        ws.merge_cells(f'A{idx}:D{idx}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def main():
    print("=" * 80)
    print("Creating Complete Receipts & Banking Workbook")
    print("=" * 80)
    
    # Connect to database
    print("\n📊 Connecting to database...")
    conn = psycopg2.connect(
        host=DB_HOST,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )
    
    try:
        # Fetch data
        print("📥 Fetching all receipts...")
        receipts = get_all_receipts(conn)
        print(f"   ✓ Found {len(receipts)} receipts")
        
        print("📥 Fetching all banking transactions...")
        banking_txns = get_all_banking(conn)
        print(f"   ✓ Found {len(banking_txns)} banking transactions")
        
        print("📥 Fetching unmatched banking transactions...")
        unmatched_txns = get_unmatched_banking_by_account(conn)
        print(f"   ✓ Found {len(unmatched_txns)} unmatched banking debits")
        
        # Create workbook
        print("\n📝 Creating Excel workbook...")
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        print("   Creating Summary sheet...")
        create_summary_sheet(wb, len(receipts), len(banking_txns), len(unmatched_txns))
        
        print("   Creating All Receipts sheet...")
        receipt_count = create_receipts_sheet(wb, receipts)
        print(f"      ✓ Added {receipt_count} receipts")
        
        print("   Creating All Banking sheet...")
        banking_count = create_banking_sheet(wb, banking_txns)
        print(f"      ✓ Added {banking_count} transactions")
        
        print("   Creating Unmatched Banking sheet...")
        unmatched_count = create_unmatched_banking_sheet(wb, unmatched_txns)
        print(f"      ✓ Added {unmatched_count} unmatched transactions")
        
        # Save workbook
        output_file = "reports/complete_receipts_banking_workbook.xlsx"
        os.makedirs("reports", exist_ok=True)
        wb.save(output_file)
        
        # Get file size
        file_size = os.path.getsize(output_file) / 1024
        
        print("\n" + "=" * 80)
        print("✅ WORKBOOK CREATED SUCCESSFULLY")
        print("=" * 80)
        print(f"\n📁 Output: {output_file}")
        print(f"📊 File Size: {file_size:.1f} KB")
        print(f"\n📈 Summary:")
        print(f"   • Total Receipts: {receipt_count}")
        print(f"   • Total Banking: {banking_count}")
        print(f"   • Unmatched Banking: {unmatched_count}")
        print(f"\n🎨 Color Coding:")
        print(f"   • Light Blue = CIBC Account")
        print(f"   • Light Green = Scotia Account")
        print(f"   • Light Yellow = Unmatched (needs attention)")
        print(f"   • White = Matched")
        print("\n" + "=" * 80)
        
    finally:
        conn.close()
        print("\n✓ Database connection closed")


if __name__ == "__main__":
    main()
