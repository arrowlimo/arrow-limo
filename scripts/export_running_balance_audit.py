import argparse
import os
from datetime import datetime
from typing import List, Dict, Any

import psycopg2
import psycopg2.extras


def get_conn():
    return psycopg2.connect(
        host=os.environ.get("DB_HOST", "localhost"),
        dbname=os.environ.get("DB_NAME", "almsdata"),
        user=os.environ.get("DB_USER", "postgres"),
        password=os.environ.get("DB_PASSWORD", "***REDACTED***"),
    )


def fetch_transactions(cur, bank_id: int = None, account_number: str = None, start: str = None, end: str = None) -> List[Dict[str, Any]]:
    if not bank_id and not account_number:
        raise ValueError("Provide either bank_id or account_number")
    clauses = []
    params: List[Any] = []
    if bank_id:
        clauses.append("bank_id = %s")
        params.append(bank_id)
    if account_number:
        clauses.append("account_number = %s")
        params.append(account_number)
    if start:
        clauses.append("transaction_date >= %s")
        params.append(start)
    if end:
        clauses.append("transaction_date < %s")
        params.append(end)
    where = " and ".join(clauses) if clauses else "TRUE"
    cur.execute(
        f"""
        select transaction_id, account_number, bank_id, transaction_date, posted_date,
               description, debit_amount, credit_amount, balance, source_file, vendor_extracted
          from banking_transactions
         where {where}
         order by transaction_date asc, transaction_id asc
        """,
        params,
    )
    return [dict(r) for r in cur.fetchall()]


def compute_expected_and_deltas(rows: List[Dict[str, Any]], anchor_opening: float = None):
    expected_key = "expected_balance"
    delta_key = "delta"
    if not rows:
        return 0, 0.0
    mismatches = 0
    total_abs_delta = 0.0
    # Anchor expected to first DB balance or provided anchor
    if anchor_opening is None:
        bal0 = rows[0]["balance"]
        rows[0][expected_key] = float(bal0) if bal0 is not None else 0.0
    else:
        rows[0][expected_key] = float(anchor_opening)
    bal0f = float(rows[0]["balance"] or 0.0)
    rows[0][delta_key] = round(bal0f - float(rows[0][expected_key] or 0.0), 2)
    if abs(rows[0][delta_key]) >= 0.01:
        mismatches += 1
        total_abs_delta += abs(rows[0][delta_key])
    # March forward using debit/credit
    for i in range(1, len(rows)):
        prev = float(rows[i - 1][expected_key] or 0.0)
        debit = float(rows[i]["debit_amount"] or 0.0)
        credit = float(rows[i]["credit_amount"] or 0.0)
        rows[i][expected_key] = round((prev - debit + credit), 2)
        db_bal = float(rows[i]["balance"] or 0.0)
        rows[i][delta_key] = round(db_bal - rows[i][expected_key], 2)
        if abs(rows[i][delta_key]) >= 0.01:
            mismatches += 1
            total_abs_delta += abs(rows[i][delta_key])
    return mismatches, round(total_abs_delta, 2)


def apply_fixes(cur, rows: List[Dict[str, Any]]):
    # Update only rows where delta != 0
    updates = 0
    for r in rows:
        if abs(r.get("delta") or 0.0) >= 0.01:
            cur.execute(
                "update banking_transactions set balance=%s where transaction_id=%s",
                (r["expected_balance"], r["transaction_id"]),
            )
            updates += 1
    return updates


def write_report(out_path: str, rows: List[Dict[str, Any]]):
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        headers = [
            "transaction_id",
            "transaction_date",
            "posted_date",
            "description",
            "debit_amount",
            "credit_amount",
            "db_balance",
            "expected_balance",
            "delta",
            "account_number",
            "bank_id",
            "source_file",
            "vendor_extracted",
        ]
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get("transaction_id"),
                r.get("transaction_date"),
                r.get("posted_date"),
                r.get("description"),
                r.get("debit_amount"),
                r.get("credit_amount"),
                r.get("balance"),
                r.get("expected_balance"),
                r.get("delta"),
                r.get("account_number"),
                r.get("bank_id"),
                r.get("source_file"),
                r.get("vendor_extracted"),
            ])
        # autosize
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        wb.save(out_path)
        return out_path, "xlsx"
    except Exception as e:
        # Fallback to CSV
        import csv

        csv_path = os.path.splitext(out_path)[0] + ".csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "transaction_id",
                "transaction_date",
                "posted_date",
                "description",
                "debit_amount",
                "credit_amount",
                "db_balance",
                "expected_balance",
                "delta",
                "account_number",
                "bank_id",
                "source_file",
                "vendor_extracted",
            ])
            for r in rows:
                writer.writerow([
                    r.get("transaction_id"),
                    r.get("transaction_date"),
                    r.get("posted_date"),
                    r.get("description"),
                    r.get("debit_amount"),
                    r.get("credit_amount"),
                    r.get("balance"),
                    r.get("expected_balance"),
                    r.get("delta"),
                    r.get("account_number"),
                    r.get("bank_id"),
                    r.get("source_file"),
                    r.get("vendor_extracted"),
                ])
        return csv_path, "csv"


def main():
    ap = argparse.ArgumentParser(description="Recompute and audit running balances; export to XLSX/CSV")
    ap.add_argument("--bank-id", type=int, dest="bank_id")
    ap.add_argument("--account-number", dest="acct")
    ap.add_argument("--start", default="2012-01-01")
    ap.add_argument("--end", default="2018-01-01")
    ap.add_argument("--outfile", required=True)
    ap.add_argument("--write", action="store_true", help="Update DB balances to expected")
    ap.add_argument("--anchor-opening", type=float, default=None, help="Optional opening balance anchor")
    args = ap.parse_args()

    conn = get_conn()
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    print("Connected.")

    rows = fetch_transactions(cur, bank_id=args.bank_id, account_number=args.acct, start=args.start, end=args.end)
    print(f"Fetched {len(rows)} transactions.")
    mismatches, total_abs_delta = compute_expected_and_deltas(rows, anchor_opening=args.anchor_opening)
    print(f"Mismatches: {mismatches}, total_abs_delta: {total_abs_delta}")

    updates = 0
    if args.write and rows:
        updates = apply_fixes(cur, rows)
        conn.commit()
        print(f"Updated {updates} rows to match expected balances.")

    out_path, fmt = write_report(args.outfile, rows)
    print(f"Exported {len(rows)} rows to {out_path} ({fmt}).")

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
