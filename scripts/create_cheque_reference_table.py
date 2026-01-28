#!/usr/bin/env python3
"""
Create Excel table of all cheques for manual vendor name entry.
Separate sheets for CIBC and Scotia Bank.
"""

import psycopg2
import os
import re
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "***REMOVED***")

conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
cur = conn.cursor()

print("=" * 80)
print("CREATING CHEQUE REFERENCE TABLE FOR MANUAL ENTRY")
print("=" * 80)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Header styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")

# Process each bank
banks = [
    (1, "CIBC 0228362", "CIBC"),
    (2, "Scotia 903990106011", "SCOTIA")
]

for bank_id, sheet_name, bank_display in banks:
    print(f"\nProcessing {bank_display}...")
    
    # Query REAL BANK cheques (must have cheque number pattern)
    # Look for: CHQ ###, CHEQUE ###, CHECK ###, CHQ #, etc.
    cur.execute("""
        SELECT transaction_id, transaction_date, debit_amount, description,
               (regexp_match(description, '(CHQ|CHEQUE|CHECK)\\s*#?\\s*(\\d+)', 'i'))[2]::integer as chq_num
        FROM banking_transactions
        WHERE bank_id = %s
          AND debit_amount IS NOT NULL
          AND (
              description ~* 'CHQ\\s*#?\\s*\\d+'
              OR description ~* 'CHEQUE\\s+\\d+'
              OR description ~* 'CHECK\\s+\\d+'
          )
          AND description NOT ILIKE %s
          AND reconciliation_status NOT IN ('NSF', 'RETURN', 'QB_DUPLICATE')
        ORDER BY (regexp_match(description, '(CHQ|CHEQUE|CHECK)\\s*#?\\s*(\\d+)', 'i'))[2]::integer NULLS LAST, 
                 transaction_date
    """, (bank_id, '%Cheque Expense%'))
    
    transactions = cur.fetchall()
    print(f"  Found {len(transactions)} cheques")
    
    if transactions:
        first_chq = transactions[0][4] if transactions[0][4] else '???'
        last_chq = transactions[-1][4] if transactions[-1][4] else '???'
        print(f"  Cheque range: {first_chq} to {last_chq}")
    
    if not transactions:
        continue
    
    # Create sheet
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ["Cheque #", "Date", "Amount", "Current Description", "Vendor Name (ENTER HERE)", "TX ID", "Notes"]
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Add data
    prev_chq_num = 0
    for tx_id, date, amount, desc, chq_num in transactions:
        # Extract cheque number (already extracted in query)
        chq_number = str(chq_num) if chq_num else "???"
        
        # Detect gaps in sequence
        gap_note = ""
        if chq_num and prev_chq_num and chq_num > prev_chq_num + 1:
            missing_count = chq_num - prev_chq_num - 1
            if missing_count == 1:
                gap_note = f"MISSING: {prev_chq_num + 1}"
            else:
                gap_note = f"MISSING: {prev_chq_num + 1} to {chq_num - 1}"
        if chq_num:
            prev_chq_num = chq_num
        
        # Try to extract vendor name from description
        vendor_hint = ""
        
        # Remove cheque number and reference numbers
        desc_clean = re.sub(r'\b(CHEQUE|CHQ|CHECK)\s*#?\s*\d+\b', '', desc, flags=re.IGNORECASE)
        desc_clean = re.sub(r'\b\d{8,}\b', '', desc_clean)  # Remove long numbers
        desc_clean = re.sub(r'^\s*[-,;:]\s*', '', desc_clean)
        desc_clean = desc_clean.strip()
        
        # Check if there's remaining text that might be a vendor name
        if len(desc_clean) > 3:
            # Clean up common prefixes
            desc_clean = re.sub(r'^(Cheque\s+Expense\s*-\s*)', '', desc_clean, flags=re.IGNORECASE)
            vendor_hint = desc_clean.strip()
        
        # Add row
        row_data = [
            chq_number,
            date.strftime("%Y-%m-%d") if date else "",
            amount,
            desc,
            vendor_hint,  # Pre-fill if found, otherwise blank for manual entry
            tx_id,
            gap_note  # Show missing cheque numbers
        ]
        ws.append(row_data)
    
    # Format columns
    ws.column_dimensions['A'].width = 12  # Cheque #
    ws.column_dimensions['B'].width = 12  # Date
    ws.column_dimensions['C'].width = 12  # Amount
    ws.column_dimensions['D'].width = 50  # Current Description
    ws.column_dimensions['E'].width = 30  # Vendor Name (ENTER HERE)
    ws.column_dimensions['F'].width = 10  # TX ID
    ws.column_dimensions['G'].width = 30  # Notes
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Format amount column as currency
    for row_num in range(2, len(transactions) + 2):
        amount_cell = ws.cell(row=row_num, column=3)
        amount_cell.number_format = '$#,##0.00'
    
    # Highlight rows without vendor names (for manual entry)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for gaps
    
    for row_num in range(2, len(transactions) + 2):
        vendor_cell = ws.cell(row=row_num, column=5)
        gap_cell = ws.cell(row=row_num, column=7)
        
        # Highlight gaps in red
        if gap_cell.value and str(gap_cell.value).startswith("MISSING"):
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).fill = red_fill
        # Highlight missing vendor names in yellow
        elif not vendor_cell.value or len(str(vendor_cell.value).strip()) < 3:
            for col_num in range(1, 8):
                ws.cell(row=row_num, column=col_num).fill = yellow_fill

# Add summary sheet
summary = wb.create_sheet("Summary", 0)
summary.append(["Arrow Limousine - Cheque Reference Table"])
summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
summary.append([])
summary.append(["Instructions:"])
summary.append(["1. Cheques sorted by NUMBER (not date) to show sequence"])
summary.append(["2. RED rows = Missing cheque numbers (void/lost/skipped)"])
summary.append(["3. YELLOW rows = Need vendor names"])
summary.append(["4. Compare against pay stubs and enter vendor names"])
summary.append(["5. Save file when complete"])
summary.append([])
summary.append(["Notes:"])
summary.append(["- Cheque numbers match your physical cheque books"])
summary.append(["- CIBC and Scotia have separate cheque books"])
summary.append(["- Some vendors may have slight spelling variations"])

# Style summary
summary.cell(1, 1).font = Font(bold=True, size=14)
summary.column_dimensions['A'].width = 60

# Save file
output_file = "L:/limo/reports/cheque_vendor_reference.xlsx"
wb.save(output_file)

print()
print("=" * 80)
print("✅ EXCEL FILE CREATED")
print("=" * 80)
print(f"Location: {output_file}")
print()
print("Sheets created:")
for bank_id, sheet_name, bank_display in banks:
    print(f"  - {sheet_name}")
print()
print("Yellow rows = Need vendor names")
print("White rows = Already have names (verify if correct)")

cur.close()
conn.close()
