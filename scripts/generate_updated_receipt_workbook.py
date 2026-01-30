#!/usr/bin/env python3
"""
Generate comprehensive receipt Excel workbook with parent-child split receipt support.
Includes fuel receipts with vehicle/charter linking, payment method tracking, and account matching.
"""

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

conn = psycopg2.connect(
    host=os.getenv('DB_HOST', 'localhost'),
    database=os.getenv('DB_NAME', 'almsdata'),
    user=os.getenv('DB_USER', 'postgres'),
    password=os.getenv('DB_PASSWORD', '***REDACTED***')
)
cur = conn.cursor()

print('='*80)
print('GENERATING UPDATED RECEIPT WORKBOOK WITH SPLIT SUPPORT')
print('='*80)
print()

wb = Workbook()

# ============================================================================
# SHEET 1: FUEL RECEIPTS (with parent-child split support)
# ============================================================================
print('Creating FUEL RECEIPTS sheet...')
ws_fuel = wb.active
ws_fuel.title = 'Fuel Receipts'

fuel_headers = [
    'Receipt ID',
    'Parent Receipt #',
    'Receipt Date',
    'Vendor',
    'Amount',
    'GST',
    'Net',
    'Fuel Type',
    'Litres',
    'Price/Litre',
    'Vehicle ID/Unit',
    'Vehicle Type',
    'Odometer',
    'Charter #',
    'Driver',
    'Card Last 4',
    'Bank Account',
    'Category',
    'Business/Personal',
    'Notes'
]

# Add header row with formatting
for col_num, header in enumerate(fuel_headers, 1):
    cell = ws_fuel.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Query for fuel receipts
cur.execute("""
    SELECT 
        r.receipt_id,
        r.parent_receipt_id,
        r.receipt_date,
        r.vendor_name,
        r.gross_amount,
        r.gst_amount,
        r.net_amount,
        r.category,
        r.vehicle_id,
        r.vehicle_number,
        r.description,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC 0228362'
            WHEN bt.account_number = '903990106011' THEN 'Scotia 903990106011'
            ELSE NULL
        END as bank_account,
        CASE WHEN r.is_personal_purchase THEN 'Personal' ELSE 'Business' END as business_personal
    FROM receipts r
    LEFT JOIN banking_transactions bt ON bt.transaction_id = r.banking_transaction_id
    WHERE r.category IN ('fuel', 'vehicle_fuel')
    OR r.fuel_amount > 0
    OR r.fuel > 0
    ORDER BY r.receipt_date DESC
    LIMIT 50
""")

# Add fuel data
fuel_data = cur.fetchall()
print(f'  Loading {len(fuel_data)} fuel receipts...')

for row_idx, row_data in enumerate(fuel_data, start=2):
    receipt_id, parent_id, receipt_date, vendor, amount, gst, net, category, vehicle_id, vehicle_number, description, bank_account, biz_personal = row_data
    
    # Fill in known data
    ws_fuel.cell(row=row_idx, column=1, value=receipt_id)
    ws_fuel.cell(row=row_idx, column=2, value=parent_id if parent_id else '')
    ws_fuel.cell(row=row_idx, column=3, value=receipt_date)
    ws_fuel.cell(row=row_idx, column=4, value=vendor)
    ws_fuel.cell(row=row_idx, column=5, value=float(amount) if amount else None)
    ws_fuel.cell(row=row_idx, column=6, value=float(gst) if gst else None)
    ws_fuel.cell(row=row_idx, column=7, value=float(net) if net else None)
    ws_fuel.cell(row=row_idx, column=18, value=category or 'fuel')
    ws_fuel.cell(row=row_idx, column=19, value=biz_personal or 'Business')
    
    # Vehicle info if available
    if vehicle_id or vehicle_number:
        ws_fuel.cell(row=row_idx, column=11, value=vehicle_number or vehicle_id)
    
    # Bank account (color coded)
    if bank_account:
        cell = ws_fuel.cell(row=row_idx, column=17, value=bank_account)
        if 'CIBC' in bank_account:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        elif 'Scotia' in bank_account:
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# Auto-size columns
for col in range(1, len(fuel_headers) + 1):
    ws_fuel.column_dimensions[get_column_letter(col)].width = 15

# ============================================================================
# SHEET 2: EXPENSE RECEIPTS (with parent-child split support)
# ============================================================================
print('Creating EXPENSE RECEIPTS sheet...')
ws_expense = wb.create_sheet('Expense Receipts')

expense_headers = [
    'Receipt ID',
    'Parent Receipt #',
    'Receipt Date',
    'Vendor',
    'Amount',
    'GST',
    'Net',
    'Category',
    'Payment Method',
    'Card/Cash/NSF',
    'Bank Account',
    'Employee',
    'Charter #',
    'Vehicle ID',
    'Business/Personal',
    'Description',
    'Notes'
]

# Add header row with formatting
for col_num, header in enumerate(expense_headers, 1):
    cell = ws_expense.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Query for expense receipts
cur.execute("""
    SELECT 
        r.receipt_id,
        r.parent_receipt_id,
        r.receipt_date,
        r.vendor_name,
        r.gross_amount,
        r.gst_amount,
        r.net_amount,
        r.category,
        r.payment_method,
        r.description,
        r.receipt_source,
        r.display_color,
        CASE 
            WHEN bt.account_number = '0228362' THEN 'CIBC 0228362'
            WHEN bt.account_number = '903990106011' THEN 'Scotia 903990106011'
            ELSE NULL
        END as bank_account,
        CASE WHEN r.is_personal_purchase THEN 'Personal' ELSE 'Business' END as business_personal
    FROM receipts r
    LEFT JOIN banking_transactions bt ON bt.transaction_id = r.banking_transaction_id
    WHERE r.category NOT IN ('fuel', 'vehicle_fuel')
    AND (r.fuel_amount IS NULL OR r.fuel_amount = 0)
    AND (r.fuel IS NULL OR r.fuel = 0)
    ORDER BY r.receipt_date DESC
    LIMIT 100
""")

# Add expense data
expense_data = cur.fetchall()
print(f'  Loading {len(expense_data)} expense receipts...')

for row_idx, row_data in enumerate(expense_data, start=2):
    receipt_id, parent_id, receipt_date, vendor, amount, gst, net, category, payment_method, description, receipt_source, display_color, bank_account, biz_personal = row_data
    
    # Fill in known data
    ws_expense.cell(row=row_idx, column=1, value=receipt_id)
    ws_expense.cell(row=row_idx, column=2, value=parent_id if parent_id else '')
    ws_expense.cell(row=row_idx, column=3, value=receipt_date)
    ws_expense.cell(row=row_idx, column=4, value=vendor)
    ws_expense.cell(row=row_idx, column=5, value=float(amount) if amount else None)
    ws_expense.cell(row=row_idx, column=6, value=float(gst) if gst else None)
    ws_expense.cell(row=row_idx, column=7, value=float(net) if net else None)
    ws_expense.cell(row=row_idx, column=8, value=category)
    ws_expense.cell(row=row_idx, column=9, value=payment_method)
    ws_expense.cell(row=row_idx, column=15, value=biz_personal or 'Business')
    
    # Description
    if description:
        ws_expense.cell(row=row_idx, column=16, value=description[:100])
    
    # Bank account (color coded)
    if bank_account:
        cell = ws_expense.cell(row=row_idx, column=11, value=bank_account)
        if 'CIBC' in bank_account:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        elif 'Scotia' in bank_account:
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    # Color code row by receipt source
    if display_color:
        color_map = {
            'GREEN': 'C6EFCE',
            'YELLOW': 'FFEB9C',
            'ORANGE': 'FFC7CE',
            'BLUE': 'D9E1F2',
            'RED': 'FFC7CE'
        }
        if display_color in color_map:
            for col in range(1, len(expense_headers) + 1):
                cell = ws_expense.cell(row=row_idx, column=col)
                if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(start_color=color_map[display_color], end_color=color_map[display_color], fill_type='solid')

# Auto-size columns
for col in range(1, len(expense_headers) + 1):
    ws_expense.column_dimensions[get_column_letter(col)].width = 15

# ============================================================================
# SHEET 3: REFERENCE DATA
# ============================================================================
print('Creating REFERENCE DATA sheet...')
ws_ref = wb.create_sheet('Reference Data')

# Vehicles section
cur.execute("""
    SELECT vehicle_id, unit_number, vehicle_type, make, model, license_plate
    FROM vehicles
    ORDER BY unit_number
    LIMIT 30
""")
vehicles = cur.fetchall()

ws_ref.cell(row=1, column=1, value='VEHICLES').font = Font(bold=True, size=14)
ws_ref.cell(row=2, column=1, value='Vehicle ID')
ws_ref.cell(row=2, column=2, value='Unit Number')
ws_ref.cell(row=2, column=3, value='Type')
ws_ref.cell(row=2, column=4, value='Make')
ws_ref.cell(row=2, column=5, value='Model')
ws_ref.cell(row=2, column=6, value='License Plate')

for row_idx, (vid, unit, vtype, make, model, plate) in enumerate(vehicles, start=3):
    ws_ref.cell(row=row_idx, column=1, value=vid)
    ws_ref.cell(row=row_idx, column=2, value=unit)
    ws_ref.cell(row=row_idx, column=3, value=vtype)
    ws_ref.cell(row=row_idx, column=4, value=make)
    ws_ref.cell(row=row_idx, column=5, value=model)
    ws_ref.cell(row=row_idx, column=6, value=plate)

# Payment methods section
payment_start_row = len(vehicles) + 5
ws_ref.cell(row=payment_start_row, column=1, value='PAYMENT METHODS').font = Font(bold=True, size=14)
ws_ref.cell(row=payment_start_row + 1, column=1, value='Method')
ws_ref.cell(row=payment_start_row + 1, column=2, value='Description')

payment_methods = [
    ('CIBC Card 4506', 'CIBC business card ending 4506'),
    ('Scotia Card', 'Scotia business card'),
    ('Cash', 'Cash payment from cash box'),
    ('NSF', 'Non-sufficient funds / bounced payment'),
    ('E-Transfer', 'Electronic bank transfer'),
    ('Check', 'Paper check payment')
]

for idx, (method, desc) in enumerate(payment_methods, start=payment_start_row + 2):
    ws_ref.cell(row=idx, column=1, value=method)
    ws_ref.cell(row=idx, column=2, value=desc)

# Bank accounts section
bank_start_row = payment_start_row + len(payment_methods) + 3
ws_ref.cell(row=bank_start_row, column=1, value='BANK ACCOUNTS').font = Font(bold=True, size=14)
ws_ref.cell(row=bank_start_row + 1, column=1, value='Account')
ws_ref.cell(row=bank_start_row + 1, column=2, value='Number')

banks = [
    ('CIBC', '0228362'),
    ('Scotia', '903990106011')
]

for idx, (bank, number) in enumerate(banks, start=bank_start_row + 2):
    ws_ref.cell(row=idx, column=1, value=bank)
    ws_ref.cell(row=idx, column=2, value=number)

# ============================================================================
# SHEET 4: COLOR LEGEND & SPLIT INSTRUCTIONS
# ============================================================================
print('Creating COLOR LEGEND & SPLIT INSTRUCTIONS sheet...')
ws_legend = wb.create_sheet('Instructions')

legend_data = [
    ['RECEIPT WORKBOOK INSTRUCTIONS', '', ''],
    ['', '', ''],
    ['COLOR CODING', '', ''],
    ['', '', ''],
    ['Bank Accounts:', '', ''],
    ['CIBC 0228362', 'Light Green', ''],
    ['Scotia 903990106011', 'Light Orange', ''],
    ['', '', ''],
    ['Receipt Status:', '', ''],
    ['✅ GREEN', 'Matched to banking transaction', ''],
    ['💰 YELLOW', 'Cash payment (no banking expected)', ''],
    ['👤 ORANGE', 'Employee reimbursement', ''],
    ['📝 BLUE', 'Manually entered', ''],
    ['❌ RED', 'Needs review/matching', ''],
    ['', '', ''],
    ['', '', ''],
    ['SPLIT RECEIPT INSTRUCTIONS', '', ''],
    ['', '', ''],
    ['Example: Costco receipt for $200 (mixed business/personal)', '', ''],
    ['', '', ''],
    ['Step 1: Enter the FULL receipt first', '', ''],
    ['  Receipt Date: 2025-12-05', '', ''],
    ['  Vendor: Costco', '', ''],
    ['  Amount: $200.00', '', ''],
    ['  Parent Receipt #: [BLANK]', '', ''],
    ['  (System will assign Receipt ID, e.g., 12345)', '', ''],
    ['', '', ''],
    ['Step 2: Enter split lines linking to parent Receipt ID', '', ''],
    ['  Line 1: $60.00 | Parent: 12345 | Category: fuel | Business', '', ''],
    ['  Line 2: $80.00 | Parent: 12345 | Category: office | Business', '', ''],
    ['  Line 3: $60.00 | Parent: 12345 | Category: personal | Personal', '', ''],
    ['', '', ''],
    ['Important Notes:', '', ''],
    ['- Leave Parent Receipt # BLANK for normal (non-split) receipts', '', ''],
    ['- Fill in Parent Receipt # ONLY for split portions', '', ''],
    ['- GST is INCLUDED in amounts (use $20 total, not $19.05 + $0.95)', '', ''],
    ['- Price per litre = Amount ÷ Litres (GST already included)', '', ''],
    ['- System validates split total matches parent amount', '', ''],
    ['', '', ''],
    ['', '', ''],
    ['GST CALCULATION (Alberta 5%)', '', ''],
    ['', '', ''],
    ['If you enter $20.00:', '', ''],
    ['  GST (included): $20.00 × 0.05 ÷ 1.05 = $0.95', '', ''],
    ['  Net amount: $20.00 - $0.95 = $19.05', '', ''],
    ['', '', ''],
    ['For 50 litres costing $20.00:', '', ''],
    ['  Price per litre: $20.00 ÷ 50 = $0.40/L (includes GST)', '', ''],
    ['  Net price per litre: $19.05 ÷ 50 = $0.381/L', '', ''],
]

for row_idx, row_data in enumerate(legend_data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_legend.cell(row=row_idx, column=col_idx, value=value)
        if 'INSTRUCTIONS' in str(value) or 'SPLIT RECEIPT' in str(value) or 'GST CALCULATION' in str(value) or 'COLOR CODING' in str(value):
            cell.font = Font(bold=True, size=14)

# Color code the bank account examples
ws_legend.cell(row=6, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
ws_legend.cell(row=7, column=1).fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# Save workbook
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
output_file = f'l:/limo/reports/receipts_workbook_with_splits_{timestamp}.xlsx'
wb.save(output_file)

cur.close()
conn.close()

print()
print('='*80)
print('WORKBOOK CREATED SUCCESSFULLY')
print('='*80)
print(f'Location: {output_file}')
print()
print('SHEETS INCLUDED:')
print('  1. Fuel Receipts - 50 fuel samples with parent-child split support')
print('  2. Expense Receipts - 100 expense samples with parent-child split support')
print('  3. Reference Data - vehicles, payment methods, bank accounts')
print('  4. Instructions - color legend and split receipt instructions')
print()
print('NEW FEATURES:')
print('  ✅ Receipt ID column (auto-assigned by database)')
print('  ✅ Parent Receipt # column (link child receipts to parent)')
print('  ✅ Business/Personal indicator')
print('  ✅ GST/Net amount calculations')
print('  ✅ Split receipt instructions and examples')
print('  ✅ Bank account color coding')
print()
