"""
CIBC 8362 2018 extractor - MULTI-LINE DESCRIPTION HANDLER + BALANCE AUDIT
Handles text-wrapped descriptions across multiple lines
"""
import re
from pathlib import Path
from decimal import Decimal
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

pdf_path = Path(r"L:\limo\pdf\2018\CIBC 8362 2018.pdf")
output = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

trans = []
opening = None
month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
             'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}

print("EXTRACTING CIBC 8362 2018 - MULTI-LINE DESCRIPTION PARSER")
print("=" * 80)

with pdfplumber.open(pdf_path) as pdf:
    for pg_num, pg in enumerate(pdf.pages, 1):
        txt = pg.extract_text() or ""
        lines = txt.split('\n')
        
        # Find opening balance
        if opening is None:
            for line in lines:
                if 'forward' in line.lower() or 'opening' in line.lower():
                    m = re.search(r'\$?([\d,]+\.\d{2})', line)
                    if m:
                        opening = Decimal(m.group(1).replace(',', ''))
                        print(f"Found opening balance: ${opening:,.2f}")
                        break
        
        # Build multi-line transactions
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Check if this line starts a transaction (month + day)
            m = re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\s+(.+)', line, re.I)
            if not m:
                i += 1
                continue
            
            month_str, day, first_part = m.groups()
            month = month_map[month_str.lower()]
            
            # Accumulate all text until next transaction starts
            full_text = first_part
            i += 1
            
            # Look ahead for continuation lines (lines that don't start with a month)
            while i < len(lines):
                next_line = lines[i].strip()
                # Stop if we hit another transaction date or empty line
                if not next_line:
                    i += 1
                    break
                if re.match(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}\s+', next_line, re.I):
                    break
                # Skip header/footer lines
                if 'CIBC' in next_line or 'Account' in next_line or 'Page' in next_line:
                    i += 1
                    break
                
                full_text += " " + next_line
                i += 1
            
            # Clean up the accumulated text
            full_text = re.sub(r'\s+', ' ', full_text).strip()
            
            # Extract all numbers from the full text
            nums = re.findall(r'([\d,]+\.\d{2})', full_text)
            if len(nums) < 1:
                continue
            
            # Last number is typically the balance (after ~)
            # Look for pattern: amount ~ balance or amount balance
            balance = None
            amount = None
            
            # Try to find balance after ~
            tilde_match = re.search(r'~\s*([\d,]+\.\d{2})', full_text)
            if tilde_match:
                balance = Decimal(tilde_match.group(1).replace(',', ''))
                # Amount is the last number before the tilde
                pre_tilde = full_text[:full_text.index('~')]
                pre_nums = re.findall(r'([\d,]+\.\d{2})', pre_tilde)
                if pre_nums:
                    amount = Decimal(pre_nums[-1].replace(',', ''))
            elif len(nums) >= 2:
                # No tilde: last is balance, second-to-last is amount
                balance = Decimal(nums[-1].replace(',', ''))
                amount = Decimal(nums[-2].replace(',', ''))
            elif len(nums) == 1:
                # Only one number - might be balance forward or fee
                if 'forward' in full_text.lower():
                    i += 1
                    continue
                # Assume it's an amount with balance missing
                amount = Decimal(nums[0].replace(',', ''))
            
            if amount is None or balance is None:
                continue
            
            # Get description (text before the amount numbers)
            desc_end = full_text.find(nums[0])
            desc = full_text[:desc_end].strip() if desc_end > 0 else full_text[:50]
            
            # Determine withdrawal vs deposit
            # Check column headers or keywords
            is_withdrawal = any(kw in full_text.upper() for kw in [
                'PURCHASE', 'DEBIT', 'FEE', 'WITHDRAWAL', 'PAYMENT', 
                'PREAUTHORIZED', 'ABM', 'ATM', 'RETAIL'
            ])
            
            if is_withdrawal:
                amount = -abs(amount)
            else:
                amount = abs(amount)
            
            trans.append({
                'date': f'2018-{month:02d}-{int(day):02d}',
                'desc': desc[:100],
                'amount': amount,
                'balance': balance,
                'page': pg_num
            })

print(f"Extracted {len(trans)} transactions")
if not trans:
    print("No transactions found")
    exit(1)

# Audit balance
print("\nAuditing balances...")
calc = opening if opening else (trans[0]['balance'] - trans[0]['amount'])
errors = []

for i, t in enumerate(trans):
    calc += t['amount']
    diff = abs(calc - t['balance'])
    t['calc'] = calc
    t['diff'] = diff
    t['err'] = diff > Decimal('0.01')
    if t['err']:
        errors.append(i)

print(f"Opening: ${opening:,.2f}")
print(f"Closing (calc): ${calc:,.2f}")
print(f"Errors: {len(errors)}")

# Write Excel
wb = Workbook()
ws = wb.active
ws.title = "Transactions"
ws.append(['Date', 'Description', 'Amount', 'Stated Bal', 'Calc Bal', 'Diff', 'Error', 'Page'])

for c in range(1, 9):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

err_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for t in trans:
    row_data = [t['date'], t['desc'], float(t['amount']), float(t['balance']),
                float(t['calc']), float(t['diff']), 'YES' if t['err'] else '', t['page']]
    ws.append(row_data)
    if t['err']:
        for c in range(1, 9):
            ws.cell(ws.max_row, c).fill = err_fill

for c in [3, 4, 5, 6]:
    for r in range(2, len(trans) + 2):
        ws.cell(r, c).number_format = '$#,##0.00'

ws.column_dimensions['B'].width = 50

wb.save(output)
print(f"\n✅ Saved: {output}")
