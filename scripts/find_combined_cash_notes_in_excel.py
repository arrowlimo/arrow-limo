#!/usr/bin/env python3
"""Find vendor names with notes about combined cash/CIBC transactions in Excel."""

import openpyxl
from pathlib import Path

# Check both Excel files
files_to_check = [
    Path("reports/2012_receipts_and_banking.xlsx"),
    Path("reports/receipt_lookup_and_entry_2012.xlsx"),
    Path("reports/complete_receipts_workbook_20251205_162410.xlsx"),
]

for file_path in files_to_check:
    if not file_path.exists():
        print(f"Skipping (not found): {file_path}")
        continue
    
    print(f"\n{'='*80}")
    print(f"Checking: {file_path.name}")
    print(f"{'='*80}")
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name} ({ws.max_row} rows)")
        
        # Look for vendor names with keywords
        keywords = ['combined', 'cash cibc', 'cibc money', 'cibc cash', 'partial cibc', 'split', 'partly']
        
        found_count = 0
        for row_idx in range(2, min(ws.max_row + 1, 5000)):  # Limit to 5000 rows
            # Check vendor column (usually column 3 or nearby)
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    for kw in keywords:
                        if kw in cell_lower:
                            # Print the full row context
                            row_data = [ws.cell(row_idx, c).value for c in range(1, min(ws.max_column + 1, 10))]
                            print(f"    Row {row_idx}: {row_data}")
                            found_count += 1
                            break
                if found_count >= 5:
                    break
            if found_count >= 5:
                break
        
        if found_count == 0:
            print(f"    No matches found in {sheet_name}")
    
    wb.close()

print("\n" + "="*80)
