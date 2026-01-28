"""
Extract CIBC 8362 2018 PDF to Excel - TRANSACTIONAL DATA ONLY + BALANCE AUDIT
Parses PDF, extracts transactions, audits running balance for errors.
"""

import sys
import re
from pathlib import Path
from decimal import Decimal

try:
    import pdfplumber
except ImportError:
    print("Installing pdfplumber...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber"])
    import pdfplumber

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Color
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Color

pdf_path = Path(r"L:\limo\pdf\2018\CIBC 8362 2018.pdf")
output_path = Path(r"L:\limo\reports\CIBC_8362_2018_extracted.xlsx")

print("EXTRACTING CIBC 8362 2018 PDF - TRANSACTIONAL DATA + BALANCE AUDIT")
print("=" * 80)
print(f"Input:  {pdf_path}")
print(f"Output: {output_path}")
print()

transactions = []
opening_balance = None

# Multiple pattern attempts for CIBC statement formats
# CIBC 2018 format: Date Description Withdrawals Deposits Balance (multi-line with ~)
pattern = r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{1,2})\s+(.+?)(?:\s+([\d,]+\.\d{2}))?(?:\s+([\d,]+\.\d{2}))?(?:\s+~?\s*([\d,]+\.\d{2}))?\s*$'

with pdfplumber.open(pdf_path) as pdf:
    print(f"PDF has {len(pdf.pages)} pages\n")
    
    current_month = None
    current_year = 2018
    
    for page_num, page in enumerate(pdf.pages, 1):
        text = page.extract_text()
        if not text:
            continue
        
        lines = text.split('\n')
        
        # Look for statement period to get year
        for line in lines:
            if 'Jan 1 to' in line or 'Account Statement' in line:
                # Try to extract year from statement header
                year_match = re.search(r'20\d{2}', line)
                if year_match:
                    current_year = int(year_match.group())
        
        # Look for opening balance
        for line in lines:
            if 'balance forward' in line.lower() or 'opening balance' in line.lower() or 'previous balance' in line.lower():
                balance_match = re.search(r'([\d,]+\.\d{2})', line)
                if balance_match and opening_balance is None:
                    opening_balance = Decimal(balance_match.group(1).replace(',', ''))
                    print(f"Found opening balance: ${opening_balance:,.2f}")
        
        # Extract transactions - look for lines starting with month
        for i, line in enumerate(lines):
            line = line.strip()
            if not line or len(line) < 5:
                continue
            
            match = re.match(pattern, line, re.IGNORECASE)
            if match:
                month_str, day_str, description, withdrawal_str, deposit_str, balance_str = match.groups()
                
                # Month name to number
                month_map = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}
                month_num = month_map.get(month_str.lower(), 1)
                
                # Handle year boundary
                if current_month and month_num < current_month:
                    current_year += 1
                current_month = month_num
                
                trans_date = f"{current_year}-{month_num:02d}-{int(day_str):02d}"
                
                # Determine amount and type
                if withdrawal_str:
                    amount = -Decimal(withdrawal_str.replace(',', ''))
                    trans_type = 'withdrawal'
                elif deposit_str:
                    amount = Decimal(deposit_str.replace(',', ''))
                    trans_type = 'deposit'
                else:
                    # No amount on this line, might be continued description
                    continue
                
                # Balance may be on same line or next line (with ~)
                if balance_str:
                    balance = Decimal(balance_str.replace(',', ''))
                else:
                    # Look ahead for balance on next line
                    if i + 1 < len(lines):
                        next_line = lines[i + 1].strip()
                        balance_match = re.search(r'~?\s*([\d,]+\.\d{2})\s*$', next_line)
                        if balance_match:
                            balance = Decimal(balance_match.group(1).replace(',', ''))
                        else:
                            continue
                    else:
                        continue
                
                transactions.append({
                    'date': trans_date,
                    'description': description.strip(),
                    'amount': amount,
                    'type': trans_type,
                    'stated_balance': balance,
                    'page': page_num
                })

print(f"Extracted {len(transactions)} transactions\n")

if not transactions:
    print("⚠️  No transactions found. Extracting raw text for manual review...")
    with pdfplumber.open(pdf_path) as pdf:
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Text"
        ws['A1'] = "Page"
        ws['B1'] = "Content"
        for col in ['A1', 'B1']:
            ws[col].font = Font(bold=True)
        for page_num, page in enumerate(pdf.pages, 1):
            ws.append([page_num, page.extract_text() or ""])
        wb.save(output_path)
        print(f"✅ Saved raw text to: {output_path}")
    sys.exit(0)

# AUDIT RUNNING BALANCE
print("AUDITING RUNNING BALANCE")
print("=" * 80)

errors = []
calculated_balance = opening_balance if opening_balance else Decimal('0.00')

if opening_balance is None:
    print("⚠️  No opening balance found, assuming first stated balance as opening")
    if transactions:
        # Work backwards from first transaction
        first_trans = transactions[0]
        calculated_balance = first_trans['stated_balance'] - first_trans['amount']
        opening_balance = calculated_balance
        print(f"   Inferred opening balance: ${opening_balance:,.2f}\n")

for idx, trans in enumerate(transactions):
    calculated_balance += trans['amount']
    stated_balance = trans['stated_balance']
    
    diff = abs(calculated_balance - stated_balance)
    
    trans['calculated_balance'] = calculated_balance
    trans['balance_diff'] = diff
    trans['balance_error'] = diff > Decimal('0.01')
    
    if trans['balance_error']:
        errors.append({
            'row': idx + 1,
            'date': trans['date'],
            'description': trans['description'],
            'stated': stated_balance,
            'calculated': calculated_balance,
            'diff': diff
        })

if errors:
    print(f"❌ Found {len(errors)} balance errors:\n")
    for err in errors[:10]:  # Show first 10
        print(f"   Row {err['row']:>4} | {err['date']} | {err['description'][:40]:<40}")
        print(f"            Stated: ${err['stated']:>12,.2f}  Calculated: ${err['calculated']:>12,.2f}  Diff: ${err['diff']:>10,.2f}")
    if len(errors) > 10:
        print(f"\n   ... and {len(errors) - 10} more errors (see Excel)")
else:
    print("✅ All balances reconcile correctly!")

# CREATE EXCEL WITH AUDIT RESULTS
wb = Workbook()
ws = wb.active
ws.title = "Transactions + Audit"

headers = ['Date', 'Description', 'Type', 'Amount', 'Stated Balance', 'Calculated Balance', 'Diff', 'Error?', 'Page']
ws.append(headers)

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for trans in transactions:
    row_data = [
        trans['date'],
        trans['description'],
        trans['type'],
        float(trans['amount']),
        float(trans['stated_balance']),
        float(trans['calculated_balance']),
        float(trans['balance_diff']),
        'YES' if trans['balance_error'] else '',
        trans['page']
    ]
    ws.append(row_data)
    
    if trans['balance_error']:
        row_num = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = error_fill

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 8
ws.column_dimensions['I'].width = 8

for row in range(2, len(transactions) + 2):
    for col in [4, 5, 6, 7]:
        ws.cell(row=row, column=col).number_format = '$#,##0.00'

# Summary sheet
summary = wb.create_sheet("Summary")
summary['A1'] = "CIBC 8362 2018 Balance Audit Summary"
summary['A1'].font = Font(bold=True, size=14)

summary['A3'] = "Opening Balance:"
summary['B3'] = float(opening_balance) if opening_balance else 0
summary['B3'].number_format = '$#,##0.00'

summary['A4'] = "Total Transactions:"
summary['B4'] = len(transactions)

summary['A5'] = "Total Deposits:"
total_deposits = sum(t['amount'] for t in transactions if t['amount'] > 0)
summary['B5'] = float(total_deposits)
summary['B5'].number_format = '$#,##0.00'

summary['A6'] = "Total Withdrawals:"
total_withdrawals = sum(abs(t['amount']) for t in transactions if t['amount'] < 0)
summary['B6'] = float(total_withdrawals)
summary['B6'].number_format = '$#,##0.00'

summary['A7'] = "Net Change:"
summary['B7'] = float(total_deposits - total_withdrawals)
summary['B7'].number_format = '$#,##0.00'

summary['A8'] = "Closing Balance (calculated):"
summary['B8'] = float(calculated_balance)
summary['B8'].number_format = '$#,##0.00'

summary['A10'] = "Balance Errors:"
summary['B10'] = len(errors)
if errors:
    summary['B10'].font = Font(color="FF0000", bold=True)

wb.save(output_path)

print(f"\n✅ Created Excel file: {output_path}")
print(f"   {len(transactions)} transactions")
print(f"   Opening: ${opening_balance:,.2f}  Closing: ${calculated_balance:,.2f}")
print(f"   Errors: {len(errors)}")
if errors:
    print(f"\n   ⚠️  Review highlighted rows in Excel for balance discrepancies")
