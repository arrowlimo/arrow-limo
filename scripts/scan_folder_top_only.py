import os, sys, hashlib, json, csv
from pathlib import Path
from datetime import datetime

KEYWORD_CATEGORIES = [
    ("cra", "CRA/Taxes"),
    ("t4", "CRA/Taxes"),
    ("pd7a", "CRA/Taxes"),
    ("pdta", "CRA/Taxes"),
    ("vacation pay", "Payroll"),
    ("payroll", "Payroll"),
    ("pay cheque", "Payroll"),
    ("pay stubs", "Payroll"),
    ("roe", "HR/ROE"),
    ("employee", "HR"),
    ("invoice", "Invoices/Receipts"),
    ("receipt", "Invoices/Receipts"),
    ("reconcile", "Accounting"),
    ("gst", "Accounting"),
    ("hst", "Accounting"),
    ("qbo", "QuickBooks"),
    ("quickbooks", "QuickBooks"),
    ("register", "QuickBooks"),
    ("cibc", "Banking/CIBC"),
    ("scotia", "Banking/Scotia"),
    ("statement", "Banking"),
    ("bank", "Banking"),
    ("coverage", "Insurance"),
    ("policy", "Insurance"),
    ("renewal", "Insurance"),
    ("heffner", "Leasing/Heffner"),
    ("lease", "Leasing"),
    ("vehicle", "Vehicles"),
    ("fleet", "Vehicles"),
]

TEXT_EXTS = {'.txt', '.csv', '.tsv', '.eml', '.rtf'}
EXCEL_EXTS = {'.xlsx', '.xlsm'}
# We won't parse legacy .xls due to dependency limits; we will fingerprint only


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def categorize(name_lower: str, text_snippet: str = '') -> str:
    hay = f"{name_lower} {text_snippet.lower()}"
    for kw, cat in KEYWORD_CATEGORIES:
        if kw in hay:
            return cat
    return "Uncategorized"


def extract_text_snippet(path: Path, limit_chars: int = 4000) -> str:
    ext = path.suffix.lower()
    try:
        if ext in TEXT_EXTS:
            with path.open('r', encoding='utf-8', errors='ignore') as f:
                return f.read(limit_chars)
        elif ext in EXCEL_EXTS:
            try:
                import openpyxl  # type: ignore
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                sheets = wb.sheetnames
                snippet = f"sheets={sheets[:5]}"
                # try to read header row of first sheet
                ws = wb[sheets[0]]
                first = next(ws.iter_rows(max_row=1, values_only=True), None)
                if first:
                    snippet += f" header={list(first)}"
                return snippet
            except Exception:
                return ''
        else:
            return ''
    except Exception:
        return ''


def main():
    if len(sys.argv) < 2:
        print("Usage: python scan_folder_top_only.py <folder> [<out_dir>]")
        sys.exit(1)
    folder = Path(sys.argv[1])
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('exports/docs/verify_this_data_new_folder')
    out_dir.mkdir(parents=True, exist_ok=True)

    entries = []
    duplicates = {}
    dup_groups = {}
    oversized = []
    unreadable = []

    for entry in os.scandir(folder):
        if not entry.is_file():
            continue  # top-level files only
        p = Path(entry.path)
        try:
            stat = p.stat()
            size = stat.st_size
            mtime = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec='seconds')
            ext = p.suffix.lower()
            name_lower = p.name.lower()
            sha = sha256_file(p)
            snippet = extract_text_snippet(p)
            category = categorize(name_lower, snippet)
            rec = {
                'file': p.name,
                'ext': ext,
                'size': size,
                'mtime': mtime,
                'sha256': sha,
                'category': category,
                'snippet': snippet[:500],
            }
            entries.append(rec)
            duplicates.setdefault(sha, []).append(p.name)
            if size > 25 * 1024 * 1024:
                oversized.append(p.name)
        except Exception as e:
            unreadable.append({ 'file': p.name, 'error': str(e) })

    # Build duplicate groups
    for h, files in duplicates.items():
        if len(files) > 1:
            dup_groups[h] = files

    # Write manifest.csv
    manifest_csv = out_dir / 'manifest.csv'
    with manifest_csv.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['file','ext','size','mtime','sha256','category','snippet'])
        w.writeheader()
        for r in sorted(entries, key=lambda x: (x['category'], x['file'])):
            w.writerow(r)

    # Write anomalies.json
    anomalies = {
        'duplicate_groups': dup_groups,
        'oversized_files': oversized,
        'unreadable_files': unreadable,
    }
    with (out_dir / 'anomalies.json').open('w', encoding='utf-8') as f:
        json.dump(anomalies, f, indent=2)

    # Summary
    counts = {}
    for r in entries:
        counts[r['category']] = counts.get(r['category'], 0) + 1
    summary_lines = ["# Verify-this-data New folder scan (top-level only)", "", f"Scanned folder: {folder}", f"Files scanned: {len(entries)}", ""]
    summary_lines.append("## By category")
    for cat, cnt in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        summary_lines.append(f"- {cat}: {cnt}")
    summary_lines.append("")
    summary_lines.append(f"Duplicate groups: {len(dup_groups)} (see anomalies.json)")
    summary_lines.append(f"Oversized files (>25MB): {len(oversized)}")
    summary_lines.append("")
    # Quick highlights for CRA/Banking/Leasing/Insurance
    def find_files(cat_prefix):
        return [r['file'] for r in entries if r['category'].startswith(cat_prefix)]
    highlights = {
        'CRA/Taxes': find_files('CRA/Taxes')[:10],
        'Banking': [r['file'] for r in entries if r['category'] in ('Banking','Banking/CIBC','Banking/Scotia')][:10],
        'Leasing': [r['file'] for r in entries if r['category'].startswith('Leasing')][:10],
        'Insurance': find_files('Insurance')[:10],
        'Payroll': find_files('Payroll')[:10],
        'QuickBooks': find_files('QuickBooks')[:10],
        'Invoices/Receipts': find_files('Invoices/Receipts')[:10],
    }
    summary_lines.append("## Highlights (top 10 per group)")
    for k, v in highlights.items():
        summary_lines.append(f"- {k}: {len(v)} shown")
        for fn in v:
            summary_lines.append(f"  - {fn}")
    with (out_dir / 'summary.md').open('w', encoding='utf-8') as f:
        f.write('\n'.join(summary_lines))

    print(json.dumps({
        'scanned': len(entries),
        'manifest': str(manifest_csv),
        'anomalies': str(out_dir / 'anomalies.json'),
        'summary': str(out_dir / 'summary.md'),
        'duplicate_groups': len(dup_groups),
    }, indent=2))

if __name__ == '__main__':
    main()
