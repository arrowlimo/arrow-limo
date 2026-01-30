#!/usr/bin/env python3
"""Generate comprehensive Scotia Bank 2012 Excel report."""

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

conn = psycopg2.connect(
    host='localhost',
    database='almsdata',
    user='postgres',
    password='***REDACTED***'
)

cur = conn.cursor()

# Get all 2012 Scotia Bank transactions
cur.execute("""
    SELECT 
        transaction_id,
        transaction_date,
        description,
        debit_amount,
        credit_amount,
        balance,
        source_file,
        created_at
    FROM banking_transactions
    WHERE account_number = '903990106011'
      AND transaction_date >= '2012-01-01'
      AND transaction_date <= '2012-12-31'
    ORDER BY transaction_date, transaction_id
""")

transactions = cur.fetchall()

# Get monthly summaries
cur.execute("""
    SELECT 
        TO_CHAR(transaction_date, 'YYYY-MM') as month,
        TO_CHAR(transaction_date, 'Month YYYY') as month_name,
        COUNT(*) as count,
        SUM(COALESCE(debit_amount, 0)) as total_debits,
        SUM(COALESCE(credit_amount, 0)) as total_credits,
        SUM(COALESCE(credit_amount, 0)) - SUM(COALESCE(debit_amount, 0)) as net_change
    FROM banking_transactions
    WHERE account_number = '903990106011'
      AND transaction_date >= '2012-01-01'
      AND transaction_date <= '2012-12-31'
    GROUP BY TO_CHAR(transaction_date, 'YYYY-MM'), TO_CHAR(transaction_date, 'Month YYYY')
    ORDER BY TO_CHAR(transaction_date, 'YYYY-MM')
""")

monthly_summaries = cur.fetchall()

cur.close()
conn.close()

# Create Excel workbook
wb = openpyxl.Workbook()

# Style definitions
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
total_font = Font(bold=True, size=11)
month_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
month_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet 1: All Transactions
ws1 = wb.active
ws1.title = "All Transactions"

# Headers
headers = ["Transaction ID", "Date", "Description", "Debit", "Credit", "Balance", "Source", "Created"]
ws1.append(headers)

for col_num, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Data rows
for trans in transactions:
    trans_id, trans_date, desc, debit, credit, balance, source, created = trans
    ws1.append([
        trans_id,
        trans_date,
        desc,
        float(debit) if debit else None,
        float(credit) if credit else None,
        float(balance) if balance else None,
        source,
        created
    ])

# Format columns
ws1.column_dimensions['A'].width = 12  # Transaction ID
ws1.column_dimensions['B'].width = 12  # Date
ws1.column_dimensions['C'].width = 60  # Description
ws1.column_dimensions['D'].width = 12  # Debit
ws1.column_dimensions['E'].width = 12  # Credit
ws1.column_dimensions['F'].width = 12  # Balance
ws1.column_dimensions['G'].width = 25  # Source
ws1.column_dimensions['H'].width = 20  # Created

# Format numbers as currency
for row in range(2, len(transactions) + 2):
    ws1.cell(row=row, column=2).number_format = 'YYYY-MM-DD'  # Date
    ws1.cell(row=row, column=4).number_format = '$#,##0.00'   # Debit
    ws1.cell(row=row, column=5).number_format = '$#,##0.00'   # Credit
    ws1.cell(row=row, column=6).number_format = '$#,##0.00'   # Balance
    ws1.cell(row=row, column=8).number_format = 'YYYY-MM-DD HH:MM:SS'  # Created

# Add totals row
total_row = len(transactions) + 2
ws1.cell(total_row, 1).value = "TOTALS:"
ws1.cell(total_row, 1).font = total_font
ws1.cell(total_row, 2).value = f"{len(transactions)} transactions"
ws1.cell(total_row, 2).font = total_font
ws1.cell(total_row, 4).value = f'=SUM(D2:D{total_row-1})'
ws1.cell(total_row, 4).number_format = '$#,##0.00'
ws1.cell(total_row, 4).font = total_font
ws1.cell(total_row, 4).fill = total_fill
ws1.cell(total_row, 5).value = f'=SUM(E2:E{total_row-1})'
ws1.cell(total_row, 5).number_format = '$#,##0.00'
ws1.cell(total_row, 5).font = total_font
ws1.cell(total_row, 5).fill = total_fill

# Sheet 2: Monthly Summary
ws2 = wb.create_sheet("Monthly Summary")

# Headers
summary_headers = ["Month", "Transaction Count", "Total Debits", "Total Credits", "Net Change"]
ws2.append(summary_headers)

for col_num, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Monthly data
year_debits = 0
year_credits = 0
year_count = 0

for month_code, month_name, count, debits, credits, net in monthly_summaries:
    ws2.append([
        month_name.strip(),
        count,
        float(debits),
        float(credits),
        float(net)
    ])
    year_debits += float(debits)
    year_credits += float(credits)
    year_count += count

# Format columns
ws2.column_dimensions['A'].width = 20  # Month
ws2.column_dimensions['B'].width = 18  # Count
ws2.column_dimensions['C'].width = 15  # Debits
ws2.column_dimensions['D'].width = 15  # Credits
ws2.column_dimensions['E'].width = 15  # Net Change

# Format numbers
for row in range(2, len(monthly_summaries) + 2):
    ws2.cell(row=row, column=2).number_format = '#,##0'
    ws2.cell(row=row, column=3).number_format = '$#,##0.00'
    ws2.cell(row=row, column=4).number_format = '$#,##0.00'
    ws2.cell(row=row, column=5).number_format = '$#,##0.00'
    
    # Color negative net changes red
    if ws2.cell(row=row, column=5).value < 0:
        ws2.cell(row=row, column=5).font = Font(color="FF0000")

# Add year totals
total_row = len(monthly_summaries) + 2
ws2.cell(total_row, 1).value = "2012 TOTALS:"
ws2.cell(total_row, 1).font = total_font
ws2.cell(total_row, 2).value = year_count
ws2.cell(total_row, 2).font = total_font
ws2.cell(total_row, 2).number_format = '#,##0'
ws2.cell(total_row, 3).value = year_debits
ws2.cell(total_row, 3).font = total_font
ws2.cell(total_row, 3).number_format = '$#,##0.00'
ws2.cell(total_row, 3).fill = total_fill
ws2.cell(total_row, 4).value = year_credits
ws2.cell(total_row, 4).font = total_font
ws2.cell(total_row, 4).number_format = '$#,##0.00'
ws2.cell(total_row, 4).fill = total_fill
ws2.cell(total_row, 5).value = year_credits - year_debits
ws2.cell(total_row, 5).font = total_font
ws2.cell(total_row, 5).number_format = '$#,##0.00'
ws2.cell(total_row, 5).fill = total_fill

# Sheet 3: Report Info
ws3 = wb.create_sheet("Report Info")
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 50

info_data = [
    ["Scotia Bank 2012 Report", ""],
    ["", ""],
    ["Account Number:", "903990106011"],
    ["Account Name:", "Scotia Bank Main"],
    ["Statement Period:", "January 1, 2012 - December 31, 2012"],
    ["", ""],
    ["Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ["Total Transactions:", len(transactions)],
    ["Total Debits:", f"${year_debits:,.2f}"],
    ["Total Credits:", f"${year_credits:,.2f}"],
    ["Net Change:", f"${year_credits - year_debits:,.2f}"],
    ["", ""],
    ["Data Source:", "Arrow Limousine Banking Database"],
    ["Database:", "almsdata"],
    ["", ""],
    ["Notes:", "All amounts in Canadian Dollars (CAD)"],
    ["", "This report includes all reconstructed transactions"],
    ["", "from physical bank statements processed via screenshots"],
]

for row_idx, (label, value) in enumerate(info_data, 1):
    ws3.cell(row=row_idx, column=1).value = label
    ws3.cell(row=row_idx, column=2).value = value
    
    if row_idx == 1:
        ws3.cell(row=row_idx, column=1).font = Font(bold=True, size=16, color="366092")
    elif label and label.endswith(':'):
        ws3.cell(row=row_idx, column=1).font = Font(bold=True)

# Save workbook
output_file = "reports/Scotia_Bank_2012_Complete_Report.xlsx"
wb.save(output_file)

print(f"✓ Scotia Bank 2012 Excel report generated: {output_file}")
print(f"  - {len(transactions)} transactions")
print(f"  - {len(monthly_summaries)} months")
print(f"  - Total Debits: ${year_debits:,.2f}")
print(f"  - Total Credits: ${year_credits:,.2f}")
print(f"  - Net Change: ${year_credits - year_debits:,.2f}")
