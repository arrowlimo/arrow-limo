#!/usr/bin/env python3
"""
Create inter-account transfer receipts for CIBC→Scotia split deposits.
Export to Excel with bright yellow highlighting for review.
"""

import psycopg2
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import hashlib
from datetime import datetime

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        database=os.getenv('DB_NAME', 'almsdata'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', '***REDACTED***')
    )

def generate_hash(date, description, amount):
    """Generate deterministic hash for deduplication."""
    hash_input = f"{date}|{description}|{amount:.2f}".encode('utf-8')
    return hashlib.sha256(hash_input).hexdigest()

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Create receipts for CIBC→Scotia inter-account transfers')
    parser.add_argument('--write', action='store_true', help='Write receipts to database')
    parser.add_argument('--excel-output', type=Path, default=Path('reports/cibc_scotia_split_deposits.xlsx'),
                        help='Path to Excel output')
    args = parser.parse_args()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    print("\n" + "="*80)
    print("CREATE RECEIPTS FOR CIBC→SCOTIA INTER-ACCOUNT TRANSFERS")
    print("="*80)
    print(f"Mode: {'WRITE' if args.write else 'DRY RUN'}")
    
    # Matched pairs from analysis
    matched_pairs = [
        {
            'date': '2012-07-16',
            'scotia_tx_id': 63676,
            'cibc_tx_id': 57731,
            'cibc_amount': 400.00,
            'scotia_amount': 400.00,
            'cash_amount': 0.00,
            'cibc_desc': 'Cheque 277 000000017540778',
            'scotia_desc': 'DEPOSIT FROM CIBC',
            'notes': 'Full transfer - 100% from CIBC'
        },
        {
            'date': '2012-10-24',
            'scotia_tx_id': 64069,
            'cibc_tx_id': 57885,
            'cibc_amount': 1000.00,
            'scotia_amount': 1700.00,
            'cash_amount': 700.00,
            'cibc_desc': 'Cheque 280 000000044210185',
            'scotia_desc': 'DEPOSIT (1000 FROM CIBC)',
            'notes': 'Split deposit - $1,000 CIBC + $700 cash'
        },
    ]
    
    print(f"\nProcessing {len(matched_pairs)} matched CIBC→Scotia transfer pairs...\n")
    
    receipts_to_create = []
    
    for pair in matched_pairs:
        date = pair['date']
        cibc_amt = pair['cibc_amount']
        scotia_amt = pair['scotia_amount']
        cash_amt = pair['cash_amount']
        
        print(f"{'-'*80}")
        print(f"Date: {date}")
        print(f"  CIBC TX {pair['cibc_tx_id']}: ${cibc_amt:,.2f} - {pair['cibc_desc']}")
        print(f"  Scotia TX {pair['scotia_tx_id']}: ${scotia_amt:,.2f} - {pair['scotia_desc']}")
        print(f"  Cash component: ${cash_amt:,.2f}")
        print(f"  Notes: {pair['notes']}")
        
        # Create receipt for inter-account transfer
        source_hash = generate_hash(date, f"Inter-account transfer CIBC→Scotia {pair['cibc_desc']}", cibc_amt)
        
        receipt_data = {
            'date': date,
            'vendor': 'Inter-Account Transfer (CIBC→Scotia)',
            'gross_amount': cibc_amt,
            'gst_amount': 0.00,
            'net_amount': cibc_amt,
            'category': 'inter_account_transfer',
            'description': f"Transfer from CIBC to Scotia - {pair['notes']}. CIBC: {pair['cibc_desc']}. Scotia: {pair['scotia_desc']}",
            'source_hash': source_hash,
            'cibc_tx_id': pair['cibc_tx_id'],
            'scotia_tx_id': pair['scotia_tx_id'],
            'business_personal': 'Business',
            'created_from_banking': True,
            'notes': pair['notes']
        }
        
        receipts_to_create.append(receipt_data)
        print(f"  ✓ Receipt prepared: {receipt_data['vendor']} ${cibc_amt:,.2f}")
    
    # Export to Excel with bright yellow highlighting
    print(f"\nExporting to Excel: {args.excel_output}")
    args.excel_output.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CIBC→Scotia Transfers"
    
    headers = [
        "Date", "Vendor", "Amount", "Category", "Description", 
        "CIBC TX ID", "Scotia TX ID", "Notes", "Source Hash"
    ]
    ws.append(headers)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Bright yellow fill for all data rows
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for receipt in receipts_to_create:
        ws.append([
            receipt['date'],
            receipt['vendor'],
            receipt['gross_amount'],
            receipt['category'],
            receipt['description'],
            receipt['cibc_tx_id'],
            receipt['scotia_tx_id'],
            receipt['notes'],
            receipt['source_hash']
        ])
        
        # Apply bright yellow to entire row
        for cell in ws[ws.max_row]:
            cell.fill = yellow_fill
        
        # Currency format
        ws.cell(ws.max_row, 3).number_format = '#,##0.00'
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 20
    
    wb.save(args.excel_output)
    print(f"✓ Excel file created with {len(receipts_to_create)} bright yellow rows")
    
    # Write to database if requested
    if args.write:
        print(f"\nWriting {len(receipts_to_create)} receipts to database...")
        
        created_count = 0
        skipped_count = 0
        
        for receipt in receipts_to_create:
            # Check if receipt already exists
            cur.execute("SELECT receipt_id FROM receipts WHERE source_hash = %s", (receipt['source_hash'],))
            existing = cur.fetchone()
            
            if existing:
                print(f"  Skipped (already exists): {receipt['vendor']} - {receipt['date']}")
                skipped_count += 1
                continue
            
            # Create receipt
            cur.execute("""
                INSERT INTO receipts (
                    receipt_date, vendor_name, gross_amount, gst_amount, net_amount,
                    category, description, business_personal, source_hash,
                    created_from_banking, mapped_bank_account_id, created_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, 1, CURRENT_TIMESTAMP
                )
                RETURNING receipt_id
            """, (
                receipt['date'], receipt['vendor'], receipt['gross_amount'],
                receipt['gst_amount'], receipt['net_amount'], receipt['category'],
                receipt['description'], receipt['business_personal'], receipt['source_hash']
            ))
            
            receipt_id = cur.fetchone()[0]
            created_count += 1
            
            # Link to CIBC transaction
            cur.execute("""
                INSERT INTO banking_receipt_matching_ledger (
                    banking_transaction_id, receipt_id, match_type, match_confidence,
                    match_status, match_date, notes
                ) VALUES (%s, %s, 'inter_account_transfer', '100', 'matched', CURRENT_TIMESTAMP, %s)
            """, (receipt['cibc_tx_id'], receipt_id, 'CIBC side of inter-account transfer'))
            
            # Link to Scotia transaction
            cur.execute("""
                INSERT INTO banking_receipt_matching_ledger (
                    banking_transaction_id, receipt_id, match_type, match_confidence,
                    match_status, match_date, notes
                ) VALUES (%s, %s, 'inter_account_transfer', '100', 'matched', CURRENT_TIMESTAMP, %s)
            """, (receipt['scotia_tx_id'], receipt_id, 'Scotia side of inter-account transfer'))
            
            print(f"  ✓ Created receipt {receipt_id}: {receipt['vendor']} - {receipt['date']}")
        
        conn.commit()
        print(f"\n✓ Created {created_count} receipts, skipped {skipped_count} duplicates")
    else:
        print("\nDRY RUN - No database changes made")
        print("Run with --write to create receipts")
    
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Matched pairs processed: {len(matched_pairs)}")
    print(f"Receipts prepared: {len(receipts_to_create)}")
    print(f"Excel output: {args.excel_output}")
    print("\n💡 These inter-account transfers should NOT be counted as expenses.")
    print("   They represent money movement between CIBC and Scotia accounts.")
    print("="*80 + "\n")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
