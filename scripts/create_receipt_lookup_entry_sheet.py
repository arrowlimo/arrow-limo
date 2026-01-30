#!/usr/bin/env python3
"""
Create an Excel workbook to look up existing receipts (2012 by default) and provide a blank entry sheet
for adding new receipts when not found.
"""

import os
from datetime import datetime
from typing import Tuple

import pandas as pd
import psycopg2
from openpyxl.utils import get_column_letter

DB_SETTINGS = {
    "host": os.getenv("DB_HOST", "localhost"),
    "database": os.getenv("DB_NAME", "almsdata"),
    "user": os.getenv("DB_USER", "postgres"),
    "password": os.getenv("DB_PASSWORD", "***REDACTED***"),
}

OUTPUT_PATH = os.path.join("reports", "receipt_lookup_and_entry_2012.xlsx")
START_DATE = "2012-01-01"
END_DATE = "2012-12-31"

LOOKUP_COLUMNS = [
    "receipt_id",
    "receipt_date",
    "vendor_name",
    "description",
    "gross_amount",
    "gst_amount",
    "net_amount",
    "category",
    "mapped_bank_account_id",
    "display_color",
    "banking_transaction_id",
    # Optional columns below may not exist in all schemas; they will be skipped if absent
    "notes",
    "created_at",
    "updated_at",
]

ENTRY_HEADERS = [
    "receipt_date",
    "vendor_name",
    "description",
    "gross_amount",
    "gst_amount",
    "net_amount",
    "category",
    "mapped_bank_account_id",
    "display_color",
    "banking_transaction_id",
    "notes",
    "source_file",
]


def fetch_receipts(date_range: Tuple[str, str]) -> pd.DataFrame:
    # Build column list that actually exists
    conn = psycopg2.connect(**DB_SETTINGS)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'receipts'
        """
    )
    existing_cols = {row[0] for row in cur.fetchall()}
    selected_cols = [c for c in LOOKUP_COLUMNS if c in existing_cols]
    sql = f"""
        SELECT {', '.join(selected_cols)}
        FROM receipts
        WHERE receipt_date >= %s AND receipt_date <= %s
        ORDER BY receipt_date, receipt_id
    """
    try:
        df = pd.read_sql(sql, conn, params=date_range)
        return df
    finally:
        conn.close()


def strip_timezones(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            if getattr(out[col].dtype, "tz", None) is not None:
                out[col] = out[col].dt.tz_convert(None)
        elif out[col].apply(lambda x: hasattr(x, "tzinfo") and x.tzinfo is not None if pd.notna(x) else False).any():
            out[col] = out[col].apply(
                lambda x: x.tz_convert(None)
                if hasattr(x, "tz_convert")
                else (x.replace(tzinfo=None) if hasattr(x, "tzinfo") else x)
            )
    return out


def autosize(ws, df: pd.DataFrame, max_width: int = 50) -> None:
    for idx, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), df[col].astype(str).map(len).max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_width, max_len + 2)


def add_entry_template(ws) -> None:
    ws.append(["Fill new receipts here when lookup is empty"])
    ws.append([])
    ws.append(ENTRY_HEADERS)
    ws.freeze_panes = "A4"
    # Add a few blank rows for convenience
    for _ in range(25):
        ws.append([None] * len(ENTRY_HEADERS))
    autosize(ws, pd.DataFrame(columns=ENTRY_HEADERS))


def main():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    print(f"Fetching receipts from {START_DATE} to {END_DATE}...")
    receipts_df = strip_timezones(fetch_receipts((START_DATE, END_DATE)))
    print(f"  ✓ {len(receipts_df):,} rows")

    print(f"Writing workbook -> {OUTPUT_PATH}...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        # Lookup sheet with auto-filter
        receipts_df.to_excel(writer, sheet_name="Lookup", index=False)
        ws_lookup = writer.sheets["Lookup"]
        ws_lookup.auto_filter.ref = ws_lookup.dimensions
        ws_lookup.freeze_panes = "A2"
        autosize(ws_lookup, receipts_df)

        # Entry sheet template
        ws_entry = writer.book.create_sheet("Add Receipt")
        add_entry_template(ws_entry)

    print("Done.")
    print(f"Lookup rows: {len(receipts_df):,}")
    if not receipts_df.empty:
        print(f"Date range: {receipts_df['receipt_date'].min()} to {receipts_df['receipt_date'].max()}")


if __name__ == "__main__":
    main()
