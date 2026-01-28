#!/usr/bin/env python3
"""
Generate WCB 2011 report with proper payment/invoice matching.
Payments are withdrawals (negative or flagged) made within 1-2 days of invoice dates.
"""
import psycopg2
import os
from datetime import datetime, timedelta
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_NAME = os.environ.get("DB_NAME", "almsdata")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASSWORD = os.environ.get("DB_PASSWORD", "***REMOVED***")

try:
    conn = psycopg2.connect(host=DB_HOST, database=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    cur = conn.cursor()
    
    print("\n" + "="*80)
    print("GENERATING WCB 2011 REPORT")
    print("="*80)
    
    # Get all WCB invoices from 2011
    print("\nQuerying WCB 2011 invoices...")
    cur.execute("""
        SELECT 
            receipt_id,
            receipt_date,
            source_reference,
            description,
            gross_amount,
            vendor_name,
            banking_transaction_id
        FROM receipts
        WHERE vendor_name ILIKE '%wcb%'
          AND EXTRACT(YEAR FROM receipt_date) = 2011
          AND gross_amount > 0
        ORDER BY receipt_date ASC
    """)
    
    invoices = cur.fetchall()
    print(f"Found {len(invoices)} invoices")
    
    # Get banking transactions for 2011
    print("Querying banking transactions for 2011...")
    cur.execute("""
        SELECT 
            transaction_id,
            transaction_date,
            check_number,
            description,
            debit_amount,
            credit_amount
        FROM banking_transactions
        WHERE (description ILIKE '%wcb%' OR description ILIKE '%workers%')
          AND EXTRACT(YEAR FROM transaction_date) = 2011
        ORDER BY transaction_date ASC
    """)
    
    banking_txs = cur.fetchall()
    print(f"Found {len(banking_txs)} banking transactions")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WCB 2011"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    invoice_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    payment_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    matched_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    unmatched_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Date", "Type", "Ref #", "Description", "Amount", "Payment Date", "Payment Ref", "Payment Amount", "Matched", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 30
    
    # Process invoices and match to payments
    row_num = 2
    matched_payments = set()
    
    for invoice in invoices:
        rec_id, rec_date, ref, desc, amount, vendor, banking_id = invoice
        amount_f = float(amount)
        
        # Format date consistently
        if isinstance(rec_date, str):
            inv_date_str = rec_date
            inv_date_obj = datetime.strptime(rec_date, '%Y-%m-%d').date()
        else:
            inv_date_obj = rec_date
            inv_date_str = rec_date.strftime('%Y-%m-%d')
        
        # Write invoice row
        ws.cell(row=row_num, column=1).value = inv_date_str
        ws.cell(row=row_num, column=2).value = "INVOICE"
        ws.cell(row=row_num, column=3).value = str(ref) if ref else ""
        ws.cell(row=row_num, column=4).value = (desc[:50] if desc else "")
        ws.cell(row=row_num, column=5).value = amount_f
        
        # Try to find matching payment (within 1-2 days)
        matched = False
        for tx_idx, banking_tx in enumerate(banking_txs):
            if tx_idx in matched_payments:
                continue
            
            tx_id, tx_date, check, tx_desc, debit, credit = banking_tx
            
            if isinstance(tx_date, str):
                tx_date_obj = datetime.strptime(tx_date, '%Y-%m-%d').date()
            else:
                tx_date_obj = tx_date
            
            # Check if amount matches (debit or credit)
            payment_amount = debit if debit and float(debit) > 0 else (credit if credit else 0)
            if not payment_amount:
                continue
            
            payment_amount_f = float(payment_amount)
            
            # Check date within 1-2 days and amount matches
            day_diff = (tx_date_obj - inv_date_obj).days
            if 0 <= day_diff <= 2 and abs(payment_amount_f - amount_f) < 0.01:
                # Match found
                ws.cell(row=row_num, column=6).value = tx_date_obj.strftime('%Y-%m-%d')
                ws.cell(row=row_num, column=7).value = str(check) if check else str(tx_id)
                ws.cell(row=row_num, column=8).value = payment_amount_f
                ws.cell(row=row_num, column=9).value = "YES"
                ws.cell(row=row_num, column=10).value = f"Matched to TX {tx_id}"
                
                matched = True
                matched_payments.add(tx_idx)
                break
        
        if not matched:
            ws.cell(row=row_num, column=9).value = "NO"
            ws.cell(row=row_num, column=10).value = "No matching payment found"
        
        # Apply styles
        for col in range(1, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = matched_fill if matched else unmatched_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if col in [5, 8]:  # Amount columns
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        row_num += 1
    
    # Add unmatched payments
    print("\nAdding unmatched payments...")
    for tx_idx, banking_tx in enumerate(banking_txs):
        if tx_idx in matched_payments:
            continue
        
        tx_id, tx_date, check, tx_desc, debit, credit = banking_tx
        payment_amount = debit if debit and float(debit) > 0 else (credit if credit else 0)
        
        if not payment_amount:
            continue
        
        if isinstance(tx_date, str):
            tx_date_str = tx_date
        else:
            tx_date_str = tx_date.strftime('%Y-%m-%d')
        
        ws.cell(row=row_num, column=1).value = tx_date_str
        ws.cell(row=row_num, column=2).value = "PAYMENT"
        ws.cell(row=row_num, column=3).value = str(check) if check else str(tx_id)
        ws.cell(row=row_num, column=4).value = (tx_desc[:50] if tx_desc else "")
        ws.cell(row=row_num, column=6).value = tx_date_str
        ws.cell(row=row_num, column=7).value = str(check) if check else str(tx_id)
        ws.cell(row=row_num, column=8).value = float(payment_amount)
        ws.cell(row=row_num, column=9).value = "NO"
        ws.cell(row=row_num, column=10).value = "Unmatched payment"
        
        # Apply styles
        for col in range(1, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = unmatched_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if col == 8:
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
        
        row_num += 1
    
    # Save file
    output_path = r"L:\limo\reports\WCB_2011.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"\n" + "="*80)
    print("REPORT GENERATED")
    print("="*80)
    print(f"\nFile saved: {output_path}")
    print(f"\nInvoices: {len(invoices)}")
    print(f"Banking transactions: {len(banking_txs)}")
    print(f"Matched: {len(matched_payments)}")
    print(f"Unmatched payments: {len(banking_txs) - len(matched_payments)}")
    
    cur.close()
    conn.close()
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()
