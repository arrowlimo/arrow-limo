#!/usr/bin/env python3
"""Fix the $116.00 entry by adding date and description."""

import openpyxl
import pandas as pd

xlsx_path = r"l:\limo\data\2012_scotia_transactions_for_editing.xlsx"

# Load workbook
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

print("=" * 80)
print("FIXING $116.00 ENTRY")
print("=" * 80)

# Find the row with $116.00 debit (no date)
df = pd.read_excel(xlsx_path, sheet_name=0)
row_116_idx = df[df['debit/withdrawal'] == 116.0].index[0]
excel_row = row_116_idx + 2  # +2 because Excel is 1-indexed and has header

print(f"\nFound $116.00 entry at Excel row {excel_row}")
print(f"Current values:")
print(f"  Date: {ws.cell(excel_row, 1).value}")
print(f"  Description: {ws.cell(excel_row, 2).value}")
print(f"  Debit: {ws.cell(excel_row, 3).value}")

# Update the date and description
from datetime import datetime
ws.cell(excel_row, 1).value = datetime(2012, 12, 3)  # Column A: date
ws.cell(excel_row, 2).value = "Run'N On Empty"       # Column B: Description

print(f"\nUpdated values:")
print(f"  Date: 2012-12-03")
print(f"  Description: Run'N On Empty")
print(f"  Debit: $116.00")

# Save the file
wb.save(xlsx_path)
print(f"\n✓ Saved changes to: {xlsx_path}")

# Verify the fix
df_updated = pd.read_excel(xlsx_path, sheet_name=0)
row_updated = df_updated.iloc[row_116_idx]
print(f"\nVerification:")
print(f"  Date: {row_updated['date']}")
print(f"  Description: {row_updated['Description']}")
print(f"  Debit: ${row_updated['debit/withdrawal']:.2f}")
print(f"  Balance: ${row_updated['balance']:.2f}")

print("\n" + "=" * 80)
print("✓ FIX COMPLETE - Ready to re-import Scotia data")
print("=" * 80)

wb.close()
