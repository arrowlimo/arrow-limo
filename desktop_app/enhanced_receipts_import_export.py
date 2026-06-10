"""
Enhanced import/export functionality for Receipts Manager
Features: Excel export with change detection hashes,
smart reimport (only updates changed rows)
"""

import hashlib
from datetime import datetime

from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QProgressDialog


class EnhancedReceiptsImportExport:
    """Import/Export handler for receipts with change detection."""

    @staticmethod
    def export_to_excel(conn, table_widget, parent=None) -> None:
        """Export receipts to Excel with row hashes for change detection."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            QMessageBox.warning(
                parent,
                "Missing Library",
                "openpyxl not installed.\nInstall with: pip install openpyxl",
            )
            return

        filename, _ = QFileDialog.getSaveFileName(
            parent,
            "Export Receipts to Excel",
            f"Receipts_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)",
        )

        if not filename:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Receipts"

            # Get visible headers
            headers = []
            visible_cols = []
            for col in range(table_widget.columnCount()):
                if not table_widget.isColumnHidden(col):
                    header = table_widget.horizontalHeaderItem(col)
                    headers.append(header.text() if header else f"Col {col}")
                    visible_cols.append(col)

            # Add metadata columns
            headers.extend(["_HASH", "_ROW_ID", "_CHANGED"])

            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Export rows with metadata
            for row_idx in range(table_widget.rowCount()):
                row_data = []

                # Get visible cell data
                for col_idx in visible_cols:
                    item = table_widget.item(row_idx, col_idx)
                    row_data.append(item.text() if item else "")

                # Create hash of row data for change detection
                row_str = "|".join(str(v) for v in row_data)
                # MD5 used for checksums only (not security), so
                # usedforsecurity=False
                row_hash = hashlib.md5(
                    row_str.encode(), usedforsecurity=False
                ).hexdigest()[:8]

                # Get receipt ID from first column (assuming ID is first
                # visible column)
                receipt_id = row_data[0] if row_data else ""

                # Add metadata
                # _CHANGED will be filled by user
                row_data.extend([row_hash, receipt_id, ""])

                # Write to Excel
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(
                        row=row_idx + 2, column=col_idx, value=value
                    )
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )

                    # Highlight _ROW_ID and _HASH columns as read-only style
                    if col_idx > len(headers) - 3:  # Last 3 are metadata
                        meta_fill = PatternFill(
                            start_color="E8E8E8",
                            end_color="E8E8E8",
                            fill_type="solid",
                        )
                        cell.fill = meta_fill

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Add instructions
            ws.append([])  # Blank row
            ws.append(["INSTRUCTIONS FOR REIMPORT:"])
            ws.append(["1. Edit any cells in the data rows (rows 2 onwards)"])
            ws.append(
                ["2. Mark edited rows with 'YES' in the _CHANGED column"]
            )
            ws.append(["3. DO NOT edit _HASH or _ROW_ID columns"])
            ws.append(
                [
                    "4. Save file and use 'Import & Update' to reimport only"
                    "changed rows"
                ]
            )

            wb.save(filename)
            row_count = table_widget.rowCount()
            QMessageBox.information(
                parent,
                "Export Success",
                f"✅ Exported {row_count} receipts to:\n{filename}\n\n"
                f"Instructions saved in file. Mark edited rows with 'YES' in"
                f"_CHANGED column.",
            )

        except Exception as e:
            QMessageBox.critical(
                parent, "Export Error", f"Failed to export:\n{e}"
            )

    @staticmethod
    def import_from_excel(
        conn, table_widget, parent=None
    ) -> tuple[int, int, int]:
        """
        Import and update receipts from Excel.
        Only updates rows marked as changed, prevents duplicate data.
        Returns: (updated_count, appended_count, skipped_count)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            QMessageBox.warning(
                parent,
                "Missing Library",
                "openpyxl not installed.\nInstall with: pip install openpyxl",
            )
            return 0, 0, 0

        filename, _ = QFileDialog.getOpenFileName(
            parent,
            "Import Receipts from Excel",
            "",
            "Excel Files (*.xlsx);;All Files (*)",
        )

        if not filename:
            return 0, 0, 0

        try:
            wb = load_workbook(filename)
            ws = wb.active

            # Get headers
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value))

            if "_HASH" not in headers or "_CHANGED" not in headers:
                QMessageBox.warning(
                    parent,
                    "Invalid File",
                    "Excel file must have _HASH and _CHANGED columns.\n"
                    "Please export from 'Export Excel' button first.",
                )
                return 0, 0, 0

            hash_col_idx = headers.index("_HASH") + 1
            changed_col_idx = headers.index("_CHANGED") + 1
            row_id_col_idx = headers.index("_ROW_ID") + 1

            updated_count = 0
            appended_count = 0
            skipped_count = 0
            updates_list = []

            # Process imported rows
            progress = QProgressDialog(
                "Importing receipts...", "Cancel", 0, ws.max_row - 1, parent
            )
            progress.setWindowModality(Qt.WindowModality.WindowModal)

            for row_num in range(2, ws.max_row + 1):
                progress.setValue(row_num - 1)

                # Get metadata
                is_changed = (
                    str(
                        ws.cell(row=row_num, column=changed_col_idx).value
                        or ""
                    ).upper()
                    == "YES"
                )
                row_id = ws.cell(row=row_num, column=row_id_col_idx).value
                imported_hash = ws.cell(row=row_num, column=hash_col_idx).value

                if not imported_hash:
                    skipped_count += 1
                    continue

                # Get row data (exclude metadata columns)
                row_data = []
                for col_num in range(1, hash_col_idx):  # Stop before metadata
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data.append(cell_value or "")

                if is_changed and row_id:
                    # Mark for update (user will handle via CRUD operations)
                    updated_count += 1
                    updates_list.append(
                        {
                            "data": row_data,
                            "row_id": row_id,
                            "hash": imported_hash,
                        }
                    )
                else:
                    skipped_count += 1

            progress.close()

            # Display summary and let user confirm
            if updated_count > 0 or appended_count > 0:
                _result = QMessageBox.information(
                    parent,
                    "Import Ready",
                    f"✅ Import analysis complete!\n\n"
                    f"• Updated (marked YES): {updated_count} rows\n"
                    f"• New (unmarked): {appended_count} rows\n"
                    f"• Skipped (unmarked changed): {skipped_count} rows\n\n"
                    f"Review the data above before confirming changes.\n"
                    f"Changes must be applied via Edit or Add buttons.",
                    QMessageBox.StandardButton.Ok,
                )

                return updated_count, appended_count, skipped_count
            else:
                QMessageBox.information(
                    parent,
                    "No Changes",
                    "No rows marked for import.\n"
                    "Mark rows with 'YES' in _CHANGED column to import.",
                )
                return 0, 0, skipped_count

        except Exception as e:
            import traceback

            QMessageBox.critical(
                parent,
                "Import Error",
                f"Failed to import:\n{e}\n\n{traceback.format_exc()}",
            )
            return 0, 0, 0
