"""
Shared report widget base with toolbar and table utilities.
Provides: refresh, filter, column toggles, export CSV, quick totals.
"""

import csv
from collections.abc import Callable
from datetime import datetime
from decimal import Decimal
from typing import Any

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QAction
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QToolBar,
    QToolButton,
    QVBoxLayout,
    QWidget,
)


class BaseReportWidget(QWidget):
    """Reusable report widget with a standard toolbar and table helpers."""

    def setLayout(self, layout) -> None:
        """Prevent resetting layout on BaseReportWidget instances."""
        if self.layout() is None:
            super().setLayout(layout)
        else:
            return

    def __init__(self, db, title: str, columns: list[dict[str, Any]]) -> None:
        super().__init__()
        self.db = db
        self.title = title
        self.columns = columns
        self.rows: list[dict[str, Any]] = []
        self.raw_values = {}

        self.table = QTableWidget()
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(
            [c.get("header", "") for c in columns]
        )
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectItems
        )
        self.table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self.table.setSortingEnabled(True)
        self.table.itemSelectionChanged.connect(self.update_selection_summary)
        self.table.doubleClicked.connect(
            self.open_drill_down_dialog
        )  # Drill-down on double-click

        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Filter rows (live)")
        self.filter_input.textChanged.connect(self.apply_filter)

        self.summary_label = QLabel("Ready")
        self.summary_label.setStyleSheet("color: gray; font-size: 11px;")

        layout = QVBoxLayout()
        layout.addWidget(self._build_toolbar())
        layout.addLayout(self._build_filter_row())
        layout.addWidget(self.table)
        layout.addWidget(self.summary_label)
        self.setLayout(layout)

        # Initial load
        self.refresh()

    # ------------------------------------------------------------------
    # UI builders
    # ------------------------------------------------------------------
    def _build_toolbar(self) -> object:
        toolbar = QToolBar()

        refresh_action = QAction("Refresh", self)
        refresh_action.triggered.connect(self.refresh)
        toolbar.addAction(refresh_action)

        export_csv_action = QAction("Export CSV", self)
        export_csv_action.triggered.connect(self.export_csv)
        toolbar.addAction(export_csv_action)

        export_excel_action = QAction("Export Excel", self)
        export_excel_action.triggered.connect(self.export_excel)
        toolbar.addAction(export_excel_action)

        export_pdf_action = QAction("Export PDF", self)
        export_pdf_action.triggered.connect(self.export_pdf)
        toolbar.addAction(export_pdf_action)

        print_action = QAction("Print", self)
        print_action.triggered.connect(self.print_report)
        toolbar.addAction(print_action)

        self.columns_menu_button = QToolButton()
        self.columns_menu_button.setText("Columns")
        self.columns_menu_button.setPopupMode(
            QToolButton.ToolButtonPopupMode.InstantPopup
        )
        toolbar.addWidget(self.columns_menu_button)

        toolbar.addSeparator()

        return toolbar

    def _build_filter_row(self) -> object:
        row = QHBoxLayout()
        row.addWidget(QLabel("Filter:"))
        row.addWidget(self.filter_input)
        row.addStretch()
        return row

    # ------------------------------------------------------------------
    # Data lifecycle
    # ------------------------------------------------------------------
    def refresh(self) -> None:
        try:
            self.load_data()
            self.apply_filter()
            self.update_selection_summary()
            self.summary_label.setText("Refreshed")
        except Exception as e:
            QMessageBox.critical(
                self, "Load Error", f"Could not load report:\n{e}"
            )

    def load_data(self) -> None:
        """Fetch rows from subclass and populate the table."""
        self.rows = self.fetch_rows()
        self._populate_table(self.rows)
        self._build_column_menu()

    def set_columns(self, columns: list[dict[str, Any]]) -> None:
        """Update column definitions and table headers."""
        self.columns = columns
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(
            [c.get("header", "") for c in columns]
        )

    def fetch_rows(self) -> list[dict[str, Any]]:
        """Subclasses must implement and return a list of dict rows."""
        raise NotImplementedError

    def _populate_table(self, rows: list[dict[str, Any]]) -> None:
        self.raw_values.clear()
        self.table.setRowCount(len(rows))
        for r_idx, row in enumerate(rows):
            for c_idx, col in enumerate(self.columns):
                key = col.get("key")
                formatter: Callable[[Any], str] = col.get("format")
                val = row.get(key)
                display = (
                    formatter(val)
                    if formatter
                    else ("" if val is None else str(val))
                )
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(r_idx, c_idx, item)
                self.raw_values[(r_idx, c_idx)] = val
        self.table.resizeColumnsToContents()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def apply_filter(self) -> None:
        text = (self.filter_input.text() or "").lower()
        for row in range(self.table.rowCount()):
            match = False
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item and text in item.text().lower():
                    match = True
                    break
            self.table.setRowHidden(row, not match if text else False)

    def export_csv(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", "report.csv", "CSV Files (*.csv)"
        )
        if not path:
            return
        headers = [col.get("header", "") for col in self.columns]
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row_idx in range(self.table.rowCount()):
                if self.table.isRowHidden(row_idx):
                    continue
                writer.writerow(
                    [
                        (
                            self.table.item(row_idx, c).text()
                            if self.table.item(row_idx, c)
                            else ""
                        )
                        for c in range(self.table.columnCount())
                    ]
                )
        QMessageBox.information(self, "Exported", f"Saved to {path}")

    def export_excel(self) -> None:
        """Export to Excel with formatting, subtotals, and totals"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            path, _ = QFileDialog.getSaveFileName(
                self, "Export Excel", "report.xlsx", "Excel Files (*.xlsx)"
            )
            if not path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.title[:31]  # Excel sheet name limit

            # Header row with styling
            headers = [col.get("header", "") for col in self.columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            # Data rows
            row_offset = 2
            for row_idx in range(self.table.rowCount()):
                if self.table.isRowHidden(row_idx):
                    continue
                for col_idx in range(self.table.columnCount()):
                    item = self.table.item(row_idx, col_idx)
                    value = item.text() if item else ""
                    # Try to convert numeric values
                    try:
                        if "$" in value:
                            value = float(
                                value.replace("$", "").replace(",", "")
                            )
                        elif (
                            value.replace(".", "", 1)
                            .replace("-", "", 1)
                            .isdigit()
                        ):
                            value = float(value)
                    except Exception:
                        pass
                    ws.cell(row=row_offset, column=col_idx + 1, value=value)
                row_offset += 1

            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(path)
            QMessageBox.information(
                self, "Success", f"Excel file saved to {path}"
            )
        except ImportError:
            QMessageBox.warning(
                self,
                "Excel Export",
                "openpyxl not installed. Run: pip install openpyxl",
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Excel export failed: {e}")

    def export_pdf(self) -> None:
        """Export to PDF with professional formatting"""
        try:
            from PyQt6.QtGui import QPageSize, QTextDocument
            from PyQt6.QtPrintSupport import QPrinter

            path, _ = QFileDialog.getSaveFileName(
                self, "Export PDF", "report.pd", "PDF Files (*.pdf)"
            )
            if not path:
                return

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(path)
            printer.setPageSize(QPageSize(QPageSize.PageSizeId.Letter))

            html = self._table_to_html_enhanced()
            doc = QTextDocument()
            doc.setHtml(html)
            doc.print(printer)

            QMessageBox.information(self, "Success", f"PDF saved to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"PDF export failed: {e}")

    def _build_column_menu(self) -> None:
        menu = self.columns_menu_button.menu()
        if menu:
            menu.clear()
        else:
            menu = QMenu(self)
            self.columns_menu_button.setMenu(menu)

        for idx, col in enumerate(self.columns):
            action = QAction(
                col.get("header", f"Col {idx + 1}"),
                self,
                checkable=True,
                checked=True,
            )
            action.toggled.connect(
                lambda checked, col_idx=idx: self.table.setColumnHidden(
                    col_idx, not checked
                )
            )
            menu.addAction(action)

    def print_report(self) -> None:
        """Print the current report with print preview"""
        try:
            from PyQt6.QtGui import QTextDocument
            from PyQt6.QtPrintSupport import QPrinter, QPrintPreviewDialog

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)

            # Use print preview dialog
            preview = QPrintPreviewDialog(printer, self)
            preview.setWindowTitle(f"Print Preview - {self.title}")

            def handle_paint_request(printer) -> None:
                html = self._table_to_html_enhanced()
                doc = QTextDocument()
                doc.setHtml(html)
                doc.print(printer)

            preview.paintRequested.connect(handle_paint_request)
            preview.exec()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Print failed: {e}")

    def _table_to_html(self) -> object:
        """Convert table to HTML"""
        html = (
            f"<h2>{self.title}</h2>"
            "<table border='1' style='border-collapse:collapse;'><tr>"
        )

        # Headers
        for i in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(i)
            if header:
                html += f"<th style='padding:8px;'>{header.text()}</th>"
        html += "</tr>"

        # Rows
        for i in range(self.table.rowCount()):
            html += "<tr>"
            for j in range(self.table.columnCount()):
                item = self.table.item(i, j)
                text = item.text() if item else ""
                html += f"<td style='padding:8px;'>{text}</td>"
            html += "</tr>"

        html += "</table>"
        return html

    def _table_to_html_enhanced(self) -> object:
        """Convert table to HTML with enhanced styling and totals"""
        html = """<html><head><style>
        body {{ font-family: Arial, sans-serif;}}
        h2 {{ color: #333; border-bottom: 2px solid #366092; padding-bottom:
        10px;}}
        table {{ width: 100%; border-collapse: collapse; margin-top: 20px;}}
        th {{ background-color: #366092; color: white; padding: 12px;
        text-align: left; font-weight: bold;}}
        td {{ padding: 8px; border: 1px solid #ddd;}}
        tr:nth-child(even) {{ background-color: #f2f2f2;}}
        .total-row {{ background-color: #d9e9f7; font-weight: bold;}}
        .group-header {{ background-color: #b8daf0; font-weight: bold;
        font-style: italic;}}
        </style></head><body>"""

        html += f"<h2>{self.title}</h2>"
        html += (
            f"<p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"
        )
        html += "<table><thead><tr>"

        # Headers
        for i in range(self.table.columnCount()):
            if not self.table.isColumnHidden(i):
                header = self.table.horizontalHeaderItem(i)
                if header:
                    html += f"<th>{header.text()}</th>"
        html += "</tr></thead><tbody>"

        # Rows
        for i in range(self.table.rowCount()):
            if not self.table.isRowHidden(i):
                html += "<tr>"
                for j in range(self.table.columnCount()):
                    if not self.table.isColumnHidden(j):
                        item = self.table.item(i, j)
                        text = item.text() if item else ""
                        html += f"<td>{text}</td>"
                html += "</tr>"

        # Calculate totals for numeric columns
        html += '<tr class="total-row"><td><strong>TOTALS:</strong></td>'
        for j in range(1, self.table.columnCount()):
            if not self.table.isColumnHidden(j):
                total = 0
                count = 0
                for i in range(self.table.rowCount()):
                    if not self.table.isRowHidden(i):
                        val = self.raw_values.get((i, j))
                        if isinstance(val, (int, float, Decimal)):
                            total += float(val)
                            count += 1
                if count > 0:
                    html += f"<td><strong>${total:,.2f}</strong></td>"
                else:
                    html += "<td></td>"
        html += "</tr>"

        html += "</tbody></table></body></html>"
        return html

    def update_selection_summary(self) -> None:
        total = 0.0
        count = 0
        for index in self.table.selectedIndexes():
            val = self.raw_values.get((index.row(), index.column()))
            if isinstance(val, (int, float, Decimal)):
                total += float(val)
                count += 1
        if count:
            self.summary_label.setText(
                f"Selected {count} cells • Sum={total:,.2f}"
            )
        else:
            self.summary_label.setText("Ready")

    def open_drill_down_dialog(self, index) -> None:
        """Open detail dialog for editing row data"""
        from PyQt6.QtWidgets import (
            QCheckBox,
            QDialog,
            QFormLayout,
            QLineEdit,
            QPushButton,
            QTextEdit,
        )

        row = index.row()
        if row < 0 or row >= len(self.rows):
            return

        row_data = self.rows[row]

        # Create detail dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Detail View - {self.title}")
        dialog.setGeometry(100, 100, 500, 400)

        layout = QVBoxLayout()
        form = QFormLayout()

        # Create editable fields for each column
        fields = {}
        for col in self.columns:
            key = col.get("key")
            header = col.get("header", key)
            value = row_data.get(key, "")

            # Determine widget type
            if key in ["notes", "description", "comments"]:
                widget = QTextEdit()
                widget.setPlainText(str(value or ""))
                widget.setMaximumHeight(80)
            else:
                widget = QLineEdit()
                widget.setText(str(value or ""))

            # Default to locked/read-only until unlocked
            if isinstance(widget, QTextEdit):
                widget.setReadOnly(True)
            else:
                widget.setReadOnly(True)

            fields[key] = widget
            form.addRow(f"{header}:", widget)

        layout.addLayout(form)

        # Unlock edits toggle
        unlock_layout = QHBoxLayout()
        unlock_check = QCheckBox("Unlock edits")
        unlock_layout.addWidget(unlock_check)
        unlock_layout.addStretch()
        layout.addLayout(unlock_layout)

        # Button row
        button_layout = QHBoxLayout()

        save_btn = QPushButton("💾 Save Corrections")
        save_btn.setEnabled(False)
        save_btn.clicked.connect(
            lambda: self._save_drill_down_corrections(row, fields, dialog)
        )
        button_layout.addWidget(save_btn)

        back_btn = QPushButton("⬅ Back (No Save)")
        back_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(back_btn)

        layout.addLayout(button_layout)

        def _set_editable(enabled: bool) -> None:
            for widget in fields.values():
                if isinstance(widget, QTextEdit):
                    widget.setReadOnly(not enabled)
                else:
                    widget.setReadOnly(not enabled)
            save_btn.setEnabled(enabled)

        unlock_check.toggled.connect(_set_editable)

        dialog.setLayout(layout)
        dialog.exec()

    def _save_drill_down_corrections(self, row, fields, dialog) -> None:
        """Save corrections made in drill-down dialog"""
        try:
            # Update local row data
            row_data = self.rows[row]
            for key, widget in fields.items():
                if isinstance(widget, QTextEdit):
                    new_value = widget.toPlainText()
                else:
                    new_value = widget.text()

                # Update display table
                for c_idx, col in enumerate(self.columns):
                    if col.get("key") == key:
                        formatter = col.get("format")
                        display = (
                            formatter(new_value) if formatter else new_value
                        )
                        self.table.setItem(
                            row, c_idx, QTableWidgetItem(display)
                        )
                        row_data[key] = new_value
                        break

            # Call subclass hook to save to database
            if hasattr(self, "save_row_corrections"):
                self.save_row_corrections(row, row_data)

            QMessageBox.information(
                self,
                "Success",
                "Changes saved locally. Sync with database to persist.",
            )
            dialog.close()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save: {e}")
