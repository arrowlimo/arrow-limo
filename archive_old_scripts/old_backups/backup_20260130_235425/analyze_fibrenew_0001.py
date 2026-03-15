#!/usr/bin/env python3
"""
Analyze all 3 sheets from fibrenew_0001.xlsx to understand the full payment history.
"""

import pandas as pd
import openpyxl

file_path = r'L:\limo\receipts\fibrenew_0001.xlsx'

print("ANALYZING FIBRENEW_0001.XLSX - ALL SHEETS")
print("=" * 80)

# Load workbook
wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print("=" * 80)
    
    ws = wb[sheet_name]
    
    # Print all cell values
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = ' | '.join([str(cell) if cell is not None else '' for cell in row])
        if row_str.strip():
            print(f"Row {row_idx:2d}: {row_str}")

print("\n" + "=" * 80)
print("SUMMARY")
print("=" * 80)
print("Sheet1: Payment Receipt - $700 cash payment on 03/26/2019")
print("        Paying invoices 7598 ($472.50) and 7848 ($227.50)")
print("Sheet2: Invoice 7848 dated 11/29/2017")
print("Sheet3: Invoice 7598 dated 10/01/2017")
