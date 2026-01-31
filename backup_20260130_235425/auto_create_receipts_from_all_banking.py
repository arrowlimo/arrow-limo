#!/usr/bin/env python3
"""
Create receipts for ALL unmatched banking transactions from CIBC and Scotia.
Flags created receipts with created_from_banking = TRUE.
"""

import psycopg2
import os
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        database=os.getenv('DB_NAME', 'almsdata'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', '***REDACTED***')
    )

def generate_hash(date, description, amount):
    """Generate deterministic hash for deduplication."""
    hash_input = f"{date}|{description}|{amount:.2f}".encode('utf-8')
    return hashlib.sha256(hash_input).hexdigest()

def categorize_transaction(description):
    """Categorize transaction based on description."""
    desc_upper = description.upper()
    
    # Fuel stations
    if any(x in desc_upper for x in ['CENTEX', 'FAS GAS', 'FASGAS', 'SHELL', 'ESSO', 
                                      'PETRO', 'HUSKY', 'CO-OP', 'COOP']):
        return 'fuel'
    
    # Office supplies
    if any(x in desc_upper for x in ['STAPLES', 'OFFICE DEPOT']):
        return 'office_supplies'
    
    # Maintenance
    if any(x in desc_upper for x in ['CANADIAN TIRE', 'CDN TIRE', 'MIDAS', 'JIFFY']):
        return 'maintenance'
    
    # Communications
    if any(x in desc_upper for x in ['TELUS', 'ROGERS', 'BELL', 'SASKTEL']):
        return 'communication'
    
    # Bank fees
    if any(x in desc_upper for x in ['FEE', 'NSF', 'OVERDRAFT', 'SERVICE CHARGE', 'S/C']):
        return 'bank_fees'
    
    # General Journal (accounting adjustments, not vendor invoices)
    if any(x in desc_upper for x in ['GENERAL JOURNAL', 'GEN J', 'G J']):
        return 'general_journal'
    
    # Insurance
    if any(x in desc_upper for x in ['INSURANCE', 'SGI', 'AVIVA', 'JEVCO']):
        return 'insurance'
    
    # Equipment/Vehicle lease
    if any(x in desc_upper for x in ['HEFFNER', 'LEASE', 'FINANCING']):
        return 'equipment_lease'
    
    # Government
    if any(x in desc_upper for x in ['CRA', 'CANADA REVENUE', 'WCB', 'RECEIVER GENERAL']):
        return 'government_fees'
    
    # Credit card payments
    if any(x in desc_upper for x in ['MCC PAYMENT', 'CREDIT CARD', 'AMEX', 'VISA', 'MASTERCARD']):
        return 'credit_card_payment'
    
    # Rent
    if any(x in desc_upper for x in ['RENT', 'LANDLORD']):
        return 'rent'
    
    # Liquor/hospitality
    if any(x in desc_upper for x in ['LIQUOR', 'BAR', 'PUB', 'BEVERAGE']):
        return 'hospitality_supplies'
    
    # Meals
    if any(x in desc_upper for x in ['RESTAURANT', 'FOOD', 'DINING', 'TIM HORTONS', 'MCDONALDS']):
        return 'meals_entertainment'
    
    # Cheques
    if 'CHEQUE' in desc_upper or 'CHQ' in desc_upper:
        return 'cheque_payment'
    
    # Default
    return 'uncategorized'

def extract_vendor_from_description(description):
    """Extract vendor name from banking transaction description."""
    desc = description.strip()
    
    # Remove common prefixes
    prefixes = ['PURCHASE', 'DEBIT MEMO', 'PRE-AUTH', 'PREAUTH', 'PAD', 'TRANSFER', 
                'WITHDRAWAL', 'E-TRANSFER', 'ETRANSFER']
    
    for prefix in prefixes:
        if desc.upper().startswith(prefix):
            # Remove prefix and any trailing numbers/symbols
            desc = desc[len(prefix):].strip()
            desc = desc.lstrip('#0123456789').strip()
            break
    
    # Take first meaningful part (up to 50 chars)
    vendor = desc[:50].strip()
    
    # If vendor is too short or empty, use full description
    if len(vendor) < 3:
        vendor = description[:50]
    
    return vendor

def calculate_gst(gross_amount, tax_rate=0.05):
    """Calculate GST included in amount."""
    gst_amount = gross_amount * tax_rate / (1 + tax_rate)
    net_amount = gross_amount - gst_amount
    return round(gst_amount, 2), round(net_amount, 2)


def parse_iso_date(date_str):
    """Parse YYYY-MM-DD strings into date objects."""
    return datetime.strptime(date_str, "%Y-%m-%d").date()


def export_receipts_to_excel(receipts: Iterable[dict], output_path: Path):
    """Export prepared receipts to Excel and highlight them bright yellow."""
    receipts = list(receipts)
    if not receipts:
        print("\n   No receipts to export; skipping Excel output")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Receipts to Create"

    headers = [
        "transaction_id",
        "transaction_date",
        "account_number",
        "vendor",
        "description",
        "gross_amount",
        "gst_amount",
        "net_amount",
        "category",
        "source_hash",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for receipt in receipts:
        ws.append([
            receipt.get("transaction_id"),
            receipt.get("date"),
            receipt.get("account_number"),
            receipt.get("vendor"),
            receipt.get("description"),
            receipt.get("gross_amount"),
            receipt.get("gst_amount"),
            receipt.get("net_amount"),
            receipt.get("category"),
            receipt.get("source_hash"),
        ])

        # Highlight the entire row bright yellow so reviewers can spot auto-created receipts.
        for cell in ws[ws.max_row]:
            cell.fill = highlight_fill

        # Apply currency formatting to monetary columns.
        for col_letter in ["F", "G", "H"]:
            ws[f"{col_letter}{ws.max_row}"].number_format = "#,##0.00"

    wb.save(output_path)
    print(f"\n   Excel summary written to: {output_path}")
    print("   All rows are highlighted bright yellow for quick review")

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Create receipts for unmatched banking transactions')
    parser.add_argument('--write', action='store_true', help='Write receipts to database')
    parser.add_argument('--account', help='Specific account number (0228362=CIBC, 903990106011=Scotia)')
    parser.add_argument('--start-date', type=parse_iso_date, help='Only process banking transactions on/after this YYYY-MM-DD date')
    parser.add_argument('--end-date', type=parse_iso_date, help='Only process banking transactions on/before this YYYY-MM-DD date')
    parser.add_argument('--excel-output', type=Path, default=Path('reports/auto_created_receipts_unmatched_banking.xlsx'),
                        help='Path to Excel summary of auto-created receipts (highlighted bright yellow)')
    args = parser.parse_args()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    print("\n" + "="*80)
    print("CREATE RECEIPTS FROM UNMATCHED BANKING TRANSACTIONS")
    print("="*80)
    print(f"Mode: {'WRITE' if args.write else 'DRY RUN'}")
    if args.start_date or args.end_date:
        print(f"Date filter: {args.start_date or 'MIN'} to {args.end_date or 'MAX'}")
    if args.account:
        print(f"Account filter: {args.account}")
    
    # Find all banking transactions (debits only - expenses) not linked to receipts
    filters = ["bt.account_number IN ('0228362', '903990106011')", "bt.debit_amount IS NOT NULL", "bt.debit_amount > 0"]
    params = []

    if args.account:
        filters.append("bt.account_number = %s")
        params.append(args.account)

    if args.start_date:
        filters.append("bt.transaction_date >= %s")
        params.append(args.start_date)

    if args.end_date:
        filters.append("bt.transaction_date <= %s")
        params.append(args.end_date)

    where_clause = " AND ".join(filters)

    print("\n1. Finding unmatched banking transactions...")
    cur.execute(f"""
        SELECT 
            bt.transaction_id,
            bt.transaction_date,
            bt.description,
            bt.debit_amount,
            bt.account_number
        FROM banking_transactions bt
        WHERE {where_clause}
        AND NOT EXISTS (
            SELECT 1 FROM banking_receipt_matching_ledger bm
            WHERE bm.banking_transaction_id = bt.transaction_id
        )
        ORDER BY bt.transaction_date, bt.transaction_id
    """, params)
    
    unmatched_banking = cur.fetchall()
    print(f"   Found {len(unmatched_banking)} unmatched banking debits")
    
    if not unmatched_banking:
        print("\n   No unmatched transactions to process!")
        cur.close()
        conn.close()
        return
    
    # Process each transaction
    print("\n2. Analyzing banking transactions...")
    
    # Get all existing source hashes in one query for speed
    cur.execute("SELECT source_hash FROM receipts WHERE source_hash IS NOT NULL")
    existing_hashes = {row[0] for row in cur.fetchall()}
    print(f"   Found {len(existing_hashes)} existing receipt hashes")
    
    created_receipts = []
    duplicate_hashes = 0
    error_count = 0
    
    print("\n3. Preparing new receipts...")
    for i, bt in enumerate(unmatched_banking):
        if (i + 1) % 100 == 0:
            print(f"   Processing {i+1}/{len(unmatched_banking)}...")
        
        tx_id, tx_date, description, debit_amount, account_number = bt
        
        # Generate hash for deduplication
        source_hash = generate_hash(tx_date, description, float(debit_amount))
        
        # Check if receipt with this hash already exists
        if source_hash in existing_hashes:
            duplicate_hashes += 1
            continue
        
        # Extract details
        vendor = extract_vendor_from_description(description)
        category = categorize_transaction(description)
        gst_amount, net_amount = calculate_gst(float(debit_amount))
        account_name = 'CIBC' if account_number == '0228362' else 'Scotia'
        
        receipt_data = {
            'transaction_id': tx_id,
            'date': tx_date,
            'vendor': vendor,
            'gross_amount': float(debit_amount),
            'gst_amount': gst_amount,
            'net_amount': net_amount,
            'category': category,
            'description': f"Auto-generated from {account_name} banking: {description}",
            'source_hash': source_hash,
            'account_number': account_number
        }
        
        created_receipts.append(receipt_data)
    
    print(f"   Prepared {len(created_receipts)} new receipts")
    print(f"   Skipped {duplicate_hashes} duplicates (hash collision)")
    
    # Display sample
    if created_receipts:
        print("\n4. Sample receipts (first 20):")
        print(f"   {'Date':>12} {'Vendor':>30} {'Amount':>12} {'Category':>20} {'Account':>10}")
        print(f"   {'-'*12} {'-'*30} {'-'*12} {'-'*20} {'-'*10}")
        
        for receipt in created_receipts[:20]:
            vendor_display = receipt['vendor'][:30]
            account_display = 'CIBC' if receipt['account_number'] == '0228362' else 'Scotia'
            print(f"   {str(receipt['date']):12} {vendor_display:30} "
                  f"${receipt['gross_amount']:10.2f} {receipt['category']:20} {account_display:10}")
    
    # Category breakdown
    print("\n5. Category breakdown:")
    category_totals = {}
    for receipt in created_receipts:
        cat = receipt['category']
        category_totals[cat] = category_totals.get(cat, 0) + receipt['gross_amount']
    
    for cat, total in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
        count = sum(1 for r in created_receipts if r['category'] == cat)
        print(f"   {cat:30} {count:5} receipts  ${total:12,.2f}")
    
    # Write to database
    if args.write and created_receipts:
        print(f"\n6. Writing {len(created_receipts)} receipts to database...")
        
        created_count = 0
        linked_count = 0
        skipped_existing = 0
        
        for receipt in created_receipts:
            try:
                # Double-check if receipt with this hash already exists (might have been created in a previous partial run)
                cur.execute("SELECT receipt_id FROM receipts WHERE source_hash = %s", (receipt['source_hash'],))
                existing_receipt = cur.fetchone()
                
                if existing_receipt:
                    # Receipt already exists, just create the link if it doesn't exist
                    receipt_id = existing_receipt[0]
                    skipped_existing += 1
                    
                    # Check if link already exists
                    cur.execute("""
                        SELECT 1 FROM banking_receipt_matching_ledger 
                        WHERE banking_transaction_id = %s AND receipt_id = %s
                    """, (receipt['transaction_id'], receipt_id))
                    
                    if not cur.fetchone():
                        # Create link
                        cur.execute("""
                            INSERT INTO banking_receipt_matching_ledger (
                                banking_transaction_id,
                                receipt_id,
                                match_date,
                                match_type,
                                match_status,
                                match_confidence,
                                notes,
                                created_by
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            receipt['transaction_id'],
                            receipt_id,
                            datetime.now().date(),
                            'auto_generated',
                            'matched',
                            '100',
                            'Auto-linked existing receipt to banking transaction',
                            'auto_create_receipts_from_all_banking.py'
                        ))
                        linked_count += 1
                    continue
                
                # Create receipt with created_from_banking flag
                cur.execute("""
                    INSERT INTO receipts (
                        receipt_date,
                        vendor_name,
                        gross_amount,
                        gst_amount,
                        net_amount,
                        category,
                        description,
                        business_personal,
                        source_hash,
                        created_from_banking,
                        mapped_bank_account_id,
                        created_at
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, 'Business', %s, TRUE, 
                        CASE WHEN %s = '0228362' THEN 1 ELSE 2 END,
                        CURRENT_TIMESTAMP
                    )
                    RETURNING receipt_id
                """, (
                    receipt['date'],
                    receipt['vendor'],
                    receipt['gross_amount'],
                    receipt['gst_amount'],
                    receipt['net_amount'],
                    receipt['category'],
                    receipt['description'],
                    receipt['source_hash'],
                    receipt['account_number']
                ))
                
                receipt_id = cur.fetchone()[0]
                created_count += 1
                
                # Create link in junction table
                cur.execute("""
                    INSERT INTO banking_receipt_matching_ledger (
                        receipt_id,
                        banking_transaction_id,
                        match_type,
                        match_confidence,
                        match_status,
                        match_date
                    ) VALUES (%s, %s, 'auto_generated', '100', 'matched', CURRENT_TIMESTAMP)
                """, (receipt_id, receipt['transaction_id']))
                
                linked_count += 1
                
            except Exception as e:
                error_count += 1
                if error_count <= 5:
                    print(f"   Error creating receipt: {e}")
        
        conn.commit()
        print(f"   Created {created_count} receipts")
        print(f"   Skipped {skipped_existing} receipts (already exist)")
        print(f"   Linked {linked_count} to banking transactions")
        if error_count > 0:
            print(f"   Errors: {error_count}")
    
    # Summary
    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Total unmatched banking debits: {len(unmatched_banking)}")
    print(f"Receipts prepared: {len(created_receipts)}")
    print(f"Duplicate hashes skipped: {duplicate_hashes}")
    
    total_amount = sum(r['gross_amount'] for r in created_receipts)
    print(f"Total amount: ${total_amount:,.2f}")

    # Always produce an Excel summary for review (even in dry run) so reviewers can see the bright-yellow rows.
    export_receipts_to_excel(created_receipts, args.excel_output)
    
    if args.write:
        print(f"\nReceipts created: {created_count}")
        print(f"Receipts skipped (already existed): {skipped_existing if 'skipped_existing' in locals() else 0}")
        print(f"Banking links created: {linked_count}")
        print("\nAll created receipts are flagged with:")
        print("  - created_from_banking = TRUE")
        print("  - mapped_bank_account_id = 1 (CIBC) or 2 (Scotia)")
        print("  - match_type = 'auto_generated' in junction table")
    else:
        print("\nDRY RUN - No changes made")
        print("Run with --write to create receipts")
    
    print("\n" + "="*80 + "\n")
    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()
