#!/usr/bin/env python3
"""
Export & Reporting Utilities for Arrow Limousine Dashboard System
Handles CSV, Excel, and PDF export for all dashboards
"""

import csv
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import Qt


class ExportManager:
    """Manage export of dashboard data to various formats"""
    
    @staticmethod
    def export_table_to_csv(parent: QWidget, table_widget, default_filename: str = "export.csv") -> bool:
        """Export QTableWidget data to CSV"""
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Export to CSV",
            f"{default_filename}",
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return False
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write headers
                headers = []
                for col in range(table_widget.columnCount()):
                    headers.append(table_widget.horizontalHeaderItem(col).text() if table_widget.horizontalHeaderItem(col) else "")
                writer.writerow(headers)
                
                # Write data rows
                for row in range(table_widget.rowCount()):
                    row_data = []
                    for col in range(table_widget.columnCount()):
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            
            QMessageBox.information(parent, "Success", f"Data exported to:\n{file_path}")
            return True
        except Exception as e:
            QMessageBox.critical(parent, "Export Failed", f"Could not export data:\n{str(e)}")
            return False
    
    @staticmethod
    def export_table_to_excel(parent: QWidget, table_widget, default_filename: str = "export.xlsx") -> bool:
        """Export QTableWidget data to Excel (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(
                parent,
                "Missing Dependency",
                "Excel export requires openpyxl.\n\nInstall with: pip install openpyxl"
            )
            return False
        
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Export to Excel",
            f"{default_filename}",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return False
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Write headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in range(table_widget.columnCount()):
                cell = ws.cell(row=1, column=col + 1)
                header_text = table_widget.horizontalHeaderItem(col).text() if table_widget.horizontalHeaderItem(col) else ""
                cell.value = header_text
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data rows
            for row in range(table_widget.rowCount()):
                for col in range(table_widget.columnCount()):
                    cell = ws.cell(row=row + 2, column=col + 1)
                    item = table_widget.item(row, col)
                    cell.value = item.text() if item else ""
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-fit columns
            for col in range(table_widget.columnCount()):
                max_width = 20
                for row in range(table_widget.rowCount() + 1):
                    cell = ws.cell(row=row + 1, column=col + 1)
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col + 1)].width = min(max_width + 2, 50)
            
            wb.save(file_path)
            QMessageBox.information(parent, "Success", f"Data exported to:\n{file_path}")
            return True
        except Exception as e:
            QMessageBox.critical(parent, "Export Failed", f"Could not export to Excel:\n{str(e)}")
            return False
    
    @staticmethod
    def export_to_pdf(parent: QWidget, table_widget, title: str = "Report", default_filename: str = "export.pdf") -> bool:
        """Export QTableWidget data to PDF (requires reportlab)"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
        except ImportError:
            QMessageBox.warning(
                parent,
                "Missing Dependency",
                "PDF export requires reportlab.\n\nInstall with: pip install reportlab"
            )
            return False
        
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Export to PDF",
            f"{default_filename}",
            "PDF Files (*.pdf);;All Files (*)"
        )
        
        if not file_path:
            return False
        
        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(title, styles['Heading1']))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Spacer(1, 0.2 * inch))
            
            # Prepare table data
            table_data = []
            
            # Headers
            headers = []
            for col in range(table_widget.columnCount()):
                header_text = table_widget.horizontalHeaderItem(col).text() if table_widget.horizontalHeaderItem(col) else ""
                headers.append(header_text)
            table_data.append(headers)
            
            # Data rows
            for row in range(table_widget.rowCount()):
                row_data = []
                for col in range(table_widget.columnCount()):
                    item = table_widget.item(row, col)
                    row_data.append(item.text() if item else "")
                table_data.append(row_data)
            
            # Create table
            if table_data:
                tbl = Table(table_data)
                tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                story.append(tbl)
            
            doc.build(story)
            QMessageBox.information(parent, "Success", f"PDF exported to:\n{file_path}")
            return True
        except Exception as e:
            QMessageBox.critical(parent, "Export Failed", f"Could not export to PDF:\n{str(e)}")
            return False


class DashboardPrintTemplate:
    """Generate printable/PDF templates for dashboard reports"""
    
    @staticmethod
    def generate_fleet_report(db) -> dict:
        """Generate comprehensive fleet management report"""
        cur = db.get_cursor()
        cur.execute("""
            SELECT v.vehicle_number, v.make, v.model, v.year,
                   COALESCE(SUM(CASE WHEN r.description ILIKE '%fuel%' THEN r.gross_amount ELSE 0 END), 0) as fuel,
                   COALESCE(SUM(CASE WHEN r.description ILIKE '%maint%' THEN r.gross_amount ELSE 0 END), 0) as maint,
                   COALESCE(SUM(CASE WHEN r.description ILIKE '%insur%' THEN r.gross_amount ELSE 0 END), 0) as insur,
                   0 as total_miles, v.purchase_price, COALESCE(v.purchase_price, 0) as current_value
            FROM vehicles v
            LEFT JOIN receipts r ON v.vehicle_id = r.vehicle_id
            GROUP BY v.vehicle_id, v.vehicle_number, v.make, v.model, v.year, v.purchase_price
            ORDER BY v.vehicle_number
        """)
        
        rows = cur.fetchall()
        return {
            "title": "Fleet Management Report",
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "vehicles": rows,
            "total_vehicles": len(rows)
        }
    
    @staticmethod
    def generate_payroll_report(db, year: int = None) -> dict:
        """Generate comprehensive payroll report"""
        if year is None:
            year = datetime.now().year
        
        cur = db.get_cursor()
        cur.execute("""
            SELECT e.full_name,
                   SUM(dp.gross_pay) as gross,
                   0 as cpp,
                   0 as ei,
                   0 as tax,
                   SUM(dp.total_deductions) as deductions,
                   SUM(dp.net_pay) as net
            FROM employees e
            LEFT JOIN driver_payroll dp ON e.employee_id = dp.employee_id
            WHERE EXTRACT(YEAR FROM dp.pay_date) = %s
            GROUP BY e.employee_id, e.full_name
            ORDER BY gross DESC
        """, (year,))
        
        rows = cur.fetchall()
        return {
            "title": f"Payroll Report - {year}",
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "employees": rows,
            "year": year
        }
