# -*- coding: utf-8 -*-
from pathlib import Path
import csv
import re
import openpyxl

DOCS = Path(r"L:\limo\docs")
XLSX = DOCS / "2012_qbb_recon_allcols_20260417.xlsx"
BATCH_RE = re.compile(r"^2012_qbb_recon_batch_(\d{3})_20260417\.csv$")
BATCH_SIZE = 200


def key_for_row(date_s, ref, payee, amount_s):
    return (
        (date_s or "").strip(),
        (ref or "").strip().lower(),
        (payee or "").strip().lower(),
        (amount_s or "").strip(),
    )

existing = []
for p in DOCS.glob("2012_qbb_recon_batch_*_20260417.csv"):
    m = BATCH_RE.match(p.name)
    if m:
        existing.append((int(m.group(1)), p))
existing.sort(key=lambda x: x[0])

next_idx = (existing[-1][0] + 1) if existing else 1
out_csv = DOCS / f"2012_qbb_recon_batch_{next_idx:03d}_20260417.csv"

seen = set()
for _, p in existing:
    try:
        with p.open("r", newline="", encoding="utf-8") as f:
            rd = csv.DictReader(f)
            for r in rd:
                seen.add(key_for_row(r.get("date", ""), r.get("ref", ""), r.get("payee", ""), str(r.get("amount", ""))))
    except Exception:
        pass

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["QBB_CIBC_2012"] if "QBB_CIBC_2012" in wb.sheetnames else wb.active
headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

required = ["date", "ref", "payee", "amount"]
missing = [c for c in required if c not in idx]
if missing:
    raise RuntimeError(f"Missing columns in workbook: {missing}. Found: {headers}")

batch_rows = []
for rownum, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    date_v = row[idx["date"]]
    ref_v = row[idx["ref"]]
    payee_v = row[idx["payee"]]
    amt_v = row[idx["amount"]]

    date_s = str(date_v).strip() if date_v is not None else ""
    ref_s = str(ref_v).strip() if ref_v is not None else ""
    payee_s = str(payee_v).strip() if payee_v is not None else ""
    amt_s = "" if amt_v is None else str(round(float(amt_v), 2))

    if not date_s and not ref_s and not payee_s and not amt_s:
        continue

    k = key_for_row(date_s, ref_s, payee_s, amt_s)
    if k in seen:
        continue

    batch_rows.append({
        "source_row": rownum,
        "date": date_s,
        "ref": ref_s,
        "payee": payee_s,
        "amount": amt_s,
    })
    seen.add(k)

    if len(batch_rows) >= BATCH_SIZE:
        break

with out_csv.open("w", newline="", encoding="utf-8") as f:
    wr = csv.DictWriter(f, fieldnames=["source_row", "date", "ref", "payee", "amount"])
    wr.writeheader()
    wr.writerows(batch_rows)

print(f"OUT={out_csv}")
print(f"ROWS={len(batch_rows)}")
