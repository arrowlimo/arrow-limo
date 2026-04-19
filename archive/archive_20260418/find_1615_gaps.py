"""
Gap analysis for CIBC 74-61615 using the VERIFIED xlsx running balance column.
Finds transactions in the Excel that are MISSING from banking_transactions (bank_id=4).
Also shows any DB rows NOT in Excel (extras).
"""

import openpyxl
import psycopg2
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

XLSX_PATH = r"L:\CIBC_7461615_2012_2017_VERIFIED.xlsx"
DB_CONN = "host=localhost port=5432 dbname=almsdata user=postgres password=ArrowLimousine"
ACCT_NUM = '1615'
YEAR_START = 2012
YEAR_END = 2014


def parse_date(val):
    """Parse a date value from Excel — handles datetime objects, date objects, and multiple string formats."""
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y",
                    "%m/%d/%y", "%d/%m/%y", "%B %d, %Y", "%b %d, %Y",
                    "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        return None
    if isinstance(val, (int, float)):
        # Excel serial date
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(val).date()
        except Exception:
            return None
    return None


def parse_amount(val):
    """Parse amount to Decimal, return None if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return Decimal(str(round(val, 2)))
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("$", "")
        if not val or val in ("-", ""):
            return None
        try:
            return Decimal(val)
        except InvalidOperation:
            return None
    return None


def load_excel():
    print(f"Loading {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title!r}, rows={ws.max_row}, cols={ws.max_column}")

    # Print first 10 rows raw to identify column layout
    print("\n--- First 10 rows (raw) ---")
    rows_raw = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 10:
            print(f"  Row {i+1}: {row}")
        rows_raw.append(row)
    print("--- Last 5 rows (raw) ---")
    all_rows = list(ws.iter_rows(values_only=True))
    for row in all_rows[-5:]:
        print(f"  {row}")

    return all_rows


def detect_columns(rows):
    """Try to auto-detect which column index is date, debit, credit, balance."""
    for i, row in enumerate(rows[:10]):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        # Check if any cell CONTAINS a keyword (handles underscored names like transaction_date)
        if any(any(k in cell for k in ["date", "debit", "credit", "balance", "withdrawal", "deposit"])
               for cell in row_lower):
            print(f"\nHeader row detected at row {i+1}: {row}")
            col_map = {}
            for j, name in enumerate(row_lower):
                if "date" in name:
                    col_map["date"] = j
                elif "withdraw" in name or "debit" in name:
                    col_map["debit"] = j
                elif "deposit" in name or "credit" in name:
                    col_map["credit"] = j
                elif "balance" in name or "running" in name:
                    col_map["balance"] = j
            return i + 1, col_map  # data starts after header
    print("\nNo header found — will try to auto-detect from data")
    return 0, {}


def load_db_transactions():
    print(f"\nLoading DB transactions for account_number={ACCT_NUM}, {YEAR_START}-{YEAR_END} ...")
    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()
    cur.execute("""
        SELECT transaction_date, debit_amount, credit_amount, description, transaction_id
        FROM banking_transactions
        WHERE account_number = %s
          AND transaction_date BETWEEN %s AND %s
        ORDER BY transaction_date, transaction_id
    """, (ACCT_NUM, f"{YEAR_START}-01-01", f"{YEAR_END}-12-31"))
    rows = cur.fetchall()
    conn.close()
    print(f"  DB rows: {len(rows)}")
    return rows


def main():
    all_rows = load_excel()
    header_skip, col_map = detect_columns(all_rows)

    print(f"\nColumn map: {col_map}")

    # If auto-detect failed, print all unique column patterns from first 20 data rows
    if len(col_map) < 3:
        print("\nCould not auto-detect columns. Printing rows 1–20 for manual inspection:")
        for i, row in enumerate(all_rows[:20]):
            print(f"  [{i}] {row}")
        return

    date_col = col_map.get("date")
    debit_col = col_map.get("debit")
    credit_col = col_map.get("credit")
    balance_col = col_map.get("balance")

    # Parse Excel transactions
    xl_txns = []
    date_format_issues = []
    desc_col = col_map.get("desc")
    for i, row in enumerate(all_rows[header_skip:], start=header_skip + 1):
        raw_date = row[date_col] if date_col is not None else None
        d = parse_date(raw_date)
        debit = parse_amount(row[debit_col]) if debit_col is not None else None
        credit = parse_amount(row[credit_col]) if credit_col is not None else None
        balance = parse_amount(row[balance_col]) if balance_col is not None else None
        # Grab description from any remaining column not mapped
        xl_desc = ""
        for ci, cell in enumerate(row):
            if ci not in (date_col, debit_col, credit_col, balance_col) and cell is not None:
                xl_desc = str(cell)
                break

        if raw_date is not None and d is None:
            date_format_issues.append((i, raw_date))

        if d is None:
            continue
        if d.year < YEAR_START or d.year > YEAR_END:
            continue

        xl_txns.append({
            "row": i,
            "date": d,
            "debit": debit,
            "credit": credit,
            "balance": balance,
            "desc": xl_desc,
            "raw_date": raw_date,
        })

    print(f"\nExcel rows in {YEAR_START}-{YEAR_END}: {len(xl_txns)}")
    if date_format_issues:
        print(f"Date parse failures ({len(date_format_issues)}):")
        for r, v in date_format_issues[:20]:
            print(f"  Row {r}: {v!r} (type={type(v).__name__})")

    # Load DB
    db_rows = load_db_transactions()

    # Build a lookup from DB: (date, amount) -> list of txn_ids
    from collections import defaultdict
    db_lookup = defaultdict(list)
    for txn_date, debit, credit, desc, tid in db_rows:
        amt = debit if (debit and debit > 0) else -(credit if credit else Decimal("0"))
        db_lookup[(txn_date, abs(debit or Decimal("0")), abs(credit or Decimal("0")))].append(tid)

    # Compare Excel rows to DB
    print("\n--- Gap Analysis (MISSING only) ---")

    missing = []
    matched = 0
    for t in xl_txns:
        d = t["date"]
        debit = t["debit"] or Decimal("0")
        credit = t["credit"] or Decimal("0")

        key = (d, debit, credit)
        if db_lookup.get(key):
            db_lookup[key].pop(0)
            matched += 1
        else:
            missing.append(t)

    print(f"\nSummary: {len(xl_txns)} Excel rows | {matched} matched in DB | {len(missing)} MISSING from DB")

    if missing:
        print(f"\n{'Date':<12} {'Dir':<7} {'Amount':>10}  Description")
        print("-" * 70)
        for t in missing:
            d = t["date"]
            direction = "CREDIT" if (t["credit"] and t["credit"] > 0) else "DEBIT"
            amt = t["credit"] if direction == "CREDIT" else t["debit"]
            print(f"  {d.strftime('%Y-%m-%d')}  {direction}  ${float(amt or 0):<10.2f}  {t.get('desc','')}")

    # Show DB rows not in Excel (potential duplicates/extras)
    extra = [(d, db, cr, desc, tid) for d, db, cr, desc, tid in db_rows]
    remaining_db_keys = {k: v for k, v in db_lookup.items() if v}
    if remaining_db_keys:
        print(f"\n=== DB ROWS NOT FOUND IN EXCEL ({sum(len(v) for v in remaining_db_keys.values())}) ===")
        # Find the actual DB rows
        for (d, deb, cre), tids in sorted(remaining_db_keys.items()):
            for tid in tids:
                # find matching row
                row_match = [(rd, rdb, rcr, rdesc) for rd, rdb, rcr, rdesc, rtid in db_rows if rtid == tid]
                if row_match:
                    rd, rdb, rcr, rdesc = row_match[0]
                    print(f"  [{tid}] {rd}  debit={rdb}  credit={rcr}  {rdesc[:60]}")


if __name__ == "__main__":
    main()
