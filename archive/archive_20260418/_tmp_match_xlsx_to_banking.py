# -*- coding: utf-8 -*-
import re
import csv
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl
import psycopg2

IN_XLSX = Path(r"L:\limo\docs\2012_qbb_recon_allcols_20260417.xlsx")
OUT_CSV = Path(r"L:\limo\docs\2012_qbb_xlsx_vs_banking_match_20260417.csv")
ACCOUNTS = ['1615', '0228362', '903990106011']
DROP_TOKENS = r"\b(chq|cheque|check|dd|wd|wld|tsf|emt|auto|nsf|return|bill|pmt|deposit|journal|general)\b"


def norm_desc(s: str) -> str:
    s = (s or '').lower().strip()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(DROP_TOKENS, " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_date_mmddyyyy(s):
    try:
        return datetime.strptime(str(s), "%m/%d/%Y").date()
    except Exception:
        return None


def parse_amount(v):
    if v is None or str(v).strip() == '':
        return None
    try:
        return round(float(v), 2)
    except Exception:
        return None


wb = openpyxl.load_workbook(IN_XLSX, data_only=True)
ws = wb['QBB_CIBC_2012'] if 'QBB_CIBC_2012' in wb.sheetnames else wb.active
headers = [str(c.value).strip().lower() if c.value is not None else '' for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

required = ['date', 'ref', 'payee', 'cleared', 'amount']
missing = [k for k in required if k not in idx]
if missing:
    raise RuntimeError(f"Missing required columns in XLSX: {missing}")

src = []
null_amount = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    d = parse_date_mmddyyyy(row[idx['date']])
    if d is None:
        continue
    amt = parse_amount(row[idx['amount']])
    if amt is None:
        null_amount += 1
    ref = str(row[idx['ref']] or '').strip()
    payee = str(row[idx['payee']] or '').strip()
    cleared = str(row[idx['cleared']] or '').strip()
    desc = payee if payee else ref
    chk = ref if re.match(r'^\d+$', ref) else ''
    n = norm_desc(desc + ' ' + chk)
    src.append({'date': d, 'amt': amt, 'ref': ref, 'payee': payee, 'cleared': cleared, 'norm': n})

conn = psycopg2.connect(host='localhost', port=5432, dbname='almsdata', user='postgres', password='ArrowLimousine')
cur = conn.cursor()

bank_rows = {}
for acct in ACCOUNTS:
    cur.execute(
        """
        SELECT transaction_id,
               transaction_date::date,
               COALESCE(description,''),
               COALESCE(check_number::text,''),
               COALESCE(debit_amount,0)::numeric AS debit,
               COALESCE(credit_amount,0)::numeric AS credit
        FROM banking_transactions
        WHERE account_number = %s
          AND transaction_date >= DATE '2012-01-01'
          AND transaction_date < DATE '2013-01-01'
        ORDER BY transaction_date, transaction_id
        """,
        (acct,),
    )
    rows = []
    for tid, dt, desc, chk, debit, credit in cur.fetchall():
        amt = round(float(credit) - float(debit), 2)
        rows.append({'id': tid, 'date': dt, 'desc': desc, 'chk': chk, 'amt': amt, 'norm': norm_desc(desc + ' ' + (chk or ''))})
    bank_rows[acct] = rows

cur.close()
conn.close()

indices = {}
for acct, rows in bank_rows.items():
    idx_exact_date = defaultdict(list)
    idx_exact_norm = defaultdict(list)
    idx_amount = defaultdict(list)
    for r in rows:
        idx_exact_date[(r['date'], r['amt'], r['norm'])].append(r)
        idx_exact_norm[(r['amt'], r['norm'])].append(r)
        idx_amount[r['amt']].append(r)
    indices[acct] = (idx_exact_date, idx_exact_norm, idx_amount)

PASS_RANK = {'pass1_exact_date_amount_norm': 0, 'pass2_amount_norm': 1, 'pass3_amount_only': 2, 'no_match': 9}


def best_in_account(s, acct):
    idx_exact_date, idx_exact_norm, idx_amount = indices[acct]
    if s['amt'] is None:
        return None, None, 'no_match'
    p1 = idx_exact_date.get((s['date'], s['amt'], s['norm']), []) if s['norm'] else []
    if p1:
        p1.sort(key=lambda x: x['id'])
        return p1[0], 0, 'pass1_exact_date_amount_norm'
    p2 = idx_exact_norm.get((s['amt'], s['norm']), []) if s['norm'] else []
    if p2:
        p2.sort(key=lambda x: (abs((x['date'] - s['date']).days), x['id']))
        b = p2[0]
        return b, abs((b['date'] - s['date']).days), 'pass2_amount_norm'
    p3 = idx_amount.get(s['amt'], [])
    if p3:
        p3.sort(key=lambda x: (abs((x['date'] - s['date']).days), x['id']))
        b = p3[0]
        return b, abs((b['date'] - s['date']).days), 'pass3_amount_only'
    return None, None, 'no_match'

results = []
summary = defaultdict(int)
summary_cleared = defaultdict(int)
summary_blank = defaultdict(int)

for s in src:
    best = None
    for acct in ACCOUNTS:
        row, lag, p = best_in_account(s, acct)
        cand = {'acct': acct, 'row': row, 'lag': lag, 'pass': p}
        if row is None:
            continue
        if best is None:
            best = cand
            continue
        if PASS_RANK[cand['pass']] < PASS_RANK[best['pass']]:
            best = cand
        elif PASS_RANK[cand['pass']] == PASS_RANK[best['pass']] and (cand['lag'] or 999999) < (best['lag'] or 999999):
            best = cand

    if best is None:
        mt = 'no_match'
        rec = {
            'pdf_date': s['date'].strftime('%m/%d/%Y'), 'pdf_ref': s['ref'], 'pdf_payee': s['payee'],
            'pdf_cleared': s['cleared'], 'pdf_amount': s['amt'], 'pdf_norm': s['norm'],
            'match_type': mt, 'match_acct': '', 'bank_id': '', 'bank_date': '',
            'bank_desc': '', 'bank_chk': '', 'bank_amount': '', 'day_delta': ''
        }
    else:
        mt = best['pass']
        b = best['row']
        rec = {
            'pdf_date': s['date'].strftime('%m/%d/%Y'), 'pdf_ref': s['ref'], 'pdf_payee': s['payee'],
            'pdf_cleared': s['cleared'], 'pdf_amount': s['amt'], 'pdf_norm': s['norm'],
            'match_type': mt, 'match_acct': best['acct'], 'bank_id': b['id'],
            'bank_date': b['date'].strftime('%m/%d/%Y'), 'bank_desc': b['desc'], 'bank_chk': b['chk'],
            'bank_amount': b['amt'], 'day_delta': best['lag']
        }
    results.append(rec)
    summary[mt] += 1
    if s['cleared'] == 'X':
        summary_cleared[mt] += 1
    else:
        summary_blank[mt] += 1

OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
fields = [
    'pdf_date', 'pdf_ref', 'pdf_payee', 'pdf_cleared', 'pdf_amount', 'pdf_norm',
    'match_type', 'match_acct', 'bank_id', 'bank_date', 'bank_desc', 'bank_chk',
    'bank_amount', 'day_delta'
]
with OUT_CSV.open('w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=fields)
    w.writeheader()
    w.writerows(results)


def print_block(title, dct, total):
    print(title)
    for k in ['pass1_exact_date_amount_norm', 'pass2_amount_norm', 'pass3_amount_only', 'no_match']:
        c = dct.get(k, 0)
        p = (c / total * 100) if total else 0.0
        print(f"  {k:<30} {c:>5} ({p:5.1f}%)")


n_total = len(src)
n_x = sum(1 for r in src if r['cleared'] == 'X')
n_b = n_total - n_x

print(f"XLSX_ROWS={n_total}")
print(f"XLSX_AMOUNT_NULL={null_amount}")
print(f"XLSX_CLEARED_X={n_x}")
print(f"XLSX_CLEARED_BLANK={n_b}")
print(f"OUT={OUT_CSV}")
print_block("SUMMARY_ALL", summary, n_total)
print_block("SUMMARY_CLEARED_X", summary_cleared, n_x)
print_block("SUMMARY_CLEARED_BLANK", summary_blank, n_b)

print("UNMATCHED_CLEARED_X_TOP20")
shown = 0
for r in results:
    if r['pdf_cleared'] == 'X' and r['match_type'] == 'no_match':
        print(f"  {r['pdf_date']} | ref={r['pdf_ref']} | payee={r['pdf_payee']} | amt={r['pdf_amount']}")
        shown += 1
        if shown >= 20:
            break
